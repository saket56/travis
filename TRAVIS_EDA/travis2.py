# -*- coding: utf-8 -*-
"""
Created on Wed Oct 26 07:02:40 2022

@author: AH97759
"""

import os
import random
import sys
import threading
import time
import tkinter as tk
import warnings
# from tkinter import ttk
from tkinter import *
from tkinter import filedialog as fd
from tkinter import font, messagebox
from turtle import bgcolor, st, width

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill

warnings.filterwarnings("ignore")

import datetime
import difflib
import os
import time
from itertools import chain, combinations

import pandas as pd
import paramiko
import pyodbc
import snowflake.connector
import ttkbootstrap as ttk
import xlwt
from dateutil import parser
# from mysql import connector
from ttkbootstrap.constants import *
from ttkbootstrap.scrolled import ScrolledFrame
from PIL import Image, ImageTk
import base64
import psutil
from decimal import Decimal

class App(ttk.Window,threading.Thread):

    def __init__(self):
        super().__init__()
        threading.Thread.__init__(self)

    
    def setFontSize(self,eventObj=np.NAN):
        fs=int(str(self.selectOperationDropDownvar.get()))
         
         
        style = ttk.Style("superhero")
        style.configure('primary.TButton',font=("Arial",fs))
        style.configure('success.TButton',font=("Arial",fs))
        style.configure('danger.TButton',font=("Arial",fs))
        
    def setOperation(self,eventObj=np.NAN):
        fs= (str(self.selectOperationDropDownvar.get()))
        
        messagebox.showinfo('Operation selected',fs)

    def run(self):
        # self = tk.Tk()
            
            self.title('TRAVIS_EDA')

             
            self.tok                       = ttk.StringVar() 
            self.tok.set('Deloitte Version')
            
            
            # self.geometry('1400x900')
            self.fontSize=10
            style = ttk.Style("superhero")
            style.configure('primary.TButton',font=("Arial",self.fontSize))
            style.configure('success.TButton',font=("Arial",self.fontSize))
            style.configure('danger.TButton',font=("Arial",self.fontSize))
            screen_width = self.winfo_screenwidth()
            screen_height = self.winfo_screenheight() -25
            self.geometry('{w}x{h}+0+0'.format(w=screen_width, h=screen_height))
            self.grid_columnconfigure(0,weight=1)
            # print(screen_width,screen_height)
            # self.configure(bg="#0C7A79")

            
            # self.sbf=ScrolledFrame( self, autohide=False,height=screen_height,width=screen_width)
            # self.sbf.grid(row=1,columnspan=2)

            self.frame01=Frame(self, highlightcolor="yellow" , borderwidth=10, relief=None,padx=1,pady=5,width=screen_width)
            self.frame01.grid(row=1)
            
            
             
             

            self.frame0=Frame(self, highlightcolor="yellow" , borderwidth=10,  padx=1,pady=5,width=screen_width,height=140)
            self.frame0.grid_propagate(0)
            self.frame0.grid_columnconfigure(0, weight=1, uniform='a')
            # self.selected = tk.StringVar()
            # r1 = ttk.Radiobutton(self.frame0, text='Batch Comparison  ', value='Batch Comparison', variable=self.selected,command=self.selectTestMethod).grid(row=1,column=0,sticky='e')
            # r2 = ttk.Radiobutton(self.frame0, text='Individual Comparison  ', value='Individual Comparison', variable=self.selected,command=self.selectTestMethod).grid(row=1,column=1)
            self.__travis_deloitte_dlogo_bstream =  "iVBORw0KGgoAAAANSUhEUgAAAFMAAABSCAIAAACnjH/yAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAAFiUAABYlAUlSJPAAAAXRSURBVHhe7ZnbSxxXHMd3Lrtr4q6ul1rrtTaxiSFog0lVUhoLliItFgq1qLU2Gh8kFTH+A6WYBoJ5UPqUp7xYfMjDakFUWEWFqF3SGqVeEE1RV6Nuuq4xu+u4O9Ofzi8Xm9nZjU7cSXo+CDO/M2dm+Z5zfpdz1BAIBAKBQCAQCAQCgUAgEAgEAkFxKLweAIqiwsPDWZZFWxZBEHie9/l8HMd5vV5sDQUKKDcYDDdu3MjKygI92OQHGCPo43K5HA7H8vLy7OzsxMTE4uKizWZ7/PgxdnqNMJlMVqsVJnN/3L9//9atWxUVFWlpaeIHYYAA8V7VgPI7d+6gjgMAH7ly5UpMTAx+V/0opVzEYrHk5+fjp1WOssoBcPvy8nL8uppRXDngdDqrqqrwB1TLq1AOrKysFBYW4m+8Ami8qo+4uLjGxsakpCS0lYbB6wEICwsrKytLTk5G+wUgaXd1dY2Ojt67d29qamp1dRVeMRqN+Ng/8fHxbre7t7dXzHMAPlAJAVd7S0sLFHmgFoCyJyEh4ezZs3V1dTAi2MM/MzMzmZmZ8CuvpfKmpibsupeMjAyz2Yyd/NPQ0CD2V1b5Yfg5TdMMI+FWk5OT9fX1Q0NDaPuhoKAAfB4N5QhxhIPSFXwBKnm0pcjJyUlNTUVDOUIf2zs6OiD4oSFFVFTUsWPH4AZWvtiiCKFXDhM+MDAAW1e0pThx4gS4DBoKEXrlAGz1PB4PGlKkpKTo9Xo0FEIVysHb5ZVDYn8zldvtdvnzGSh7gjzzCR5VKN/Y2JA/z9FqtYqXMapQDrLl43bADvtAFcp1Op186IYoIB/894EqlEdGRkoWeU959OhRwOPNl0UVyiF0gyejIcXDhw85jkNDIVShPD09HbZxaEixsLCwtbWFhkKoQnleXp688unp6TfQz6Ojo3Nzc9GQApwcSh242d2hUzRcn/2DhGKofeb50CsvLi6GjToaUsB+xmazoaHR8BpIb8J7p0xFP7xbee39qmvHK348fqE4IeKtlyvyQqwctiKXLl2Sr0wHBweXlpbEe8jqeh1T+H3qtz+lXfgmOvOToyc/1md9Gv5ZdfTFxrTTH8WK3YIhlMohsF2/fj07OxttKRwOh8ViEZ0cZLNauuC7xILK2Ih4jtt2uTZdHpfH4+Zo1pdymv2yLv7U+SjxxYAchnK32/18NoYEBprLy8vb2tqKioqw1Q89PT0jIyNoaDQnc6LOfx3H09vbWz5hZzjEwk7gfcKW2xuTKOSXxBpjg1r2ChTDJpOps7MT4jPaL9DX12c2m2HeoFyJiIiALSfM85kzZ8SnMJP+anKn0wlRAMSLpv4o+1V98odfGLbcvCBAPStGe3gXX6cZGANtR8vaUPsDsUWGw1C+b5qbmxsaGp6ul7gUw8WfU6MSOYHXg/KdYLfDHgnhRt3g7fXbTfO4GvwT+tjuj/7+/qtXr+5xE53GEEnxwlOpz2b7CeABwlFDmFanwwb/qFT52NhYbW3t2toa2rvwvG+b80Ey3923Sc8pTVFeTvB5Axf5alQOIa2ysnJ8fBztJ2w6fWs2jmUhNPgRRml4H7P+gOOD2N6oS7nX621tbS0tLb179y42PcemY3v6dxfDsAIlfYCj0zOOZe+0dQNtWVSkHKb68uXL1dXVc3Nz2LQXgRf+tDjmRr1HjmhhyWPrE2ioYnnmD8v63F/rwQRuZZTL767lge1nd3d3TU1NSUnJzZs3IfmL9Tk+3otj2WP+ZWlxktfrWYalII2Jf6yOZintyG8bvb8+gHyHvWUJYnACAVmtvb0dslowB0bQB3C5XPPz8xMTE8PDw1ardXZ29j/BTJ6YZH3e529/kG8Mj96ZOcHLrP7ND3evjFrsns3AHi6igHKoyc6dOwf6AyqHmfR4PCDSbrfDfhumF8BnLwnN0MZYNuadMN0R3cY/7vUVzrW+jc8ODX8rU+WoMavtBxj813L8CQQCgUAgEAgEAoFAIPwf0Wj+BRvddEsN/46HAAAAAElFTkSuQmCC"
            self.__travis_deloitte_dlogo_image = ImageTk.PhotoImage(self.__load_deloitte_dlogo_image())
            logo_img_label  = ttk.Label(self.frame0, image=self.__travis_deloitte_dlogo_image)
            logo_img_label.grid(row=0, column=0, sticky="w", padx=10, pady=10) 

            dqslabel  = ttk.Label(self.frame0, text='Data Quality Studio')
            dqslabel.grid(row=0, column=0, sticky="w", padx=110, pady=10) 



            

        

            # populate theme combobox menu 
            self.__populate_token_header()

            # populate IP Address of the machine currently used for running travis 
            self.__populate_ip_address_header()

            # populate CPU meter
            self.__populate_cpu_count_header()

            # populate memory available meter
            self.__populate_available_memory_header()

            # populate number of concurrent users using the machine 
            self.__populate_logged_usercount_header()
            
            separator = ttk.Separator(self.frame0)
            separator.grid(row=2, column=0, columnspan=20, padx=(10,10), pady=(10, 10), sticky="ew")        
            
            # r2 = ttk.Radiobutton(self.frame0,text='TableStats  ', value='TableStats', variable=self.selected,command=self.selectTestMethod).grid(row=1,column=5)
            # r2 = ttk.Radiobutton(self.frame0, text='Utilities  ', value='Utilities', variable=self.selected,command=self.selectTestMethod).grid(row=1,column=7)
            # r2 = ttk.Radiobutton(self.frame0, text='PdfCompare  ', value='PdfCompare', variable=self.selected,command=self.selectTestMethod).grid(row=1,column=9)
            
            # r2 = ttk.Radiobutton(self.frame0, text='DB Connect  ', value='DB Connect', variable=self.selected,command=self.selectTestMethod).grid(row=1,column=4,padx=(0,0),pady=0)
            # ttk.Radiobutton(self.frame0, text='DDL Validation  ', value='DDL validation', variable=self.selected,command=self.selectTestMethod).grid(row=1,column=6)
            # ttk.Radiobutton(self.frame0, text='Sqoop Batch Execution  ', value='Sqoop batch execution', variable=self.selected,command=self.selectTestMethod).grid(row=1,column=8)
            

            self.theme_label                    = ttk.Label(self.frame0, 
                                                        text="Select an Operation: ")
            self.theme_label.grid(row=0, column=2, padx=(20,10), sticky="e") 

            self.selected = tk.StringVar()
            self.selectOperationDropDown = ttk.Combobox(self.frame0, textvariable=self.selected)
            self.selectOperationDropDown['state'] = 'readonly'
            self.selectOperationDropDown.grid(row=0,column=3,pady=10,padx=10)
            self.selectOperationDropDown['value']=['Batch Comparison','Individual Comparison','DB Connect','DDL Validation',
                                                  'TableStats','Utilities','Sqoop Batch Execution','PdfCompare']
            self.selectOperationDropDown.set('Operation Name')
            self.selectOperationDropDown.bind('<<ComboboxSelected>>', self.selectTestMethod)


            self.selected.set('Batch Comparison')

            self.frame0.grid(row=0)

            # tk.Label(self.frame0, text="Select DB",bg="white",fg="red", font=("Arial", self.fontSize-1)).grid(row=1,column=6,padx=(10,0),pady=0)
            self.frame_database_connect=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10,padx=1,pady=1)
            # self.frame_database_connect=ScrolledFrame( self, autohide=False,height=(screen_height-50),width=(screen_width-50))

            self.selectDatabaseDropDownvar = tk.StringVar()
            self.selectDataBaseDropDown = ttk.Combobox(self.frame_database_connect, textvariable=self.selectDatabaseDropDownvar,font = ('Arial', '10'))
            self.selectDataBaseDropDown['state'] = 'readonly'
            self.selectDataBaseDropDown.grid(row=1,columnspan=2,pady=5,padx=(0,0),ipadx=30)
            self.selectDataBaseDropDown['value']=['Snowflake','Snowflake_Prod','Teradata','Oracle','Hive','SSH','MySQL WorkBench','DB2','AWS RDS','PostgresSQL','MySQL','MS-SQL Server','MongoDB','Amazon Athena','AWS Aurora','DynamoDB']
            self.selectDataBaseDropDown.set('Select Database To Connect')
            self.selectDataBaseDropDown.bind('<<ComboboxSelected>>',self.selectTestMethod)

            
            
            # sf =  ScrolledFrame(self, autohide=True)
            # sf.grid(row=1)
            self.frameIndividual= ScrolledFrame( self.frame01, autohide=False,height=screen_height,width=screen_width, scrollheight=None)
            self.frameBatch=tk.LabelFrame(self.frame01,   borderwidth=10,text='Batch' )
            # self.frameIndividual=tk.LabelFrame(self.frame01,text='Individual Table/File Testing',padx=10,pady=10)
            self.sf_frame=ttk.Labelframe(self.frame_database_connect, text='Snowflake DB Connection')
            self.sf_prod_frame=Frame(self.frame_database_connect, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=20,pady=1)

            self.td_frame=ttk.Labelframe(self.frame_database_connect, text='Teradata Database')
            self.Oracle_frame=Frame(self.frame_database_connect, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=20,pady=1)
            # self.Hive_frame=Frame(self.frame_database_connect, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=20,pady=1)
            self.Hive_frame=ScrolledFrame( self.frame_database_connect, autohide=False,height=screen_height,width=screen_width, scrollheight=None)
            self.SSH_frame=Frame(self.frame_database_connect, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=20,pady=1)
            self.MySQLWorkBench_frame=Frame(self.frame_database_connect, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=20,pady=1)
            self.mssql_frame=Frame(self.frame_database_connect, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=20,pady=1)
            self.PostgresSQL_frame=Frame(self.frame_database_connect, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=20,pady=1)
           
            

            self.frameUtilities=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=2,pady=1)
            self.framePdfCompare=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=2,pady=1)

            self.frame_stats=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10,padx=50,pady=1)
            self.DDL_validation=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10,padx=2,pady=1)
            self.Sqoop_batch_exection=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=2,pady=1)


            # ddl validation

            self.selectDDL_validationDropDownvar = tk.StringVar()
            self.selectDDL_validationDropDown = ttk.Combobox(self.DDL_validation, textvariable=self.selectDDL_validationDropDownvar,font = ('Arial', '10'))
            self.selectDDL_validationDropDown['state'] = 'readonly'
            self.selectDDL_validationDropDown.grid(row=0,columnspan=2,pady=5,padx=(0,0),ipadx=50)
            self.selectDDL_validationDropDown['value']=['Hive-SF DDL check','SF-SF Schema & DDL check','SF-SF DDL check','TD-SF DDL check','MSSQL-SF DDL check','MSSQL-TD DDL check','TD-SF casted col validation','Tokenization Check','SF View Validation']
            
            self.selectDDL_validationDropDown.set('Select Database Pair For DDL check')
            self.selectDDL_validationDropDown.bind('<<ComboboxSelected>>',self.selectDDLPair)
             

            # table-stats

            self.selectframe_statsDropDownvar = tk.StringVar()
            self.selectframe_statsDropDown = ttk.Combobox(self.frame_stats, textvariable=self.selectframe_statsDropDownvar,font = ('Arial', '10'))
            self.selectframe_statsDropDown['state'] = 'readonly'
            self.selectframe_statsDropDown.grid(row=1,columnspan=2,pady=5,padx=(0,0),ipadx=50)
            self.selectframe_statsDropDown['value']=['TD-SF stats','SF-SF stats','MSSQL-SF stats','Hive-SF stats','Hive-SF PostProd Check (DDL, Rowcount,SumOfValues,Rowcount)','TD-SF PostProd Check (DDL, Rowcount,SumOfValues,Rowcount)','MSSQL-SF PostProd Check (DDL, Rowcount,SumOfValues,Rowcount)','SF-SF PostProd Check (DDL, Rowcount,SumOfValues,Rowcount)']
            
            self.selectframe_statsDropDown.set('Select Database Pair For Table-Stats')
            self.selectframe_statsDropDown.bind('<<ComboboxSelected>>',self.selectTableStatsPair)
             

            
            

            # 
            
            #selecting which database
            self.initiateUi_Batch()
            self.initiateUi_Individual()
            self.initiateUi_SF()
            self.initiateUi_SF_prod()
            self.initiateUi_TD()
            self.initiateUi_Oracle()
            self.initiateUi_Hive()
            self.initiateUi_SSH()
            self.initiateUi_MySQLWorkBench()
            self.initiateUi_PostgresSQL()
            self.initiateUi_mssql()
            self.initiateUi_Stats()
            self.initiateUi_DDL_validation()
            self.initiateUi_Utilities()
            self.initiateUi_PdfCompare()
            self.initiateUi_SqoopBatchExecution()
            
            #self.selectTestDatabase()
            
            
            self.selectTestMethod()

            self.resetButton = ttk.Button(
                self,
                text='Reset All Fields',
                # bg="red",
                # background="white",
                # fg="red",
                command=self.resetEverything,
                # height= 1, 
                width=20, 
                bootstyle='danger'
                # font=("Arial", self.fontSize+5)
            )

            # self.resetButton.grid(row=0,column=1,pady=20,padx=50,sticky='e')


            # 
   
 

    def __populate_token_header(self) -> None:
        
        """ Populate combobox on header frame with theme select """

        

        # create labels 
        self.theme_label                    = ttk.Label(self.frame0, 
                                                        text="Valid Till: ")
        self.theme_label.grid(row=0, column=4, padx=10, sticky="e") 

        # # create theme combo box
        # self.__theme_names=['superhero']
        # self.theme_combo                    = ttk.Combobox(self.frame0, 
        #                                                    values=self.__theme_names)
        

        self.token_entry                       = ttk.Entry(self.frame0,
                                                        textvariable=self.tok,
                                                        justify=CENTER,
                                                        state=DISABLED)
        self.token_entry.grid(row=0, column=5, padx=10, sticky="w")
        # self.theme_combo.current(self.__theme_names.index(self.__style_name.theme.name))
        # self.theme_combo.bind("<<ComboboxSelected>>", self.__change_travis_theme)
        self.token_entry.configure(state="readonly")


    def __populate_ip_address_header(self) -> None:
        """ Populate ip address of this machine on the screen """

        import socket

        # create an entry box with IP Address 
        self.ip_label                       = ttk.Label(self.frame0, 
                                                        text="IP Address: ")
        self.ip_label.grid(row=0, column=6, padx=10, sticky="e")

        # create ip value string 
        ip_address                          = socket.gethostbyname(socket.gethostname())
        self.ip_value                       = ttk.StringVar() 
        self.ip_value.set(ip_address)

        self.ip_entry                       = ttk.Entry(self.frame0,
                                                        textvariable=self.ip_value,
                                                        justify=CENTER,
                                                        state=DISABLED)
        self.ip_entry.grid(row=0, column=7, padx=10, sticky="w")


    def __populate_cpu_count_header(self) -> None:
        """ Populate Number of cores available on this machine"""

        

        # place meter widget for cpu count 
        cpu_count                           = os.cpu_count()
        self.cpu_meter                      = ttk.Meter(self.frame0, 
                                                        bootstyle="default", 
                                                        amounttotal=cpu_count, 
                                                        metersize=100, 
                                                        amountused=cpu_count,
                                                        textright="Logical",
                                                        subtext="processors",
                                                        textfont="-size 10 -weight bold",
                                                        subtextfont="-size 7",
                                                        interactive=False)                
        self.cpu_meter.grid(row=0, column=8, sticky="e", padx=10)


    def __populate_available_memory_header(self) -> None:
        """ Populate memory avaialble for execution on this machine """

        
        
        # place meter widget for memory count 
        memory_count                        = psutil.virtual_memory()
        self.memory_meter                   = ttk.Meter(self.frame0, 
                                                        bootstyle="success", 
                                                        amounttotal=(memory_count.total // (1024 ** 3)), 
                                                        metersize=100, 
                                                        amountused=(memory_count.used // (1024 ** 3)),
                                                        textright="gb",
                                                        subtext=f"memory in use",
                                                        textfont="-size 10 -weight bold",
                                                        subtextfont="-size 7",
                                                        interactive=False)            
        self.memory_meter.grid(row=0, column=9, sticky="e", padx=10)
        self.memory_meter.after(1000, self.__update_memory_meter)

    def __update_memory_meter(self):
        """ update memory meter after every 1 second """
        # logging.info(friday_reusable.get_function_name())

        memory_count                        = psutil.virtual_memory()
        self.memory_meter.configure(amounttotal=(memory_count.total // (1024 ** 3)),
                                    amountused=(memory_count.used // (1024 ** 3)))
        
        if not self.debugger_is_active():
            self.memory_meter.after(1000, self.__update_memory_meter)

    def debugger_is_active(self):
        """ check if running in debugger """

        # logging.info(friday_reusable.get_function_name())
        return hasattr(sys, 'gettrace') and sys.gettrace() is not None


    def __populate_logged_usercount_header(self) -> None:
        """ Populate number of concurrent users working on this machine """

        

        # place meter widget with number of active users 
        user_count                          = len(psutil.users())
        self.users_meter                    = ttk.Meter(self.frame0, 
                                                        bootstyle="danger",
                                                        amounttotal=user_count,
                                                        metersize=100,
                                                        amountused=user_count,
                                                        textright="user(s)",
                                                        subtext="logged in",
                                                        textfont="-size 10 -weight bold",
                                                        subtextfont="-size 7",
                                                        interactive=False)
        self.users_meter.grid(row=0, column=10, sticky="e", padx=10)
        self.users_meter.after(1000, self.__update_user_meter)

    def __update_user_meter(self):
        """ update user count meter after every one second """
        # logging.info(friday_reusable.get_function_name())

        user_count                        = len(psutil.users())
        self.users_meter.configure(amounttotal=user_count,
                                   amountused=user_count)
        
        if not self.debugger_is_active():
            self.users_meter.after(1000, self.__update_user_meter) 

           
    def __load_deloitte_dlogo_image(self) -> Image:
        """ Load deloitte image """
        import io
        travis_dlogo_bytes                    = base64.b64decode(self.__travis_deloitte_dlogo_bstream.encode())
        travis_dlogo_stream                   = io.BytesIO(travis_dlogo_bytes)
        travis_dlogo_pil                      = Image.open(travis_dlogo_stream)

        return travis_dlogo_pil
         
        # Btn.grid(row=10,column=2)
    def initiateUi_SqoopBatchExecution(self):
        tk.Label(self.Sqoop_batch_exection, text=" Download from semicolon separated query file (Preq. SSH connection) ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=5,padx=(100,0),sticky='w')

        tk.Label(self.Sqoop_batch_exection, text="Enter Sqoop Command : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize+3)).grid(row=6,padx=(100,0))
        self.sqoop_command_batch= ttk.Entry(self.Sqoop_batch_exection, font="Arial 11 ")
        self.sqoop_command_batch.grid(row=6,column=1,pady=5,ipadx=200)

        self.file_location_btn_SqoopBatch = ttk.Button(
            self.Sqoop_batch_exection,
            text='Select Result Location ',
            command=self.select_download_location_SqoopBatch,
            width=25, bootstyle='primary' 
            )

        self.file_location_btn_SqoopBatch.grid(row=7,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_SqoopBatch = ttk.Button(
            self.Sqoop_batch_exection,
            text='Select query file',
            command=self.select_query_file_SqoopBatch,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_SqoopBatch.grid(row=8,column=0,padx=(70,0),pady=5)

        
        
        
        self.download_path_SqoopBatch = ttk.Entry(self.Sqoop_batch_exection, font="Arial 11 ")
        self.download_path_SqoopBatch.grid(row=7,column=1,pady=5,ipadx=200)
        

        self.query_SqoopBatch = ttk.Entry(self.Sqoop_batch_exection, font="Arial 11 ")
        self.query_SqoopBatch.grid(row=8,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_SqoopBatch = ttk.Button(
            self.Sqoop_batch_exection,
            text='Start Query Execution',
            command=self.start_download_SqoopBatch,
            width=25, bootstyle='success'
            )

        self.executeQueryFileBtn_SqoopBatch.grid(row=11,column=1,pady=5,padx=5)
        

    def initiateUi_PdfCompare(self):
        #pdf comparison

            tk.Label(self.framePdfCompare, text="PDF Comparison",bg="white",fg="red", font=("Arial", self.fontSize+5)).grid(row=4,padx=(0,0),pady=15)
            # selecting result folder
            self.selectPdfComparisonResultFolderBtn = ttk.Button(
            self.framePdfCompare,
            text='Select Result Folder',
            command=self.select_pdf_comparison_result_folder_location,
            width=20, bootstyle='primary'
            )

            self.selectPdfComparisonResultFolderBtn.grid(row=5,column=0,pady=5,padx=50)

            self.pdfComparisonResultFolder_entry = ttk.Entry(self.framePdfCompare, font="Arial 10 ")
            self.pdfComparisonResultFolder_entry.grid(row=5,column=1,ipadx=220,pady=10,padx=50)

            #selecting files
            self.selectPdfComparisonSourceBtn = ttk.Button(
            self.framePdfCompare,
            text='Select Source File',
            command=self.select_pdf_comparison_Source,
            width=20, bootstyle='primary'
            )

            self.selectPdfComparisonSourceBtn.grid(row=6,column=0,pady=5,padx=50)

            self.pdfComparisonSource_entry = ttk.Entry(self.framePdfCompare, font="Arial 10 ")
            self.pdfComparisonSource_entry.grid(row=6,column=1,ipadx=220,pady=10,padx=50)

            self.selectPdfComparisonTargetBtn = ttk.Button(
            self.framePdfCompare,
            text='Select Target File',
            command=self.select_pdf_comparison_Target,
            width=20, bootstyle='primary'
            )

            self.selectPdfComparisonTargetBtn.grid(row=7,column=0,pady=5,padx=50)

            self.pdfComparisonTarget_entry = ttk.Entry(self.framePdfCompare, font="Arial 10 ")
            self.pdfComparisonTarget_entry.grid(row=7,column=1,ipadx=220,pady=10,padx=50)

            #execute
            self.comparePdfFilesBtn = ttk.Button(
            self.framePdfCompare,
            text='Compare PDF Files',
            command=self.comparePdfFiles,
            width=20, bootstyle='success'
            )

            self.comparePdfFilesBtn.grid(row=8,column=1,pady=5,padx=50)

        

    def initiateUi_Utilities(self):

            tk.Label(self.frameUtilities, text="Excel to Csv Conversion",bg="white",fg="red", font=("Arial", self.fontSize+5)).grid(row=1,padx=(50,0),pady=15)
            self.selectExcelToCsvFolderBtn = ttk.Button(
            self.frameUtilities,
            text='Select Excel Folder',
            command=self.select_batch_Excel_folder_location,
            width=25, bootstyle='primary'
            )

            self.selectExcelToCsvFolderBtn.grid(row=2,column=0,padx=(70,0),pady=5)

            self.batch_ExcelFolder_entry = ttk.Entry(self.frameUtilities, font="Arial 10 ")
            self.batch_ExcelFolder_entry.grid(row=2,column=1,ipadx=220,pady=20,padx=50)

            self.convertExcelToCsvFolderBtn = ttk.Button(
            self.frameUtilities,
            text='Convert Excel to CSV',
            command=self.convert_batch_Excel_to_csv,
            width=30, bootstyle='success'
            )

            self.convertExcelToCsvFolderBtn.grid(row=3,column=1,pady=5,padx=50)     

            
            
              


            # get tok cols from aedl to paste into travis compare   
            tk.Label(self.frameUtilities, text="Tokenized Cols from AEDL SIT/UAT env",bg="white",fg="red", font=("Arial", self.fontSize+5)).grid(row=4,padx=(50,0),pady=15)
            
            self.tok_cols_aedl_template_btn=ttk.Button(
                self.frameUtilities,
                text=u'\u2193'+'Download Template',
                command=lambda: self.download_template('Tok-Cols AEDL',['TableName','AppCd','Toknzd in SIT','Toknzd in PROD','sit/uat']),
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
            self.tok_cols_aedl_template_btn.grid(row=4,column=1,pady=5,padx=5,sticky='e')

            self.file_location_btn_tok_cols_aedl = ttk.Button(
                self.frameUtilities,
                text='Select Result Location ',
                command=lambda: self.select_folder(self.download_path_tok_cols_aedl),
                width=25, bootstyle='primary'
                )

            self.file_location_btn_tok_cols_aedl.grid(row=5,column=0,padx=(70,0),pady=5)
            

            self.download_path_tok_cols_aedl = ttk.Entry(self.frameUtilities, font="Arial 11 ")
            self.download_path_tok_cols_aedl.grid(row=5,column=1,pady=5,ipadx=200)
            
            
            
            #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
            self.queryFile_btn_tok_cols_aedl = ttk.Button(
                self.frameUtilities,
                text='Select Template file',
                command=lambda: self.select_file('Tok-Cols Template',self.query_excel_file_tok_cols_aedl),
                width=25, bootstyle='primary'
                )

            self.queryFile_btn_tok_cols_aedl.grid(row=6,column=0,padx=(70,0),pady=5)

            

            self.query_excel_file_tok_cols_aedl = ttk.Entry(self.frameUtilities, font="Arial 11 ")
            self.query_excel_file_tok_cols_aedl.grid(row=6,column=1,pady=5,ipadx=200)

            self.executeQueryFileBtn_tok_cols_aedl = ttk.Button(
                self.frameUtilities,
                text='Get Tok-Cols From AEDL',
                command=self.get_tok_cols,
                width=30,
                  bootstyle="success"
                )

            self.executeQueryFileBtn_tok_cols_aedl.grid(row=7,column=1,pady=5,padx=5)


            

    def get_tok_cols(self):
        print('Getting tok cols...')
        try:
            import mysql.connector

            hostname_MySQLWorkBench='antm-mysqldb-cluster.cluster-ro-csntho9gpvhy.us-east-1.rds.amazonaws.com'
            username='edl_rds_user'
            password='EDLTemp123'

            cnxn_MySQLWorkBench = mysql.connector.connect(
                                    host=hostname_MySQLWorkBench,
                                    user=username,
                                    passwd=password
                                    )

            mycursor = cnxn_MySQLWorkBench.cursor()

            
            

            result_loc=str(self.download_path_tok_cols_aedl.get()).strip()
            df_loc=str(self.query_excel_file_tok_cols_aedl.get()).strip()
            os.chdir(result_loc)

            df_temp=pd.read_excel(df_loc)
            df=pd.DataFrame(columns=['TableName','Tok Columns From AEDL'])

            for row in (df_temp.itertuples()):
                table=str(row[1]).strip()
                app_cd=str(row[2]).strip()
                tknzd_in_sit=str(row[3]).strip()
                tknzd_in_prod=str(row[4]).strip()
                env=str(row(5)).strip().lower()


                
                
                query="select atrb_nm from {env}_audt_cntrl.edl_tknztn_mtdta where UPPER(tbl_nm) = '{table_aedl}'".format(table_aedl=table.upper(),env=env)

                if(str(app_cd).lower()!='nan'):
                    query +=" and UPPER(aplctn_cd) = '{app_cd}'".format(app_cd=app_cd.upper())

                if(str(tknzd_in_sit).lower()!='nan'):
                    query +=" and UPPER(tknzd_in_test_ind) = '{tnnz_sit}'".format(tnnz_sit=tknzd_in_sit.upper())
                if(str(tknzd_in_prod).lower()!='nan'):
                    query+=" and UPPER(tknzd_in_prodn_ind) = '{tknz_prod}'".format(tknz_prod=tknzd_in_prod.upper())
                

                mycursor.execute(query )
                myresult = mycursor.fetchall()
                tokenized_cols_aedl= [str(x[0]).lower().strip() for x in myresult]
                tokenized_cols_aedl.sort()
                tok_cols=",".join(tokenized_cols_aedl)
            #     print(tok_cols)
                df.loc[len(df.index)]=[table,tok_cols]

            df.to_csv('AEDL_Tokenized_Columns_'+str(time.time())+'.csv',index=False)
            print('Tok cols analysis done.')
            messagebox.showinfo('Done','Tok-Cols from AEDL are written in result folder.')
        except Exception as err:
            print(str(err))
            messagebox.showerror('Error During Tokenization Check',str(err))
        
    
    def comparePdfFiles(self):
        from PyPDF2 import PdfReader
        
        try:
            result_path=str(self.pdfComparisonResultFolder_entry.get()).strip()
            os.chdir(result_path)

            Source_file_pdf=self.pdfComparisonSource_entry.get()
            Release_file_pdf=self.pdfComparisonTarget_entry.get()
            sfn=Source_file_pdf.split("/")[-1].split('.')[0]
            tfn=Release_file_pdf.split("/")[-1].split('.')[0]

            # # creating a pdf reader object
            reader_source = PdfReader(Source_file_pdf)
            reader_target = PdfReader(Release_file_pdf)
            
            # # printing number of pages in pdf file
            
            totalpage_source = int(len(reader_source.pages))
            # # getting a specific page from the pdf file
            totalpage_target = int(len(reader_target.pages))
            
            temp_source_fn = "Source_file.txt"
            temp_target_fn = "Release_file.txt"
            
            for i in range(totalpage_source):
                page = reader_source.pages[i]
                text = str(page.extract_text())
                f= open(temp_source_fn,'a', encoding="utf-8")
                f.write(str(text))
                f.close()
                
            for j in range(totalpage_target):
                page = reader_target.pages[j]
                text = str(page.extract_text())
                f= open(temp_target_fn,'a', encoding="utf-8")
                f.write(str(text))
                f.close()
            
            
            source_read =  open(temp_source_fn,encoding="utf-8").readlines()
            target_read =  open(temp_target_fn,encoding="utf-8").readlines()
            
            difference =  difflib.HtmlDiff().make_file(source_read,target_read,temp_source_fn,temp_target_fn)
            difference_report = open(sfn+"__"+tfn+"__diffreport_"+ str(datetime.datetime.now().timestamp()) +".html" ,'w',encoding="utf-8")
            difference_report.write(difference)
            difference_report.close()
            
            try:
                os.remove(temp_source_fn)
                os.remove(temp_target_fn)
            except: pass
            messagebox.showinfo('Done','Result stored in: '+result_path)

        except Exception as err:
            messagebox.showerror('Error','Error while pdf comparison: '+str(err))


    def initiateUi_DDL_validation(self):
        
         
        # hive-sf datatype
        self.DDL_validation_hive_sf=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=2,pady=1)
        tk.Label(self.DDL_validation_hive_sf, text="Hive-SF datatype comparison : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,padx=(100,0),sticky='w')

        self.hive_sf_datatype_template_btn=ttk.Button(
                self.DDL_validation_hive_sf,
                text=u'\u2193'+'Download Template',
                command=self.download_hive_sf_datatype_comparison_template,
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
        self.hive_sf_datatype_template_btn.grid(row=0,column=1,pady=5,padx=5,sticky='e')

        self.file_location_btn_table_details_datatype_hive_sf = ttk.Button(
            self.DDL_validation_hive_sf,
            text='Select Result Location ',
            command=self.select_download_location_datatype_hive_sf,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_table_details_datatype_hive_sf.grid(row=1,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_table_details_datatype_hive_sf = ttk.Button(
            self.DDL_validation_hive_sf,
            text='Select Template file',
            command=self.select_query_file_table_details_datatype_hive_sf,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_table_details_datatype_hive_sf.grid(row=2,column=0,padx=(70,0),pady=5)

        
        
        
        self.download_path_datatype_hive_sf = ttk.Entry(self.DDL_validation_hive_sf, font="Arial 11 ")
        self.download_path_datatype_hive_sf.grid(row=1,column=1,pady=5,ipadx=200)
        

        self.query_excel_file_datatype_hive_sf = ttk.Entry(self.DDL_validation_hive_sf, font="Arial 11 ")
        self.query_excel_file_datatype_hive_sf.grid(row=2,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_table_details_datatype_hive_sf = ttk.Button(
            self.DDL_validation_hive_sf,
            text='Start Hive-SF Datatype Comparison',
            command=self.datatype_check_hive_sf,
            width=30,
                  bootstyle="success"
            )

        self.executeQueryFileBtn_table_details_datatype_hive_sf.grid(row=3,column=1,pady=5,padx=5)


         # Schema&DDL validation datatype
        self.DDL_validation_ddl_schema_sf_sf=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=2,pady=1)
        tk.Label(self.DDL_validation_ddl_schema_sf_sf, text="Schema & DDL checks : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=4,padx=(100,0),sticky='w')

        self.hive_sf_sf_schema_ddl_template_btn=ttk.Button(
                self.DDL_validation_ddl_schema_sf_sf,
                text=u'\u2193'+'Download Template',
                command=self.download_sf_sf_schema_ddl_comparison_template,
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
        self.hive_sf_sf_schema_ddl_template_btn.grid(row=4,column=1,pady=5,padx=5,sticky='e')

        self.file_location_btn_table_details_sf_sf_schema_ddl = ttk.Button(
            self.DDL_validation_ddl_schema_sf_sf,
            text='Select Result Location ',
            command=self.select_download_location_sf_sf_schema_ddl,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_table_details_sf_sf_schema_ddl.grid(row=5,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_table_details_sf_sf_schema_ddl = ttk.Button(
            self.DDL_validation_ddl_schema_sf_sf,
            text='Select Template file',
            command=self.select_query_file_table_details_sf_sf_schema_ddl,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_table_details_sf_sf_schema_ddl.grid(row=6,column=0,padx=(70,0),pady=5)

        
        
        
        self.download_path_sf_sf_schema_ddl = ttk.Entry(self.DDL_validation_ddl_schema_sf_sf, font="Arial 11 ")
        self.download_path_sf_sf_schema_ddl.grid(row=5,column=1,pady=5,ipadx=200)
        

        self.query_excel_file_sf_sf_schema_ddl = ttk.Entry(self.DDL_validation_ddl_schema_sf_sf, font="Arial 11 ")
        self.query_excel_file_sf_sf_schema_ddl.grid(row=6,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_table_details_sf_sf_schema_ddl = ttk.Button(
            self.DDL_validation_ddl_schema_sf_sf,
            text='Start Schema & DDL Comparison',
            command=self.schemaAndDDLverification_sf,
            width=30,
                  bootstyle="primary"
            )

        self.executeQueryFileBtn_table_details_sf_sf_schema_ddl.grid(row=7,column=1,pady=5,padx=5)


        # td-sf datatype
        self.DDL_validation_ddl_td_sf=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=2,pady=1)
        tk.Label(self.DDL_validation_ddl_td_sf, text="TD-SF datatype comparison : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=8,padx=(100,0),sticky='w')

        self.td_sf_datatype_template_btn=ttk.Button(
                self.DDL_validation_ddl_td_sf,
                text=u'\u2193'+'Download Template',
                command=self.download_td_sf_datatype_comparison_template,
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
        self.td_sf_datatype_template_btn.grid(row=8,column=1,pady=5,padx=5,sticky='e')

        self.file_location_btn_table_details_datatype_td_sf = ttk.Button(
            self.DDL_validation_ddl_td_sf,
            text='Select Result Location ',
            command=self.select_download_location_datatype_td_sf,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_table_details_datatype_td_sf.grid(row=9,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_table_details_datatype_td_sf = ttk.Button(
            self.DDL_validation_ddl_td_sf,
            text='Select Template file',
            command=self.select_query_file_table_details_datatype_td_sf,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_table_details_datatype_td_sf.grid(row=10,column=0,padx=(70,0),pady=5)

        
        
        
        self.download_path_datatype_td_sf = ttk.Entry(self.DDL_validation_ddl_td_sf, font="Arial 11 ")
        self.download_path_datatype_td_sf.grid(row=9,column=1,pady=5,ipadx=200)
        

        self.query_excel_file_datatype_td_sf = ttk.Entry(self.DDL_validation_ddl_td_sf, font="Arial 11 ")
        self.query_excel_file_datatype_td_sf.grid(row=10,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_table_details_datatype_td_sf = ttk.Button(
            self.DDL_validation_ddl_td_sf,
            text='Start TD-SF Datatype Comparison',
            command=self.datatype_check_td_sf,
            width=30,
                  bootstyle="success"
            )

        self.executeQueryFileBtn_table_details_datatype_td_sf.grid(row=11,column=1,pady=5,padx=5)

        # mssql -sf datatype
        self.DDL_validation_ddl_mssql_sf=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=2,pady=1)
        tk.Label(self.DDL_validation_ddl_mssql_sf, text="MS SQL - SF datatype comparison : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=12,padx=(100,0),sticky='w')

        self.mssql_sf_datatype_template_btn=ttk.Button(
                self.DDL_validation_ddl_mssql_sf,
                text=u'\u2193'+'Download Template',
                command=self.download_mssql_sf_datatype_comparison_template,
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
        self.mssql_sf_datatype_template_btn.grid(row=12,column=1,pady=5,padx=5,sticky='e')

        self.file_location_btn_table_details_datatype_mssql_sf = ttk.Button(
            self.DDL_validation_ddl_mssql_sf,
            text='Select Result Location ',
            command=self.select_download_location_datatype_mssql_sf,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_table_details_datatype_mssql_sf.grid(row=13,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_table_details_datatype_mssql_sf = ttk.Button(
            self.DDL_validation_ddl_mssql_sf,
            text='Select Template file',
            command=self.select_query_file_table_details_datatype_mssql_sf,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_table_details_datatype_mssql_sf.grid(row=14,column=0,padx=(70,0),pady=5)

        
        
        
        self.download_path_datatype_mssql_sf = ttk.Entry(self.DDL_validation_ddl_mssql_sf, font="Arial 11 ")
        self.download_path_datatype_mssql_sf.grid(row=13,column=1,pady=5,ipadx=200)
        

        self.query_excel_file_datatype_mssql_sf = ttk.Entry(self.DDL_validation_ddl_mssql_sf, font="Arial 11 ")
        self.query_excel_file_datatype_mssql_sf.grid(row=14,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_table_details_datatype_mssql_sf = ttk.Button(
            self.DDL_validation_ddl_mssql_sf,
            text='Start MS-SQL vs SF Datatype Comparison',
            command=self.datatype_check_mssql_sf,
            width=32, bootstyle='success'
            )

        self.executeQueryFileBtn_table_details_datatype_mssql_sf.grid(row=15,column=1,pady=5,padx=5)

        # checkbox for sqoop
        self.useSqoopCommand=IntVar()
        self.c8 = Checkbutton(self.DDL_validation_ddl_mssql_sf, text = "Use Sqoop",variable=self.useSqoopCommand,bg="#0C7A79", font=("Arial", self.fontSize),command=self.showSqoopInputBar)
        self.c8.grid(row=16,column=0,padx=(70,0),pady=5)

        self.sqoop_command_base = ttk.Entry(self.DDL_validation_ddl_mssql_sf, font="Arial 11 ")
        

        #-sf-sf ddl validation start
        # sf -sf datatype
        self.DDL_validation_ddl_sf_sf=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=2,pady=1)
        tk.Label(self.DDL_validation_ddl_sf_sf, text="SF - SF datatype comparison : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=12,padx=(100,0),sticky='w')

        self.sf_sf_datatype_template_btn=ttk.Button(
                self.DDL_validation_ddl_sf_sf,
                text=u'\u2193'+'Download Template',
                command=self.download_sf_sf_datatype_comparison_template,
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
        self.sf_sf_datatype_template_btn.grid(row=12,column=1,pady=5,padx=5,sticky='e')

        self.file_location_btn_table_details_datatype_sf_sf = ttk.Button(
            self.DDL_validation_ddl_sf_sf,
            text='Select Result Location ',
            command=self.select_download_location_datatype_sf_sf,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_table_details_datatype_sf_sf.grid(row=13,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_table_details_datatype_sf_sf = ttk.Button(
            self.DDL_validation_ddl_sf_sf,
            text='Select Template file',
            command=self.select_query_file_table_details_datatype_sf_sf,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_table_details_datatype_sf_sf.grid(row=14,column=0,padx=(70,0),pady=5)

        
        
        
        self.download_path_datatype_sf_sf = ttk.Entry(self.DDL_validation_ddl_sf_sf, font="Arial 11 ")
        self.download_path_datatype_sf_sf.grid(row=13,column=1,pady=5,ipadx=200)
        

        self.query_excel_file_datatype_sf_sf = ttk.Entry(self.DDL_validation_ddl_sf_sf, font="Arial 11 ")
        self.query_excel_file_datatype_sf_sf.grid(row=14,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_table_details_datatype_sf_sf = ttk.Button(
            self.DDL_validation_ddl_sf_sf,
            text='Start SF vs SF Datatype Comparison',
            command=self.datatype_check_sf_sf_only_desc,
            width=32, bootstyle='success'
            )

        self.executeQueryFileBtn_table_details_datatype_sf_sf.grid(row=15,column=1,pady=5,padx=5)

        # sf-sf ddl validation end
        

         # td-sf casted check
        self.td_sf_casted_cols_frame=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=2,pady=1)
        tk.Label(self.td_sf_casted_cols_frame, text="TD- SF casted cols comparison : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=12,padx=(100,0),sticky='w')
        t_parameters=['TD db.table','SF db.schema.table']
        self.td_sf_casted_cols_template_btn=ttk.Button(
                self.td_sf_casted_cols_frame,
                text=u'\u2193'+'Download Template',
                command=lambda: self.download_template('TD-SF casting check template',t_parameters),
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
        self.td_sf_casted_cols_template_btn.grid(row=12,column=1,pady=5,padx=5,sticky='e')

        self.file_location_btn_table_details_td_sf_casted_cols = ttk.Button(
            self.td_sf_casted_cols_frame,
            text='Select Result Location ',
            command=lambda: self.select_folder(self.download_path_td_sf_casted_cols),
            width=25, bootstyle='primary'
            )

        self.file_location_btn_table_details_td_sf_casted_cols.grid(row=13,column=0,padx=(70,0),pady=5)
        
        
        self.download_path_td_sf_casted_cols = ttk.Entry(self.td_sf_casted_cols_frame, font="Arial 11 ")
        self.download_path_td_sf_casted_cols.grid(row=13,column=1,pady=5,ipadx=200)
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_table_details_td_sf_casted_cols = ttk.Button(
            self.td_sf_casted_cols_frame,
            text='Select Template file',
            command=lambda: self.select_file('Select Template file',self.query_excel_file_td_sf_casted_cols),
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_table_details_td_sf_casted_cols.grid(row=14,column=0,padx=(70,0),pady=5)

        
        
        

        self.query_excel_file_td_sf_casted_cols = ttk.Entry(self.td_sf_casted_cols_frame, font="Arial 11 ")
        self.query_excel_file_td_sf_casted_cols.grid(row=14,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_table_details_td_sf_casted_cols = ttk.Button(
            self.td_sf_casted_cols_frame,
            text='Start TD-SF casted cols Comparison',
            command=self.check_casted_cols_td_sf,
            width=32, bootstyle='primary'
            )

        self.executeQueryFileBtn_table_details_td_sf_casted_cols.grid(row=15,column=1,pady=5,padx=5)

        # td-sf casted check validation end
        

        # mssql-teradat datatype
        self.DDL_validation_ddl_mssql_td=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=2,pady=1)
        tk.Label(self.DDL_validation_ddl_mssql_td, text="MS SQL - TD datatype comparison : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=16,padx=(100,0),sticky='w')

        self.mssql_td_datatype_template_btn=ttk.Button(
                self.DDL_validation_ddl_mssql_td,
                text=u'\u2193'+'Download Template',
                command=self.download_mssql_td_datatype_comparison_template,
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
        self.mssql_td_datatype_template_btn.grid(row=16,column=1,pady=5,padx=5,sticky='e')

        self.file_location_btn_table_details_datatype_mssql_td = ttk.Button(
            self.DDL_validation_ddl_mssql_td,
            text='Select Result Location ',
            command=self.select_download_location_datatype_mssql_td,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_table_details_datatype_mssql_td.grid(row=17,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_table_details_datatype_mssql_td = ttk.Button(
            self.DDL_validation_ddl_mssql_td,
            text='Select Template file',
            command=self.select_query_file_table_details_datatype_mssql_td,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_table_details_datatype_mssql_td.grid(row=18,column=0,padx=(70,0),pady=5)

        
        
        
        self.download_path_datatype_mssql_td = ttk.Entry(self.DDL_validation_ddl_mssql_td, font="Arial 11 ")
        self.download_path_datatype_mssql_td.grid(row=17,column=1,pady=5,ipadx=200)
        

        self.query_excel_file_datatype_mssql_td = ttk.Entry(self.DDL_validation_ddl_mssql_td, font="Arial 11 ")
        self.query_excel_file_datatype_mssql_td.grid(row=18,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_table_details_datatype_mssql_td = ttk.Button(
            self.DDL_validation_ddl_mssql_td,
            text='Start MS-SQL vs td Datatype Comparison',
            command=self.datatype_check_mssql_td,
            width=32, bootstyle='success'
            )

        self.executeQueryFileBtn_table_details_datatype_mssql_td.grid(row=19,column=1,pady=5,padx=5)
        
         # Tokenization check  
        self.tokenization_check_frame=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=2,pady=1)
        
        tk.Label(self.tokenization_check_frame, text="Tokenization check : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,padx=(100,0),sticky='w')

        self.tokenization_check_template_btn=ttk.Button(
                self.tokenization_check_frame,
                text=u'\u2193'+'Download Template',
                command=self.download_tokenization_check_template,
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
        self.tokenization_check_template_btn.grid(row=0,column=1,pady=5,padx=5,sticky='e')

        self.file_location_btn_table_details_tokenization_check = ttk.Button(
            self.tokenization_check_frame,
            text='Select Result Location ',
            command=self.select_download_location_tokenization_check,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_table_details_tokenization_check.grid(row=1,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_table_details_tokenization_check = ttk.Button(
            self.tokenization_check_frame,
            text='Select Template file',
            command=self.select_query_file_table_details_tokenization_check,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_table_details_tokenization_check.grid(row=2,column=0,padx=(70,0),pady=5)

        
        
        
        self.download_path_tokenization_check = ttk.Entry(self.tokenization_check_frame, font="Arial 11 ")
        self.download_path_tokenization_check.grid(row=1,column=1,pady=5,ipadx=200)
        

        self.query_excel_file_tokenization_check = ttk.Entry(self.tokenization_check_frame, font="Arial 11 ")
        self.query_excel_file_tokenization_check.grid(row=2,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_table_details_tokenization_check = ttk.Button(
            self.tokenization_check_frame,
            text='Start Tokenization Check',
            command=self.tokenization_check_fn,
            width=30,
                  bootstyle="primary"
            )

        self.executeQueryFileBtn_table_details_tokenization_check.grid(row=3,column=1,pady=5,padx=5)


        # view validation sf
         
        self.View_Validation_SF=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=2,pady=1)
        tk.Label(self.View_Validation_SF, text="View Validation SF  : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,padx=(100,0),sticky='w')

        self.View_Validation_SF_template_btn=ttk.Button(
                self.View_Validation_SF,
                text=u'\u2193'+'Download Template',
                command=self.download_View_Validation_SF_comparison_template,
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
        self.View_Validation_SF_template_btn.grid(row=0,column=1,pady=5,padx=5,sticky='e')

        self.file_location_btn_table_details_View_Validation_SF = ttk.Button(
            self.View_Validation_SF,
            text='Select Result Location ',
            command=self.select_download_location_View_Validation_SF,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_table_details_View_Validation_SF.grid(row=1,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_table_details_View_Validation_SF = ttk.Button(
            self.View_Validation_SF,
            text='Select Template file',
            command=self.select_query_file_table_details_View_Validation_SF,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_table_details_View_Validation_SF.grid(row=2,column=0,padx=(70,0),pady=5)

        
        
        
        self.download_path_View_Validation_SF = ttk.Entry(self.View_Validation_SF, font="Arial 11 ")
        self.download_path_View_Validation_SF.grid(row=1,column=1,pady=5,ipadx=200)
        

        self.query_excel_file_View_Validation_SF = ttk.Entry(self.View_Validation_SF, font="Arial 11 ")
        self.query_excel_file_View_Validation_SF.grid(row=2,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_table_details_View_Validation_SF = ttk.Button(
            self.View_Validation_SF,
            text='Start SF View Validation ',
            command=self.view_validation_sf,
            width=30,
                  bootstyle="primary"
            )

        self.executeQueryFileBtn_table_details_View_Validation_SF.grid(row=3,column=1,pady=5,padx=5)

         # view validation sf rowcount + schema +timestamp
         
        
        tk.Label(self.View_Validation_SF, text="View Validation SF [Rowcount+Schemas+Timestamp]  : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=4,padx=(100,0),sticky='w')

        # self.View_Validation_SF_rtc_template_btn=ttk.Button(
        #         self.View_Validation_SF,
        #         text=u'\u2193'+'Download Template',
        #         command=self.download_View_Validation_SF_rtc_comparison_template,
        #         height= 1, width=20, font=("Arial bold", self.fontSize),
        #         fg='white',
        #         bg='grey'

        #     )
        # self.View_Validation_SF_rtc_template_btn.grid(row=4,column=1,pady=5,padx=5,sticky='e')

        self.file_location_btn_table_details_View_Validation_SF_rtc = ttk.Button(
            self.View_Validation_SF,
            text='Select Result Location ',
            command=lambda: self.select_folder(self.download_path_View_Validation_SF_rtc),
            width=25, bootstyle='primary'
            )

        self.file_location_btn_table_details_View_Validation_SF_rtc.grid(row=5,column=0,padx=(70,0),pady=5)
        
        
        self.download_path_View_Validation_SF_rtc = ttk.Entry(self.View_Validation_SF, font="Arial 11 ")
        self.download_path_View_Validation_SF_rtc.grid(row=5,column=1,pady=5,ipadx=200)
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_table_details_View_Validation_SF_rtc = ttk.Button(
            self.View_Validation_SF,
            text='Select Template file',
            command=lambda: self.select_file('Select Template file',self.query_excel_file_View_Validation_SF_rtc),
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_table_details_View_Validation_SF_rtc.grid(row=6,column=0,padx=(70,0),pady=5)

        self.query_excel_file_View_Validation_SF_rtc = ttk.Entry(self.View_Validation_SF, font="Arial 11 ")
        self.query_excel_file_View_Validation_SF_rtc.grid(row=6,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_table_details_View_Validation_SF_rtc = ttk.Button(
            self.View_Validation_SF,
            text='Start SF View[RTC] Validation ',
            command=self.view_validation_sf_rtc_stg,
            width=30,
                  bootstyle="primary"
            )

        self.executeQueryFileBtn_table_details_View_Validation_SF_rtc.grid(row=7,column=1,pady=5,padx=5)

        
         # view validation sf rowcount + schema +timestamp
         
        
        tk.Label(self.View_Validation_SF, text="View Validation SF [Rowcount+Schemas+Timestamp]-NonSTG  : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=8,padx=(100,0),sticky='w')

        # self.View_Validation_SF_rtc_template_btn=ttk.Button(
        #         self.View_Validation_SF,
        #         text=u'\u2193'+'Download Template',
        #         command=self.download_View_Validation_SF_rtc_comparison_template,
        #         height= 1, width=20, font=("Arial bold", self.fontSize),
        #         fg='white',
        #         bg='grey'

        #     )
        # self.View_Validation_SF_rtc_template_btn.grid(row=4,column=1,pady=5,padx=5,sticky='e')

        self.file_location_btn_table_details_View_Validation_SF_rtc_non_stg = ttk.Button(
            self.View_Validation_SF,
            text='Select Result Location ',
            command=lambda: self.select_folder(self.download_path_View_Validation_SF_rtc_non_stg),
            width=25, bootstyle='primary'
            )

        self.file_location_btn_table_details_View_Validation_SF_rtc_non_stg.grid(row=9,column=0,padx=(70,0),pady=5)
        
        
        self.download_path_View_Validation_SF_rtc_non_stg = ttk.Entry(self.View_Validation_SF, font="Arial 11 ")
        self.download_path_View_Validation_SF_rtc_non_stg.grid(row=9,column=1,pady=5,ipadx=200)
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_table_details_View_Validation_SF_rtc_non_stg = ttk.Button(
            self.View_Validation_SF,
            text='Select Template file',
            command=lambda: self.select_file('Select Template file',self.query_excel_file_View_Validation_SF_rtc_non_stg),
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_table_details_View_Validation_SF_rtc_non_stg.grid(row=10,column=0,padx=(70,0),pady=5)

        self.query_excel_file_View_Validation_SF_rtc_non_stg = ttk.Entry(self.View_Validation_SF, font="Arial 11 ")
        self.query_excel_file_View_Validation_SF_rtc_non_stg.grid(row=10,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_table_details_View_Validation_SF_rtc_non_stg = ttk.Button(
            self.View_Validation_SF,
            text='Start SF View[RTC_non_stg] Validation ',
            command=self.view_validation_sf_rtc_non_stg_non_stg,
            width=30,
                  bootstyle="primary"
            )

        self.executeQueryFileBtn_table_details_View_Validation_SF_rtc_non_stg.grid(row=11,column=1,pady=5,padx=5)
        # view validation sf ends
    def get_casted_cols(self,script):
        a=script.split(',')
        b=list(filter(lambda x:"cast('xx'" in x.lower().replace(' ',''),a))
        casted_list=list()
        for ss in b:
                s3=ss.lower().strip()
                s3_base=s3.split('select')

                for s3_b in s3_base:
                        s3_all=s3_b.split('from')[0].split(',')
                        for col in s3_all:
                            
                            if "cast('xx'" not in col.replace(' ',''): continue
                            
                            if 'case' in col and 'when' in col:
                                casted_list.append(col.split('end')[-1].split()[-1])
                            else:  casted_list.append(col.split()[-1])
        return [i.strip().lower() for i in casted_list]



    def check_casted_cols_td_sf(self):

        print('TD-SF casted cols test started..')
         
        basepath=str(self.download_path_td_sf_casted_cols.get()).strip()
        os.chdir(basepath)

        df_path=(self.query_excel_file_td_sf_casted_cols.get()).strip()
        df=pd.read_excel(df_path)
        Result_DF=pd.DataFrame(columns=['Cols casted in td not in sf','Cols casted in sf not in td','Common casted cols','Casted cols in td','Casted cols in sf','TD_Table','SF_Table'])

        for row in (df.itertuples()):
            td_table=str(row[1]).strip()
            sf_table=str(row[2]).strip()
            tdq='show select * from '+td_table
            tdq_df=self.run_td_query(tdq)
            tdq_res=tdq_df.iloc[0][0]
            td_casted_cols=self.get_casted_cols(tdq_res)
            # print(td_casted_cols)

            sfq="select get_ddl('table','{tn}')".format(tn=sf_table)
            sfq_df=self.run_sf_query(sfq)
            sfq_res=sfq_df.iloc[0][0]
            sf_casted_cols=self.get_casted_cols(sfq_res)
            # print(sf_casted_cols)

            cols_in_td_not_in_sf=list()
            for col in td_casted_cols:
                if col not in sf_casted_cols:
                    cols_in_td_not_in_sf.append(col)

            common_cols=list()
            cols_in_sf_not_in_td=list()
            for col in sf_casted_cols:
                if col not in td_casted_cols:
                    cols_in_sf_not_in_td.append(col)
                else: common_cols.append(col)


            
            Result_DF.loc[len(Result_DF.index)]=[",".join(cols_in_td_not_in_sf),",".join(cols_in_sf_not_in_td),",".join(common_cols),",".join(td_casted_cols),",".join(sf_casted_cols),td_table,sf_table]



        filename='Result_Td_SF_Casted_Cols_Comparison'+str(time.time())+'.xlsx'
        Result_DF.to_excel(filename,index=False)
        messagebox.showinfo('done','done dona don')

    def view_validation_sf_rtc_non_stg_non_stg(self):
        print('SF view non-stg Validation started..')
         
        basepath=str(self.download_path_View_Validation_SF_rtc_non_stg.get()).strip()
        os.chdir(basepath)

        df_path=(self.query_excel_file_View_Validation_SF_rtc_non_stg.get()).strip()
        self.view_nonstg_table_check(df_path)

    
    def view_nonstg_table_check(self,path) :

        #path = "C:\\Users\\AL17741\\Downloads\\Security View Validation\\mar 7\EDL_RAWZ_0307_0309.xlsx"
        wrkbk = openpyxl.load_workbook(path)
        
        sh = wrkbk.get_sheet_by_name('Sheet1')
        establish = self.con.cursor()
        
            
        print("Max column :" , sh.max_column)
        print("Max row :" , sh.max_row)
        #iterate through excel and display data
        for i in range(2, sh.max_column):  
            #for j in range(2, 3):
            for j in range(2, sh.max_row+1):
                print("column no :" , i)
                print("row no:" , j)
                cell_obj = sh.cell(row=j, column=2)
                cell_obj_col_val = sh.cell(row=1, column=i+1)
                DW_val = sh.cell(row=j, column=1)
                print("table name :  "  , cell_obj.value)
                print("schema name :  "  , cell_obj_col_val.value)
                print("Datawarehouse name :  "  , DW_val.value)
                try:
                    if i == 2:
                        print("show tables like '" + cell_obj.value + "' in " + DW_val.value +"."+ cell_obj_col_val.value)
                        establish.execute("show tables like '" + cell_obj.value + "' in " + DW_val.value +"."+ cell_obj_col_val.value)
                        
                    else:
                        print("show views like '" + cell_obj.value + "' in " + DW_val.value +"."+ cell_obj_col_val.value)
                        establish.execute("show views like '" + cell_obj.value + "' in " + DW_val.value +"."+ cell_obj_col_val.value)
                    c2 = sh.cell(row = j, column = i+1)
                    print(j,i+1)
                    getcount1 = establish.fetchone()
                    getval = str(getcount1[0])
                    c2.value =  getval
                    print(getval)                  
                except:
                    c2 = sh.cell(row = j, column = i+1)
                    c2.value = '0'
                    fill_cell = PatternFill(patternType='solid', fgColor= 'ffe6e6')
                    sh.cell(row = j, column = i+1).fill = fill_cell
        
                wrkbk.save(path)
        
        sh_count = wrkbk.get_sheet_by_name('Count')
        
        print("Max column :" , sh_count.max_column)
        print("Max row :" , sh_count.max_row)
        #iterate through excel and display data
        for i in range(2, sh_count.max_column):  
            #for j in range(2, 3):
            for j in range(2, sh_count.max_row+1):
                print("column no :" , i)
                print("row no:" , j)
                cell_obj = sh_count.cell(row=j, column=2)
                cell_obj_col_val = sh_count.cell(row=1, column=i+1)
                DW_val = sh_count.cell(row=j, column=1)
                print("table name :  "  , cell_obj.value)
                print("schema name :  "  , cell_obj_col_val.value)
                print("Datawarehouse name :  "  , DW_val.value)
                try:
                    if i == 2:
                        print("select count(*) from " + DW_val.value +"."+ cell_obj_col_val.value +"."+ cell_obj.value)
                        establish.execute("select count(*) from " + DW_val.value +"."+ cell_obj_col_val.value +"."+ cell_obj.value)               
                    else:
                        print("select count(*) from " + DW_val.value +"."+ cell_obj_col_val.value +"."+ cell_obj.value)
                        establish.execute("select count(*) from " + DW_val.value +"."+ cell_obj_col_val.value +"."+ cell_obj.value)
                    c2 = sh_count.cell(row = j, column = i+1)
                    print(j,i+1)
                    getcount1 = establish.fetchone()
                    getval = str(getcount1[0])
                    c2.value =  getval
                    print(getval)                  
                except:
                    c2 = sh.cell(row = j, column = i+1)
                    c2.value = '0'
                    fill_cell = PatternFill(patternType='solid', fgColor= 'ffe6e6')
                    sh.cell(row = j, column = i+1).fill = fill_cell
                wrkbk.save(path)
                
        
    def view_validation_sf_rtc_stg(self):
        print('SF view stg Validation started..')
         
        basepath=str(self.download_path_View_Validation_SF_rtc.get()).strip()
        os.chdir(basepath)

        df_path=(self.query_excel_file_View_Validation_SF_rtc.get()).strip()
        
        self.security_view_check(df_path,"Time_Stamp_Check")
        self.security_view_check(df_path,"Count_Check")
         

    
    def security_view_check(self,path,sheet_name) :


        wrkbk = openpyxl.load_workbook(path)
        sh = wrkbk[sheet_name]
        establish = self.con.cursor()
        
        print(" Security View Validation for " + sheet_name + " Started")
        #iterate through excel and display data
        for i in range(3, sh.max_column):  
            for j in range(2, sh.max_row+1):
                table_nm = sh.cell(row=j, column=2)
                base_schema_nm = sh.cell(row=j, column=3)
                if sh.cell(row=1, column=i+1).value == 'Base_Schema' :
                    schema_nm = str(sh.cell(row=j, column=3).value)
                else:
                    schema_nm = str(sh.cell(row=j, column=3).value) +"_"+ str(sh.cell(row=1, column=i+1).value)
                Datawarehouse_nm = sh.cell(row=j, column=1)
                try:
                    if sheet_name == 'Time_Stamp_Check':
                        if i == 3:
                            establish.execute("show tables like '" + table_nm.value + "' in " + Datawarehouse_nm.value +"."+ schema_nm)                 
                        else:
                            establish.execute("show views like '" + table_nm.value + "' in " + Datawarehouse_nm.value +"."+ schema_nm)
                    elif sheet_name == 'Count_Check':
                        if i == 3:
                            establish.execute("select count(*) from " + Datawarehouse_nm.value +"."+ schema_nm +"."+ table_nm.value)              
                        else:
                            establish.execute("select count(*) from " + Datawarehouse_nm.value +"."+ schema_nm +"."+ table_nm.value)
                    c2 = sh.cell(row = j, column = i+1)
                    getcount1 = establish.fetchone()
                    getval = str(getcount1[0])
                    c2.value =  getval                
                except:
                    c2 = sh.cell(row = j, column = i+1)
                    c2.value = '0'
                wrkbk.save(path)
                print(" Security View Validation for " + sheet_name + " Completed")
            

    def view_validation_sf(self):
        print('Validation started..')
         
        basepath=str(self.download_path_View_Validation_SF.get()).strip()
        os.chdir(basepath)

        df_path=(self.query_excel_file_View_Validation_SF.get()).strip()
        
       
        df=pd.read_excel(df_path)
        df2=pd.read_excel(df_path,sheet_name='Module-Columns')
        
        df=df.merge(df2,on='Module')

        Schema_Present_Absent_df=pd.DataFrame(columns=['DB','Table','Present in Schemas','Absent in Schemas'])
        Cast_needed_df=pd.DataFrame(columns=['DB','Schema','Table','column cast in sf but not in list','columns in list but not casted in sf','casted_columns_in_sf','casted_columns_in_list'])
        Cast_not_needed_df=pd.DataFrame(columns=['DB','Schema','Table','Columns Casted in SF'])
        #temprorary line below
        module_used_cols_mapping=dict()

        for module in df2['Module']:
            module_used_cols_mapping[module]=list()
        

        for row in (df.itertuples()):
            db_name=row[1]
            table_name=row[2]
            cast_needed_schemas=row[3]
            cast_needed_schemas_list=[ sch.strip().lower() for sch in cast_needed_schemas.split(',')]
            cast_needed_schemas_list.sort()
            
            cast_not_needed_schemas=row[4]
            cast_not_needed_schemas_list=[ sch.strip().lower() for sch in cast_not_needed_schemas.split(',')]
            cast_not_needed_schemas_list.sort()
            module=row[5]
            


            # all_column_names_that_should_be_casted=['contactname','contactphonenumber','note','insurednumber','oldinsurednumber','hcid','memberpolicyfirstname','memberpolicylastname','memberpolicymiddleinitial','memberpolicydob','contacttypecd','phonenumber','callername','memberfirstname','membermiddlename','memberlastname','memberdob','mcid','submittercontactname','submitteremailaddress','submitterphone','submitterextension','rejectuserid','rejectusername','notes','contactnote','memberid','membersequencenumber','medicaidid','conditioncd','assessmenttitle','assessmentdesc','memberidentifier','callerfirstname','callerphonenumber','callerinsurednumber','callerhcid','nurselinecustomerid','callerlastname','callerdob','firstname','lastname','dob','hcid','address1','postalcode','faxnumber','salesrepname','questionsandanswers','membertrumppolicyid','memberpoliciesid','nguserid']
            all_column_names_that_should_be_casted=[x.strip().lower() for x in row[6].split(',')]
            all_column_names_that_should_be_casted.sort()

            


            query="SELECT TABLE_SCHEMA FROM {db}.INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME='{tn}'".format(db=db_name.upper(),tn=table_name.upper())
            schemas_where_table_is_present_df=self.run_sf_query(query)
            schemas_where_table_is_present_list=list(schemas_where_table_is_present_df['TABLE_SCHEMA'])
            schemas_where_table_is_present_list=[x.strip().lower() for x in schemas_where_table_is_present_list]
            schemas_where_table_is_present_list.sort()

            schemas_where_table_is_absent_list=list() 
            all_schema_names_list=cast_needed_schemas_list+cast_not_needed_schemas_list

            for schema in all_schema_names_list:
                if schema not in schemas_where_table_is_present_list:
                    schemas_where_table_is_absent_list.append(schema.strip().lower())

            schemas_where_table_is_absent_list.sort()
            schemas_where_table_is_absent_list=list(set(schemas_where_table_is_absent_list))
            schemas_where_table_is_present_list=list(set(schemas_where_table_is_present_list))
            Schema_Present_Absent_df.loc[len(Schema_Present_Absent_df.index)]=[db_name,table_name,','.join(schemas_where_table_is_present_list),','.join(schemas_where_table_is_absent_list)]
        
            
            
            for schema in cast_needed_schemas_list:
                if schema.strip().lower()  in schemas_where_table_is_absent_list: continue
                full_table_name=db_name.upper()+'.'+schema.upper()+'.'+table_name.upper()
                query="select get_ddl('view','{viewname}')".format(viewname=full_table_name)
                df=self.run_sf_query(query) 
                df.columns=['c1']

                s=list(df['c1'])[0]
                # all_views_defn.write(table_sf+":\n")
                # all_views_defn.write(s+"\n\n")
                # s=list(s)[0]
                 

                # a=s.split(',')
                # b=list(filter(lambda x:"cast('xx'" in x.lower(),a))
                # casted_list=list()
                casted_list=self.get_casted_cols(s)

                # for ss in b:
                #     s3=ss.lower().strip()
                #     casted_list.append(s3.split('from')[0].split()[-1])

                casted_list.sort()

                column_cast_in_sf_but_not_in_list=list()
                column_in_list_but_not_casted_in_sf=list()

                for col in casted_list:
                    l1=module_used_cols_mapping[module]
                    l1.append(col)
                    module_used_cols_mapping[module]=l1

                    if(col.strip().lower() not in all_column_names_that_should_be_casted ):
                            column_cast_in_sf_but_not_in_list.append(col)
                   
                    
                column_cast_in_sf_but_not_in_list.sort()
                all_cols=self.getCols_SF(full_table_name)

                for col in all_cols:
                    if (col.strip().lower() in all_column_names_that_should_be_casted )and (col.strip().lower() not  in casted_list):
                        column_in_list_but_not_casted_in_sf.append(col.strip().lower())

                
                column_in_list_but_not_casted_in_sf.sort()
                column_in_list_but_not_casted_in_sf=list(set(column_in_list_but_not_casted_in_sf))
                column_cast_in_sf_but_not_in_list=list(set(column_cast_in_sf_but_not_in_list))
                casted_list=list(set(casted_list))
                all_column_names_that_should_be_casted=list(set(all_column_names_that_should_be_casted))
                Cast_needed_df.loc[len(Cast_needed_df.index)]=[db_name,schema,table_name,",".join(column_cast_in_sf_but_not_in_list),",".join(column_in_list_but_not_casted_in_sf),",".join(casted_list),",".join(all_column_names_that_should_be_casted)]


            for schema in cast_not_needed_schemas_list:
                if schema.strip().lower()  in schemas_where_table_is_absent_list: continue
                full_table_name=db_name.upper()+'.'+schema.upper()+'.'+table_name.upper()
                query="select get_ddl('view','{viewname}')".format(viewname=full_table_name)
                df=self.run_sf_query(query) 
                df.columns=['c1']

                s=list(df['c1'])[0]
                # all_views_defn.write(table_sf+":\n")
                # all_views_defn.write(s+"\n\n")
                # s=list(s)[0]
                 

                # a=s.split(',')
                # b=list(filter(lambda x:"cast('xx'" in x.lower(),a))
                # casted_list=list()

                # for ss in b:
                #     s3=ss.lower().strip()
                #     casted_list.append(s3.split('from')[0].split()[-1])

                casted_list=self.get_casted_cols(s)
                casted_list.sort()
                casted_list=list(set(casted_list))
                Cast_not_needed_df.loc[len(Cast_not_needed_df.index)]=[db_name,schema,table_name,",".join(casted_list)]

        values_present_nowhere=dict()
        for module in df2['Module']:
            values_present_nowhere[module]=''


        for row in  (df2.itertuples()):
            modulename=row[1]
            # print(row)
            cast_cols=[x.strip().lower() for x in row[2].split(',')]
            
            for col in cast_cols:
                if col not in module_used_cols_mapping[modulename]:
                    values_present_nowhere[modulename]=values_present_nowhere[modulename]+','+str(col)

        df3=pd.DataFrame({'Module':values_present_nowhere.keys(),'Cols':values_present_nowhere.values()})
        # df3=df3.transpose()
        filename='Result_Sf_View_Validation_'+str(time.time())+'.xlsx'

        with pd.ExcelWriter(filename) as writer:
                    Cast_needed_df.to_excel(writer,sheet_name='Cast_Needed_Schemas',index = False,header=True) 
                    Cast_not_needed_df.to_excel(writer,sheet_name='Cast_Not_Needed_Schemas',index = False,header=True) 
                    Schema_Present_Absent_df.to_excel(writer,sheet_name='Absent tables',index = False,header=True) 
                    df3.to_excel(writer,sheet_name='Columns Never Casted',index = False,header=True) 
                
        print('Done')
        messagebox.showinfo('Done',filename+' store in '+basepath)
                



        

    def showSqoopInputBar(self):
        if(self.useSqoopCommand.get()==1):
           self.sqoop_command_base.grid(row=16,column=1,pady=5,ipadx=200)
        else: self.sqoop_command_base.grid_forget()

    
    def initiateUi_MySQLWorkBench(self):
        tk.Label(self.MySQLWorkBench_frame, text="Step 1 - Enter Connection Details: ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,padx=(100,0),sticky='w')
        # tk.Label(self.MySQLWorkBench_frame, text="Unix Hostname : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=1,padx=(10,0))
        
        tk.Label(self.MySQLWorkBench_frame, text="Hostname : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=2,padx=(10,0))
        tk.Label(self.MySQLWorkBench_frame, text="UserId : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=3,padx=(10,0))
        tk.Label(self.MySQLWorkBench_frame, text="Password : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=4,padx=(10,0))


         
        
        self.hostname_MySQLWorkBench = ttk.Entry(self.MySQLWorkBench_frame, font="Arial 11 ")
        self.hostname_MySQLWorkBench.grid(row=2,column=1,pady=5,ipadx=200)
        self.hostname_MySQLWorkBench.insert(0,'antm-mysqldb-cluster.cluster-ro-csntho9gpvhy.us-east-1.rds.amazonaws.com')
        
        self.uid_MySQLWorkBench = ttk.Entry(self.MySQLWorkBench_frame, font="Arial 11 ")
        self.uid_MySQLWorkBench.grid(row=3,column=1,pady=5,ipadx=200)
        
        self.pwd_MySQLWorkBench = ttk.Entry(self.MySQLWorkBench_frame,show='*', font="Arial 11 ")
        self.pwd_MySQLWorkBench.grid(row=4,column=1,pady=5,ipadx=200)


        self.connect_Btn_MySQLWorkBench = ttk.Button(
            self.MySQLWorkBench_frame,
            text='Establish MySQLWorkBench Connection',
            command=self.establish_connection_MySQLWorkBench,
            width=30,
                  bootstyle="success"
            )

        self.connect_Btn_MySQLWorkBench.grid(row=5,column=1,pady=5,padx=5)

        tk.Label(self.MySQLWorkBench_frame, text="Step 2 - Download from semicolon(;) separated query file : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=6,padx=(100,0),sticky='w')

        self.file_location_btn_MySQLWorkBench = ttk.Button(
            self.MySQLWorkBench_frame,
            text='Select Result Location ',
            command=self.select_download_location_MySQLWorkBench,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_MySQLWorkBench.grid(row=7,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_MySQLWorkBench = ttk.Button(
            self.MySQLWorkBench_frame,
            text='Select query file',
            command=self.select_query_file_MySQLWorkBench,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_MySQLWorkBench.grid(row=8,column=0,padx=(70,0),pady=5)

        
        
        
        self.download_path_MySQLWorkBench = ttk.Entry(self.MySQLWorkBench_frame, font="Arial 11 ")
        self.download_path_MySQLWorkBench.grid(row=7,column=1,pady=5,ipadx=200)
        

        self.query_MySQLWorkBench = ttk.Entry(self.MySQLWorkBench_frame, font="Arial 11 ")
        self.query_MySQLWorkBench.grid(row=8,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_MySQLWorkBench = ttk.Button(
            self.MySQLWorkBench_frame,
            text='Start Query Execution',
            command=self.start_download_MySQLWorkBench,
            width=25, bootstyle='success'
            )

        self.executeQueryFileBtn_MySQLWorkBench.grid(row=11,column=1,pady=5,padx=5)
    
    
    def initiateUi_PostgresSQL(self):
        tk.Label(self.PostgresSQL_frame, text="Step 1 - Enter Connection Details: ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,padx=(100,0),sticky='w')
        # tk.Label(self.PostgresSQL_frame, text="Unix Hostname : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=1,padx=(10,0))
        
        tk.Label(self.PostgresSQL_frame, text="Hostname : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=2,padx=(10,0))
        tk.Label(self.PostgresSQL_frame, text="Database : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=3,padx=(10,0))
        tk.Label(self.PostgresSQL_frame, text="Port : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=4,padx=(10,0))
        tk.Label(self.PostgresSQL_frame, text="UserId : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=5,padx=(10,0))
        tk.Label(self.PostgresSQL_frame, text="Password : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=6,padx=(10,0))


         
        
        self.hostname_PostgresSQL = ttk.Entry(self.PostgresSQL_frame, font="Arial 11 ")
        self.hostname_PostgresSQL.grid(row=2,column=1,pady=5,ipadx=200)
        self.hostname_PostgresSQL.insert(0,'')

        self.database_PostgresSQL = ttk.Entry(self.PostgresSQL_frame, font="Arial 11 ")
        self.database_PostgresSQL.grid(row=3,column=1,pady=5,ipadx=200)
        self.database_PostgresSQL.insert(0,'')

        self.port_PostgresSQL = ttk.Entry(self.PostgresSQL_frame, font="Arial 11 ")
        self.port_PostgresSQL.grid(row=4,column=1,pady=5,ipadx=200)
        self.port_PostgresSQL.insert(0,'5432')
        
        self.uid_PostgresSQL = ttk.Entry(self.PostgresSQL_frame, font="Arial 11 ")
        self.uid_PostgresSQL.grid(row=5,column=1,pady=5,ipadx=200)
        
        self.pwd_PostgresSQL = ttk.Entry(self.PostgresSQL_frame,show='*', font="Arial 11 ")
        self.pwd_PostgresSQL.grid(row=6,column=1,pady=5,ipadx=200)


        self.connect_Btn_PostgresSQL = ttk.Button(
            self.PostgresSQL_frame,
            text='Establish PostgresSQL Connection',
            command=self.establish_connection_PostgresSQL,
            width=30,
                  bootstyle="success"
            )

        self.connect_Btn_PostgresSQL.grid(row=7,column=1,pady=5,padx=5)

        tk.Label(self.PostgresSQL_frame, text="Step 2 - Download from semicolon(;) separated query file : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=8,padx=(100,0),sticky='w')

        self.file_location_btn_PostgresSQL = ttk.Button(
            self.PostgresSQL_frame,
            text='Select Result Location ',
            command=self.select_download_location_PostgresSQL,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_PostgresSQL.grid(row=9,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_PostgresSQL = ttk.Button(
            self.PostgresSQL_frame,
            text='Select query file',
            command=self.select_query_file_PostgresSQL,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_PostgresSQL.grid(row=10,column=0,padx=(70,0),pady=5)

        
        
        
        self.download_path_PostgresSQL = ttk.Entry(self.PostgresSQL_frame, font="Arial 11 ")
        self.download_path_PostgresSQL.grid(row=9,column=1,pady=5,ipadx=200)
        

        self.query_PostgresSQL = ttk.Entry(self.PostgresSQL_frame, font="Arial 11 ")
        self.query_PostgresSQL.grid(row=10,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_PostgresSQL = ttk.Button(
            self.PostgresSQL_frame,
            text='Start Query Execution',
            command=self.start_download_PostgresSQL,
            width=25, bootstyle='success'
            )

        self.executeQueryFileBtn_PostgresSQL.grid(row=11,column=1,pady=5,padx=5)
    


    def initiateUi_SSH(self):
        tk.Label(self.SSH_frame, text="Step 1 - Enter Connection Details: ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,padx=(100,0),sticky='w')
        tk.Label(self.SSH_frame, text="Unix Hostname : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=1,padx=(10,0))
        
        # tk.Label(self.SSH_frame, text="Hostname : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=2,padx=(10,0))
        tk.Label(self.SSH_frame, text="UserId : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=3,padx=(10,0))
        tk.Label(self.SSH_frame, text="Password : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=4,padx=(10,0))


        self.hostname_unix_ssh = ttk.Entry(self.SSH_frame, font="Arial 11 ")
        self.hostname_unix_ssh.grid(row=1,column=1,pady=5,ipadx=200)
        self.hostname_unix_ssh.insert(0,'bddvts2r4e2.wellpoint.com')
        
        
        
        
        self.uid_SSH = ttk.Entry(self.SSH_frame, font="Arial 11 ")
        self.uid_SSH.grid(row=3,column=1,pady=5,ipadx=200)
        
         
        
        self.pwd_SSH = ttk.Entry(self.SSH_frame,show='*', font="Arial 11 ")
        self.pwd_SSH.grid(row=4,column=1,pady=5,ipadx=200)
        

        


        self.connect_Btn_SSH = ttk.Button(
            self.SSH_frame,
            text='Establish SSH Connection',
            command=self.establish_connection_SSH,
            width=25, bootstyle='success'
            )

        self.connect_Btn_SSH.grid(row=5,column=1,pady=5,padx=5)

    def initiateUi_Hive(self):
        tk.Label(self.Hive_frame, text="Step 1 - Enter Connection Details: ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,padx=(100,0),sticky='w')
        tk.Label(self.Hive_frame, text="  ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=1000,padx=(100,0),sticky='w')
        tk.Label(self.Hive_frame, text="  ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=1001,padx=(100,0),sticky='w')
        tk.Label(self.Hive_frame, text="  ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=1002,padx=(100,0),sticky='w')
        tk.Label(self.Hive_frame, text="  ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=1003,padx=(100,0),sticky='w')
        tk.Label(self.Hive_frame, text="  ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=1004,padx=(100,0),sticky='w')
        tk.Label(self.Hive_frame, text="  ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=1005,padx=(100,0),sticky='w')
        tk.Label(self.Hive_frame, text="  ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=1006,padx=(100,0),sticky='w')
        tk.Label(self.Hive_frame, text="  ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=1007,padx=(100,0),sticky='w')
        tk.Label(self.Hive_frame, text="  ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=1008,padx=(100,0),sticky='w')
        tk.Label(self.Hive_frame, text="  ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=1009,padx=(100,0),sticky='w')
        tk.Label(self.Hive_frame, text="  ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=1010,padx=(100,0),sticky='w')
        tk.Label(self.Hive_frame, text="Unix Hostname : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=1,padx=(10,0))
        # checkbox for sqoop
        self.runFromImpala=IntVar()
        self.r0 = Checkbutton(self.Hive_frame, text = "Use Impala",variable=self.runFromImpala,bg="#0C7A79", font=("Arial", self.fontSize))
        self.r0.grid(row=2,column=2,padx=(70,0),pady=5)

        tk.Label(self.Hive_frame, text="Hostname : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=2,padx=(10,0))
        tk.Label(self.Hive_frame, text="UserId : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=3,padx=(10,0))
        tk.Label(self.Hive_frame, text="Password : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=4,padx=(10,0))


        self.hostname_unix = ttk.Entry(self.Hive_frame, font="Arial 11 ")
        self.hostname_unix.grid(row=1,column=1,pady=5,ipadx=200)
        self.hostname_unix.insert(0,'bddvts3e1v.wellpoint.com')
        # self.hostname_unix.insert(0,'bdpr3r6e1pr.wellpoint.com')
        
        self.hostname_Hive = ttk.Entry(self.Hive_frame, font="Arial 11 ")
        self.hostname_Hive.grid(row=2,column=1,pady=5,ipadx=200)
        self.hostname_Hive.insert(0,'bddvts3hs2lb')
        # self.hostname_Hive.insert(0,'bdpr3hs2lb')
        
        self.uid_Hive = ttk.Entry(self.Hive_frame, font="Arial 11 ")
        self.uid_Hive.grid(row=3,column=1,pady=5,ipadx=200)
         
        
        self.pwd_Hive = ttk.Entry(self.Hive_frame,show='*', font="Arial 11 ")
        self.pwd_Hive.grid(row=4,column=1,pady=5,ipadx=200)
        


        self.connect_Btn_Hive = ttk.Button(
            self.Hive_frame,
            text='Establish Hive Connection',
            command=self.establish_connection_Hive,
            width=25, bootstyle='success'
            )

        self.connect_Btn_Hive.grid(row=5,column=1,pady=5,padx=5)

        tk.Label(self.Hive_frame, text="Step 2 - Download from semicolon(;) separated query file : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=6,padx=(100,0),sticky='w')

        self.file_location_btn_Hive = ttk.Button(
            self.Hive_frame,
            text='Select Result Location ',
            command=self.select_download_location_Hive,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_Hive.grid(row=7,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_Hive = ttk.Button(
            self.Hive_frame,
            text='Select query file',
            command=self.select_query_file_Hive,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_Hive.grid(row=8,column=0,padx=(70,0),pady=5)

        
        
        
        self.download_path_Hive = ttk.Entry(self.Hive_frame, font="Arial 11 ")
        self.download_path_Hive.grid(row=7,column=1,pady=5,ipadx=200)
        

        self.query_Hive = ttk.Entry(self.Hive_frame, font="Arial 11 ")
        self.query_Hive.grid(row=8,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_Hive = ttk.Button(
            self.Hive_frame,
            text='Start Query Execution',
            command=self.start_download_Hive,
            width=25, bootstyle='success'
            )

        self.executeQueryFileBtn_Hive.grid(row=11,column=1,pady=5,padx=5)
        
        
        tk.Label(self.Hive_frame, text="Download tablename,datatype,maxlength of columns : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=12,padx=(100,0),sticky='w')

        self.file_location_btn_table_details_Hive = ttk.Button(
            self.Hive_frame,
            text='Select Result Location ',
            command=self.select_download_location_table_details_Hive,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_table_details_Hive.grid(row=13,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_table_details_Hive = ttk.Button(
            self.Hive_frame,
            text='Select Table names file',
            command=self.select_query_file_table_details_Hive,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_table_details_Hive.grid(row=14,column=0,padx=(70,0),pady=5)

        
        
        
        self.download_path_table_details_Hive = ttk.Entry(self.Hive_frame, font="Arial 11 ")
        self.download_path_table_details_Hive.grid(row=13,column=1,pady=5,ipadx=200)
        

        self.query_table_details_Hive = ttk.Entry(self.Hive_frame, font="Arial 11 ")
        self.query_table_details_Hive.grid(row=14,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_table_details_Hive = ttk.Button(
            self.Hive_frame,
            text='Get Table Details',
            command=self.getTableDetails_Hive,
            width=25, bootstyle='primary'
            )

        self.executeQueryFileBtn_table_details_Hive.grid(row=15,column=1,pady=5,padx=5)

        ### check db's in hive
        tk.Label(self.Hive_frame, text="Hive DB check : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=16,padx=(100,0),sticky='w')

        self.file_location_btn_find_db_Hive = ttk.Button(
            self.Hive_frame,
            text='Select Result Location ',
            command=lambda:self.select_folder(self.download_path_find_db_Hive),
            width=25, bootstyle='primary'
            )

        self.file_location_btn_find_db_Hive.grid(row=17,column=0,padx=(70,0),pady=5)
        
        self.download_path_find_db_Hive = ttk.Entry(self.Hive_frame, font="Arial 11 ")
        self.download_path_find_db_Hive.grid(row=17,column=1,pady=5,ipadx=200)


        
        # 
        self.queryFile_btn_find_db_Hive = ttk.Button(
            self.Hive_frame,
            text='Select Table names file(txt)',
            command=lambda:self.select_file("Table Names File",self.query_find_db_Hive),
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_find_db_Hive.grid(row=18,column=0,padx=(70,0),pady=5)

        
        self.query_find_db_Hive = ttk.Entry(self.Hive_frame, font="Arial 11 ")
        self.query_find_db_Hive.grid(row=18,column=1,pady=5,ipadx=200)

        self.DbNames_btn_find_db_Hive = ttk.Button(
            self.Hive_frame,
            text='Select DB names file(txt)',
            command=lambda:self.select_file("DB Names File",self.DbNames_find_db_Hive),
            width=25, bootstyle='primary'
            )

        self.DbNames_btn_find_db_Hive.grid(row=19,column=0,padx=(70,0),pady=5)

        
        self.DbNames_find_db_Hive = ttk.Entry(self.Hive_frame, font="Arial 11 ")
        self.DbNames_find_db_Hive.grid(row=19,column=1,pady=5,ipadx=200)

        
        self.Btn_find_db_Hive = ttk.Button(
            self.Hive_frame,
            text='Find DB',
            command=self.findHiveDb,
            width=25, bootstyle='primary'
            )

        self.Btn_find_db_Hive.grid(row=120,column=1,pady=5,padx=5)
    
    def findHiveDb(self):
        try:
            print("Finding Hive Db's..")
            result_path=str(self.download_path_find_db_Hive.get()).strip()
            os.chdir(result_path)

            table_Names_file_path=str(self.query_find_db_Hive.get()).strip()
            tableNames=''
            with open(table_Names_file_path,'r') as f:
                tableNames_txt=(f.read())
                tableNames=[str(tn).strip().lower() for tn in str(tableNames_txt).strip().split(',')]

            db_Names_file_path=str(self.DbNames_find_db_Hive.get()).strip()
            dbNames=''
            with open(db_Names_file_path,'r') as f:
                dbNames_txt=(f.read())

                dbNames=[str(dn).strip().lower() for dn in str(dbNames_txt).strip().split(',')]

            df=pd.DataFrame(columns=['TableName','DB in which table is present'])
            
            for table in tableNames:
                dbl=list()
                for db in dbNames:
                    try:
                        q='describe '+db+'.'+table
                        res=self.run_hive_query(q)
                        dbl.append(db)
                    except:pass
                df.loc[len(df.index)]=[table,",".join(dbl)]
            
            df.to_csv('Hive_DB_Identification_Result_'+str(time.time())+'.csv',index=False)
            print('Done')
            messagebox.showinfo('Done','DB identification  done.')
        except Exception as err:
            print(str(err))
            messagebox.showerror('Error',str(err))

        
        




    
    def initiateUi_Oracle(self):
        tk.Label(self.Oracle_frame, text="Step 1 - Enter Connection Details: ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,padx=(100,0),sticky='w')
        tk.Label(self.Oracle_frame, text="Driver : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=1,padx=(10,0))
        tk.Label(self.Oracle_frame, text="Mechanism Name : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=2,padx=(10,0))
        tk.Label(self.Oracle_frame, text="UserId : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=3,padx=(10,0))
        tk.Label(self.Oracle_frame, text="Password : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=4,padx=(10,0))
        
        
        #tk.Label(self, text="Download Path: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=8,padx=(70,0))
        
             
         
        
        self.driver_Oracle = ttk.Entry(self.Oracle_frame, font="Arial 11 ")
        self.driver_Oracle.grid(row=1,column=1,pady=5,ipadx=200)
        self.driver_Oracle.insert(0,'Teradata Database ODBC Driver 17.10')
        
        self.mechanism_name_Oracle = ttk.Entry(self.Oracle_frame, font="Arial 11 ")
        self.mechanism_name_Oracle.grid(row=2,column=1,pady=5,ipadx=200)
        self.mechanism_name_Oracle.insert(0,'DWPROD2COP1.CORP.ANTHEM.COM')
        
        self.uid_Oracle = ttk.Entry(self.Oracle_frame, font="Arial 11 ")
        self.uid_Oracle.grid(row=3,column=1,pady=5,ipadx=200)
        
        self.pwd_Oracle = ttk.Entry(self.Oracle_frame,show='*', font="Arial 11 ")
        self.pwd_Oracle.grid(row=4,column=1,pady=5,ipadx=200)
        
        
        
        
        self.connect_Btn_Oracle = ttk.Button(
            self.Oracle_frame,
            text='Establish Oracle Connection',
            command=self.establish_connection_Oracle,
            width=25, bootstyle='success'
            )

        self.connect_Btn_Oracle.grid(row=6,column=1,pady=5,padx=5)


        tk.Label(self.Oracle_frame, text="Enter Sqoop Command : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize+3)).grid(row=7,padx=(100,0))
        self.sqoop_command_oracle= ttk.Entry(self.Oracle_frame, font="Arial 11 ")
        self.sqoop_command_oracle.grid(row=7,column=1,pady=5,ipadx=200)

        self.connect_Btn_sqoop_Oracle = ttk.Button(
            self.Oracle_frame,
            text='Check Sqoop Connection',
            command=self.establish_connection_Oracle_sqoop,
            width=25, bootstyle='primary'
            )

        self.connect_Btn_sqoop_Oracle.grid(row=8,column=1,pady=5,padx=5)


        tk.Label(self.Oracle_frame, text="Step 2 - Download from semicolon(;) separated query file : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=9,padx=(100,0),sticky='w')
        
        
        self.download_path_Oracle = ttk.Entry(self.Oracle_frame, font="Arial 11 ")
        self.download_path_Oracle.grid(row=10,column=1,pady=5,ipadx=200)
        

        self.query_Oracle = ttk.Entry(self.Oracle_frame, font="Arial 11 ")
        self.query_Oracle.grid(row=11,column=1,pady=5,ipadx=200)

        self.file_location_btn_Oracle = ttk.Button(
            self.Oracle_frame,
            text='Select Result Location ',
            command=self.select_download_location_Oracle,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_Oracle.grid(row=10,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_Oracle = ttk.Button(
            self.Oracle_frame,
            text='Select query file',
            command=self.select_query_file_Oracle,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_Oracle.grid(row=11,column=0,padx=(70,0),pady=5)
         
       
        self.executeQueryFileBtn_Oracle = ttk.Button(
            self.Oracle_frame,
            text='Start Query Execution',
            command=self.start_download_Oracle,
            width=25, bootstyle='success'
            )

        self.executeQueryFileBtn_Oracle.grid(row=12,column=1,pady=5,padx=5)


        
        # tk.Label(self.Oracle_frame, text="Single Oracle Table Stats: ",bg="white",fg="red", font=("Arial", self.fontSize+3)).grid(row=11,padx=(100,0),sticky='w',pady=15)
        # self.tableStatsResulOracleownloadLocationEntryBtn_Oracle = ttk.Button(
        #     self.Oracle_frame,
        #         text='Select TableStats Result Location',
        #         command=self.select_TableStatsDownloadLocation_Oracle,
        #         width=25, bootstyle='primary'
        #         )

        # self.tableStatsResulOracleownloadLocationEntryBtn_Oracle.grid(row=12,column=0,pady=10,padx=(70,0))


        # self.tableStatsResultDownloadLocationEntry_Oracle = ttk.Entry(self.Oracle_frame, font="Arial 11 ")
        # self.tableStatsResultDownloadLocationEntry_Oracle.grid(row=12,column=1,pady=5,ipadx=200)

        # self.TableStatsFileBtn_Oracle = ttk.Button(
        #     self.Oracle_frame,
        #         text='Select Input Excel File',
        #         command=self.select_TableStatsFile_Oracle,
        #         width=25, bootstyle='primary'
        #         )

        # self.TableStatsFileBtn_Oracle.grid(row=13,column=0,pady=5,padx=(70,0))

            
        # self.startTableStatsCheckbtn_Oracle = ttk.Button(
        #         self.Oracle_frame,
        #         text='Start TableStats Checks',
        #         command=self.createRowDistinctFreqDistri_Oracle,
        #         width=25, bootstyle='primary'
        #         )

        # self.startTableStatsCheckbtn_Oracle.grid(row=14,column=1,pady=5,padx=5)
            
            
        # self.TableStatsFile_Oracle = ttk.Entry(self.Oracle_frame, font="Arial 11 ")
        # self.TableStatsFile_Oracle.grid(row=13,column=1,pady=5,ipadx=200)
            



    def initiateUi_mssql(self):
        tk.Label(self.mssql_frame, text="Step 1 - Enter Connection Details: ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,padx=(100,0),sticky='w')
        tk.Label(self.mssql_frame, text="Driver : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=1,padx=(10,0))
        tk.Label(self.mssql_frame, text="Server : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=2,padx=(10,0))
        tk.Label(self.mssql_frame, text="Database : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=3,padx=(10,0))
        tk.Label(self.mssql_frame, text="UID : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=4,padx=(10,0))
        tk.Label(self.mssql_frame, text="PWD : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=5,padx=(10,0))
        
        
        
        #tk.Label(self, text="Download Path: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=8,padx=(70,0))
        self.file_location_btn_mssql = ttk.Button(
            self.mssql_frame,
            text='Select Result Location ',
            command=self.select_download_location_mssql,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_mssql.grid(row=8,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_mssql = ttk.Button(
            self.mssql_frame,
            text='Select query file',
            command=self.select_query_file_mssql,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_mssql.grid(row=9,column=0,padx=(70,0),pady=5)
             
        
        self.driver_mssql = ttk.Entry(self.mssql_frame, font="Arial 11 ")
        self.driver_mssql.grid(row=1,column=1,pady=5,ipadx=200)
        self.driver_mssql.insert(0,'SQL Server Native Client 11.0')
        
        self.server_mssql = ttk.Entry(self.mssql_frame, font="Arial 11 ")
        self.server_mssql.grid(row=2,column=1,pady=5,ipadx=200)
        self.server_mssql.insert(0,'va33twvsql358.devad.wellpoint.com\sql01,10001')
        
        self.db_mssql = ttk.Entry(self.mssql_frame, font="Arial 11 ")
        self.db_mssql.grid(row=3,column=1,pady=5,ipadx=200)
        self.db_mssql.insert(0,'smartpcp')

        self.uid_mssql = ttk.Entry(self.mssql_frame, font="Arial 11 ")
        self.uid_mssql.grid(row=4,column=1,pady=5,ipadx=200)
        # self.db_mssql.insert(0,'smartpcp')

        self.pwd_mssql = ttk.Entry(self.mssql_frame,show='*', font="Arial 11 ")
        self.pwd_mssql.grid(row=5,column=1,pady=5,ipadx=200)
        # self.db_mssql.insert(0,'smartpcp')

             
        
        
        
        self.connect_Btn_mssql = ttk.Button(
            self.mssql_frame,
            text='Establish mssql Connection',
            command=self.establish_connection_mssql,
            width=25, bootstyle='success'
           )

        self.connect_Btn_mssql.grid(row=6,column=1,pady=5,padx=5)

        tk.Label(self.mssql_frame, text="Step 2 - Download from semicolon(;) separated query file : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=7,padx=(100,0),sticky='w')
        
        
        self.download_path_mssql = ttk.Entry(self.mssql_frame, font="Arial 11 ")
        self.download_path_mssql.grid(row=8,column=1,pady=5,ipadx=200)
        
        self.query_mssql = ttk.Entry(self.mssql_frame, font="Arial 11 ")
        self.query_mssql.grid(row=9,column=1,pady=5,ipadx=200)
         
       
        self.executeQueryFileBtn_mssql = ttk.Button(
            self.mssql_frame,
            text='Start Query Execution',
            command=self.start_download_mssql,
            width=25, bootstyle='success'
            )

        self.executeQueryFileBtn_mssql.grid(row=10,column=1,pady=5,padx=5)

    def initiateUi_TD(self):
        tk.Label(self.td_frame, text="Step 1 - Enter Connection Details: ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,padx=(100,0),sticky='w')
        tk.Label(self.td_frame, text="Driver : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=1,padx=(10,0))
        tk.Label(self.td_frame, text="Host Name : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=2,padx=(10,0))
        tk.Label(self.td_frame, text="User(Anthem id) : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=3,padx=(10,0))
        tk.Label(self.td_frame, text="Password : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=4,padx=(10,0))
        
        
        #tk.Label(self, text="Download Path: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=8,padx=(70,0))
        self.file_location_btn_TD = ttk.Button(
            self.td_frame,
            text='Select Result Location ',
            command=self.select_download_location_TD,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_TD.grid(row=8,column=0,padx=(70,0),pady=5)
        
        
        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_TD = ttk.Button(
            self.td_frame,
            text='Select query file',
            command=self.select_query_file_TD,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_TD.grid(row=9,column=0,padx=(70,0),pady=5)
             
        
        self.driver_td = ttk.Entry(self.td_frame, font="Arial 11 ")
        self.driver_td.grid(row=1,column=1,pady=5,ipadx=200)
        self.driver_td.insert(0,'Teradata Database ODBC Driver 17.10')
        
        self.hostname_td = ttk.Entry(self.td_frame, font="Arial 11 ")
        self.hostname_td.grid(row=2,column=1,pady=5,ipadx=200)
        self.hostname_td.insert(0,'DWPROD2COP1.CORP.ANTHEM.COM')
        # self.hostname_td.insert(0,'30.231.212.19')
        
        self.anthem_id_td = ttk.Entry(self.td_frame, font="Arial 11 ")
        self.anthem_id_td.grid(row=3,column=1,pady=5,ipadx=200)
        # self.anthem_id_td.insert(0,'AL17741')
        
        self.pwd_td = ttk.Entry(self.td_frame,show='*', font="Arial 11 ")
        self.pwd_td.grid(row=4,column=1,pady=5,ipadx=200)
    
        
        
        
        
        self.connect_Btn_td = ttk.Button(
            self.td_frame,
            text='Establish TD Connection',
            command=self.establish_connection_TD,
            width=25, bootstyle='success'
            )

        self.connect_Btn_td.grid(row=6,column=1,pady=5,padx=5)

        tk.Label(self.td_frame, text="Step 2 - Download from semicolon(;) separated query file : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=7,padx=(100,0),sticky='w')
        
        
        self.download_path_TD = ttk.Entry(self.td_frame, font="Arial 11 ")
        self.download_path_TD.grid(row=8,column=1,pady=5,ipadx=200)
        
        self.query_TD = ttk.Entry(self.td_frame, font="Arial 11 ")
        self.query_TD.grid(row=9,column=1,pady=5,ipadx=200)
         
       
        self.executeQueryFileBtn_TD = ttk.Button(
            self.td_frame,
            text='Start Query Execution',
           command=self.start_download_TD,
            width=25, bootstyle='success'
            )

        self.executeQueryFileBtn_TD.grid(row=10,column=1,pady=5,padx=5)


        
        # tk.Label(self.td_frame, text="Single TD Table Stats: ",bg="white",fg="red", font=("Arial", self.fontSize+3)).grid(row=11,padx=(100,0),sticky='w',pady=15)
        # self.tableStatsResultDownloadLocationEntryBtn_TD = ttk.Button(
        #     self.td_frame,
        #         text='Select TableStats Result Location',
        #         command=self.select_TableStatsDownloadLocation_TD,
        #         width=25, bootstyle='primary'
        #         )

        # self.tableStatsResultDownloadLocationEntryBtn_TD.grid(row=12,column=0,pady=10,padx=(70,0))

        # self.tableStatsResultDownloadLocationEntry_TD = ttk.Entry(self.td_frame, font="Arial 11 ")
        # self.tableStatsResultDownloadLocationEntry_TD.grid(row=12,column=1,pady=5,ipadx=200)

        # self.tableStatsResultDownloadLocationEntry_Oracle = ttk.Entry(self.Oracle_frame, font="Arial 11 ")
        # self.tableStatsResultDownloadLocationEntry_Oracle.grid(row=12,column=1,pady=5,ipadx=200)


        # self.TableStatsFileBtn_TD = ttk.Button(
        #     self.td_frame,
        #         text='Select Input Excel File',
        #         command=self.select_TableStatsFile_TD,
        #         width=25, bootstyle='primary'
        #         )

        # self.TableStatsFileBtn_TD.grid(row=13,column=0,pady=5,padx=(70,0))

            
        # self.startTableStatsCheckbtn_TD = ttk.Button(
        #         self.td_frame,
        #         text='Start TableStats Checks',
        #         command=self.createRowDistinctFreqDistri_TD,
        #         width=25, bootstyle='primary'
        #         )

        # self.startTableStatsCheckbtn_TD.grid(row=14,column=1,pady=5,padx=5)
            
            
        # self.TableStatsFile_TD = ttk.Entry(self.td_frame, font="Arial 11 ")
        # self.TableStatsFile_TD.grid(row=13,column=1,pady=5,ipadx=200)
            
    def reset_fields(self):
        self.bl.delete(0,END)
        self.hl.delete(0,END)
        self.sl.delete(0,END)
        self.resultFileName.delete(0,END)
        self.longi.delete(0,END)
        self.lati.delete(0,END)
        self.timest.delete(0,END)
        self.skipCols.delete(0,END)
        self.skipColsDuringFullCompare.delete(0,END)
        self.colsMapping.delete(0,END)
        self.pk.delete(0,END)
        print('Individual fields resetted..')
            
    def initiateUi_Individual(self):
            tk.Label(self.frameIndividual, text="Individual Table/File Comparison ",bg="white",fg="red", font=("Arial", self.fontSize+5)).grid(row=0,padx=(0,0),sticky='w')
            tk.Label(self.frameIndividual, text=" ",bg="white",fg="red", font=("Arial", self.fontSize+5)).grid(row=100,padx=(50,0),sticky='w')
            tk.Label(self.frameIndividual, text=" ",bg="white",fg="red", font=("Arial", self.fontSize+5)).grid(row=101,padx=(50,0),sticky='w')
            tk.Label(self.frameIndividual, text=" ",bg="white",fg="red", font=("Arial", self.fontSize+5)).grid(row=102,padx=(50,0),sticky='w')
            tk.Label(self.frameIndividual, text=" ",bg="white",fg="red", font=("Arial", self.fontSize+5)).grid(row=103,padx=(50,0),sticky='w')
            tk.Label(self.frameIndividual, text=" ",bg="white",fg="red", font=("Arial", self.fontSize+5)).grid(row=104,padx=(50,0),sticky='w')
            tk.Label(self.frameIndividual, text=" ",bg="white",fg="red", font=("Arial", self.fontSize+5)).grid(row=105,padx=(50,0),sticky='w')
            tk.Label(self.frameIndividual, text=" ",bg="white",fg="red", font=("Arial", self.fontSize+5)).grid(row=106,padx=(50,0),sticky='w')
            tk.Label(self.frameIndividual, text=" ",bg="white",fg="red", font=("Arial", self.fontSize+5)).grid(row=108,padx=(50,0),sticky='w')
            tk.Label(self.frameIndividual, text=" ",bg="white",fg="red", font=("Arial", self.fontSize+5)).grid(row=109,padx=(50,0),sticky='w')
            tk.Label(self.frameIndividual, text=" ",bg="white",fg="red", font=("Arial", self.fontSize+5)).grid(row=110,padx=(50,0),sticky='w')
            
            self.resetIndividualFields = ttk.Button(
            self.frameIndividual,
            text='Reset Fields',
            command=self.reset_fields,
            bootstyle="danger",
                # height= 1,
                  width=25,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'
            # height= 1, width=25, font=("Arial", self.fontSize+2), fg='red'
            )

            self.resetIndividualFields.grid(row=0,column=1,pady=5,padx=50)

            # self.v = Scrollbar(self.frameIndividual, orient='vertical')
            # self.v.grid(row=0,column=2)

            # self.frameIndividual_part1= Frame(self.frameIndividual, borderwidth=10, padx=5,pady=30)
            # self.frameIndividual_part1.grid(row=1,column=0,sticky='nw')

            # self.frameIndividual_part2= Frame(self.frameIndividual, borderwidth=10, padx=5,pady=30)
            # self.frameIndividual_part2.grid(row=2,column=1,sticky='ne')


            self.frame1 = tk.LabelFrame(self.frameIndividual, borderwidth=10, text='Step 1',padx=5,pady=5,width=1200,height=300)
            self.frame1.grid_propagate(0)
            self.frame2= tk.LabelFrame(self.frameIndividual, borderwidth=10, text='Step 2',padx=5,pady=5,width=1200,height=250)
            self.frame2.grid_propagate(0)
            
            self.frame3= tk.LabelFrame(self.frameIndividual, borderwidth=10, text='Step 3',padx=5,pady=5,width=1200,height=250)
            self.frame3.grid_propagate(0)
            self.frame4= tk.LabelFrame(self.frameIndividual, borderwidth=10, text='Step 4',padx=5,pady=5,width=550,height=550)
            self.frame4.grid_propagate(0)
            
            

            tk.Label(self.frame1, text="Column Delimiter: ",anchor='e',bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=5,padx=5,pady=5)
            tk.Label(self.frame1, text="Result File Name: ",anchor='e',bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=6,pady=5,padx=5)



            self.frame1.grid(row=1,column=0,sticky='nw')
            tk.Label(self.frame1, text="Enter File Details: ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,padx=(50,0))
            self.frame2.grid(row=1,column=0,sticky='sw')
            tk.Label(self.frame2, text="Field Mapping: ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,padx=(50,0))
            self.frame3.grid(row=3,column=0,sticky='w')
            tk.Label(self.frame3, text="Specify Latitude,Longitude,Timestamp,Skips columns(eg. Tokenized fields): ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,columnspan=2,padx=(50,0),sticky='w')
            self.frame4.grid(row=1,column=1,sticky='nw')
            tk.Label(self.frame4, text="Select tests to be done ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,padx=(50,0))

            


            
            tk.Label(self.frame3, text="Longitude : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=8,padx=(78,0))
            tk.Label(self.frame3, text="Latitude : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(84,0))
            tk.Label(self.frame3, text="Timestamp : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=10,padx=(70,0))
            tk.Label(self.frame3, text="Skip cols (Throughout) : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=11,padx=(79,0))
            tk.Label(self.frame3, text="Skip cols/Tokenized Fields (Full Comparison): ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=12,padx=(0,0))
            tk.Label(self.frame2, text="Column mapping(Source->Target):",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=14,column=0,padx=(0,0))
            # tk.Label(frame2, text="Source",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=14,column=1,padx=(0,0),sticky='w')
            # tk.Label(frame2, text="Target",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=14,column=3,padx=(0,0),sticky='w')

            # tk.Label(self, text="Source: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=14,padx=(0,0),sticky='w')
            # tk.Label(self, text="Target: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=15,padx=(0,0),sticky='w')

            



            self.bl = ttk.Entry(self.frame1, font="Arial 10 ")
            self.hl = ttk.Entry(self.frame1, font="Arial 10 ")
            self.sl = ttk.Entry(self.frame1, font="Arial 10 ")
            
            self.resultFileName = ttk.Entry(self.frame1, font="Arial 10 ")
            self.delimit=ttk.Entry(self.frame1, font="Arial 10 ")

            self.longi=ttk.Entry(self.frame3, font="Arial 10 ")
            self.lati=ttk.Entry(self.frame3, font="Arial 10 ")
            self.timest=ttk.Entry(self.frame3, font="Arial 10 ")
            self.skipCols=ttk.Entry(self.frame3, font="Arial 10 ")
            self.skipColsDuringFullCompare=ttk.Entry(self.frame3, font="Arial 10 ")
            self.colsMapping=ttk.Entry(self.frame2, font="Arial 10 ",state='')

            # tk.Label(self, text="",bg="#0C7A79").grid(row=0,column=1, padx= 50) 
            # tk.Label(self, text="",bg="#0C7A79").grid(row=1,column=1, padx= 50)
            self.bl.grid(row=2, column=1, padx= 50,pady=5,ipadx=250)
            self.hl.grid(row=3, column=1, padx= 50,pady=5,ipadx=250)
            self.sl.grid(row=4, column=1, padx= 50,pady=5,ipadx=250)
            self.delimit.grid(row=5,column=1, padx= 50,pady=5,ipadx=250) 
            self.delimit.delete(0,END)
            self.delimit.insert(0,',')
            
            self.resultFileName.grid(row=6, column=1, padx= 50,pady=5,ipadx=250)

            self.longi.grid(row=8,column=1, padx= 50,ipadx=150 )
            self.lati.grid(row=9,column=1, padx= 50,ipadx=150 )
            self.timest.grid(row=10,column=1 , padx= 50,ipadx=150)
            self.skipCols.grid(row=11,column=1,padx=50,ipadx=150)
            self.skipColsDuringFullCompare.grid(row=12,column=1,padx=50,ipadx=150)
            self.colsMapping.grid(row=15,column=1,columnspan=2,padx=(0,0),pady=5,ipadx=150)

            # tk.Label(self.frame4, text="Select tests to be done", bg="#0C7A79",fg="white",font=("Arial", self.fontSize)).grid(row=1,column=0, padx= 50,sticky='w')


            #primary key
            tk.Label(self.frame2, text="Primary key(s): ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=17,column=0,padx=(59,0))
            self.pk=ttk.Entry(self.frame2, font="Arial 10 ")
            self.pk.grid(row=17,column=1,columnspan=2, padx=60,pady=5,ipadx=150,sticky='w')
            #Create a dropdown Menu
            # primary_keys_drop= OptionMenu(self, menu,"C++", "Java","Python","JavaScript","Rust","GoLang")
            # primary_keys_drop.grid(row=6,column=2)
            self.primary_key_var = tk.StringVar()

            self.addPrimaryKeyColsBtn = ttk.Button(
            self.frame2,
            text='Add',
            command=self.add_primary_key_cols,
            #   font=("Arial", self.fontSize)
            bootstyle="primary",
                # height= 1,
                #   width=25,
            )    
                
            self.addPrimaryKeyColsBtn.grid(row=17,column=2,padx=(00,0),sticky = 'e')

            self.addAllCommonColsBtn = ttk.Button(
            self.frame2,
            text='Add all keys',
            command=self.add_all_common_cols,
                # height= 1, 
                width=11,
                #   font=("Arial", self.fontSize)
                bootstyle="primary"
            )    


            self.addAllCommonColsBtn.grid(row=17,column=3,padx=(10,5)) 

            #checkboxes


            self.commonCols_combo = ttk.Combobox(self.frame2, textvariable=self.primary_key_var,font = ('Arial', '10'))
            # commonCols_combo['state'] = 'readonly'
            self.commonCols_combo.grid(row=1,column=1,pady=5,padx=(0,0),ipadx=30)


            self.addExcelWithAllCases = ttk.Button(
            self.frameBatch,
            text='Select Input Excel File',
            command=self.startPoint,
            # height= 1,
              width=25,
              bootstyle="primary"
            #   font=("Arial", self.fontSize)
            )

            self.addExcelWithAllCases.grid(row=1,column=0,pady=5,padx=50)
        




            

            self.doAllChecks=IntVar()
            self.c0 = Checkbutton(self.frame4, text = "Check/Uncheck all tests",variable=self.doAllChecks,bg="#0C7A79", font=("Arial", self.fontSize),command=self.checkAll)
            self.c0.config(activebackground="#0C7A79")
            self.c0.config(activeforeground="white")
            self.c0.text_color="#000"

            self.c0.grid(row=2, column=0,sticky='w', padx= 50)



            self.doNullCheck=IntVar()
            self.c1 = Checkbutton(self.frame4, text = "Null check",variable=self.doNullCheck,bg="#0C7A79", font=("Arial", self.fontSize))
            self.c1.config(activebackground="#0C7A79")
            self.c1.config(activeforeground="white")
            self.c1.text_color="#000"

            self.c1.grid(row=4, column=0,sticky='w', padx= 50)

            self.doBlankCheck=IntVar()
            self.c2 = Checkbutton(self.frame4, text = "Blank Check",variable=self.doBlankCheck,bg="#0C7A79", font=("Arial", self.fontSize))
            self.c2.grid(row=5, column=0,sticky='w', padx= 50)

            self.doTimeStCheck=IntVar()
            c3= Checkbutton(self.frame4, text = "Timestamp check",variable=self.doTimeStCheck,bg="#0C7A79", font=("Arial", self.fontSize))
            c3.grid(row=6, column=0,sticky='w', padx= 50)

            self.doRowCountCheck=IntVar()
            c4 = Checkbutton(self.frame4, text = "Rowcount check",variable=self.doRowCountCheck,bg="#0C7A79", font=("Arial", self.fontSize))
            c4.grid(row=7, column=0,sticky='w', padx= 50)

            self.doDuplicateRowCheck=IntVar()
            self.c5 = Checkbutton(self.frame4, text = "Duplicate rows check",variable=self.doDuplicateRowCheck,bg="#0C7A79", font=("Arial", self.fontSize))
            self.c5.grid(row=8, column=0,sticky='w', padx= 50)

            self.doColNameCheck=IntVar()
            self.c6 = Checkbutton(self.frame4, text = "ColName(Source,Target) check",variable=self.doColNameCheck,bg="#0C7A79", font=("Arial", self.fontSize))
            self.c6.grid(row=9, column=0,sticky='w', padx= 50)

            self.doLatLongCheck=IntVar()
            self.c7 = Checkbutton(self.frame4, text = "Invalid Latitude,Longitude check",variable=self.doLatLongCheck,bg="#0C7A79", font=("Arial", self.fontSize))
            self.c7.grid(row=10, column=0,sticky='w', padx= 50)

            self.doCompareDataCheck=IntVar()
            self.c8 = Checkbutton(self.frame4, text = "Compare complete record",variable=self.doCompareDataCheck,bg="#0C7A79", font=("Arial", self.fontSize))
            self.c8.grid(row=11, column=0,sticky='w', padx= 50)

            self.doCaseInsensitiveCheck=IntVar()
            self.c8 = Checkbutton(self.frame4, text = "Case Insensitive",variable=self.doCaseInsensitiveCheck,bg="#0C7A79", font=("Arial", self.fontSize-1))
            self.c8.grid(row=11, column=1,sticky='e', padx= 0)

            self.doUniqueValueCountCheck=IntVar()
            self.c9 = Checkbutton(self.frame4, text = "Unique count check",variable=self.doUniqueValueCountCheck,bg="#0C7A79", font=("Arial", self.fontSize))
            self.c9.grid(row=12, column=0,sticky='w', padx= 50)

            self.doLeadingTrailingSpaceCheck=IntVar()
            self.c10 = Checkbutton(self.frame4, text = "Leading/Trailing space check",variable=self.doLeadingTrailingSpaceCheck,bg="#0C7A79", font=("Arial", self.fontSize))
            self.c10.grid(row=13, column=0,sticky='w', padx= 50)

            self.doCheckSumCheck=IntVar()
            self.c11 = Checkbutton(self.frame4, text = "SumOfValues",variable=self.doCheckSumCheck,bg="#0C7A79", font=("Arial", self.fontSize))
            self.c11.grid(row=14, column=0,sticky='w', padx= 50)

            #dropdown
            #Set the Menu initially
            # menu= StringVar()
            # menu.set("Select Any Language")





            self.unCommonCol_Source_var=tk.StringVar()
            self.unCommonCol_Source_combo= ttk.Combobox(self.frame2, textvariable=self.unCommonCol_Source_var,font = ('Arial', '10'))
            self.unCommonCol_Source_combo.grid(row=14,column=1,padx=0,ipadx=30)

            self.unCommonCol_Release_var=tk.StringVar()
            self.unCommonCol_Release_combo= ttk.Combobox(self.frame2, textvariable=self.unCommonCol_Release_var,font = ('Arial', '10'))
            self.unCommonCol_Release_combo.grid(row=14,column=2,pady=0,padx=(0,0),ipadx=30)

            self.commonCols_combo.bind('<KeyRelease>',self.search)


            self.Btn = ttk.Button(
            self,
            text='Get primary keys',
            command=self.getPrimaryKeys
            
            )     

            self.addCommonColsBtn = ttk.Button(
            self.frame2,
            text='Retrieve common columns',
            command=self.add_common_cols,
            # height= 1, width=21, font=("Arial", self.fontSize)
            # height= 1, 
                width=23,
                #   font=("Arial", self.fontSize)
                bootstyle="primary"
            )    


                
            self.addCommonColsBtn.grid(row=1,column=0,padx=(30,0),pady=0,sticky='e')    

            self.addLatitudeColsBtn = ttk.Button(
            self.frame3,
            text='Add',
            command=self.add_latitude_cols, 
            # font=("Arial", self.fontSize)
            # height= 1, 
                # width=10,
                #   font=("Arial", self.fontSize)
                bootstyle="primary"
            )    
                
            self.addLatitudeColsBtn.grid(row=9,column=2,padx=0,sticky = 'w')  

            self.addLongitudeColsBtn = ttk.Button(
            self.frame3,
            text='Add',
            command=self.add_longitude_cols,  bootstyle="primary"
            )    
                
            self.addLongitudeColsBtn.grid(row=8,column=2,padx=0,sticky = 'w')  

            self.addTimestampColsBtn = ttk.Button(
            self.frame3,
            text='Add',
            command=self.add_timestamp_cols,  bootstyle="primary"
            )    
                
            self.addTimestampColsBtn.grid(row=10,column=2,padx=0,sticky = 'w')      

            self.addSkipColsBtn = ttk.Button(
                self.frame3,
                text='Add',
                command=self.add_skip_cols,  bootstyle="primary"
            )    
                
            self.addSkipColsBtn.grid(row=11,column=2,padx=0,sticky = 'w')      

            self.addSkipColreleaseullCompareBtn = ttk.Button(
                self.frame3,
                text='Add',
                command=self.add_skipFullCompare_cols,  bootstyle="primary"
            )    
                
            self.addSkipColreleaseullCompareBtn.grid(row=12,column=2,padx=0,sticky = 'w') 

            self.addUncommonColMapping = ttk.Button(
                self.frame2,
               text='Add',
                command=self.add_col_map,  bootstyle="primary"
            )    
                
            self.addUncommonColMapping.grid(row=14,column=3,padx=0)   






            self.select_SourceBtn = ttk.Button(
                self.frame1,
                text='Select Result folder',
                command=self.select_Source_location,
                # height= 1,
                  width=20,  bootstyle="primary"
            )    
                
            self.select_SourceBtn.grid(row=2,column=0,padx=5,pady=5)
                

            self.SourceBtn = ttk.Button(
                self.frame1,
                text='Open Source file',
                command=self.open_Source_file,
                # height= 1,
                  width=20,  bootstyle="primary"
            )    
                
            self.SourceBtn.grid(row=3,column=0,padx=5,pady=5)

            self.releaseBtn = ttk.Button(
                self.frame1,
                text='Open Target file',
                command=self.open_release_file,
                # height= 1,
                  width=20,  bootstyle="primary"
            
            )    
                
            self.releaseBtn.grid(row=4,column=0,padx=5,pady=5)

            self.testBtn = ttk.Button(
                self.frame4,
                text='Start Individual Table Testing',
                command=self.initialize,
                # height= 1, 
                width=27,  bootstyle="success"
            )





            self.testBtn.grid(row=15, column=0,pady=20,padx=50)
            



    def initiateUi_Stats(self):
            tk.Label(self.frame_stats, text="Get TableStats: SumOfValues,Rowcount,DistinctCount,FrequencyDistribution    ",bg="white",fg="red", font=("Arial", self.fontSize+3)).grid(row=0,columnspan=2,padx=(0,0),pady=15)
            
            self.frame_stats_td_sf=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=50,pady=1)
            

            tk.Label(self.frame_stats_td_sf, text="TD-SF TableStats Comparison  (Preq: SF & TD connection) ",bg="white",fg="red", font=("Arial", self.fontSize+2)).grid(row=1,columnspan=1,padx=(0,0),pady=15)
            
            
            
            self.td_sf_template_btn=ttk.Button(
                self.frame_stats_td_sf,
                text=u'\u2193'+'Download Template',
                command=self.download_td_sf_stats_template,
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
            self.td_sf_template_btn.grid(row=1,column=1,pady=5,padx=5,sticky='e')


            self.frame_stats_sf_sf=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=50,pady=1)
            
            tk.Label(self.frame_stats_sf_sf, text="SF- SF TableStats Comparison (Preq: SF connection)      ",bg="white",fg="red", font=("Arial", self.fontSize+2)).grid(row=5,columnspan=1,padx=(0,0),pady=15)

            self.sf_sf_template_btn=ttk.Button(
                self.frame_stats_sf_sf,
                text=u'\u2193'+'Download Template',
                command=self.download_sf_sf_stats_template,
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
            self.sf_sf_template_btn.grid(row=5,column=1,pady=5,padx=5,sticky='e')


            self.TableStatsFileBtn_sf_to_sf = ttk.Button(
            self.frame_stats_sf_sf,
                text='Select Input Excel (SF-SF) File',
                command=self.select_TableStatsFile_sf_to_sf,
                # height= 1,
                  width=30,
                  bootstyle="primary",
                    # font=("Arial", self.fontSize)
                )

            self.TableStatsFileBtn_sf_to_sf.grid(row=7,column=0,pady=5,padx=5)

            self.startTableStatsCheckbtn_sf_to_sf = ttk.Button(
                    self.frame_stats_sf_sf,
                    text='Start TableStats(SF-SF) Checks',
                    command=self.createRowDistinctFreqDistri_sf_to_sf,
                    width=30,
                  bootstyle="success",
                    )

            self.startTableStatsCheckbtn_sf_to_sf.grid(row=8,column=1,pady=5,padx=5)

            self.TableStatsFile_sf_to_sf = ttk.Entry(self.frame_stats_sf_sf, font="Arial 11 ")
            self.TableStatsFile_sf_to_sf.grid(row=7,column=1,pady=5,ipadx=200)


            self.tableStatsResultDownloadLocationEntryBtn_sf_to_sf = ttk.Button(
            self.frame_stats_sf_sf,
                text='Select Result (SF-SF) Location',
                command=self.select_TableStatsDownloadLocation_sf_to_sf,
                width=30,
                  bootstyle="primary",
                )

            self.tableStatsResultDownloadLocationEntryBtn_sf_to_sf.grid(row=6,column=0,pady=10,padx=5)

            self.tableStatsResultDownloadLocationEntry_sf_to_sf = ttk.Entry(self.frame_stats_sf_sf, font="Arial 11 ")
            self.tableStatsResultDownloadLocationEntry_sf_to_sf.grid(row=6,column=1,pady=5,ipadx=200)
        


            self.tableStatsResultDownloadLocationEntryBtn = ttk.Button(
            self.frame_stats_td_sf,
                text='Select Result Location',
                command=self.select_TableStatsDownloadLocation,
                width=30,
                  bootstyle="primary"
                )

            self.tableStatsResultDownloadLocationEntryBtn.grid(row=2,column=0,pady=10,padx=5)

            self.tableStatsResultDownloadLocationEntry = ttk.Entry(self.frame_stats_td_sf, font="Arial 11 ")
            self.tableStatsResultDownloadLocationEntry.grid(row=2,column=1,pady=5,ipadx=200)

            

            self.TableStatsFileBtn = ttk.Button(
            self.frame_stats_td_sf,
                text='Select Input Excel File',
                command=self.select_TableStatsFile,
                width=30,
                  bootstyle="primary",
                )

            self.TableStatsFileBtn.grid(row=3,column=0,pady=5,padx=5)


            
            self.startTableStatsCheckbtn = ttk.Button(
                self.frame_stats_td_sf,
                text='Start TableStats (TD-SF) Checks',
                command=self.createRowDistinctFreqDistri,
                width=30,
                  bootstyle="success",
                )

            self.startTableStatsCheckbtn.grid(row=4,column=1,pady=5,padx=5)
            
            
            self.TableStatsFile = ttk.Entry(self.frame_stats_td_sf, font="Arial 11 ")
            self.TableStatsFile.grid(row=3,column=1,pady=5,ipadx=200)

            #hive_sf sanity check
            self.frame_stats_hive_sf_sanity=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=50,pady=1)
            
            tk.Label(self.frame_stats_hive_sf_sanity, text="Hive- SF PostProd Check[DDL,Rowcount,SumOfValues,Distinct] (Preq: SF connection)      ",bg="white",fg="red", font=("Arial", self.fontSize+2)).grid(row=9,columnspan=1,padx=(0,0),pady=15)

            self.sf_sf_template_btn=ttk.Button(
                self.frame_stats_hive_sf_sanity,
                text=u'\u2193'+'Download Template',
                command=self.download_hive_sf_sanity_check_template,
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
            self.sf_sf_template_btn.grid(row=9,column=1,pady=5,padx=5,sticky='e')

            self.sanity_checkResultDownloadLocationEntryBtn_hive_to_sf = ttk.Button(
            self.frame_stats_hive_sf_sanity,
                text='Select Result (Hive-SF) Location',
                command=self.select_sanity_checkDownloadLocation_hive_to_sf,
                width=30,
                  bootstyle="primary",
                )

            self.sanity_checkResultDownloadLocationEntryBtn_hive_to_sf.grid(row=10,column=0,pady=10,padx=5)

            self.sanity_checkResultDownloadLocationEntry_hive_to_sf = ttk.Entry(self.frame_stats_hive_sf_sanity, font="Arial 11 ")
            self.sanity_checkResultDownloadLocationEntry_hive_to_sf.grid(row=10,column=1,pady=5,ipadx=200)



            self.sanity_checkBtn_hive_to_sf = ttk.Button(
            self.frame_stats_hive_sf_sanity,
                text='Select Input Excel (Hive-SF) File',
                command=self.select_sanity_check_file_hive_to_sf,
                width=30,
                  bootstyle="primary",
                )

            self.sanity_checkBtn_hive_to_sf.grid(row=11,column=0,pady=5,padx=5)

            

            self.sanity_checkFile_hive_sf = ttk.Entry(self.frame_stats_hive_sf_sanity, font="Arial 11 ")
            self.sanity_checkFile_hive_sf.grid(row=11,column=1,pady=5,ipadx=200)


            
            self.startsanity_checkCheckbtn_hive_to_sf = ttk.Button(
                    self.frame_stats_hive_sf_sanity,
                    text='Start Hive-SF Check',
                    command=self.performSanityCheck_hive_sf,
                    width=30,
                  bootstyle="primary",
                    )

            self.startsanity_checkCheckbtn_hive_to_sf.grid(row=12,column=1,pady=5,padx=5)

            #mssql_sf sanity check
            self.frame_stats_mssql_sf_sanity=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=50,pady=1)
            
            tk.Label(self.frame_stats_mssql_sf_sanity, text="MSSQL- SF PostProd Check[DDL,Rowcount,SumOfValues,Distinct] (Preq: SF connection)      ",bg="white",fg="red", font=("Arial", self.fontSize+2)).grid(row=9,columnspan=1,padx=(0,0),pady=15)

            self.mssql_sf_template_btn=ttk.Button(
                self.frame_stats_mssql_sf_sanity,
                text=u'\u2193'+'Download Template',
                command=lambda: self.download_template('MSSQ-SF Sanity Check Template',['MSSQL_TableName','MSSQL_Condition','SF_TableName','SF_condition','DDL check(y/n)','Row count(y/n)','Distinct Check(y/n)','SumOfValues(y/n)']),
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
            self.mssql_sf_template_btn.grid(row=9,column=1,pady=5,padx=5,sticky='e')

            self.sanity_checkResultDownloadLocationEntryBtn_mssql_to_sf = ttk.Button(
            self.frame_stats_mssql_sf_sanity,
                text='Select Result (mssql-SF) Location',
                command=lambda: self.select_folder(self.sanity_checkResultDownloadLocationEntry_mssql_to_sf),
                width=30,
                  bootstyle="primary",
                )

            self.sanity_checkResultDownloadLocationEntryBtn_mssql_to_sf.grid(row=10,column=0,pady=10,padx=5)

            self.sanity_checkResultDownloadLocationEntry_mssql_to_sf = ttk.Entry(self.frame_stats_mssql_sf_sanity, font="Arial 11 ")
            self.sanity_checkResultDownloadLocationEntry_mssql_to_sf.grid(row=10,column=1,pady=5,ipadx=200)



            self.sanity_checkBtn_mssql_to_sf = ttk.Button(
            self.frame_stats_mssql_sf_sanity,
                text='Select Input Excel (mssql-SF) File',
                command=lambda: self.select_file('Input-File',self.sanity_checkFile_mssql_sf),
                width=30,
                  bootstyle="primary",
                )

            self.sanity_checkBtn_mssql_to_sf.grid(row=11,column=0,pady=5,padx=5)

            

            self.sanity_checkFile_mssql_sf = ttk.Entry(self.frame_stats_mssql_sf_sanity, font="Arial 11 ")
            self.sanity_checkFile_mssql_sf.grid(row=11,column=1,pady=5,ipadx=200)


            
            self.startsanity_checkCheckbtn_mssql_to_sf = ttk.Button(
                    self.frame_stats_mssql_sf_sanity,
                    text='Start mssql-SF Check',
                    command=self.performSanityCheck_mssql_sf,
                    width=30,
                  bootstyle="primary",
                    )

            self.startsanity_checkCheckbtn_mssql_to_sf.grid(row=12,column=1,pady=5,padx=5)
            
            ####
             #sf1_sf2 sanity check
            self.frame_stats_sf1_sf2_sanity=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=50,pady=1)
            
            tk.Label(self.frame_stats_sf1_sf2_sanity, text="SF-SF PostProd Check[DDL,Rowcount,SumOfValues,Distinct] (Preq: sf2 connection)      ",bg="white",fg="red", font=("Arial", self.fontSize+2)).grid(row=9,columnspan=1,padx=(0,0),pady=15)

            self.sf1_sf2_template_btn=ttk.Button(
                self.frame_stats_sf1_sf2_sanity,
                text=u'\u2193'+'Download Template',
                command=lambda: self.download_template('SF-SF Sanity Check Template',['sf1_TableName','sf1_Condition','sf2_TableName','sf2_condition','DDL check(y/n)','Row count(y/n)','Distinct Check(y/n)','SumOfValues(y/n)','Null count(y/n)']),
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
            self.sf1_sf2_template_btn.grid(row=9,column=1,pady=5,padx=5,sticky='e')

            self.sanity_checkResultDownloadLocationEntryBtn_sf1_to_sf2 = ttk.Button(
            self.frame_stats_sf1_sf2_sanity,
                text='Select Result (sf1-sf2) Location',
                command=lambda: self.select_folder(self.sanity_checkResultDownloadLocationEntry_sf1_to_sf2),
                width=30,
                  bootstyle="primary",
                )

            self.sanity_checkResultDownloadLocationEntryBtn_sf1_to_sf2.grid(row=10,column=0,pady=10,padx=5)

            self.sanity_checkResultDownloadLocationEntry_sf1_to_sf2 = ttk.Entry(self.frame_stats_sf1_sf2_sanity, font="Arial 11 ")
            self.sanity_checkResultDownloadLocationEntry_sf1_to_sf2.grid(row=10,column=1,pady=5,ipadx=200)



            self.sanity_checkBtn_sf1_to_sf2 = ttk.Button(
            self.frame_stats_sf1_sf2_sanity,
                text='Select Input Excel (SF-SF) File',
                command=lambda: self.select_file('Input-File',self.sanity_checkFile_sf1_sf2),
                width=30,
                  bootstyle="primary",
                )

            self.sanity_checkBtn_sf1_to_sf2.grid(row=11,column=0,pady=5,padx=5)

            

            self.sanity_checkFile_sf1_sf2 = ttk.Entry(self.frame_stats_sf1_sf2_sanity, font="Arial 11 ")
            self.sanity_checkFile_sf1_sf2.grid(row=11,column=1,pady=5,ipadx=200)


            
            self.startsanity_checkCheckbtn_sf1_to_sf2 = ttk.Button(
                    self.frame_stats_sf1_sf2_sanity,
                    text='Start sf1-sf2 Check',
                    command=self.performSanityCheck_sf1_sf2,
                    width=30,
                  bootstyle="primary",
                    )

            self.startsanity_checkCheckbtn_sf1_to_sf2.grid(row=12,column=1,pady=5,padx=5)

            ######
            #######################################
            #td_sf sanity check
            self.frame_stats_td_sf_sanity=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=50,pady=1)
            
            tk.Label(self.frame_stats_td_sf_sanity, text="TD- SF PostProd Check[DDL,Rowcount,SumOfValues,Distinct] (Preq: SF connection)      ",bg="white",fg="red", font=("Arial", self.fontSize+2)).grid(row=9,columnspan=1,padx=(0,0),pady=15)

            self.td_sf_template_btn=ttk.Button(
                self.frame_stats_td_sf_sanity,
                text=u'\u2193'+'Download Template',
                command=self.download_td_sf_sanity_check_template,
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
            self.td_sf_template_btn.grid(row=9,column=1,pady=5,padx=5,sticky='e')

            self.sanity_checkResultDownloadLocationEntryBtn_td_to_sf = ttk.Button(
            self.frame_stats_td_sf_sanity,
                text='Select Result (td-SF) Location',
                command=self.select_sanity_checkDownloadLocation_td_to_sf,
                width=30,
                  bootstyle="primary"
                )

            self.sanity_checkResultDownloadLocationEntryBtn_td_to_sf.grid(row=10,column=0,pady=10,padx=5)

            self.sanity_checkResultDownloadLocationEntry_td_to_sf = ttk.Entry(self.frame_stats_td_sf_sanity, font="Arial 11 ")
            self.sanity_checkResultDownloadLocationEntry_td_to_sf.grid(row=10,column=1,pady=5,ipadx=200)



            self.sanity_checkBtn_td_to_sf = ttk.Button(
            self.frame_stats_td_sf_sanity,
                text='Select Input Excel (td-SF) File',
                command=self.select_sanity_check_file_td_to_sf,
                width=30,
                  bootstyle="primary"
                )

            self.sanity_checkBtn_td_to_sf.grid(row=11,column=0,pady=5,padx=5)

            

            self.sanity_checkFile_td_sf = ttk.Entry(self.frame_stats_td_sf_sanity, font="Arial 11 ")
            self.sanity_checkFile_td_sf.grid(row=11,column=1,pady=5,ipadx=200)


            
            self.startsanity_checkCheckbtn_td_to_sf = ttk.Button(
                    self.frame_stats_td_sf_sanity,
                    text='Start td-SF Check',
                    command=self.performSanityCheck_td_sf,
                    width=30,
                  bootstyle="primary"
                    )

            self.startsanity_checkCheckbtn_td_to_sf.grid(row=12,column=1,pady=5,padx=5)

            ###################################

            self.frame_stats_mssql_sf=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=50,pady=1)
            
            
            tk.Label(self.frame_stats_mssql_sf, text="MSSQL- SF TableStats Comparison (Preq: SF & MSSQL connection)      ",bg="white",fg="red", font=("Arial", self.fontSize+2)).grid(row=13,columnspan=1,padx=(0,0),pady=15)

            self.mssql_sf_template_btn=ttk.Button(
                self.frame_stats_mssql_sf,
                text=u'\u2193'+'Download Template',
                command=self.download_mssql_sf_stats_template,
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
            self.mssql_sf_template_btn.grid(row=13,column=1,pady=5,padx=5,sticky='e')


            self.TableStatsFileBtn_mssql_to_sf = ttk.Button(
            self.frame_stats_mssql_sf,
                text='Select Input Excel (MSSQL-SF) File',
                command=self.select_TableStatsFile_mssql_to_sf,
                width=30,
                  bootstyle="primary"
                )

            self.TableStatsFileBtn_mssql_to_sf.grid(row=15,column=0,pady=5,padx=5)

            self.startTableStatsCheckbtn_mssql_to_sf = ttk.Button(
                    self.frame_stats_mssql_sf,
                    text='Start TableStats(MSSQL-SF) Checks',
                    command=self.createRowDistinctFreqDistri_mssql_to_sf,
                    width=30,
                  bootstyle="success"
                    )

            self.startTableStatsCheckbtn_mssql_to_sf.grid(row=16,column=1,pady=5,padx=5)

            self.TableStatsFile_mssql_to_sf = ttk.Entry(self.frame_stats_mssql_sf, font="Arial 11 ")
            self.TableStatsFile_mssql_to_sf.grid(row=15,column=1,pady=5,ipadx=200)


            self.tableStatsResultDownloadLocationEntryBtn_mssql_to_sf = ttk.Button(
            self.frame_stats_mssql_sf,
                text='Select Result (MSSQL-SF) Location',
                command=self.select_TableStatsDownloadLocation_mssql_to_sf,
                width=30,
                  bootstyle="primary"
                )

            self.tableStatsResultDownloadLocationEntryBtn_mssql_to_sf.grid(row=14,column=0,pady=10,padx=5)

            self.tableStatsResultDownloadLocationEntry_mssql_to_sf = ttk.Entry(self.frame_stats_mssql_sf, font="Arial 11 ")
            self.tableStatsResultDownloadLocationEntry_mssql_to_sf.grid(row=14,column=1,pady=5,ipadx=200)
        

            self.frame_stats_hive_sf=Frame(self.frame01, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=50,pady=1)
            
            
            tk.Label(self.frame_stats_hive_sf, text="Hive - SF TableStats Comparison (Preq: SF & hive connection)      ",bg="white",fg="red", font=("Arial", self.fontSize+2)).grid(row=13,columnspan=1,padx=(0,0),pady=15)

            self.hive_sf_template_btn=ttk.Button(
                self.frame_stats_hive_sf,
                text=u'\u2193'+'Download Template',
                command=self.download_hive_sf_stats_template,
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
            self.hive_sf_template_btn.grid(row=13,column=1,pady=5,padx=5,sticky='e')


            self.TableStatsFileBtn_hive_to_sf = ttk.Button(
            self.frame_stats_hive_sf,
                text='Select Input Excel (Hive-SF) File',
                command=self.select_TableStatsFile_hive_to_sf,
                width=30,
                  bootstyle="primary"
                )

            self.TableStatsFileBtn_hive_to_sf.grid(row=15,column=0,pady=5,padx=5)

            self.startTableStatsCheckbtn_hive_to_sf = ttk.Button(
                    self.frame_stats_hive_sf,
                    text='Start TableStats(Hive-SF) Checks',
                    command=self.createRowDistinctFreqDistri_hive_to_sf,
                    width=30,
                  bootstyle="success"
                    )

            self.startTableStatsCheckbtn_hive_to_sf.grid(row=16,column=1,pady=5,padx=5)

            self.TableStatsFile_hive_to_sf = ttk.Entry(self.frame_stats_hive_sf, font="Arial 11 ")
            self.TableStatsFile_hive_to_sf.grid(row=15,column=1,pady=5,ipadx=200)


            self.tableStatsResultDownloadLocationEntryBtn_hive_to_sf = ttk.Button(
            self.frame_stats_hive_sf,
                text='Select Result (hive-SF) Location',
                command=self.select_TableStatsDownloadLocation_hive_to_sf,
                width=30,
                  bootstyle="primary"
                )

            self.tableStatsResultDownloadLocationEntryBtn_hive_to_sf.grid(row=14,column=0,pady=10,padx=5)

            self.tableStatsResultDownloadLocationEntry_hive_to_sf = ttk.Entry(self.frame_stats_hive_sf, font="Arial 11 ")
            self.tableStatsResultDownloadLocationEntry_hive_to_sf.grid(row=14,column=1,pady=5,ipadx=200)
        



            

    def initiateUi_Batch(self):

            tk.Label(self.frameBatch, text="Batch Table/File Comparison ",bg="white",fg="red", font=("Arial", self.fontSize+5)).grid(row=0,padx=(0,0),pady=15)
            
            self.batch_template_btn=ttk.Button(
                self.frameBatch,
                text=u'\u2193'+'Download Template',
                command=self.download_batch_comparison_template,
                bootstyle="danger",
                # height= 1,
                  width=20,
                    # font=("Arial bold", self.fontSize),
                # fg='white',
                # bg='grey'

            )
            self.batch_template_btn.grid(row=0,column=1,pady=5,padx=5,sticky='e')

            self.selectBatchSourceFolderBtn = ttk.Button(
            self.frameBatch,
            text='Select Source Folder',
            command=self.select_batch_Source_folder_location,
            # height= 1, 
            bootstyle="primary",
            width=25
            # font=("Arial", self.fontSize)
            )

            self.selectBatchSourceFolderBtn.grid(row=2,column=0,pady=5,padx=50)
            self.batch_Source_entry = ttk.Entry(self.frameBatch, font="Arial 10 ")
            self.batch_Source_entry.grid(row=2,column=1,ipadx=220)

            # prefix source
            

            # tk.Label(self.frameBatch, text="Input Source Files Prefix ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize+1)).grid(row=3,column=0,pady=5,padx=50)
            
            # self.batch_SourcePrefix_entry = ttk.Entry(self.frameBatch, font="Arial 10 ")
            # self.batch_SourcePrefix_entry.grid(row=3,column=1,ipadx=220)

            # 
            

            self.selectBatchReleaseFolderBtn = ttk.Button(
            self.frameBatch,
            text='Select Target Folder',
            command=self.select_batch_release_folder_location,
            # height= 1, 
            width=25,
            bootstyle="primary"
            #  font=("Arial", self.fontSize)
            )

            self.selectBatchReleaseFolderBtn.grid(row=3,column=0,pady=5,padx=50)    

            self.batch_release_entry = ttk.Entry(self.frameBatch, font="Arial 10 ")
            self.batch_release_entry.grid(row=3,column=1,ipadx=220)
            # prefix target

            
            # tk.Label(self.frameBatch, text="Input Target Files Prefix ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize+1)).grid(row=5,column=0,pady=5,padx=50)    

             

            # self.batch_releasePrefix_entry = ttk.Entry(self.frameBatch, font="Arial 10 ")
            # self.batch_releasePrefix_entry.grid(row=5,column=1,ipadx=220)

            # 

            self.selectBatchResultFolderBtn = ttk.Button(
            self.frameBatch,
            text='Select Result Folder',
            command=self.select_batch_Result_folder_location,
            # height= 1,
              width=25, 
              bootstyle="primary"
            #   font=("Arial", self.fontSize)
            )

            self.selectBatchResultFolderBtn.grid(row=4,column=0,pady=5,padx=50)    

            self.batch_Result_entry = ttk.Entry(self.frameBatch, font="Arial 10 ")
            self.batch_Result_entry.grid(row=4,column=1,ipadx=220)


            
            self.batchExcelPath=ttk.Entry(self.frameBatch, font="Arial 13")
            self.batchExcelPath.grid(row=1,column=1, padx= 50,pady=5,ipadx=200)



            self.startExcelFileTestParsing = ttk.Button(
                self.frameBatch,
                text='Start Batch Testing',
                command=self.initializeAllFromExcel, 
                # relief=RIDGE,
                # height= 1, 
                width=25, 
                bootstyle="success"
                # font=("Arial", self.fontSize)
            )

            self.startExcelFileTestParsing.grid(row=10, column=1,pady=20,padx=50)

    def download_td_sf_stats_template(self):
        try:
            parameters=['TD_TableName','TD_Condition','SF_TableName','SF_condition','Columns For Measure Sum(comma separated)','TD Cols','SF Cols','Ignore Case for FreqDistri(y/n)','Ignore_Cols_During_Freq_Distri(Tok cols)','Top Occurring']
            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)
            df.to_excel('TD-SF_TableStats_Template'+'.xlsx',index=False)
            messagebox.showinfo('Downloaded','TD-SF_TableStats_Template'+str(time.time())+'.xlsx'+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))


    
    def download_hive_sf_sanity_check_template(self):
        try:
            parameters=['Hive_TableName','Hive_Condition','SF_TableName','SF_condition','DDL check(y/n)','Row count(y/n)','Distinct Check(y/n)','SumOfValues(y/n)','Null count(y/n)']
            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)
            df.to_excel('Hive-SF_Sanity_Check_Template'+'.xlsx',index=False)
            messagebox.showinfo('Downloaded','Hive-SF_Sanity_Check_Template'+'.xlsx'+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))

    def download_td_sf_sanity_check_template(self):
        try:
            parameters=['TD_TableName','TD_Condition','SF_TableName','SF_condition','DDL check(y/n)','Row count(y/n)','Distinct Check(y/n)','SumOfValues(y/n)']
            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)
            df.to_excel('TD-SF_Sanity_Check_Template'+'.xlsx',index=False)
            messagebox.showinfo('Downloaded','TD-SF_Sanity_Check_Template'+'.xlsx'+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))

    def download_sf_sf_stats_template(self):
        try:
            parameters=['SF1_TableName','SF1_Condition','SF2_TableName','SF2_condition','Columns For Measure Sum']
            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)
            df.to_excel('SF-SF_TableStats_Template'+'.xlsx',index=False)
            messagebox.showinfo('Downloaded','SF-SF_TableStats_Template'+'.xlsx'+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))
        
    def download_mssql_sf_stats_template(self):
        try:
            parameters=['MSSQL_TableName','MSSQL_Condition','SF_TableName','SF_condition','Columns For Measure Sum']
            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)
            df.to_excel('MSSQL-SF_TableStats_Template'+'.xlsx',index=False)
            messagebox.showinfo('Downloaded','MSSQL-SF_TableStats_Template'+'.xlsx'+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))

    def download_hive_sf_stats_template(self):
        try:
            parameters=['Hive_TableName','Hive_Condition','SF_TableName','SF_condition','Columns For Measure Sum']
            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)
            df.to_excel('Hive-SF_TableStats_Template'+'.xlsx',index=False)
            messagebox.showinfo('Downloaded','Hive-SF_TableStats_Template_'+str(time.time())+'.xlsx'+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))

    def download_batch_comparison_template(self):
        try:
            parameters=['Source File name','TargetFileName','Primary Key','Skip Cols(Completely ignore from all tests)','Tokenized(Ignore columns during full comparison)','Timestamp fields (col1,col2,col3...)','Longitude columns','Latitude columns','Column Mapping (Source_Col_Name->Target_Col_Name)','Complete Record Check(y/n)','Case Insensitive check','All Checks','Null Check','Blank Check','TimeStamp Check','RowCount Check','Duplicate Rows Check','Column Name check','Latitude/Longitude Check','Unique Count Check','Leading Trailing Space check','SumOfValues']

            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)
            filename='TravisTool_BatchComparison_Template'+'.xlsx'
            df.to_excel(filename,index=False)
            messagebox.showinfo('Downloaded',filename+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))

    

    def download_View_Validation_SF_comparison_template(self):
        try:
            parameters=['DB Name','TableName','Views_where_cast_fn_is_needed','Views_where_cast_fn_not_needed','Module']
            parameters_tab2=['Module','ColumnList']

            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)

            filename='View_Validation_SF_Template'+'.xlsx'

            # df.to_excel(filename,index=False)
            tab1=pd.DataFrame(columns=parameters)
            tab2=pd.DataFrame(columns=parameters_tab2)

            with pd.ExcelWriter(filename) as writer:
                    tab1.to_excel(writer,sheet_name='Tables',index = False,header=True) 
                    tab2.to_excel(writer,sheet_name='Module-Columns',index = False,header=True) 
                
            print('Done')
            messagebox.showinfo('Downloaded',filename+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))

    def download_template(self,title,parameters):
        try:
           
            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)
            filename=title+'.xlsx'
            df.to_excel(filename,index=False)
            messagebox.showinfo('Downloaded',filename+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))


    def download_hive_sf_datatype_comparison_template(self):
        try:
            parameters=['Hive_TableName','SF_TableName']

            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)
            filename='Hive-SF_Datatype_Comparison_Template'+'.xlsx'
            df.to_excel(filename,index=False)
            messagebox.showinfo('Downloaded',filename+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))

    def download_tokenization_check_template(self):
        try:
            parameters=['Source','TargetTable(FullName)','Toknzn in SIT','Toknzn in Prod','App Cd']

            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)
            filename='Tokenization_Check_Template'+'.xlsx'
            df.to_excel(filename,index=False)
            messagebox.showinfo('Downloaded',filename+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))

    def download_td_sf_datatype_comparison_template(self):
        try:
            parameters=['TD_TableName','SF_TableName']

            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)
            filename='TD-SF_Datatype_Comparison_Template'+'.xlsx'
            df.to_excel(filename,index=False)
            messagebox.showinfo('Downloaded',filename+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))

    def download_mssql_sf_datatype_comparison_template(self):
        try:
            parameters=['MS-SQL_TableName','SF_TableName']

            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)
            filename='MS-SQL-SF_Datatype_Comparison_Template'+'.xlsx'
            df.to_excel(filename,index=False)
            messagebox.showinfo('Downloaded',filename+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))

    def download_sf_sf_datatype_comparison_template(self):
        try:
            parameters=['SF1_TableName','SF2_TableName']

            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)
            filename='SF-SF_Datatype_Comparison_Template'+'.xlsx'
            df.to_excel(filename,index=False)
            messagebox.showinfo('Downloaded',filename+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))

    def download_mssql_td_datatype_comparison_template(self):
        try:
            parameters=['MS-SQL_TableName','TD_TableName']

            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)
            filename='MS-SQL-TD_Datatype_Comparison_Template'+'.xlsx'
            df.to_excel(filename,index=False)
            messagebox.showinfo('Downloaded',filename+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))


    def download_sf_sf_schema_ddl_comparison_template(self):
        try:
            parameters=['Base schema(DDL repository) full path'              ,'Table','DB','Schemas to check']

            df=pd.DataFrame(columns=parameters)
            pathname=os.path.join(os.path.expandvars("%userprofile%"),"Downloads")
            os.chdir(pathname)
            filename='SF-SF_Schema_DDL_Comparison_Template'+'.xlsx'
            df.to_excel(filename,index=False)
            messagebox.showinfo('Downloaded',filename+' downloaded in '+pathname)
        except Exception as err:
            messagebox.showerror('Error',str(err))

    def initiateUi_SF(self):
        tk.Label(self.sf_frame, text="Step 1 - Enter Connection Details: ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,padx=(0,0))
        tk.Label(self.sf_frame, text="HA id : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=1,padx=(70,0))
        # tk.Label(self.sf_frame, text="Database : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=3,padx=(84,0))
        tk.Label(self.sf_frame, text="Warehouse : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=2,padx=(103,0))
        # tk.Label(self.sf_frame, text="Schema : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=4,padx=(84,0))
        # tk.Label(self.sf_frame, text="Role : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=5,padx=(70,0))
        tk.Label(self.sf_frame, text="Account(base url) : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=6,padx=(135,0))

        self.ha_id = ttk.Entry(self.sf_frame, font="Arial 11 ")
        self.ha_id.grid(row=1,column=1,pady=5,ipadx=200)
        # self.ha_id.insert(0,'AN461447AD')
        self.ha_id.insert(0,'AN536766AD')
        
        # self.database_SF = ttk.Entry(self.sf_frame, font="Arial 11 ")
        # self.database_SF.grid(row=3,column=1,pady=5,ipadx=200)
        
        self.warehouse_SF = ttk.Entry(self.sf_frame, font="Arial 11 ")
        self.warehouse_SF.grid(row=2,column=1,pady=5,ipadx=200)
        # self.warehouse_SF.insert(0,'T01_VBC_FT_USER_WH_L')
        self.warehouse_SF.insert(0,'T01_CHI_USER_WH_M')
        # self.schema_SF = ttk.Entry(self.sf_frame, font="Arial 11 ")
        # self.schema_SF.grid(row=4,column=1,pady=5,ipadx=200)
        
        # self.role_SF = ttk.Entry(self.sf_frame, font="Arial 11 ")
        # self.role_SF.grid(row=5,column=1,pady=5,ipadx=200)
         
        
        self.account_SF = ttk.Entry(self.sf_frame, font="Arial 11 ")
        self.account_SF.grid(row=6,column=1,pady=5,ipadx=200)
        self.account_SF.insert(0,'carelon-eda_nonprod.privatelink')
        
        
        tk.Label(self.sf_frame, text="Step 2 - Download from semicolon(;) separated query file : ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=9,padx=(100,0),sticky='w')
        #tk.Label(self, text="Download Path: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=8,padx=(70,0))
        self.file_location_btn_SF = ttk.Button(
            self.sf_frame,
            text='Select Result Location ',
            command=self.select_download_location_SF,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_SF.grid(row=10,column=0,padx=(70,0),pady=5)

        
        #tk.Label(self, text="Query: ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=9,padx=(70,0))
        self.queryFile_btn_SF = ttk.Button(
            self.sf_frame,
            text='Select query file',
            command=self.select_query_file_SF,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_SF.grid(row=11,column=0,padx=(70,0),pady=5)
             
        
        
        self.download_path_SF = ttk.Entry(self.sf_frame, font="Arial 11 ")
        self.download_path_SF.grid(row=10,column=1,pady=5,ipadx=200)
        
        
        
        self.query_SF = ttk.Entry(self.sf_frame, font="Arial 11 ")
        self.query_SF.grid(row=11,column=1,pady=5,ipadx=200)
        
        self.downloadBtn_SF = ttk.Button(
            self.sf_frame,
            text='Establish SF Connection',
            command=self.establish_connection_SF,
            width=25, bootstyle='success'
            )

        self.downloadBtn_SF.grid(row=8,column=1,pady=5,padx=5)
        
        
        
       
        self.executeQueryFileBtn_SF = ttk.Button(
            self.sf_frame,
            text='Start Query Execution',
            command=self.start_download_SF,
            width=20, bootstyle='success'
            )

        self.executeQueryFileBtn_SF.grid(row=12,column=1,pady=5,padx=5)


        #timestamp fields
        tk.Label(self.sf_frame, text="Get Timestamp from comma(,) separated table names file: ",bg="white",fg="red", font=("Arial", self.fontSize+3)).grid(row=14,padx=(100,0),pady=5,sticky='w')
        
        self.file_location_btn_SF_timestamps = ttk.Button(
            self.sf_frame,
            text='Select Result(TS) Location ',
            command=self.select_download_location_SF_timestamps,
            width=25, bootstyle='primary'
            )

        self.file_location_btn_SF_timestamps.grid(row=15,column=0,padx=(70,0))
        
        self.download_path_SF_timestamps = ttk.Entry(self.sf_frame, font="Arial 11 ")
        self.download_path_SF_timestamps.grid(row=15,column=1,pady=5,ipadx=200)
        
        self.queryFile_btn_SF_timestamps = ttk.Button(
            self.sf_frame,
            text='Select ts query file',
            command=self.select_query_file_SF_timestamps,
            width=25, bootstyle='primary'
            )

        self.queryFile_btn_SF_timestamps.grid(row=16,column=0,padx=(70,0),pady=5)

        self.query_SF_timestamps = ttk.Entry(self.sf_frame, font="Arial 11 ")
        self.query_SF_timestamps.grid(row=16,column=1,pady=5,ipadx=200)

        self.executeQueryFileBtn_SF_timestamps = ttk.Button(
            self.sf_frame,
            text='Get Timestamps',
            command=self.getTimestampColumns,
            width=20, bootstyle='success'
            )

        self.executeQueryFileBtn_SF_timestamps.grid(row=17,column=1,pady=5,padx=5)
          

    def initiateUi_SF_prod(self):
        tk.Label(self.sf_prod_frame, text="Step 1 - Enter Connection Details (Preq. Connect to  Dev Env First if running for first time in day): ",bg="#0C7A79",fg="yellow", font=("Arial", self.fontSize+3)).grid(row=0,columnspan=2,padx=(0,0))
        tk.Label(self.sf_prod_frame, text="HA id : ",bg="#0C7A79",fg="white",anchor='e', font=("Arial", self.fontSize)).grid(row=1,padx=(20,0))
        # tk.Label(self.sf_prod_frame, text="Database : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=3,padx=(84,0))
        tk.Label(self.sf_prod_frame, text="Warehouse : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=2,padx=(50,0))
        # tk.Label(self.sf_prod_frame, text="Schema : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=4,padx=(84,0))
        # tk.Label(self.sf_prod_frame, text="Role : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=5,padx=(70,0))
        tk.Label(self.sf_prod_frame, text="Account(base url) : ",bg="#0C7A79",fg="white", font=("Arial", self.fontSize)).grid(row=6,padx=(70,0))

        self.ha_id_prod = ttk.Entry(self.sf_prod_frame, font="Arial 11 ")
        self.ha_id_prod.grid(row=1,column=1,pady=5,ipadx=200)
        self.ha_id_prod.insert(0,'AN536766AD')
        # self.database_SF_prod = ttk.Entry(self.sf_prod_frame, font="Arial 11 ")
        # self.database_SF_prod.grid(row=3,column=1,pady=5,ipadx=200)
        
        self.warehouse_SF_prod = ttk.Entry(self.sf_prod_frame, font="Arial 11 ")
        self.warehouse_SF_prod.grid(row=2,column=1,pady=5,ipadx=200)
        self.warehouse_SF_prod.insert(0,'p01_uwrate_user_wh_m')
        # self.schema_SF_prod = ttk.Entry(self.sf_prod_frame, font="Arial 11 ")
        # self.schema_SF_prod.grid(row=4,column=1,pady=5,ipadx=200)
        
        # self.role_SF_prod = ttk.Entry(self.sf_prod_frame, font="Arial 11 ")
        # self.role_SF_prod.grid(row=5,column=1,pady=5,ipadx=200)
         
        
        self.account_SF_prod = ttk.Entry(self.sf_prod_frame, font="Arial 11 ")
        self.account_SF_prod.grid(row=6,column=1,pady=5,ipadx=200)
        self.account_SF_prod.insert(0,'carelon-edaprod.privatelink')
        
        
        
        self.downloadBtn_SF = ttk.Button(
            self.sf_prod_frame,
            text='Establish Prod-SF Connection',
            command=self.establish_connection_SF_prod,
            width=25, bootstyle='success'
            )

        self.downloadBtn_SF.grid(row=8,column=1,pady=5,padx=5)
        
        



           
        
    def select_query_file_SF(self):
        query_file_name= fd.askopenfilename(title='Select Query file') 
        self.query_SF.delete(0,END)
        self.query_SF.insert(0,query_file_name)
        self.query_SF.focus()
        self.query_SF.icursor(END)
        print('SF Query File selected')


    def select_query_file_PostgresSQL(self):
            query_file_name= fd.askopenfilename(title='Select Query file') 
            self.query_PostgresSQL.delete(0,END)
            self.query_PostgresSQL.insert(0,query_file_name)
            self.query_PostgresSQL.focus()
            self.query_PostgresSQL.icursor(END)
            print('PostgresSQL Query File selected')


    def select_query_file_Hive(self):
        query_file_name= fd.askopenfilename(title='Select Query file') 
        self.query_Hive.delete(0,END)
        self.query_Hive.insert(0,query_file_name)
        self.query_Hive.focus()
        self.query_Hive.icursor(END)
        print('Hive Query File selected')

        # SqoopBatch
    def select_query_file_SqoopBatch(self):
        query_file_name= fd.askopenfilename(title='Select Query file') 
        self.query_SqoopBatch.delete(0,END)
        self.query_SqoopBatch.insert(0,query_file_name)
        self.query_SqoopBatch.focus()
        self.query_SqoopBatch.icursor(END)
        print('SqoopBatch Query File selected')

    def select_query_file_table_details_Hive(self):
        query_file_name= fd.askopenfilename(title='Select Table names file') 
        self.query_table_details_Hive.delete(0,END)
        self.query_table_details_Hive.insert(0,query_file_name)
        self.query_table_details_Hive.focus()
        self.query_table_details_Hive.icursor(END)
        print('Hive Query File selected')

    def select_query_file_table_details_View_Validation_SF(self):
        query_file_name= fd.askopenfilename(title='Select Template file') 
        self.query_excel_file_View_Validation_SF.delete(0,END)
        self.query_excel_file_View_Validation_SF.insert(0,query_file_name)
        self.query_excel_file_View_Validation_SF.focus()
        self.query_excel_file_View_Validation_SF.icursor(END)
        print('View_Validation_SF Query File selected')

        # datatype_hive_sf

    def select_query_file_table_details_datatype_hive_sf(self):
        query_file_name= fd.askopenfilename(title='Select Hive-SF Table Mapping Excel file') 
        self.query_excel_file_datatype_hive_sf.delete(0,END)
        self.query_excel_file_datatype_hive_sf.insert(0,query_file_name)
        self.query_excel_file_datatype_hive_sf.focus()
        self.query_excel_file_datatype_hive_sf.icursor(END)
        print('datatype_hive_sf Query File selected')

    def select_query_file_table_details_datatype_sf_sf(self):
        query_file_name= fd.askopenfilename(title='Select SF-SF Table Mapping Excel file') 
        self.query_excel_file_datatype_sf_sf.delete(0,END)
        self.query_excel_file_datatype_sf_sf.insert(0,query_file_name)
        self.query_excel_file_datatype_sf_sf.focus()
        self.query_excel_file_datatype_sf_sf.icursor(END)
        print('datatype_sf_sf Query File selected')

    def select_query_file_table_details_tokenization_check(self):
        query_file_name= fd.askopenfilename(title='Select Tokenization Table Mapping Excel file') 
        self.query_excel_file_tokenization_check.delete(0,END)
        self.query_excel_file_tokenization_check.insert(0,query_file_name)
        self.query_excel_file_tokenization_check.focus()
        self.query_excel_file_tokenization_check.icursor(END)
        print('tokenization_check Query File selected')

    def select_query_file_table_details_datatype_td_sf(self):
        query_file_name= fd.askopenfilename(title='Select TD-SF Table Mapping Excel file') 
        self.query_excel_file_datatype_td_sf.delete(0,END)
        self.query_excel_file_datatype_td_sf.insert(0,query_file_name)
        self.query_excel_file_datatype_td_sf.focus()
        self.query_excel_file_datatype_td_sf.icursor(END)
        print('datatype_td_sf Query File selected')

    def select_query_file_table_details_datatype_mssql_sf(self):
        query_file_name= fd.askopenfilename(title='Select MS-SQL-SF Table Mapping Excel file') 
        self.query_excel_file_datatype_mssql_sf.delete(0,END)
        self.query_excel_file_datatype_mssql_sf.insert(0,query_file_name)
        self.query_excel_file_datatype_mssql_sf.focus()
        self.query_excel_file_datatype_mssql_sf.icursor(END)
        print('datatype_mssql_sf Query File selected')


    def select_query_file_table_details_datatype_mssql_td(self):
        query_file_name= fd.askopenfilename(title='Select MS-SQL-TeraData Table Mapping Excel file') 
        self.query_excel_file_datatype_mssql_td.delete(0,END)
        self.query_excel_file_datatype_mssql_td.insert(0,query_file_name)
        self.query_excel_file_datatype_mssql_td.focus()
        self.query_excel_file_datatype_mssql_td.icursor(END)
        print('datatype_mssql_td Query File selected')

    def select_query_file_table_details_sf_sf_schema_ddl(self):
        query_file_name= fd.askopenfilename(title='Select Hive-SF Table Mapping Excel file') 
        self.query_excel_file_sf_sf_schema_ddl.delete(0,END)
        self.query_excel_file_sf_sf_schema_ddl.insert(0,query_file_name)
        self.query_excel_file_sf_sf_schema_ddl.focus()

        self.query_excel_file_sf_sf_schema_ddl.icursor(END)
        print('sf_sf_schema_ddl Query File selected')


    def select_query_file_MySQLWorkBench(self):
        query_file_name= fd.askopenfilename(title='Select Query file') 
        self.query_MySQLWorkBench.delete(0,END)
        self.query_MySQLWorkBench.insert(0,query_file_name)
        self.query_MySQLWorkBench.focus()
        self.query_MySQLWorkBench.icursor(END)
        print('MySQLWorkBench Query File selected')

        
    def select_query_file_SF_timestamps(self):
        query_file_name= fd.askopenfilename(title='Select Query file') 
        self.query_SF_timestamps.delete(0,END)
        self.query_SF_timestamps.insert(0,query_file_name)
        self.query_SF_timestamps.focus()
        self.query_SF_timestamps.icursor(END)
        print('SF Timestamp File selected')
        
    def select_TableStatsFile(self):
        query_file_name= fd.askopenfilename(title='Select TableStats file') 
        self.TableStatsFile.delete(0,END)
        self.TableStatsFile.insert(0,query_file_name)
        self.TableStatsFile.focus()
        self.TableStatsFile.icursor(END)
        print('TableStats  File selected')

        # select_sanity_check_file_hive_to_sf

    def select_sanity_check_file_hive_to_sf(self):
        query_file_name= fd.askopenfilename(title='Select sanity_check file') 
        self.sanity_checkFile_hive_sf.delete(0,END)
        self.sanity_checkFile_hive_sf.insert(0,query_file_name)
        self.sanity_checkFile_hive_sf.focus()
        self.sanity_checkFile_hive_sf.icursor(END)
        print('sanity_check  File selected')

    def select_sanity_check_file_td_to_sf(self):
        query_file_name= fd.askopenfilename(title='Select sanity_check file') 
        self.sanity_checkFile_td_sf.delete(0,END)
        self.sanity_checkFile_td_sf.insert(0,query_file_name)
        self.sanity_checkFile_td_sf.focus()
        self.sanity_checkFile_td_sf.icursor(END)
        print('sanity_check  File selected')

    def select_TableStatsFile_TD(self):
        query_file_name= fd.askopenfilename(title='Select TD TableStats file') 
        self.TableStatsFile_TD.delete(0,END)
        self.TableStatsFile_TD.insert(0,query_file_name)
        self.TableStatsFile_TD.focus()
        self.TableStatsFile_TD.icursor(END)
        print('TableStats  File selected')


    def select_TableStatsFile_Oracle(self):
        query_file_name= fd.askopenfilename(title='Select Oracle TableStats file') 
        self.TableStatsFile_Oracle.delete(0,END)
        self.TableStatsFile_Oracle.insert(0,query_file_name)
        self.TableStatsFile_Oracle.focus()
        self.TableStatsFile_Oracle.icursor(END)
        print('TableStats  File selected')
        
    def select_TableStatsFile_sf_to_sf(self):
        query_file_name= fd.askopenfilename(title='Select sf_to_sf TableStats file') 
        self.TableStatsFile_sf_to_sf.delete(0,END)
        self.TableStatsFile_sf_to_sf.insert(0,query_file_name)
        self.TableStatsFile_sf_to_sf.focus()
        self.TableStatsFile_sf_to_sf.icursor(END)
        print('TableStats  File selected')
        
    def select_TableStatsFile_mssql_to_sf(self):
        query_file_name= fd.askopenfilename(title='Select mssql_to_sf TableStats file') 
        self.TableStatsFile_mssql_to_sf.delete(0,END)
        self.TableStatsFile_mssql_to_sf.insert(0,query_file_name)
        self.TableStatsFile_mssql_to_sf.focus()
        self.TableStatsFile_mssql_to_sf.icursor(END)
        print('TableStats  File selected')

    def select_TableStatsFile_hive_to_sf(self):
        query_file_name= fd.askopenfilename(title='Select hive_to_sf TableStats file') 
        self.TableStatsFile_hive_to_sf.delete(0,END)
        self.TableStatsFile_hive_to_sf.insert(0,query_file_name)
        self.TableStatsFile_hive_to_sf.focus()
        self.TableStatsFile_hive_to_sf.icursor(END)
        print('TableStats  File selected')
          
    def select_query_file_TD(self):
        query_file_name= fd.askopenfilename(title='Select Query file') 
        self.query_TD.delete(0,END)
        self.query_TD.insert(0,query_file_name)
        self.query_TD.focus()
        self.query_TD.icursor(END)
        print('TD Query File selected')

    def select_query_file_mssql(self):
        query_file_name= fd.askopenfilename(title='Select Query file') 
        self.query_mssql.delete(0,END)
        self.query_mssql.insert(0,query_file_name)
        self.query_mssql.focus()
        self.query_mssql.icursor(END)
        print('mssql Query File selected')

    def select_query_file_Oracle(self):
        query_file_name= fd.askopenfilename(title='Select Oracle Query file') 
        self.query_Oracle.delete(0,END)
        self.query_Oracle.insert(0,query_file_name)
        self.query_Oracle.focus()
        self.query_Oracle.icursor(END)
        print('Oracle Query File selected')

    def select_download_location_Oracle(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_Oracle.delete(0,END)
        self.download_path_Oracle.insert(0,download_location_adrs)
        self.download_path_Oracle.focus()
        self.download_path_Oracle.icursor(END)
        print('Oracle Result location selected')

# SqoopBatch
    def select_download_location_SqoopBatch(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_SqoopBatch.delete(0,END)
        self.download_path_SqoopBatch.insert(0,download_location_adrs)
        self.download_path_SqoopBatch.focus()
        self.download_path_SqoopBatch.icursor(END)
        print('SqoopBatch Result location selected')

        # datatype_hive_sf


    def select_download_location_datatype_hive_sf(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_datatype_hive_sf.delete(0,END)
        self.download_path_datatype_hive_sf.insert(0,download_location_adrs)
        self.download_path_datatype_hive_sf.focus()
        self.download_path_datatype_hive_sf.icursor(END)
        print('datatype_hive_sf Result location selected')

    def select_download_location_datatype_sf_sf(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_datatype_sf_sf.delete(0,END)
        self.download_path_datatype_sf_sf.insert(0,download_location_adrs)
        self.download_path_datatype_sf_sf.focus()
        self.download_path_datatype_sf_sf.icursor(END)
        print('datatype_sf_sf Result location selected')

    def select_download_location_tokenization_check(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_tokenization_check.delete(0,END)
        self.download_path_tokenization_check.insert(0,download_location_adrs)
        self.download_path_tokenization_check.focus()
        self.download_path_tokenization_check.icursor(END)
        print('tokenization_check Result location selected')

    def select_download_location_datatype_td_sf(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_datatype_td_sf.delete(0,END)
        self.download_path_datatype_td_sf.insert(0,download_location_adrs)
        self.download_path_datatype_td_sf.focus()
        self.download_path_datatype_td_sf.icursor(END)
        print('datatype_td_sf Result location selected')

    def select_download_location_datatype_mssql_sf(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_datatype_mssql_sf.delete(0,END)
        self.download_path_datatype_mssql_sf.insert(0,download_location_adrs)
        self.download_path_datatype_mssql_sf.focus()
        self.download_path_datatype_mssql_sf.icursor(END)
        print('datatype_mssql_sf Result location selected')

    def select_download_location_datatype_mssql_td(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_datatype_mssql_td.delete(0,END)
        self.download_path_datatype_mssql_td.insert(0,download_location_adrs)
        self.download_path_datatype_mssql_td.focus()
        self.download_path_datatype_mssql_td.icursor(END)
        print('datatype_mssql_td Result location selected')


        
    def select_download_location_sf_sf_schema_ddl(self):
            download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
            self.download_path_sf_sf_schema_ddl.delete(0,END)
            self.download_path_sf_sf_schema_ddl.insert(0,download_location_adrs)
            self.download_path_sf_sf_schema_ddl.focus()
            self.download_path_sf_sf_schema_ddl.icursor(END)
            print('sf_sf_schema_ddl Result location selected')


    def select_download_location_Hive(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_Hive.delete(0,END)
        self.download_path_Hive.insert(0,download_location_adrs)
        self.download_path_Hive.focus()
        self.download_path_Hive.icursor(END)
        print('Hive Result location selected')

    def select_download_location_table_details_Hive(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_table_details_Hive.delete(0,END)
        self.download_path_table_details_Hive.insert(0,download_location_adrs)
        self.download_path_table_details_Hive.focus()
        self.download_path_table_details_Hive.icursor(END)
        print('Hive Result location selected')

        

    def select_download_location_MySQLWorkBench(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_MySQLWorkBench.delete(0,END)
        self.download_path_MySQLWorkBench.insert(0,download_location_adrs)
        self.download_path_MySQLWorkBench.focus()
        self.download_path_MySQLWorkBench.icursor(END)
        print('MySQLWorkBench Result location selected')

    def select_download_location_View_Validation_SF(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_View_Validation_SF.delete(0,END)
        self.download_path_View_Validation_SF.insert(0,download_location_adrs)
        self.download_path_View_Validation_SF.focus()
        self.download_path_View_Validation_SF.icursor(END)
        print('View_Validation_SF Result location selected')
        
        
    def select_download_location_SF(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_SF.delete(0,END)
        self.download_path_SF.insert(0,download_location_adrs)
        self.download_path_SF.focus()
        self.download_path_SF.icursor(END)
        print('SF Result location selected')


        # select_download_location_PostgresSQL

    def select_download_location_PostgresSQL(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_PostgresSQL.delete(0,END)
        self.download_path_PostgresSQL.insert(0,download_location_adrs)
        self.download_path_PostgresSQL.focus()
        self.download_path_PostgresSQL.icursor(END)
        print('PostgresSQL Result location selected')

    def select_download_location_SF_timestamps(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_SF_timestamps.delete(0,END)
        self.download_path_SF_timestamps.insert(0,download_location_adrs)
        self.download_path_SF_timestamps.focus()
        self.download_path_SF_timestamps.icursor(END)
        print('SF Result location selected')

    def select_TableStatsDownloadLocation(self):
        table_stats_download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.tableStatsResultDownloadLocationEntry.delete(0,END)
        self.tableStatsResultDownloadLocationEntry.insert(0,table_stats_download_location_adrs)
        self.tableStatsResultDownloadLocationEntry.focus()
        self.tableStatsResultDownloadLocationEntry.icursor(END)
        print('TableStats Result location selected')

    def select_TableStatsDownloadLocation_Oracle(self):
        table_stats_download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.tableStatsResultDownloadLocationEntry_Oracle.delete(0,END)
        self.tableStatsResultDownloadLocationEntry_Oracle.insert(0,table_stats_download_location_adrs)
        self.tableStatsResultDownloadLocationEntry_Oracle.focus()
        self.tableStatsResultDownloadLocationEntry_Oracle.icursor(END)
        print('TableStats Result location selected')

    def select_TableStatsDownloadLocation_TD(self):
        table_stats_download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.tableStatsResultDownloadLocationEntry_TD.delete(0,END)
        self.tableStatsResultDownloadLocationEntry_TD.insert(0,table_stats_download_location_adrs)
        self.tableStatsResultDownloadLocationEntry_TD.focus()
        self.tableStatsResultDownloadLocationEntry_TD.icursor(END)
        print('TableStats Result location selected')

    def select_TableStatsDownloadLocation_sf_to_sf(self):
        table_stats_download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.tableStatsResultDownloadLocationEntry_sf_to_sf.delete(0,END)
        self.tableStatsResultDownloadLocationEntry_sf_to_sf.insert(0,table_stats_download_location_adrs)
        self.tableStatsResultDownloadLocationEntry_sf_to_sf.focus()
        self.tableStatsResultDownloadLocationEntry_sf_to_sf.icursor(END)
        print('TableStats Result location selected')

    def select_TableStatsDownloadLocation_mssql_to_sf(self):
        table_stats_download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.tableStatsResultDownloadLocationEntry_mssql_to_sf.delete(0,END)
        self.tableStatsResultDownloadLocationEntry_mssql_to_sf.insert(0,table_stats_download_location_adrs)
        self.tableStatsResultDownloadLocationEntry_mssql_to_sf.focus()
        self.tableStatsResultDownloadLocationEntry_mssql_to_sf.icursor(END)
        print('TableStats Result location selected')
        
    def select_TableStatsDownloadLocation_hive_to_sf(self):
        table_stats_download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.tableStatsResultDownloadLocationEntry_hive_to_sf.delete(0,END)
        self.tableStatsResultDownloadLocationEntry_hive_to_sf.insert(0,table_stats_download_location_adrs)
        self.tableStatsResultDownloadLocationEntry_hive_to_sf.focus()
        self.tableStatsResultDownloadLocationEntry_hive_to_sf.icursor(END)
        print('TableStats Result location selected')


    # select_sanity_checkDownloadLocation_hive_to_sf
    def select_sanity_checkDownloadLocation_hive_to_sf(self):
        sanity_check_download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.sanity_checkResultDownloadLocationEntry_hive_to_sf.delete(0,END)
        self.sanity_checkResultDownloadLocationEntry_hive_to_sf.insert(0,sanity_check_download_location_adrs)
        self.sanity_checkResultDownloadLocationEntry_hive_to_sf.focus()
        self.sanity_checkResultDownloadLocationEntry_hive_to_sf.icursor(END)
        print('sanity_check Result location selected')

    def select_sanity_checkDownloadLocation_td_to_sf(self):
        sanity_check_download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.sanity_checkResultDownloadLocationEntry_td_to_sf.delete(0,END)
        self.sanity_checkResultDownloadLocationEntry_td_to_sf.insert(0,sanity_check_download_location_adrs)
        self.sanity_checkResultDownloadLocationEntry_td_to_sf.focus()
        self.sanity_checkResultDownloadLocationEntry_td_to_sf.icursor(END)
        print('sanity_check Result location selected')

        
    def select_download_location_TD(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_TD.delete(0,END)
        self.download_path_TD.insert(0,download_location_adrs)
        self.download_path_TD.focus()
        self.download_path_TD.icursor(END)
        print('TD Result location selected')

    def select_download_location_mssql(self):
        download_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.download_path_mssql.delete(0,END)
        self.download_path_mssql.insert(0,download_location_adrs)
        self.download_path_mssql.focus()
        self.download_path_mssql.icursor(END)
        print('mssql Result location selected')
       
    def establish_connection_SSH(self):
        try:
            hostname_unix=str(self.hostname_unix_ssh.get()).strip()
            port=22
            username=str(self.uid_SSH.get()).strip()
            password=str(self.pwd_SSH.get()).strip()
            self.SSH_ssh=paramiko.SSHClient()
            self.SSH_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            self.SSH_ssh.connect(hostname_unix,port,username,password)
            messagebox.showinfo('Connected','SSH connected successfully!')
        except Exception as err:
            messagebox.showerror('Unable To Connect',str(err))
            raise Exception(str(err))

    def establish_connection_Hive(self):
        try:
            hostname_unix=str(self.hostname_unix.get()).strip()
            port=22
            username=str(self.uid_Hive.get()).strip()
            password=str(self.pwd_Hive.get()).strip()
            self.hive_ssh=paramiko.SSHClient()
            self.hive_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            self.hive_ssh.connect(hostname_unix,port,username,password)
            messagebox.showinfo('Connected','Hive connected successfully!')
        except Exception as err:
            messagebox.showerror('Unable To Connect',str(err))
            raise Exception(str(err))


    def establish_connection_PostgresSQL(self):
        import psycopg2
        try:
            hostname_PostgresSQL=str(self.hostname_PostgresSQL.get()).strip()
            database_PostgresSQL=str(self.database_PostgresSQL.get()).strip()
            port_PostgresSQL=str(self.port_PostgresSQL.get()).strip()
            username=str(self.uid_PostgresSQL.get()).strip()
            password=str(self.pwd_PostgresSQL.get()).strip()
             
             
           
            self.cnxn_PostgresSQL =psycopg2.connect(
                                    host=hostname_PostgresSQL,
                                    database=database_PostgresSQL,
                                    port=port_PostgresSQL,
                                    user=username,
                                    password=password)
            # self.cnxn_PostgresSQL = self.cnxn_PostgresSQL.cursor()
            messagebox.showinfo('Connected','PostgresSQL connected successfully!')
        except Exception as err:
            messagebox.showerror('Unable To Connect',str(err))
            raise Exception(str(err))
    


    def establish_connection_MySQLWorkBench(self):
        
        try:
            hostname_MySQLWorkBench=str(self.hostname_MySQLWorkBench.get()).strip()
            username=str(self.uid_MySQLWorkBench.get()).strip()
            password=str(self.pwd_MySQLWorkBench.get()).strip()
             
            
           
            self.cnxn_MySQLWorkBench = mysql.connector.connect(
                        host=hostname_MySQLWorkBench,
                        user=username,
                        passwd=password
                        )
            # self.cnxn_MySQLWorkBench = self.cnxn_MySQLWorkBench.cursor()
            messagebox.showinfo('Connected','MySQLWorkBench connected successfully!')
        except Exception as err:
            messagebox.showerror('Unable To Connect',str(err))
            raise Exception(str(err))
    
    
    def establish_connection_SF_prod(self):
        
        try:
            print('trying to connect to sf prod')
            self.con_prod = snowflake.connector.connect(
                # chi 
        #           user = 'AN536766AD',
        # account='eda_prod.anthemdatalake.us-east-1.privatelink',

        # authenticator = 'externalbrowser',
        # database='R01_SPCP',
        # warehouse = 'R01_SPCP_USER_WH_M',
        # schema='SPCP_NOPHI_NOGBD',
        # role='AN536766AD_PRIVS',
             user = str(self.ha_id_prod.get()).strip(),
        account= str(self.account_SF_prod.get()).strip(),

        authenticator = 'externalbrowser',
        database='',
        warehouse = str(self.warehouse_SF_prod.get()).strip(),
        schema='',
        role='',
         
        )
            print('Connection established')
            messagebox.showinfo('Connected','Connected to Snowflake Prod')
        except Exception as err:
            print(str(err))
            messagebox.showerror('Unable to connect to Snowflake Prod',str(err))
        
    def establish_connection_SF(self):
        
        try:
            self.con = snowflake.connector.connect(
                # chi 
        #           user = 'AN536766AD',
        # account='carelon-eda_nonprod.privatelink',

        # authenticator = 'externalbrowser',
        # database='T01_UWRATE',
        # warehouse = 'P01_EDL_USER_WH_M',
        # schema='SPCP_NOPHI_NOGBD',
        # role='AN536766AD_PRIVS'
        # lgrs
        # database='R01_SPCP',
        # warehouse = 'T01_UWARTE_USER_WH_M',
        # schema='uwrate_stg_nogbd',
        # role='AN536766AD_PRIVS'
#edm3
        #     user = 'AN346509AD',
        # account='eda_prod.anthemdatalake.us-east-1.privatelink',
        # authenticator = 'externalbrowser',
        # database='R01_EDM_V3',
        # warehouse = 'R01_EDM_USER_WH_L',
        # schema='EDM_V3_NOGBD',
        # role='AN346509AD_PRIVS'
        # 
        

             user = str(self.ha_id.get()).strip(),
        account= str(self.account_SF.get()).strip(),

        authenticator = 'externalbrowser',
        database='',
        warehouse = str(self.warehouse_SF.get()).strip(),
        schema='',
        role='',
         
        )
            print('Connection established')
            messagebox.showinfo('Connected','Connected to Snowflake')
        except Exception as err:
            print(str(err))
            messagebox.showerror('Unable to connect',str(err))
        
        
        
    

    def establish_connection_mssql(self):
        
        try:
            mssql_driver = str(self.driver_mssql.get()).strip()
            mssql_server = str(self.server_mssql.get()).strip() # can be an IP too
            mssql_db = str(self.db_mssql.get()).strip()
            mssql_uid = str(self.uid_mssql.get()).strip()
            mssql_pwd = str(self.pwd_mssql.get()).strip()
            
            
            
            print('Connecting to mssql')
            cnxn_str = ("Driver={"+mssql_driver+"};"
                "Server="+mssql_server+";"
                "Database="+mssql_db+";"
                "uid="+mssql_uid+";"
                "pwd="+mssql_pwd+";"
                "Trusted_Connection=No;")
            self.cnxn_mssql = pyodbc.connect(cnxn_str)
            print('Connected to mssql')
            messagebox.showinfo('Connected','Successfully connected to mssql servers.')
        except Exception as err:
            print(str(err))
            messagebox.showerror('Login issue',str(err))

    def establish_connection_TD(self):
        # Set variables 
        #
        #str(self.hostname_td.get()).strip()
        # Set variables 
        #str(self.driver_td.get()).strip()
        #str(self.hostname_td.get()).strip()
       # str(self.anthem_id_td.get()).strip()
       try:
        teradata_driver = str(self.driver_td.get()).strip()
        teradata_host_name = str(self.hostname_td.get()).strip() # can be an IP too
        teradata_user = str(self.anthem_id_td.get()).strip()
        teradata_password =str(self.pwd_td.get()).strip()
        # If you dont know the connection mechanism check with your DBA
        teradata_authentication = 'LDAP'
        
        
        print('Connecting to TD')
        self.cnxn_TD = pyodbc.connect("DRIVER={%s};DBCNAME=%s;UID=%s;PWD=%s;authentication=%s"\
                              %(teradata_driver,teradata_host_name,teradata_user,teradata_password,teradata_authentication))
        print('Connected to TD')
        messagebox.showinfo('Connected','Successfully connected to Teradata servers.')
       except Exception as err:
            messagebox.showerror('Error',str(err))
            print('Error while connecting Teradata: ',str(err))


# sqoop_command_oracle
    def establish_connection_Oracle_sqoop(self):
        try:
            df=self.run_oracle_sqoop_query(query='SELECT CURRENT_DATE, SESSIONTIMEZONE FROM DUAL')
            messagebox.showinfo('Connected','Oracle server connected with given sqoop command')
        except:
            messagebox.showerror('Error','Not able to connect to Oracle with given sqoop command.')





    def establish_connection_Oracle(self):
        # Set variables 
        #
        #str(self.hostname_td.get()).strip()
        # Set variables 
        #str(self.driver_td.get()).strip()
        #str(self.hostname_td.get()).strip()
       # str(self.anthem_id_td.get()).strip()
        Oracle_driver = str(self.driver_Oracle.get()).strip()
        Oracle_mechanism_name = str(self.mechanism_name_Oracle.get()).strip() # can be an IP too
        Oracle_user = str(self.uid_Oracle.get()).strip()
        Oracle_password = str(self.pwd_Oracle.get()).strip()
        # If you dont know the connection mechanism check with your DBA
        Oracle_authentication = 'LDAP'
        
        
        print('Connecting to Oracle')
        self.cnxn_Oracle = pyodbc.connect("DRIVER={%s};Dbq=%s;UID=%s;PWD=%s;authentication=%s"\
                              %(Oracle_driver,Oracle_mechanism_name,Oracle_user,Oracle_password,Oracle_authentication))
        print('Connected to Oracle')
        messagebox.showinfo('Connected','Successfully connected to Oracle servers.')
        

   

    def getTableDetails_Hive(self):
               
         
        print('Table details download from hive started..')
         
        resultpath=str(self.download_path_table_details_Hive.get()).strip()
        os.chdir(resultpath)
        filename=str(self.query_table_details_Hive.get()).strip()
        final_desc_df=pd.DataFrame(columns=['col_name','data_type','comment','MaxLength','Table'])

        with open(filename,'r') as f:
            qs=f.read()
            table_names_list=qs.split(',')
             

            for table in table_names_list:
                table=table.strip()
                desc_query='describe '+table
                desc_df=self.run_hive_query(desc_query)
                desc_df.columns=[col.lower() for col in desc_df.columns]
                # desc_df=desc_df[['col_name','data_type']]


                if(self.runFromImpala.get()==0):
                         desc_df.rename({'name':'col_name','type':'data_type'},inplace=True)

                

                hive_cols=desc_df['col_name']

                hive_cols_for_max_length_query_helper=''

                for colname,dtype in zip(desc_df['col_name'],desc_df['data_type']):
                            if (dtype in ['string','char','varchar','binary']) or ('char' in dtype ) or ('varchar' in dtype):
                                hive_cols_for_max_length_query_helper += ',max( length({colName}) ) as {colName}'.format(colName=colname)


                if len(hive_cols_for_max_length_query_helper)>0:
                    hive_cols_for_max_length_query_helper=hive_cols_for_max_length_query_helper[1:]

                max_len_query='select {q} from {tn}'.format(q=hive_cols_for_max_length_query_helper,tn=table)

                res_df=self.run_hive_query(max_len_query)
                res_df = res_df.transpose().reset_index()
                # res_df.to_csv('aaaa.csv')
                res_df.columns=['name','maxlen']
                # res_df.to_csv('bbbb.csv')

                hive_max_lens=list()

                
                for col in hive_cols:
                    if col in list(res_df['name']):
                         
                        df1=res_df[res_df['name']==col]['maxlen']
                        val=df1.iloc[0]
                         
                        hive_max_lens.append(val)
                    else:
                        hive_max_lens.append(np.nan)

                 

                # for col in hive_cols:
                #     try:
                #             h_q='select max( length({colName}) ) as c from {tn}'.format(colName=col,tn=table)
                #             #    print("Running query: ",s_q)
                #             res=self.run_hive_query(h_q)
                #             hive_max_lens.append(res.iloc[:,0][0])
                #     except:
                #             hive_max_lens.append(np.nan)

                
                desc_df['MaxLength']=hive_max_lens
                desc_df['Table']=table
                # desc_df.to_csv(table+'.csv',index=False)
                final_desc_df=pd.concat([final_desc_df,desc_df])

        final_desc_df.to_csv('Hive_Table_details'+str(time.time())+'.csv',index=False)
        messagebox.showinfo('Done','Result downloaded')



    

    def datatype_check_mssql_sf(self):

        print('Download from query started..')
         
        basepath=str(self.download_path_datatype_mssql_sf.get()).strip()
        os.chdir(basepath)

        df_path=(self.query_excel_file_datatype_mssql_sf.get()).strip()
        
       
        df=pd.read_excel(df_path)

        datatypes_mssql_sf_mapping=dict()
        datatypes_mssql_sf_mapping['string']='varchar'
        datatypes_mssql_sf_mapping['timestamp']='timestamp_ntz'
        datatypes_mssql_sf_mapping['int']='number'
        datatypes_mssql_sf_mapping['decimal']='number'
        datatypes_mssql_sf_mapping['float']='number'
        datatypes_mssql_sf_mapping['date']='date'

        colnames=['mssql_COLUMN_NAME','mssql_DATA_TYPE','mssql_COLUMN_DEFAULT','mssql_IS_NULLABLE','mssql_CHARACTER_MAXIMUM_LENGTH','SF_COLUMN_NAME','SF_IS_NULLABLE','SF_COLUMN_DEFAULT','SF_CHARACTER_MAXIMUM_LENGTH','SF_DATA_TYPE']

        finalDf=pd.DataFrame(columns=colnames)

        for row in (df.itertuples()):
            try:
                table_mssql=row[1]
                table_sf=row[2]
                print("mssql table:",table_mssql," SF table:",table_sf)
                mssql_db_name,mssql_schema,mssql_tablename=table_mssql.split('.')
                #mssql
                query_mssql= "select * from {db}.information_schema.columns  where table_name = '{tn}' and table_schema='{ts}' order by ordinal_position".format(db=mssql_db_name,tn=mssql_tablename,ts=mssql_schema)
                
                if self.useSqoopCommand.get()==1:
                    query_mssql_df=self.run_mssql_sqoop_query(query_mssql)
                else:
                    query_mssql_df=self.run_mssql_query(query_mssql)
                
                
                query_mssql_df.columns=[col.strip() for col in query_mssql_df.columns]
                query_mssql_df=query_mssql_df[['COLUMN_NAME','DATA_TYPE','COLUMN_DEFAULT','IS_NULLABLE','CHARACTER_MAXIMUM_LENGTH']]
                
                query_mssql_df=query_mssql_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                

                query_mssql_df.columns=['mssql_COLUMN_NAME','mssql_DATA_TYPE','mssql_COLUMN_DEFAULT','mssql_IS_NULLABLE','mssql_CHARACTER_MAXIMUM_LENGTH']

                query_mssql_df=query_mssql_df.applymap(lambda x: x.upper() if isinstance(x, str) else x)
                # query_mssql_df=query_mssql_df.replace('(NULL)',np.NaN)

                

                #sf
                sf_db_name,sf_schema,sf_tablename=table_sf.split('.')

                datatype_query='desc table '+ table_sf
                sf_datatype_df=self.run_sf_query(datatype_query)
                sf_datatype_df=sf_datatype_df[['name','type']]
                sf_datatype_df.columns=['SF_COLUMN_NAME','SF_DATA_TYPE']
               
                info_schema_sf_query="select COLUMN_NAME as SF_COLUMN_NAME ,IS_NULLABLE as SF_IS_NULLABLE,COLUMN_DEFAULT as SF_COLUMN_DEFAULT,CHARACTER_MAXIMUM_LENGTH as SF_CHARACTER_MAXIMUM_LENGTH from  {db}.INFORMATION_SCHEMA.columns  where table_name='{tn}' and table_schema='{ts}'".format(db=sf_db_name.upper(),tn=sf_tablename.upper(),ts=sf_schema.upper())
                info_schema_sf_df=self.run_sf_query(info_schema_sf_query)

                info_schema_sf_df=info_schema_sf_df.merge(sf_datatype_df,on='SF_COLUMN_NAME')
                info_schema_sf_df=info_schema_sf_df.applymap(lambda x: x.upper() if isinstance(x, str) else x)
                
                tempDf=query_mssql_df.merge(info_schema_sf_df,left_on='mssql_COLUMN_NAME',right_on='SF_COLUMN_NAME',how='outer')
                # tempDf.to_csv('mssql_sf_merged.csv')
                
                tn=str(table_sf).split('.')[-1]
                tablename_list=list()
                
                attributes=['COLUMN_NAME','DATA_TYPE','COLUMN_DEFAULT','IS_NULLABLE','CHARACTER_MAXIMUM_LENGTH']
                mismatching_attribute_list=list()
                result_datatype=list()

                for row in (tempDf.itertuples()):
                    tablename_list.append(tn)

                    for attr in attributes:
                        mssql_col=str(getattr(row,'mssql_'+attr)).strip()
                        sf_col=str(getattr(row,'SF_'+attr)).strip()


                        try:
                           v1=str(float(mssql_col))
                           v2=str(float(sf_col))
                           mssql_col=v1
                           sf_col=v2
                                 
                        except: pass
                        # if ( ('(NULL)' in mssql_col) and ('None'==sf_col or 'nan'==sf_col) ):
                        #     continue
                        

                        if attr=='DATA_TYPE':
                            mssql_dtype=str(getattr(row,'mssql_DATA_TYPE'))
                    
                            sf_dtype=str(getattr(row,'SF_DATA_TYPE'))

                            if str(mssql_col)=='nan' or str(sf_col)=='nan':
                                mismatching_attribute_list.append('Column Missing')
                                result_datatype.append("No")
                                continue
                            
                            if not ( ( mssql_dtype==sf_dtype  ) or
                                ( ('INT' in mssql_dtype) and ('NUMBER' in sf_dtype) ) or 
                                ( ('smallint' in mssql_dtype) and ('NUMBER' in sf_dtype) ) or 
                                ( ('byteint' in mssql_dtype) and ('NUMBER' in sf_dtype) ) or 
                                ( ('DECIMAL' in mssql_dtype) and ('NUMBER' in sf_dtype) ) or 
                                ( ('NUMERIC' in mssql_dtype) and ('NUMBER' in sf_dtype) ) or 
                                ( ('FLOAT' in mssql_dtype) and ('NUMBER' in sf_dtype) ) or 
                                ( ('CHAR' in mssql_dtype) and ('VARCHAR' in sf_dtype) ) or
                                ( ('BIT' in mssql_dtype) and ('BOOLEAN' in sf_dtype) ) or 
                                ( ('DATE' in mssql_dtype) and ('DATE' in sf_dtype) ) or #(NULL)
                                
                                ( (mssql_dtype in ['CHAR','VARCHAR']) and ('VARCHAR' in sf_dtype)  ) or
                                ( ('TIMESTAMP' in mssql_dtype or 'DATETIME' in mssql_dtype) and ('TIMESTAMP' in sf_dtype) )  ):
                                    mismatching_attribute_list.append('Datatype mismatch')
                                    result_datatype.append("No")
                                    break
                        
                        elif mssql_col!=sf_col and not ( mssql_col=='(NULL)' and sf_col in ['nan','None']):
                                mismatching_attribute_list.append(attr)
                                result_datatype.append("No")
                                break
                        if attr=='CHARACTER_MAXIMUM_LENGTH':
                            mismatching_attribute_list.append('')
                            result_datatype.append("Yes")

                    
                            
                tempDf['Mismatched_Attr']=mismatching_attribute_list
                tempDf['DDL Same?']=result_datatype
                tempDf['TableName']=tablename_list
                
                
                
                finalDf=pd.concat([finalDf,tempDf],ignore_index=True)

            except Exception as err:
                print("In table ",row[0],'facing issue: ',str(err))
        res_name='Datatypes_mssql_SF_Comparison_Result_'+str(time.time())+'.csv'
        finalDf.columns=['msql_ColName','msql_Datatype','msql_DEFAULT','msql_IS_NULLABLE','msql_MaxLen','SF_ColName','SF_IS_NULLABLE','SF_DEFAULT','SF_MaxLen','SF_Datatype','Mismatched_Attr','DDL Same?','TableName']
        finalDf.to_csv(res_name,index=False)
        print('Done')  
        messagebox.showinfo('Done','Downloads stored in '+str(basepath)+' as '+res_name)      


    def datatype_check_sf_sf(self):

        print('Download from query started..')
         
        basepath=str(self.download_path_datatype_sf_sf.get()).strip()
        os.chdir(basepath)

        df_path=(self.query_excel_file_datatype_sf_sf.get()).strip()
        
       
        df=pd.read_excel(df_path)

        colnames=['sf1_COLUMN_NAME','sf1_DATA_TYPE','sf1_COLUMN_DEFAULT','sf1_IS_NULLABLE','sf1_CHARACTER_MAXIMUM_LENGTH','sf2_COLUMN_NAME','sf2_DATA_TYPE','sf2_IS_NULLABLE','sf2_COLUMN_DEFAULT','sf2_CHARACTER_MAXIMUM_LENGTH']

        finalDf=pd.DataFrame(columns=colnames)
        missing_tables=pd.DataFrame(columns=['Table'])

        for row in (df.itertuples()):
            try:
                table_sf1=str(row[1]).upper().strip()
                table_sf2=str(row[2]).upper().strip()
                print("sf table:",table_sf1," SF table:",table_sf2)
                sf_db_name1,sf_schema1,sf_tablename1=table_sf1.split('.')
                sf_db_name2,sf_schema2,sf_tablename2=table_sf2.split('.')

                try:
                #sf
                    query_sf= "select * from {db}.information_schema.columns  where table_name = '{tn}' and table_schema='{ts}' order by ordinal_position".format(db=sf_db_name1.strip().upper(),tn=sf_tablename1.strip().upper(),ts=sf_schema1.strip().upper())
                    
                    query_sf_df=self.run_sf_query(query_sf)
                    
                    
                    query_sf_df.columns=[col.strip() for col in query_sf_df.columns]
                    query_sf_df=query_sf_df[['COLUMN_NAME','DATA_TYPE','COLUMN_DEFAULT','IS_NULLABLE','CHARACTER_MAXIMUM_LENGTH']]
                    
                    query_sf_df=query_sf_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                    

                    query_sf_df.columns=['sf1_COLUMN_NAME','sf1_DATA_TYPE','sf1_COLUMN_DEFAULT','sf1_IS_NULLABLE','sf1_CHARACTER_MAXIMUM_LENGTH']

                    query_sf_df1=query_sf_df.applymap(lambda x: x.upper() if isinstance(x, str) else x)
                # query_sf_df=query_sf_df.replace('(NULL)',np.NaN)
                except Exception as err:
                    print(err)
                    missing_tables.loc[len(missing_tables.index)]=[table_sf1]

                

                #sf
                #sf2
                try:
                    query_sf= "select * from {db}.information_schema.columns  where table_name = '{tn}' and table_schema='{ts}' order by ordinal_position".format(db=sf_db_name2.strip().upper(),tn=sf_tablename2.strip().upper(),ts=sf_schema2.strip().upper())
                    
                    query_sf_df=self.run_sf_query(query_sf)
                    
                    
                    query_sf_df.columns=[col.strip() for col in query_sf_df.columns]
                    query_sf_df=query_sf_df[['COLUMN_NAME','DATA_TYPE','COLUMN_DEFAULT','IS_NULLABLE','CHARACTER_MAXIMUM_LENGTH']]
                    
                    query_sf_df=query_sf_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                    

                    query_sf_df.columns=['sf2_COLUMN_NAME','sf2_DATA_TYPE','sf2_COLUMN_DEFAULT','sf2_IS_NULLABLE','sf2_CHARACTER_MAXIMUM_LENGTH']

                    query_sf_df2=query_sf_df.applymap(lambda x: x.upper() if isinstance(x, str) else x)
                except Exception as err:
                    print(err)
                    missing_tables.loc[len(missing_tables.index)]=[table_sf2]

                
                tempDf=query_sf_df1.merge(query_sf_df2,left_on='sf1_COLUMN_NAME',right_on='sf2_COLUMN_NAME',how='outer')
                # tempDf.to_csv('sf_sf_merged.csv')
                
                tn=str(table_sf1).split('.')[-1]
                
                tablename_list=list()
                
                attributes=['COLUMN_NAME','DATA_TYPE','COLUMN_DEFAULT','IS_NULLABLE','CHARACTER_MAXIMUM_LENGTH']
                mismatching_attribute_list=list()
                result_datatype=list()

                for row in (tempDf.itertuples()):
                    tablename_list.append(tn)
                    mismatched_attrbts=list()
                    for attr in attributes:
                        sf1_col=str(getattr(row,'sf1_'+attr)).strip()
                        sf2_col=str(getattr(row,'sf2_'+attr)).strip()

                        if attr=='DATA_TYPE' and (str(sf1_col)=='nan' or str(sf2_col)=='nan'):
                            mismatched_attrbts.append('Column Missing')
                            result_datatype.append("No")
                            break

                        if sf1_col!=sf2_col:
                            mismatched_attrbts.append(attr)
                            if  attr=='CHARACTER_MAXIMUM_LENGTH':
                                result_datatype.append("No")
                        elif attr=='CHARACTER_MAXIMUM_LENGTH':
                            mismatched_attrbts.append('')
                            result_datatype.append("Yes")

                    mismatching_attribute_list.append(','.join(mismatched_attrbts))
                    
                            
                tempDf['Mismatched_Attr']=mismatching_attribute_list
                tempDf['DDL Same?']=result_datatype
                tempDf['TableName']=tn
                
                
                
                finalDf=pd.concat([finalDf,tempDf],ignore_index=True)

            except Exception as err:
                print("In table ",row[0],'facing issue: ',str(err))

        res_name='Datatypes_sf_SF_Comparison_Result_'+str(time.time())+'.xlsx'
        finalDf.columns=['sf1_ColName','sf1_Datatype','sf1_DEFAULT','sf1_NULLABLE','sf1_MaxLen','sf2_ColName','sf2_DATA_TYPE','sf2_IS_NULLABLE','sf2_DEFAULT','sf2_MaxLen','Mismatched_Attr','DDL Same?','TableName']
        # finalDf.to_csv(res_name,index=False)

        with pd.ExcelWriter(res_name) as writer:
                    finalDf.to_excel(writer,sheet_name='DDL',index = False,header=True) 
                    missing_tables.to_excel(writer,sheet_name='Absent tables',index = False,header=True) 
            
        print('Done')  
        messagebox.showinfo('Done','Downloads stored in '+str(basepath)+' as '+res_name)      


    ###temp method
    def datatype_check_sf_sf_only_desc(self):

        print('Download from query started..')
         
        basepath=str(self.download_path_datatype_sf_sf.get()).strip()
        os.chdir(basepath)

        df_path=(self.query_excel_file_datatype_sf_sf.get()).strip()
        
       
        df=pd.read_excel(df_path)

        # colnames=['sf1_COLUMN_NAME','sf1_DATA_TYPE','sf1_COLUMN_DEFAULT','sf1_IS_NULLABLE','sf1_CHARACTER_MAXIMUM_LENGTH','sf2_COLUMN_NAME','sf2_DATA_TYPE','sf2_IS_NULLABLE','sf2_COLUMN_DEFAULT','sf2_CHARACTER_MAXIMUM_LENGTH']
        # colnames=['sf1_ColName','sf1_Datatype','sf1_DEFAULT','sf1_IS_NULLABLE','sf2_ColName','sf2_Datatype','sf2_DEFAULT','sf2_IS_NULLABLE','Mismatched_Attr','DDL Same?','TableName']
        finalDf=pd.DataFrame()
        missing_tables=pd.DataFrame(columns=['Table'])

        for row in (df.itertuples()):
            try:
                table_sf1=str(row[1]).upper().strip()
                table_sf2=str(row[2]).upper().strip()
                print("sf table:",table_sf1," SF table:",table_sf2)
                sf_db_name1,sf_schema1,sf_tablename1=table_sf1.split('.')
                sf_db_name2,sf_schema2,sf_tablename2=table_sf2.split('.')
                stop=False

                try:
                #sf
#                     query_sf= "select * from {db}.information_schema.columns  where table_name = '{tn}' and table_schema='{ts}' order by ordinal_position".format(db=sf_db_name1.strip().upper(),tn=sf_tablename1.strip().upper(),ts=sf_schema1.strip().upper())
                    query_sf="desc table "+ table_sf1
                    query_sf_df=self.run_sf_query(query_sf)
                    
                    
                    query_sf_df.columns=[col.strip() for col in query_sf_df.columns]
                    query_sf_df=query_sf_df[['name','type','null?','default']]
                    
                    query_sf_df=query_sf_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                    

                    query_sf_df.columns=['sf1_COLUMN_NAME','sf1_DATA_TYPE','sf1_IS_NULLABLE','sf1_COLUMN_DEFAULT']

                    query_sf_df1=query_sf_df.applymap(lambda x: x.upper() if isinstance(x, str) else x)
                # query_sf_df=query_sf_df.replace('(NULL)',np.NaN)
                except Exception as err:
                    print(err)
                    stop=True
                    missing_tables.loc[len(missing_tables.index)]=[table_sf1]

                

                #sf
                #sf2
                try:
#                     query_sf= "select * from {db}.information_schema.columns  where table_name = '{tn}' and table_schema='{ts}' order by ordinal_position".format(db=sf_db_name2.strip().upper(),tn=sf_tablename2.strip().upper(),ts=sf_schema2.strip().upper())
                    query_sf="desc table "+ table_sf2
                    
                    query_sf_df=self.run_sf_query(query_sf)
                    
                    
                    query_sf_df.columns=[col.strip() for col in query_sf_df.columns]
                    query_sf_df=query_sf_df[['name','type','null?','default']]
                    
                    query_sf_df=query_sf_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                    

                    query_sf_df.columns=['sf2_COLUMN_NAME','sf2_DATA_TYPE','sf2_IS_NULLABLE','sf2_COLUMN_DEFAULT']

                    query_sf_df2=query_sf_df.applymap(lambda x: x.upper() if isinstance(x, str) else x)
                except Exception as err:
                    print(err)
                    stop=True
                    missing_tables.loc[len(missing_tables.index)]=[table_sf2]

                if(stop==True): continue
                tempDf=query_sf_df1.merge(query_sf_df2,left_on='sf1_COLUMN_NAME',right_on='sf2_COLUMN_NAME',how='outer')
                # tempDf.to_csv('sf_sf_merged.csv')
                
                tn=str(table_sf1).split('.')[-1]
                
                tablename_list=list()
                
                attributes=['COLUMN_NAME','DATA_TYPE','COLUMN_DEFAULT','IS_NULLABLE']
                mismatching_attribute_list=list()
                result_datatype=list()
                tn=table_sf1 +" --- " + table_sf2
                for row in (tempDf.itertuples()):
                    tablename_list.append(tn)
                    mismatched_attrbts=list()
                    for attr in attributes:
                        sf1_col=str(getattr(row,'sf1_'+attr)).strip()
                        sf2_col=str(getattr(row,'sf2_'+attr)).strip()

                        if attr=='DATA_TYPE' and (str(sf1_col)=='nan' or str(sf2_col)=='nan'):
                            mismatched_attrbts.append('Column Missing')
                            #result_datatype.append("No")
                            break

                        if sf1_col!=sf2_col:
                            mismatched_attrbts.append(attr)
                            
                        elif attr=='IS_NULLABLE':
                            mismatched_attrbts.append('')
                            
                     
                    mismatching_attribute_list.append(','.join(mismatched_attrbts))

                    if(mismatching_attribute_list[-1]!=''): result_datatype.append("No")
                    else: result_datatype.append("Yes")
                    
                            
                tempDf['Mismatched_Attr']=mismatching_attribute_list
                tempDf['DDL Same?']=result_datatype
                tempDf['TableName']=tn
                
                
                
                finalDf=pd.concat([finalDf,tempDf],ignore_index=True)

            except Exception as err:
                print("In table ",row[0],'facing issue: ',str(err))

        res_name='Datatypes_sf_SF_Comparison_Result_'+str(time.time())+'.xlsx'
        # finalDf.columns=['sf1_ColName','sf1_Datatype','sf1_DEFAULT','sf1_IS_NULLABLE','sf2_ColName','sf2_Datatype','sf2_DEFAULT','sf2_IS_NULLABLE','Mismatched_Attr','DDL Same?','TableName']
        finalDf.to_csv("res.csv",index=False)
        missing_tables.to_csv('missing.csv')

        with pd.ExcelWriter(res_name) as writer:
                    finalDf.to_excel(writer,sheet_name='DDL',index = False,header=True) 
                    missing_tables.to_excel(writer,sheet_name='Absent tables',index = False,header=True) 
            
        print('Done')  
        messagebox.showinfo('Done','Downloads stored in '+str(basepath)+' as '+res_name)      





    def datatype_check_mssql_td(self):

        print('Download from query started..')
         
        basepath=str(self.download_path_datatype_mssql_td.get()).strip()
        os.chdir(basepath)

        df_path=(self.query_excel_file_datatype_mssql_td.get()).strip()
        
       
        df=pd.read_excel(df_path)

        datatypes_mssql_td_mapping=dict()
        datatypes_mssql_td_mapping['string']='varchar'
        datatypes_mssql_td_mapping['timestamp']='timestamp_ntz'
        datatypes_mssql_td_mapping['int']='number'
        datatypes_mssql_td_mapping['decimal']='number'
        datatypes_mssql_td_mapping['float']='number'
        datatypes_mssql_td_mapping['date']='date'

        colnames=['mssql_COLUMN_NAME','mssql_DATA_TYPE','mssql_IS_NULLABLE','mssql_CHARACTER_MAXIMUM_LENGTH','td_COLUMN_NAME','td_IS_NULLABLE','td_CHARACTER_MAXIMUM_LENGTH','td_DATA_TYPE','Mismatched_Attr','DDL Same?','TableName']

        finalDf=pd.DataFrame(columns=colnames)

        for row in (df.itertuples()):
            try:
                table_mssql=row[1]
                table_td=row[2]
                print("mssql table:",table_mssql," td table:",table_td)
                mssql_db_name,mssql_schema,mssql_tablename=table_mssql.split('.')
                #mssql
                query_mssql= "select * from {db}.information_schema.columns  where table_name = '{tn}' and table_schema='{ts}' order by ordinal_position".format(db=mssql_db_name,tn=mssql_tablename,ts=mssql_schema)
                query_mssql_df=self.run_mssql_query(query_mssql)
                query_mssql_df=query_mssql_df[['COLUMN_NAME','DATA_TYPE','IS_NULLABLE','CHARACTER_MAXIMUM_LENGTH']]
                query_mssql_df.columns=['mssql_COLUMN_NAME','mssql_DATA_TYPE','mssql_IS_NULLABLE','mssql_CHARACTER_MAXIMUM_LENGTH']

                query_mssql_df=query_mssql_df.applymap(lambda x: x.upper() if isinstance(x, str) else x)

                

                
                #td
                colLengthIndexed_query_td='help column '+ table_td+'.*'
                colLengthIndexed_query_td_df=self.run_td_query(colLengthIndexed_query_td)
                colLengthIndexed_query_td_df=colLengthIndexed_query_td_df[['Column Dictionary Name','Max Length','Nullable']]
                colLengthIndexed_query_td_df.columns=['td_COLUMN_NAME','td_CHARACTER_MAXIMUM_LENGTH','td_IS_NULLABLE']


                query_creation_for_td_datatype=''
                for col in list(colLengthIndexed_query_td_df['td_COLUMN_NAME']):
                    query_creation_for_td_datatype +=', TYPE({colName}) as  {colName}'.format(colName=col)

                if(len(query_creation_for_td_datatype)>0):
                    query_creation_for_td_datatype=query_creation_for_td_datatype[1:]


                td_datatype_query='Select Distinct {query} from {tn}'.format(query=query_creation_for_td_datatype,tn=table_td)
                td_datatype_query_df=self.run_td_query(td_datatype_query)
                td_datatype_query_df=td_datatype_query_df.transpose().reset_index()
                td_datatype_query_df.columns=['td_COLUMN_NAME','td_DATA_TYPE']

                td_datatype_query_df=td_datatype_query_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                
                colLengthIndexed_query_td_df=colLengthIndexed_query_td_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                
                colLengthIndexed_query_td_df=colLengthIndexed_query_td_df.merge(colLengthIndexed_query_td_df,left_on='td_COLUMN_NAME',right_on='td_COLUMN_NAME',how='outer')
                colLengthIndexed_query_td_df=colLengthIndexed_query_td_df[['td_COLUMN_NAME','td_DATA_TYPE','td_IS_NULLABLE','td_CHARACTER_MAXIMUM_LENGTH']]
               

                tempDf=query_mssql_df.merge(colLengthIndexed_query_td_df,left_on='mssql_COLUMN_NAME',right_on='td_COLUMN_NAME',how='outer')
                
                # tempDf.to_csv('mssql_td_merged.csv')
                
                tn=str(table_td).split('.')[-1]
                tablename_list=list()
                
                attributes=['COLUMN_NAME','DATA_TYPE','IS_NULLABLE','CHARACTER_MAXIMUM_LENGTH']
                mismatching_attribute_list=list()
                result_datatype=list()

                for row in (tempDf.itertuples()):
                    tablename_list.append(tn)

                    for attr in attributes:
                        mssql_col=str(getattr(row,'mssql_'+attr))
                        td_col=str(getattr(row,'td_'+attr))
                        if str(mssql_col)=='nan' or str(td_col)=='nan':
                            mismatching_attribute_list.append('Column Missing')
                            result_datatype.append("No")
                            break

                        elif attr=='DATA_TYPE':
                            mssql_dtype=str(getattr(row,'mssql_DATA_TYPE'))
                    
                            td_dtype=str(getattr(row,'td_DATA_TYPE'))
                            
                            if not ( ( mssql_dtype==td_dtype  ) or
                                ( ('INT' in mssql_dtype) and ('NUMBER' in td_dtype) ) or 
                                ( ('smallint' in mssql_dtype) and ('NUMBER' in td_dtype) ) or 
                                ( ('byteint' in mssql_dtype) and ('NUMBER' in td_dtype) ) or 
                                ( ('DECIMAL' in mssql_dtype) and ('NUMBER' in td_dtype) ) or 
                                ( ('NUMERIC' in mssql_dtype) and ('NUMBER' in td_dtype) ) or 
                                ( ('FLOAT' in mssql_dtype) and ('NUMBER' in td_dtype) ) or 
                                ( ('CHAR' in mssql_dtype) and ('VARCHAR' in td_dtype) ) or 
                                ( ('DATE' in mssql_dtype) and ('DATE' in td_dtype) ) or 
                                ( (mssql_dtype in ['CHAR','VARCHAR']) and ('VARCHAR' in td_dtype)  ) or
                                ( ('TIMESTAMP' in mssql_dtype) and ('TIMESTAMP' in td_dtype) )  ):
                                    mismatching_attribute_list.append('Datatype mismatch')
                                    result_datatype.append("No")
                                    break
                        
                        elif mssql_col!=td_col:
                                mismatching_attribute_list.append(attr)
                                result_datatype.append("No")
                                break
                        if attr=='CHARACTER_MAXIMUM_LENGTH':
                            mismatching_attribute_list.append('')
                            result_datatype.append("Yes")

                    
                            
                tempDf['Mismatched_Attr']=mismatching_attribute_list
                tempDf['DDL Same?']=result_datatype
                tempDf['TableName']=tablename_list
                
                
                
                finalDf=pd.concat([finalDf,tempDf],ignore_index=True)

            except Exception as err:
                print("In table ",row[0],'facing issue: ',str(err))
        res_name='Datatypes_mssql_td_Comparison_Result_'+str(time.time())+'.csv'
        finalDf.columns=['msql_ColName','msql_Datatype','msql_IS_NULLABLE','msql_MaxLen','td_ColName','td_IS_NULLABLE','td_MaxLen','td_Datatype','Mismatched_Attr','DDL Same?','TableName']
        finalDf.to_csv(res_name,index=False)
        print('Done')  
        messagebox.showinfo('Done','Downloads stored in '+str(basepath)+' as '+res_name)      


    def datatype_check_td_sf(self):

        print('Download from query started..')
         
        basepath=str(self.download_path_datatype_td_sf.get()).strip()
        os.chdir(basepath)

        df_path=(self.query_excel_file_datatype_td_sf.get()).strip()
        
       
        df=pd.read_excel(df_path)

        datatypes_td_sf_mapping=dict()
        datatypes_td_sf_mapping['string']='varchar'
        datatypes_td_sf_mapping['timestamp']='timestamp_ntz'
        datatypes_td_sf_mapping['int']='number'
        datatypes_td_sf_mapping['decimal']='number'
        datatypes_td_sf_mapping['float']='number'
        datatypes_td_sf_mapping['date']='date'


        colnames=['td_COLUMN_NAME','td_DATA_TYPE','td_IS_NULLABLE','td_CHARACTER_MAXIMUM_LENGTH','SF_COLUMN_NAME','SF_IS_NULLABLE','SF_CHARACTER_MAXIMUM_LENGTH','SF_DATA_TYPE','Mismatched_Attr','DDL Same?','TableName']

        finalDf=pd.DataFrame(columns=colnames)
        tablesNotFound=pd.DataFrame(columns=['TableName','Found','DB'])

        for row in (df.itertuples()):
            try:
                table_td=row[1]
                table_sf=row[2]
                print("td table:",table_td," SF table:",table_sf)
                

                #td
                colLengthIndexed_query_td_df=''
                try:
                    colLengthIndexed_query_td='help column '+ table_td+'.*'
                    colLengthIndexed_query_td_df=self.run_td_query(colLengthIndexed_query_td)
                    colLengthIndexed_query_td_df=colLengthIndexed_query_td_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

                    colLengthIndexed_query_td_df=colLengthIndexed_query_td_df[['Column Dictionary Name','Max Length','Nullable']]
                    pq=''' Select ColumnName,
                            CASE WHEN (ColumnType = 'CF' OR ColumnType = 'CV') THEN 'VARCHAR('||TRIM(ColumnLength)||')'
                            WHEN ColumnType = 'D' THEN 'NUMBER('||TRIM(DecimalTotalDigits)||','||trim(DecimalFractionalDigits)||')'
                            WHEN ColumnType = 'I8' or ColumnType='I' THEN 'NUMBER(38,0)'
                            WHEN ColumnType='TS' THEN 'TIMESTAMP_NTZ(6)'
                            WHEN ColumnType='DA' THEN 'DATE'
                            ELSE TRIM(ColumnType) END AS DataType
                             
                            FROM DBC.COLUMNS
                            WHERE tablename='{tn}' AND DataBaseName = '{dbname}'
                    '''.format(tn=table_td.split('.')[1].upper(),dbname=table_td.split('.')[0].upper())

                    extra_info_td_df=self.run_td_query(pq)
                    extra_info_td_df.columns=['Column Dictionary Name','Type']
                    extra_info_td_df=extra_info_td_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                    colLengthIndexed_query_td_df=colLengthIndexed_query_td_df.merge(extra_info_td_df,on='Column Dictionary Name',how='outer')
                    colLengthIndexed_query_td_df=colLengthIndexed_query_td_df[['Column Dictionary Name','Type','Max Length','Nullable']]
                    # colLengthIndexed_query_td_df=colLengthIndexed_query_td_df[['Column Dictionary Name','Type','Max Length','Nullable']]
                    # colLengthIndexed_query_td_df['Type']=colLengthIndexed_query_td_df['Type'].replace({'A1':'ARRAY (one dimensional)','AN':'ARRAY (multidimensional)','I8':'BIGINT','BO':'BINARY LARGE OBJECT','BF':'BYTE','BV':'BYTE VARYING','I1':'BYTEINT','CF':'CHARACTER (fixed)','CV':'CHARACTER (varying)','CO':'CHARACTER LARGE OBJECT','D':'DECIMAL','DA':'DATE','F':'FLOAT','I':'INTEGER','DY':'INTERVAL DAY','DH':'INTERVAL DAY TO HOUR','DM':'INTERVAL DAY TO MINUTE','DS':'INTERVAL DAY TO SECOND','HR':'INTERVAL HOUR','HM':'INTERVAL HOUR TO MINUTE','HS':'INTERVAL HOUR TO SECOND','MI':'INTERVAL MINUTE','MS':'INTERVAL MINUTE TO SECOND','MO':'INTERVAL MONTH','SC':'INTERVAL SECOND','YR':'INTERVAL YEAR','YM':'INTERVAL YEAR TO MONTH','N':'NUMBER','D':'NUMERIC','PD':'PERIOD(DATE)','PT':'PERIOD(TIME(n))','PZ':'PERIOD(TIME(n) WITH TIME ZONE)','PS':'PERIOD(TIMESTAMP(n))','PM':'PERIOD(TIMESTAMP(n) WITH TIME ZONE)','F':'FLOAT','I2':'SMALLINT','AT':'TIME','TS':'TIMESTAMP','TZ':'TIME WITH TIME ZONE','SZ':'TIMESTAMP WITH TIME ZONE','UT':'USER DEFINED TYPE (all types)','XM':'XML'})
                    # colLengthIndexed_query_td_df['Type']=colLengthIndexed_query_td_df['Type'].replace({'A1':'ARRAY (one dimensional)','AN':'ARRAY (multidimensional)','I8':'BIGINT',
                    #                                                                                    'BO':'BINARY LARGE OBJECT','BF':'BYTE','BV':'BYTE VARYING','I1':'BYTEINT',
                    #                                                                                    'CF':'VARCHAR (fixed)','CV':'CHARACTER (varying)','CO':'CHARACTER LARGE OBJECT',
                    #                                                                                    'D':'DECIMAL','DA':'DATE','F':'FLOAT','I':'INTEGER','DY':'INTERVAL DAY',
                    #                                                                                    'DH':'INTERVAL DAY TO HOUR','DM':'INTERVAL DAY TO MINUTE','DS':'INTERVAL DAY TO SECOND',
                    #                                                                                    'HR':'INTERVAL HOUR','HM':'INTERVAL HOUR TO MINUTE','HS':'INTERVAL HOUR TO SECOND','MI':'INTERVAL MINUTE',
                    #                                                                                    'MS':'INTERVAL MINUTE TO SECOND','MO':'INTERVAL MONTH','SC':'INTERVAL SECOND','YR':'INTERVAL YEAR',
                    #                                                                                    'YM':'INTERVAL YEAR TO MONTH','N':'NUMBER','D':'NUMERIC','PD':'PERIOD(DATE)','PT':'PERIOD(TIME(n))',
                    #                                                                                    'PZ':'PERIOD(TIME(n) WITH TIME ZONE)','PS':'PERIOD(TIMESTAMP(n))','PM':'PERIOD(TIMESTAMP(n) WITH TIME ZONE)',
                    #                                                                                    'F':'FLOAT','I2':'SMALLINT','AT':'TIME','TS':'TIMESTAMP','TZ':'TIME WITH TIME ZONE','SZ':'TIMESTAMP WITH TIME ZONE',
                    #                                                                                    'UT':'USER DEFINED TYPE (all types)','XM':'XML'})
                    colLengthIndexed_query_td_df.columns=['td_COLUMN_NAME','td_DATA_TYPE','td_CHARACTER_MAXIMUM_LENGTH','td_IS_NULLABLE']
                except:
                    tablesNotFound.loc[len(tablesNotFound.index)]=[table_td,'No','TD']


               
      
                
                #sf
                sf_db_name,sf_schema,sf_tablename=table_sf.split('.')

                info_schema_sf_df=''
                try:
                    datatype_query='desc table '+ table_sf
                    sf_datatype_df=self.run_sf_query(datatype_query)
                    sf_datatype_df=sf_datatype_df[['name','type']]
                    sf_datatype_df.columns=['SF_COLUMN_NAME','SF_DATA_TYPE']
                
                    info_schema_sf_query="select COLUMN_NAME as SF_COLUMN_NAME ,IS_NULLABLE as SF_IS_NULLABLE,CHARACTER_MAXIMUM_LENGTH as SF_CHARACTER_MAXIMUM_LENGTH from  {db}.INFORMATION_SCHEMA.columns  where table_name='{tn}' and table_schema='{ts}'".format(db=sf_db_name.strip().upper(),tn=sf_tablename.strip().upper(),ts=sf_schema.strip().upper())
                    info_schema_sf_df=self.run_sf_query(info_schema_sf_query)

                    info_schema_sf_df=info_schema_sf_df.merge(sf_datatype_df,on='SF_COLUMN_NAME')
                    info_schema_sf_df=info_schema_sf_df.applymap(lambda x: x.upper() if isinstance(x, str) else x)
                except Exception as err:
                    print(err) 
                    tablesNotFound.loc[len(tablesNotFound.index)]=[table_sf,'No','SF']
                
                tempDf=colLengthIndexed_query_td_df.merge(info_schema_sf_df,left_on='td_COLUMN_NAME',right_on='SF_COLUMN_NAME',how='outer')
                # tempDf.to_csv('td_sf_merged.csv')
                
                tn=str(table_sf).split('.')[-1]
                tablename_list=list()
                
                attributes=['COLUMN_NAME','DATA_TYPE','IS_NULLABLE','CHARACTER_MAXIMUM_LENGTH']
                mismatching_attribute_list=list()
                result_datatype=list()

                for row in (tempDf.itertuples()):
                    tablename_list.append(tn)
                    mismatches=list()
                    for attr in attributes:
                        td_col=str(getattr(row,'td_'+attr))
                        sf_col=str(getattr(row,'SF_'+attr))
                        

                        if attr=='DATA_TYPE':
                            td_dtype=td_col
                            sf_dtype=sf_col
                            if str(td_col)=='nan' or str(sf_col)=='nan':
                                mismatching_attribute_list.append('Column Missing')
                                result_datatype.append("No")
                                break
                            
                            if str(td_dtype).strip()!=(str(sf_dtype).split('COLLATE')[0]).strip():
                                mismatches.append('Datatype mismatch')

                            # if not ( ( td_dtype==sf_dtype  ) or
                            #     ( ('INT' in td_dtype) and ('NUMBER' in sf_dtype) ) or 
                            #     ( ('SMALLINT' in td_dtype) and ('NUMBER' in sf_dtype) ) or 
                            #     ( ('BYTEINT' in td_dtype) and ('NUMBER' in sf_dtype) ) or 
                            #     ( ('DECIMAL' in td_dtype) and ('NUMBER' in sf_dtype) ) or 
                            #     ( ('NUMERIC' in td_dtype) and ('NUMBER' in sf_dtype) ) or 
                            #     ( ('FLOAT' in td_dtype) and ('NUMBER' in sf_dtype) ) or 
                            #     ( ('CHAR' in td_dtype) and ('VARCHAR' in sf_dtype) ) or 
                            #     ( ('DATE' in td_dtype) and ('DATE' in sf_dtype) ) or 
                            #     ( (td_dtype in ['CHAR','VARCHAR']) and ('VARCHAR' in sf_dtype)  ) or
                            #     ( ('TIMESTAMP' in td_dtype) and ('TIMESTAMP' in sf_dtype) )  ):
                            #         mismatches.append('Datatype mismatch')
                                     
                            elif 'NUMBER' in sf_dtype:  
                                tempDf.at[row.Index,'SF_CHARACTER_MAXIMUM_LENGTH']=sf_dtype.split(',')[0].split('(')[1]
                                # tempDf.set_value(row.Index,'SF_CHARACTER_MAXIMUM_LENGTH',sf_dtype.split(',')[0].split('(')[1])
                                # row.SF_CHARACTER_MAXIMUM_LENGTH=sf_dtype.split(',')[0].split('(')[1]
                                # row['SF_'+'CHARACTER_MAXIMUM_LENGTH']=sf_dtype.split(',')[0].split('(')[1]

                        
                        elif attr=='IS_NULLABLE':
                                if (td_col.upper()=='Y' and sf_col.upper()=='YES') or (td_col.upper()=='N' and sf_col.upper()=='NO'):
                                    continue
                                mismatches.append(attr)

                        elif  attr=='CHARACTER_MAXIMUM_LENGTH':
                            try:
                                a=int(td_col)
                                b=int(sf_col)
                                if a!=b:
                                     mismatches.append(attr)
                            except:
                                if td_col.lower()=='nan' and sf_col.lower()!='nan':
                                    mismatches.append(attr)
                                elif td_col.lower()!='nan' and sf_col.lower()=='nan':
                                    mismatches.append(attr)


                                    

                        elif td_col!=sf_col :
                            
                                
                                mismatches.append(attr)
                                 
                        if attr=='CHARACTER_MAXIMUM_LENGTH':
                            mismatching_attribute_list.append(','.join(mismatches))
                            if len(mismatches)>0:
                                result_datatype.append("No")
                            else:
                                result_datatype.append("Yes")

                    
                            
                tempDf['Mismatched_Attr']=mismatching_attribute_list
                tempDf['DDL Same?']=result_datatype
                tempDf['TableName']=tablename_list
                
                
                
                finalDf=pd.concat([finalDf,tempDf],ignore_index=True)

            except Exception as err:
                print("In table ",row[0],'facing issue: ',str(err))

        res_name='Datatypes_TD_SF_Comparison_Result_'+str(time.time())+'.xlsx'
        finalDf.columns=['TD_ColName','TD_Datatype','TD_IS_NULLABLE','TD_MaxLen','SF_ColName','SF_IS_NULLABLE','SF_MaxLen','SF_Datatype','Mismatched_Attr','DDL Same?','TableName']
        # finalDf.to_csv(res_name,index=False)
        # res_name='Datatypes_Hive_SF_Comparison_Result_'+str(time.time())+'.xlsx'

        with pd.ExcelWriter(res_name) as writer:
                    finalDf.to_excel(writer,sheet_name='DDL',index = False,header=True) 
                    tablesNotFound.to_excel(writer,sheet_name='Absent tables',index = False,header=True) 
                
        print('Done')  

        messagebox.showinfo('Done','Downloads stored in '+str(basepath)+' as '+res_name)      


    def tokenization_check_fn(self):

        print('Download from query started..')
         
        basepath=str(self.download_path_tokenization_check.get()).strip()
        os.chdir(basepath)

        df_path=(self.query_excel_file_tokenization_check.get()).strip()
        
       
        df=pd.read_excel(df_path)

        name_vieww_all="viewDefn"+str(time.time())+".txt"
        all_views_defn= open(name_vieww_all,"w+")


        finalDf=pd.DataFrame(columns=['TableName','Tokenzd in AEDL not in SF','Tokenzd in SF not in AEDL','Invalid Y_N format','Invalid Y_Y format','tokenized cols in aedl','tokenized cols in sf'])
        tablesNotFound=pd.DataFrame(columns=['Tablename','Found'])
        for row in (df.itertuples()):
            try:
                table_aedl=row[1]
                table_sf=row[2]
                tknzd_sit=row[3]
                tknzd_prod=row[4]
                app_cd=row[5]
                print("aedl table:",table_aedl," SF table:",table_sf)
                error_In_query=False
                
                mycursor = self.cnxn_MySQLWorkBench.cursor()
                 
                query="select atrb_nm,tknzd_in_test_ind,tknzd_in_prodn_ind from sit_audt_cntrl.edl_tknztn_mtdta where UPPER(tbl_nm) = '{table_aedl}'".format(table_aedl=table_aedl.upper())

                if(str(app_cd).lower()!='nan'):
                    query +=" and UPPER(aplctn_cd) = '{app_cd}'".format(app_cd=app_cd.upper())

                # if(str(tknzd_sit).lower()!='nan'):
                #     query +=" and UPPER(tknzd_in_test_ind) = '{tnnz_sit}'".format(tnnz_sit=tknzd_sit.upper())
                # if(str(tknzd_prod).lower()!='nan'):
                #    query+=" and UPPER(tknzd_in_prodn_ind) = '{tknz_prod}'".format(tknz_prod=tknzd_prod.upper())

                print("running sql workbench query: ",query)
                mycursor.execute(query )
                myresult = mycursor.fetchall()
                ###
                ind_sit_prod=[[str(x[0]),str(x[2])+"_"+str(x[2])]   for x in myresult]
                sit_prod_map=dict()
                for [x,y] in ind_sit_prod:
                    sit_prod_map[x]=y
                ###
                tokenized_cols_aedl= [str(x[0]).lower() for x in myresult]
                tokenized_cols_aedl.sort()

                try:
                    query="select get_ddl('view','{viewname}')".format(viewname=table_sf.strip().upper())
                    df=self.run_sf_query(query) 
                    df.columns=['c1']
                except Exception as err:
                    print(err)
                    tablesNotFound.loc[len(tablesNotFound.index)]=[table_sf,'No']

                s=list(df['c1'])[0]
                all_views_defn.write(table_sf+":\n")
                all_views_defn.write(s+"\n\n")
                # s=list(s)[0]
                a=s.split(',')
                b=list(filter(lambda x:'protegrity' in x.lower(),a))
                
                

                tokenized_cols_sf=list()

                for lines in b:
                    arr=lines.split(';')[0]

                    arr=arr.split('from')[0].split()
                    if(arr[-1]==';'): 
                        tokenized_cols_sf.append(arr[-2])
                    else:
                        tokenized_cols_sf.append(arr[-1])

                tokenized_cols_sf.sort()

                
                common_cols_sf_aedl=list()
                for col in tokenized_cols_sf:
                    if col.lower() in tokenized_cols_aedl:
                        common_cols_sf_aedl.append(col.lower())
                        
                
# invalid y_y 
                format_y_n=list(filter(lambda x:'t01_protegrity.scrty_acs_cntrl.' in x.lower() and '_detok(' not in x.lower(), common_cols_sf_aedl))
                invalid_y_y=list()
                clst=list()
                
                for lines in format_y_n:
                    arr=lines.split(';')[0]

                    arr=arr.split('from')[0].split()
                    if(arr[-1]==';'): 
                        clst.append(arr[-2])
                    else:
                        clst.append(arr[-1])

                for val in clst:
                    mapping_sit_prod=sit_prod_map[val.lower()]
                    if mapping_sit_prod=='y_y': 
                        invalid_y_y.append(val)

                invalid_y_y.sort()

# invalid y_n
                format_y_y=list(filter(lambda x:'t01_protegrity.scrty_acs_cntrl.' in x.lower() and '_detok('  in x.lower(), common_cols_sf_aedl))
                invalid_y_n=list()
                clst2=list()

                for lines in format_y_y:
                    arr=lines.split(';')[0]

                    arr=arr.split('from')[0].split()
                    if(arr[-1]==';'): 
                        clst2.append(arr[-2])
                    else:
                        clst2.append(arr[-1])

                for val in clst2:
                    mapping_sit_prod=sit_prod_map[val.lower()]
                    if mapping_sit_prod=='y_n': 
                        invalid_y_n.append(val)

                invalid_y_n.sort()

                cols_not_tokenized_in_sf=[]
                for col in tokenized_cols_aedl:
                    if col not in tokenized_cols_sf:
                        cols_not_tokenized_in_sf.append(col)

                cols_not_tokenized_in_aedl=[]
                for col in tokenized_cols_sf:
                    if col not in tokenized_cols_aedl:
                        cols_not_tokenized_in_aedl.append(col)
                 

                finalDf.loc[len(finalDf.index)]=[table_sf,",".join(cols_not_tokenized_in_sf),
                                                 ",".join(cols_not_tokenized_in_aedl),
                                                 ",".join(invalid_y_n),
                                                 ",".join(invalid_y_y),
                                                 ",".join(tokenized_cols_aedl),
                                                 ",".join(tokenized_cols_sf)
                                                 ]

            except Exception as err:
                print("In table ",row[0],'facing issue: ',str(err))


        res_name='Tokenization_Result_'+str(time.time())+'.xlsx'
        # finalDf.to_csv(res_name,index=False)

        with pd.ExcelWriter(res_name) as writer:
                    finalDf.to_excel(writer,sheet_name='Tokenization',index = False,header=True) 
                    tablesNotFound.to_excel(writer,sheet_name='Absent tables',index = False,header=True) 
                   
        
        all_views_defn.close()
        print('Done')  
        messagebox.showinfo('Done','Downloads stored in '+str(basepath)+' as '+res_name)      






    def datatype_check_hive_sf(self):

        print('Download from query started..')
         
        basepath=str(self.download_path_datatype_hive_sf.get()).strip()
        os.chdir(basepath)

        df_path=(self.query_excel_file_datatype_hive_sf.get()).strip()
        
       
        df=pd.read_excel(df_path)

        datatypes_hive_sf_mapping=dict()
        datatypes_hive_sf_mapping['string']='varchar'
        datatypes_hive_sf_mapping['timestamp']='timestamp_ntz'
        datatypes_hive_sf_mapping['int']='number'
        datatypes_hive_sf_mapping['decimal']='number'
        datatypes_hive_sf_mapping['float']='number'
        datatypes_hive_sf_mapping['date']='date'
        datatypes_hive_sf_mapping['decimal']='date'
        # datatypes_hive_sf_mapping['varchar']=['char','varchar','string']
        # datatypes_hive_sf_mapping['number']=['int','smallint','bigint','float','decimal']


        finalDf=pd.DataFrame(columns=['Hive_ColName','Hive_Datatype','SF_ColName','SF_Datatype','Datatypes Same?','Hive_TableName','SF_TableName'])
        tablesNotFound=pd.DataFrame(columns=['Tablename','Schema','Found','DB'])
        for row in (df.itertuples()):
            try:
                table_hive=row[1]
                table_sf=row[2]
                print("Hive table:",table_hive," SF table:",table_sf)
                error_In_query=False
                try:
                #hive
                    desc_query='describe '+table_hive
                    desc_df=self.run_hive_query(desc_query)
                    desc_df.columns=[col.lower() for col in desc_df.columns]

                    if(self.runFromImpala.get()==1):
                         desc_df=desc_df[['name','type']] 
                    else:
                         desc_df=desc_df[['col_name','data_type']] 

                    desc_df.columns=['Hive_ColName','Hive_Datatype']
                    desc_df=desc_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                except:
                    error_In_query=True
                    tablesNotFound.loc[len(tablesNotFound.index)]=[table_hive.split('.')[1],table_hive.split('.')[0],'No','Hive']

                try:
                #sf
                    datatype_query='desc table '+ table_sf
                    sf_datatype_df=self.run_sf_query(datatype_query)
                    sf_datatype_df=sf_datatype_df[['name','type']]
                    sf_datatype_df.columns=['SF_ColName','SF_Datatype']
                    sf_datatype_df=sf_datatype_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                except:
                    error_In_query=TRUE
                    tablesNotFound.loc[len(tablesNotFound.index)]=[table_sf.split('.')[2],table_sf.split('.')[0]+table_sf.split('.')[1],'No','SF']

                if(error_In_query):
                    raise Exception('Not able to run for table: '+table_sf)

                tempDf=desc_df.merge(sf_datatype_df,left_on='Hive_ColName',right_on='SF_ColName',how='outer')

                result_datatype=list()
                tn=str(table_sf).split('.')[-1]
                tablename_list=list()

                for row in (tempDf.itertuples()):
                    tablename_list.append(tn)
                    hive_dtype=(str(getattr(row,'Hive_Datatype'))).lower().strip()
                    
                    sf_dtype=(str(getattr(row,'SF_Datatype'))).lower().strip()
                    
                    if(hive_dtype=='nan' or sf_dtype=='nan'):
                         result_datatype.append("No")

                    elif    ( ( hive_dtype==sf_dtype  ) or
                                ( ('int' in hive_dtype) and ('number' in sf_dtype) ) or 
                                ( ('smallint' in hive_dtype) and ('number' in sf_dtype) ) or 
                                ( ('byteint' in hive_dtype) and ('number' in sf_dtype) ) or 
                                ( ('decimal' in hive_dtype) and ('number' in sf_dtype) ) or 
                                ( ('numeric' in hive_dtype) and ('number' in sf_dtype) ) or 
                                ( ('float' in hive_dtype) and ('number' in sf_dtype) ) or 
                                ( ('float' in hive_dtype) and ('decimal' in sf_dtype) ) or 
                                ( ('float' in hive_dtype) and ('double' in sf_dtype) ) or 
                                ( ('double' in hive_dtype) and ('decimal' in sf_dtype) ) or 
                                ( ('char' in hive_dtype) and ('varchar' in sf_dtype) ) or
                                ( ('bit' in hive_dtype) and ('boolean' in sf_dtype) ) or 
                                ( ('bigint' in hive_dtype) and ('number' in sf_dtype) ) or #(NULL)
                                
                                ( (hive_dtype in ['char','varchar']) and ('varchar' in sf_dtype)  ) or
                                ( ('varchar' in hive_dtype) and (('varchar' in sf_dtype)) ) or
                                ( ('timestamp' in hive_dtype or 'datetime' in hive_dtype) and ('timestamp' in sf_dtype) )  ):result_datatype.append("Yes")
                                    

                    elif hive_dtype in  datatypes_hive_sf_mapping :
                        if datatypes_hive_sf_mapping[hive_dtype] in sf_dtype:
                            result_datatype.append("Yes")
                        else:
                            result_datatype.append("No")
                    else: result_datatype.append("Mapping not found")        

                tempDf['Datatypes Same?']=result_datatype
                tempDf['Hive_TableName']=table_hive
                tempDf['SF_TableName']=table_sf

                
                
                
                finalDf=pd.concat([finalDf,tempDf],ignore_index=True)

            except Exception as err:
                print("In table ",row[0],'facing issue: ',str(err))


        res_name='Datatypes_Hive_SF_Comparison_Result_'+str(time.time())+'.xlsx'

        with pd.ExcelWriter(res_name) as writer:
                    finalDf.to_excel(writer,sheet_name='DDL',index = False,header=True) 
                    tablesNotFound.to_excel(writer,sheet_name='Absent tables',index = False,header=True) 
                   
        
         
        print('Done')  
        messagebox.showinfo('Done','Downloads stored in '+str(basepath)+' as '+res_name)      



    def start_download_SqoopBatch(self):
         print('Download from query started..')
         
         basepath=str(self.download_path_SqoopBatch.get()).strip()
         os.chdir(basepath)
         filename=str(self.query_SqoopBatch.get()).strip()
         with open(filename,'r') as f:
            qs=f.read()
            query_list=qs.split(';')
            count=1
            
            for query in query_list:
                if len(str(query).strip())==0: continue
                # print('Running query: ',query)
                try:
                    Sqoop_df = self.run_sqoop_query(query)
                    # Sqoop_df = Sqoop_df.iloc[: , 1:]
                    # Sqoop_df.reset_index(inplace=True)
                    # print(Sqoop_df.head())
                    # fn='Sqoop_queryRes_'+str(count)+'.csv'
                    # Sqoop_df.to_csv(fn)
                    fn='Sqoop_queryRes_'+str(count)+'.xlsx'
                    
                    # fn=query.split('from')[1].split(' ')[1]
                    Sqoop_df.to_excel(fn,index=False)
                except Exception as err:
                    print('Not able to generate result for query no.',count)

                count +=1
         print('Querying done..')
         messagebox.showinfo('Done','Downloads stored in '+str(basepath))

    def start_download_Hive(self):
         #self.download_table_SF(query=)
         print('Download from query started..')
         
         basepath=str(self.download_path_Hive.get()).strip()
         os.chdir(basepath)
         filename=str(self.query_Hive.get()).strip()
         with open(filename,'r') as f:
            qs=f.read()
            query_list=qs.split(';')
            count=1
            
            for query in query_list:
                if len(str(query).strip())==0: continue
                # print('Running query: ',query)
                try:
                    
                    # Hive_df = Hive_df.iloc[: , 1:]
                    # Hive_df.reset_index(inplace=True)
                    # print(Hive_df.head())
                    # fn='Hive_queryRes_'+str(count)+'.csv'
                    # Hive_df.to_csv(fn)
                    try:
                        table_name=str(query.split('#')[0]).strip()
                        query=str(query.split('#')[1]).strip()
                    except:table_name='Hive_queryRes_'

                    # Hive_df =
                    self.run_hive_query_direct_download(query,table_name+str(count))
                    # fn=table_name+str(count)+'.xlsx'
                    
                    # fn=query.split('from')[1].split(' ')[1]
                    # Hive_df.to_excel(fn,index=False)

                except Exception as err:
                    print('Not able to generate result for query no.',count)

                count +=1
         print('Querying done..')
         messagebox.showinfo('Done','Downloads stored in '+str(basepath))
    
    def start_download_MySQLWorkBench(self):
         #self.download_table_SF(query=)
         print('Download from query started..')
         
         basepath=str(self.download_path_MySQLWorkBench.get()).strip()
         os.chdir(basepath)
         filename=str(self.query_MySQLWorkBench.get()).strip()
         with open(filename,'r') as f:
            qs=f.read()
            query_list=qs.split(';')
            count=1
            
            for query in query_list:
                if len(str(query).strip())==0: continue
                # print('Running query: ',query)
                try:
                    MySQLWorkBench_df = self.run_MySQLWorkBench_query(query)
                    # MySQLWorkBench_df = MySQLWorkBench_df.iloc[: , 1:]
                    # MySQLWorkBench_df.reset_index(inplace=True)
                    # print(MySQLWorkBench_df.head())
                    fn='MySQLWorkBench_queryRes_'+str(count)+'.csv'
                    MySQLWorkBench_df.to_csv(fn,index=False)
                except Exception as err:
                    print('Not able to generate result for query no.',count,'due to msg: ',str(err))

                count +=1
         print('Querying done..')
         messagebox.showinfo('Done','Downloads stored in '+str(basepath))
    
# start_download_PostgresSQL
    def start_download_PostgresSQL(self):
         #self.download_table_SF(query=)
         print('Dowload from query started..')
         basepath=str(self.download_path_PostgresSQL.get()).strip()
         filename=str(self.query_PostgresSQL.get()).strip()
         with open(filename,'r') as f:
            qs=f.read()
            query_list=qs.split(';')
            count=1
            
            for query in query_list:
                if len(str(query).strip())==0: continue
                try:
                    table_name=query.split('#')[0]+"_"+'.csv'
                    query=query.split('#')[1]
                except:table_name='PostgresSQL_queryRes_'+str(count)+"_"+'.csv'
                # print('Running query: ',query)
                self.download_table_PostgresSQL(query,table_name,basepath)
                count +=1
         print('Querying done..')
         messagebox.showinfo('Done','Downloads stored in '+str(basepath))
    

    def start_download_SF(self):
         #self.download_table_SF(query=)
         print('Dowload from query started..')
         basepath=str(self.download_path_SF.get()).strip()
         filename=str(self.query_SF.get()).strip()
         with open(filename,'r') as f:
            qs=f.read()
            query_list=qs.split(';')
            count=1
            
            for query in query_list:
                if len(str(query).strip())==0: continue
                try:
                    table_name=str(query.split('#')[0]).strip()+'.xlsx'
                    query=str(query.split('#')[1]).strip()
                except:table_name='sf_queryRes_'+str(count)+"_"+'.xlsx'
                # print('Running query: ',query)
                self.download_table_SF(query,table_name,basepath)
                count +=1
         print('Querying done..')
         messagebox.showinfo('Done','Downloads stored in '+str(basepath))
    

    
    
    def download_table_PostgresSQL(self,query,tableName='table_postgres_sql_3.csv',path=r'C:\Users\AH97759\Downloads'):
        establish = self.cnxn_PostgresSQL.cursor()
        establish.execute(query)
        os.chdir(path)
        all_rows = establish.fetchall()
        num_fields = len(establish.description)
        field_names = [i[0] for i in establish.description]
        df = pd.DataFrame(all_rows,columns=field_names)
        # print(all_rows)
        # print("**",field_names)
        # df.columns = field_names
        #return establish.fetch_pandas_all()
       
        df.reset_index(inplace=True)
        df = df.iloc[: , 1:]
        df.to_csv(tableName,index=False)

    def download_table_SF(self,query,tableName='table_sf_3.xlsx',path=r'C:\Users\AH97759\Downloads'):

        print('Executing query: ',query)
        try:
            establish = self.con.cursor()
            establish.execute(query)
        except:
            establish = self.con_prod.cursor()
            establish.execute(query)


        os.chdir(path)
        all_rows = establish.fetchall()
        num_fields = len(establish.description)
        field_names = [i[0] for i in establish.description]
        df = pd.DataFrame(all_rows,columns=field_names)
        # print(all_rows)
        # print("**",field_names)
        # df.columns = field_names
        #return establish.fetch_pandas_all()
       
        df.reset_index(inplace=True)
        df = df.iloc[: , 1:]
        df.to_excel(tableName,index=False)
    
    

    
    
    def start_download_Oracle(self):
         #self.download_table_SF(query=)
         print('Download from query started..')
         
         basepath=str(self.download_path_Oracle.get()).strip()
         os.chdir(basepath)
         filename=str(self.query_Oracle.get()).strip()
         with open(filename,'r') as f:
            qs=f.read()
            query_list=qs.split(';')
            count=1
            
            for query in query_list:
                if len(str(query).strip())==0: continue
                # print('Running query: ',query)
                teradata_df = self.run_Oracle_query(query)
                # teradata_df = teradata_df.iloc[: , 1:]
                # teradata_df.reset_index(inplace=True)
                # print(teradata_df.head())
                fn='Oracle_queryRes_'+str(count)+'.xlsx'
                teradata_df.to_excel(fn,index=False)
                count +=1
         print('Querying done..')
         messagebox.showinfo('Done','Downloads stored in '+str(basepath))
        
    
    def start_download_mssql(self):
         #self.download_table_SF(query=)
         print('Download from query started..')
         
         basepath=str(self.download_path_mssql.get()).strip()
         os.chdir(basepath)
         filename=str(self.query_mssql.get()).strip()
         with open(filename,'r') as f:
            qs=f.read()
            query_list=qs.split(';')
            count=1
            
            for query in query_list:
                if len(str(query).strip())==0: continue
                print('Running query: ',query)
                mssql_df = self.run_mssql_query(query.strip())
                # mssql_df = mssql_df.iloc[: , 1:]
                # mssql_df.reset_index(inplace=True)
                # print(mssql_df.head())
                fn='mssql_queryRes_'+str(count)+'.csv'
                mssql_df.to_csv(fn,index=False)
                count +=1
         print('Querying done..')
         messagebox.showinfo('Done','Downloads stored in '+str(basepath))
        
        
    


    def start_download_TD(self):
         #self.download_table_SF(query=)
         print('Download from query started..')
         
         basepath=str(self.download_path_TD.get()).strip()
         os.chdir(basepath)
         filename=str(self.query_TD.get()).strip()
         with open(filename,'r') as f:
            qs=f.read()
            query_list=qs.split(';')
            count=1
            
            for query in query_list:
                if len(str(query).strip())==0: continue
                print('Running query: ',query)
                
                # teradata_df = teradata_df.iloc[: , 1:]
                # teradata_df.reset_index(inplace=True)
                # print(teradata_df.head())
                try:
                        table_name=query.split('#')[0]
                        query=query.split('#')[1]
                except:table_name='TD_queryRes_'

                # teradata_df = pd.read_sql(query,self.cnxn_TD)
                teradata_df=self.run_td_query(query)
               

                fn=table_name+str(count)+'.csv'
                teradata_df.to_csv(fn,index=False)
                count +=1
         print('Querying done..')
         messagebox.showinfo('Done','Downloads stored in '+str(basepath))
        
        
    
    def schemaAndDDLverification_sf(self):
        # result_location=r'C:\\Users\\AH97759\\Downloads'
        result_location=str(self.download_path_sf_sf_schema_ddl.get()).strip()
        os.chdir(result_location)

        template_filepath=str(self.query_excel_file_sf_sf_schema_ddl.get()).strip()
        df=pd.read_excel(template_filepath)
        schema_result_df=pd.DataFrame(columns=['TableName','Schema','Present'])
        full_result_df=pd.DataFrame()
        dataComparisonDF=pd.DataFrame(columns=['DB','TableName','Source_Schema','Target_Schema','ColumnName','Attribute','Base_Value','Target_Value'])
        
        for row in df.itertuples():
            ddl_repo=str(row[1]).strip()
            db_name=str(row[3]).strip()
            target_table_name=str(row[2]).strip()
            schema_names_to_check_list=[(sch.strip()).upper() for sch in str(row[4]).split(',')]
            # base_table_name=ddl_repo+'.'+table_name
            base_table_name=ddl_repo.split('.')[2]
            base_db_name=ddl_repo.split('.')[0]
            base_schema_name=ddl_repo.split('.')[1]
            try:
                base_table_df=self.run_sf_query('desc table '+ddl_repo)
                
            except Exception as err:
                schema_result_df.loc[len(schema_result_df.index)]=[base_table_name,ddl_repo,"No"]
                print(str(err))
                continue

            

            schemas_present_base_table_df=self.run_sf_query("SELECT DISTINCT table_schema from {db}.INFORMATION_SCHEMA.columns  where table_name='{tn}' ".format(db=db_name.upper(),tn=base_table_name.upper()))
            schemas_present_base_table_list=[sch.upper().strip() for sch in  list(schemas_present_base_table_df['TABLE_SCHEMA'])]
            schema_list_ans=list()

            for sch in schema_names_to_check_list:
                    if sch in schemas_present_base_table_list:
                        schema_list_ans.append('Yes')
                    else:
                        schema_list_ans.append('No')
            
            
            schema_df_temp=pd.DataFrame({'TableName':base_table_name,'Schema':schema_names_to_check_list,'Present':schema_list_ans})
            schema_result_df=pd.concat([schema_result_df,schema_df_temp],ignore_index=True)

             

            full_info_schema_base_table_query="select TABLE_CATALOG,TABLE_NAME,COLUMN_NAME,ORDINAL_POSITION,COLUMN_DEFAULT,IS_NULLABLE,DATA_TYPE,CHARACTER_MAXIMUM_LENGTH,CHARACTER_OCTET_LENGTH,NUMERIC_PRECISION,NUMERIC_PRECISION_RADIX from {db}.INFORMATION_SCHEMA.columns  where table_name='{tn}' and table_schema='{schm}' ".format(db=base_db_name.upper(),tn=base_table_name.upper(),schm=base_schema_name.upper())
            info_schema_base_df=self.run_sf_query(full_info_schema_base_table_query)

            name_type_base_df=base_table_df[['name','type']]
            map_name_type=dict()
            for nam,typ in zip(name_type_base_df['name'],name_type_base_df['type']):
                map_name_type[nam]=typ

            name_type_list=list()
            
            for name in list(info_schema_base_df['COLUMN_NAME']):
                if name in map_name_type:
                    name_type_list.append(map_name_type[name])
                else:
                    name_type_list.append('Column missing in base')
            
            info_schema_base_df['DATA_TYPE']=name_type_list

            cols_all=info_schema_base_df.columns
            info_schema_base_df.columns=[col+'_Base' for col in info_schema_base_df.columns] 
            # info_schema_base_df.to_csv('abc123_base.csv')
            # name_type_base_df.to_csv('abc123_base_base_desc.csv')
            
            for schema in schema_df_temp[schema_df_temp['Present']=='Yes']['Schema']:
                schema=schema.upper()
                curr_table_name=db_name+'.'+schema+'.'+target_table_name
                curr_table_df=self.run_sf_query('desc table '+curr_table_name)
                full_info_schema_curr_table_query="select TABLE_CATALOG,TABLE_NAME,COLUMN_NAME,ORDINAL_POSITION,COLUMN_DEFAULT,IS_NULLABLE,DATA_TYPE,CHARACTER_MAXIMUM_LENGTH,CHARACTER_OCTET_LENGTH,NUMERIC_PRECISION,NUMERIC_PRECISION_RADIX from {db}.INFORMATION_SCHEMA.columns  where table_name='{tn}' and table_schema='{schm}'".format(db=db_name.upper(),tn=target_table_name.upper(),schm=schema.strip().upper())
                info_schema_curr_df=self.run_sf_query(full_info_schema_curr_table_query)

                name_type_curr_df=curr_table_df[['name','type']]
                map_name_type=dict()
                for nam,typ in zip(name_type_curr_df['name'],name_type_curr_df['type']):
                    map_name_type[nam]=typ

                name_type_list=list()

                for name in list(info_schema_curr_df['COLUMN_NAME']):
                    if name in map_name_type:
                        name_type_list.append(map_name_type[name])
                    else:
                        name_type_list.append('Column missing in target')

                info_schema_curr_df['DATA_TYPE']=name_type_list

                info_schema_curr_df.columns=[col+'_'+schema for col in info_schema_curr_df.columns]

                merged_df=info_schema_base_df.merge(info_schema_curr_df,left_on='COLUMN_NAME_Base',right_on='COLUMN_NAME'+'_'+schema,indicator=True,suffixes=['_Base','_'+schema],how='outer')
                merged_df.to_csv('merged_'+str(time.time())+'.csv')
                 
                # dfLenId=1
                for row in (merged_df.itertuples()):
                    mergeIndex=len(merged_df.columns)
                    # [database,tablename,sourceschema,targetschema,colname,atttr,base_val,target_val]
                    # print("val->",row[mergeIndex])
                    # print(mergeIndex,row)
                    
                    if(row[mergeIndex]!='both'): 
                        if(row[mergeIndex]=='left_only'):
                            
                            dataComparisonDF.loc[len(dataComparisonDF.index)]=[db_name,target_table_name,base_table_name,schema,str(getattr(row,'COLUMN_NAME_Base')),'Column Missing in Target Schema','Present','Absent']
                            # dfLenId=dfLenId+1
                        else:
                            
                            dataComparisonDF.loc[len(dataComparisonDF.index)]=[db_name,target_table_name,base_table_name,schema,str(getattr(row,'COLUMN_NAME_'+schema)),'Column Missing in Base Schema','Absent','Present']
                            # dfLenId=dfLenId+1
                    else:
                        for col in cols_all:
                            try:
                                val_base=str(getattr(row,col+'_Base')).strip()
                                val_curr=str(getattr(row,col+'_'+schema)).strip()

                                if val_base!=val_curr:
                                    dataComparisonDF.loc[len(dataComparisonDF.index)]=[db_name,target_table_name,base_table_name,schema,str(getattr(row,'COLUMN_NAME_'+schema)),col,val_base,val_curr]
                                    dfLenId=dfLenId+1
                            except: pass




                full_result_df=pd.concat([merged_df,full_result_df])

        # merged_df.to_csv('abc2423.csv')
        schema_result_df.to_csv('schema_srgeg.csv')
        dataComparisonDF.to_csv('Final result.csv')

        s=str(datetime.datetime.now())
        s=s.replace(' ','__')
        s=s.replace('-','__')
        s=s.replace(':','_')
        fileName='Result_Schema_DDL_validation+'+s+'.xlsx'

        with pd.ExcelWriter(fileName) as writer:
                    dataComparisonDF.to_excel(writer,sheet_name='DDL validation',index = False,header=True) 
                    schema_result_df.to_excel(writer,sheet_name='Schema validation',index = False,header=True) 
                     
        print('Done')
        messagebox.showinfo('Done','Final result store in result location')





    '''
    def getFilesFromSourceAndDestination(self):
        try:
           
            source_base_location=r''
            target_base_location=r''
            
            source_files_list=list()
            target_files_list=list()
            
            all_ff_source=os.listdir(source_base_location)
            for file in all_ff_source:
                if file.split('.')[-1] in ['xlsx','csv'] :
                    source_files_list.append(file)
                    
            all_ff_target=os.listdir(target_base_location)
            for file in all_ff_target:
                if file.split('.')[-1] in ['xlsx','csv'] :
                    target_files_list.append(file)

            if(len(source_files_list)!=len(target_files_list)):
                print('Different number of files present in source and destination')
                raise Exception('Different number of files present in source and destination')
                
            for i in range(len(source_files_list)):
                df_source=pd.read_csv(source_files_list)
            

        except Exception as err:
            print('Error while reading source files from prefix ',str(err))
    
    '''
    def run_MySQLWorkBench_query(self,query):
        try:
            print('Running MySQLWorkbench: ',query)
            return  pd.read_sql(query, self.cnxn_MySQLWorkBench)
        except Exception as err:
            print(err)
            raise Exception(str(err))
        
    
    def run_sqoop_query(self,query):
        try:
            print('Running sqoop',query)
            if str(self.sqoop_command_batch.get()).strip()!='':
                cmd_pre=str(self.sqoop_command_batch.get()).strip()
                cmd_pre=cmd_pre.split('--query')[0]
            else: cmd_pre='sqoop eval -Dmapreduce.job.queuename=root.bdf.bdf_yarn -Dhadoop.security.credential.provider.path=jceks://hdfs/user/srcbdfrsczbthts/sql_hip_ingest.jceks --connect "jdbc:jtds:sqlserver://VA10DWVSQL306.us.ad.wellpoint.com:10003;domain=US;databaseName=LandingZone;ssl=require;" --driver net.sourceforge.jtds.jdbc.Driver --connection-manager org.apache.sqoop.manager.SQLServerManager --username SRCHIPLZBDF --password-alias hip-ingest-password.alias'

            cmd='{cmd_pre} --query "{q}"'.format(cmd_pre=cmd_pre,q=query)
            stdin,stdout,stderr=self.SSH_ssh.exec_command(cmd)
            outlines=stdout.readlines()
            resp=''.join(outlines)
            print('Result generated')
             
            with open('temp#12$32_sql@.txt', 'w', encoding="utf-8") as f:
                f.write(resp)

            df=pd.read_csv('temp#12$32_sql@.txt',delimiter='|',encoding='utf8',skiprows=[0,2],na_values=[''], keep_default_na=False,dtype=str)
            df=df.iloc[0:-1,1:-1]

            try:
                os.remove('temp#12$32_sql@.txt')
            except: pass
            return df
        except Exception as err:
            print('during query error:',str(err))
            raise Exception(str(err))

    def run_mssql_sqoop_query(self,query):
        try:
            print('Running mssql-sqoop',query)
            if str(self.sqoop_command_base.get()).strip()!='':
                cmd_pre=str(self.sqoop_command_base.get()).strip()
                cmd_pre=cmd_pre.split('--query')[0]
            else: cmd_pre='sqoop eval -Dmapreduce.job.queuename=root.bdf.bdf_yarn -Dhadoop.security.credential.provider.path=jceks://hdfs/user/srcbdfrsczbthts/sql_hip_ingest.jceks --connect "jdbc:jtds:sqlserver://VA10DWVSQL306.us.ad.wellpoint.com:10003;domain=US;databaseName=LandingZone;ssl=require;" --driver net.sourceforge.jtds.jdbc.Driver --connection-manager org.apache.sqoop.manager.SQLServerManager --username SRCHIPLZBDF --password-alias hip-ingest-password.alias'

            cmd='{cmd_pre} --query "{q}"'.format(cmd_pre=cmd_pre,q=query)
            stdin,stdout,stderr=self.SSH_ssh.exec_command(cmd)
            outlines=stdout.readlines()
            resp=''.join(outlines)
            print('Result generated')
             
            with open('temp#12$32_sql@.txt', 'w', encoding="utf-8") as f:
                f.write(resp)

            df=pd.read_csv('temp#12$32_sql@.txt',delimiter='|',encoding='utf8',skiprows=[0,2],na_values=[''], keep_default_na=False,dtype=str)
            df=df.iloc[0:-1,1:-1]

            try:
                os.remove('temp#12$32_sql@.txt')
            except: pass
            return df
        except Exception as err:
            print('during query error:',str(err))
            raise Exception(str(err))

    def run_hive_query_direct_download(self,query,tablename):
        try:
            print('Running hive: ', query)
            hostname_hive=str(self.hostname_Hive.get()).strip()
            query=str(query).strip()
            cmd='beeline -u "jdbc:hive2://{host_hive}.wellpoint.com:10000/default;principal=hive/_HOST@US.AD.WELLPOINT.COM;ssl=true" set mapred.job.queue.name=root.bdf.bdf_yarn --outputformat=tsv2 -e "{q}"'.format(host_hive=hostname_hive,q=query)
            
            if(self.runFromImpala.get()==1):
                cmd='impala-shell -i {host_impala}.wellpoint.com -k --ssl --ca_cert=/opt/cloudera/security/truststore/ca-truststore.pem -B --print_header  -q "{q}"'.format(host_impala=hostname_hive,q=query) 
                print('Running impala')        # cmd='beeline -u "jdbc:hive2://{host_hive}.wellpoint.com:10000/default;principal=hive/_HOST@US.AD.WELLPOINT.COM;ssl=true" set mapred.job.queue.name=cii_yarn; set hive.execution.engine=tez  --outputformat=tsv2 -e "{q}"'.format(host_hive=hostname_hive,q=query)
            
             
            print(str(datetime.datetime.now())+'  Running beeline: ',cmd)
            stdin,stdout,stderr=self.hive_ssh.exec_command(cmd)
            outlines=stdout.readlines()
            resp=''.join(outlines)
            print('Result')
            with open(tablename+'.tsv', 'w', encoding="utf-8") as f:
                f.write(resp)
        except Exception as err:
            print('During hive direct download ',str(err))
            



    def run_hive_query(self,query):
        try:
            print('Running hive: ', query)
            hostname_hive=str(self.hostname_Hive.get()).strip()
            
            cmd='beeline -u "jdbc:hive2://{host_hive}.wellpoint.com:10000/default;principal=hive/_HOST@US.AD.WELLPOINT.COM;ssl=true" set mapred.job.queue.name=root.bdf.bdf_yarn --outputformat=tsv2 -e "{q}"'.format(host_hive=hostname_hive,q=query)
            
            if(self.runFromImpala.get()==1):
                cmd='impala-shell -i {host_impala}.wellpoint.com -k --ssl --ca_cert=/opt/cloudera/security/truststore/ca-truststore.pem -B --print_header  -q "{q}"'.format(host_impala=hostname_hive,q=query) 
                print('Running impala')        # cmd='beeline -u "jdbc:hive2://{host_hive}.wellpoint.com:10000/default;principal=hive/_HOST@US.AD.WELLPOINT.COM;ssl=true" set mapred.job.queue.name=cii_yarn; set hive.execution.engine=tez  --outputformat=tsv2 -e "{q}"'.format(host_hive=hostname_hive,q=query)
            
            print('Running cmd: ',cmd)
            stdin,stdout,stderr=self.hive_ssh.exec_command(cmd)
            outlines=stdout.readlines()
            resp=''.join(outlines)
            print('Result')
            with open('temp#12$32@.txt', 'w', encoding="utf-8") as f:
                f.write(resp)



            df=pd.read_csv('temp#12$32@.txt',delimiter='\t')
            try:
                os.remove('temp#12$32@.txt')
            except: pass
            return df
        except Exception as err:
            print('during query error:',str(err))
            raise Exception(str(err))


    def run_sf_prod_query(self,query):
        establish = self.con_prod.cursor()
        # print(query)
        print(str(datetime.datetime.now())+'  Running prod SF: ',query)
        establish.execute(query)
        all_rows = establish.fetchall()
        num_fields = len(establish.description)
        field_names = [i[0] for i in establish.description]
        df = pd.DataFrame(all_rows,columns=field_names)
        # print(all_rows)
        # print("**",field_names)
        # df.columns = field_names
        #return establish.fetch_pandas_all()
        return df

    def run_sf_query(self,query):
        try:
            establish = self.con.cursor()
            # print(query)
            print(str(datetime.datetime.now())+'  Running SF: ',query)
            establish.execute(query)
            all_rows = establish.fetchall()
            num_fields = len(establish.description)
            field_names = [i[0] for i in establish.description]
            df = pd.DataFrame(all_rows,columns=field_names)
            # print(all_rows)
            # print("**",field_names)
            # df.columns = field_names
            #return establish.fetch_pandas_all()
            return df
        except Exception as err:
            print('Trying to run on prod--',str(err))
            return self.run_sf_prod_query(query)
        
    def run_td_query(self,query):
        # try:
            print(str(datetime.datetime.now())+'  Running td: ',query)
            with open('f1_temp.txt','w',encoding='utf-8') as f:
                f.write(query)
            
            df=pd.DataFrame()
            with open('f1_temp.txt','r') as  f:
                df= pd.read_sql(f.read(),self.cnxn_TD)
            # try:
            os.remove('f1_temp.txt')
            # except:pass
            return df
        # except Exception as err:
        #     raise Exception(str(err))

    def run_mssql_query(self,query):
        try:
            print(str(datetime.datetime.now())+'  Running mssql: ',query)
            return pd.read_sql(query,self.cnxn_mssql)
        except Exception as err:
            raise Exception(str(err))


    
    def run_oracle_sqoop_query(self,query):
        try:
            print('Running sqoop',query)
            if str(self.sqoop_command_oracle.get()).strip()!='':
                cmd_pre=str(self.sqoop_command_oracle.get()).strip()
                cmd_pre=cmd_pre.split('--query')[0]
            else: cmd_pre='sqoop eval -Dmapreduce.job.queuename=bdf_yarn -Dhadoop.security.credential.provider.path=jceks://hdfs/user/srcbdfrsczbthts/ORA_MCCCOMP_ecdreportuser.jceks --connect jdbc:oracle:thin:@//va33dx14-scan1.wellpoint.com:1525/xeccdat --driver oracle.jdbc.driver.OracleDriver --username ECDREPORTUSER --password-alias ORA_MCCCOMP_ecdreportuser.alias '

            cmd='{cmd_pre} --query "{q}"'.format(cmd_pre=cmd_pre,q=query)
            stdin,stdout,stderr=self.SSH_ssh.exec_command(cmd)
            outlines=stdout.readlines()
            resp=''.join(outlines)
            print('Result generated')
             
            with open('temp#12$32_oracle@.txt', 'w', encoding="utf-8") as f:
                f.write(resp)

            df=pd.read_csv('temp#12$32_oracle@.txt',delimiter='|',encoding='utf8',skiprows=[0,2],na_values=[''], keep_default_na=False,dtype=str)
            df=df.iloc[0:-1,1:-1]

            try:
                os.remove('temp#12$32_oracle@.txt')
            except: pass
            return df
        except Exception as err:
            print('during sqoop query error:',str(err))
            raise Exception(str(err))


    def run_Oracle_query(self,query):
        try:
            print(str(datetime.datetime.now())+'  Running oracle query: ',query)
            return pd.read_sql(query,self.cnxn_Oracle)
        except Exception as err:
            try:
                return self.run_oracle_sqoop_query(query)
            except Exception as err2:
                print("exception occurred during ",str(err2))
                raise Exception(str(err2))
            
            
        
    
    def getCols_SF(self,table,cols='*'):
       

        try:
            
            q='desc table '+table
            df=self.run_sf_query(q)
            return list(df['name'])
        except:
            
            q='select '+cols+' from '+table +' limit 1'
            df=self.run_sf_query(q)
            return list(df.columns)

    def getCols_mssql(self,table,cols='*'):
       

        try:
            try:
                table=table.split('where')[0]
            except:pass
            q='desc table '+table
            df=self.run_mssql_query(q)
            return list(df['name'])
        except:
            
            q='select top 1 '+cols+' from '+table 
            df=self.run_mssql_query(q)
            return list(df.columns)

    def getCols_hive_forStats(self,table,cols='*'):
    
    
            q='select '+cols+' from '+table +' limit 1'
            df=self.run_hive_query(q)
            return list(df.columns)
        
        
        
    
    def getCols_Hive(self,table):
        # q='select * from '+table +' limit 1'
        try:
            table=table.split('where')[0]
            q='describe '+table
            df=self.run_hive_query(q)
            try:
                return list(df['col_name'])
            except:
                return list(df['name'])
        except:
            q='select * from '+table +' limit 1' 
            df=self.run_hive_query(q)
            return list(df.columns)

    
    
    def getCols_TD(self,table,cols='*'):
        
        try:
            table=(table.lower().split('where')[0]).strip()
            q_td='help column '+ table+'.*'
            cols_td_df=self.run_td_query(q_td)
            return list(cols_td_df['Column Dictionary Name'])
        except:
            q='select top 1 '+cols+' from '+table 
            df=self.run_td_query(q)
            return df.columns

    def getCols_Oracle(self,table):
        q='select top 1 * from '+table 
        df=self.run_Oracle_query(q)
        return df.columns
    
    
    def createRowDistinctFreqDistri(self):
        
        df_path=(self.TableStatsFile.get()).strip()
        
       
        df=pd.read_excel(df_path)
        
        for row in (df.itertuples()):
         ########################## BASIC INFO #################################
             
            try:
              cols_to_sum=[]
              
              try:
                if (str(row[5])!='nan'):
                        cols_to_sum=[(col.lower()).strip() for col in str(row[5]).split(',')]
              except: pass
              selected_cols_td='*'
              if(str(row[6]).lower()!='nan'):
                selected_cols_td=row[6].strip()

              selected_cols_sf='*'
              if(str(row[7]).lower()!='nan'):
                selected_cols_sf=row[7].strip()
            
              if(str(row[8]).lower()=='y'):
                ignore_case_during_freqDistri=True
              else: ignore_case_during_freqDistri=False

              if(str(row[9]).lower()!=''):
                ignore_cols_during_freq_distri=row[9]
              else: ignore_cols_during_freq_distri=[]

               

              self.DownloadAndGetCount(row[1],row[2],row[3],row[4],cols_to_sum,selected_cols_td,selected_cols_sf,ignore_case_during_freqDistri,ignore_cols_during_freq_distri)
            except Exception as err:
                print(row[1],str(err))

    def createRowDistinctFreqDistri_hive_to_sf(self):
        
        df_path=(self.TableStatsFile_hive_to_sf.get()).strip()
        
       
        df=pd.read_excel(df_path)
        
        for row in (df.itertuples()):
         ########################## BASIC INFO #################################
            # print(row[0])
            # print(row[1])
            # print(row[2])
            # print(row[3])
            # print(row[4])
            try:
              cols_to_sum=[]
              
              try:
                if (str(row[5])!='nan'):
                        cols_to_sum=[(col.lower()).strip() for col in str(row[5]).split(',')]
              except: pass
              selected_cols_hive='*'
              try:
                if(str(row[6]).lower()!='nan'):
                    selected_cols_hive=row[6].strip()
              except: pass

              selected_cols_sf='*'

              try:
                if(str(row[7]).lower()!='nan'):
                    selected_cols_sf=row[7].strip()
              except: pass

              self.DownloadAndGetCount_hive_sf(row[1],row[2],row[3],row[4],cols_to_sum,selected_cols_hive,selected_cols_sf)
            except Exception as err:
                print(row[1],str(err))

    def createRowDistinctFreqDistri_TD(self):
        
        df_path=(self.TableStatsFile_TD.get()).strip()
        print(df_path)
       
        df=pd.read_excel(df_path)
        
        for row in (df.itertuples()):
         ########################## BASIC INFO #################################
            # print(row[0])
            # print(row[1])
            # print(row[2])
            # print(row[3])
            # print(row[4])
            try:
               cols_to_ignore=[]
               try:
                if (str(row[5])!='nan'):
                        cols_to_ignore=[(col.lower()).strip() for col in str(row[5]).split(',')]
               except:pass
               self.DownloadAndGetCount_TD(row[1],row[2],cols_to_ignore)
            except Exception as err:
                print("table ",row[1]," where ",row[2]," getting error: ",str(err))
         

    def createRowDistinctFreqDistri_Oracle(self):
        
        df_path=(self.TableStatsFile_Oracle.get()).strip()
        print(df_path)
       
        df=pd.read_excel(df_path)
        
        for row in (df.itertuples()):
         ########################## BASIC INFO #################################
            # print(row[0])
            # print(row[1])
            # print(row[2])
            # print(row[3])
            # print(row[4])
            try:
               cols_to_ignore=[]
               try:
                if (str(row[5])!='nan'):
                        cols_to_ignore=[(col.lower()).strip() for col in str(row[5]).split(',')]
               except:pass
               self.DownloadAndGetCount_Oracle(row[1],row[2],cols_to_ignore)
            except Exception as err:
                print("table ",row[1]," where ",row[2]," getting error: ",str(err))

    
    

    
    def DownloadAndGetCount_Oracle(self,Oracle_table,Oracle_condition,cols_to_ignore=[]):
        path=self.tableStatsResulOracleownloadLocationEntry_Oracle.get()
        os.chdir(path)
        try:
            ########################## BASIC INFO #################################
        
            #Oracle_table=row[1]
            #Oracle_condition=row[2]
            
            if (str(Oracle_condition)!='nan'):
                        Oracle_table +=' where '+Oracle_condition
            
            #sf_table=row[3]
            #sf_condition=row[4]
            
            
            Oracle_table_cols=self.getCols_Oracle(Oracle_table)
            
            
            print('***Basic info checked for:',Oracle_table)
            
            ########################## TAB-1 SumOfValues #################################
            try:
                Oracle_checksum=list()
                measure_cols_Oracle=''
                measure_cols_Oracle_names_list=list()
                for col in Oracle_table_cols:
                    try:
                        t_q='select cast(sum({colName}) as double) as c from {tn}'.format(colName=col,tn=Oracle_table)
                        # print("Running query: ",t_q)
                        res=self.run_Oracle_query(t_q)
                        Oracle_checksum.append(res.iloc[:,0][0])
                        measure_cols_Oracle +=', cast(sum({colName}) as double) as Oracle_Sum_{colName} '.format(colName=col)
                        measure_cols_Oracle_names_list.append('Oracle_Sum_'+col)
                    except:
                        Oracle_checksum.append(np.nan)
                

                
                        
                tab1_df=pd.concat([pd.DataFrame({'Oracle_ColName':Oracle_table_cols}),pd.DataFrame({'Oracle_CheckSum':Oracle_checksum})],axis=1)
                

                tab1_df.to_csv('Tab1_'+Oracle_table+'.csv',index=False)
                print('SumOfValues Tab (1st) created')
            except Exception as err:
                print('Not able to complete tab1 due to: ',str(err)) 

            ########################## TAB-2 Total Row counts #################################

            try:
                tq='select cast(count(*) as double) as Oracle_ROW_COUNT from  '+ Oracle_table
                # print("Running query: ",tq)
                row_counts=self.run_Oracle_query(tq)
                
                    
                
                #result
                row_counts.to_csv('Tab2_'+Oracle_table+'.csv',index=False)    
                print('Row Count Tab (2nd) created')
            except Exception as err:
                print('Not able to complete tab2 due to: ',str(err))
            ########################## TAB-3 DISTINCT COUNT #################################
            try:
                #Oracle
                Oracle_distinct_query_list=list()
                for col in Oracle_table_cols:
                    Oracle_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                
                Oracle_distinct_query=",".join(Oracle_distinct_query_list)
                
                finalDistinct_query='select ' + Oracle_distinct_query + " from "+ Oracle_table
                # print("Running query: ",finalDistinct_query)
                final_Distinct=self.run_Oracle_query(finalDistinct_query)
                final_Distinct=final_Distinct.transpose().reset_index()
                final_Distinct.columns=['Oracle_ColName','Oracle_DistinctCount']
                
                    
                    
                
                
                    
                final_Distinct.to_csv('Tab3_'+Oracle_table+'.csv',index=False) 
                print('Distinct Counts Tab (3rd) Created')
            except Exception as err:
                print('Not able to complete tab3 due to: ',str(err))
                
            ########################## TAB-4 FREQUENCY DISTRIBUTION #################################
            
            #Oracle
            
            
            try:
                df_Oracle=pd.DataFrame()
                
                

                print("****Oracle queries")
                for col in Oracle_table_cols and col not in cols_to_ignore:
                    query='select top 1000 {colName},cast(count({colName}) as double) as Oracle_val_row_count {measureCols}  from {tname} group by {colName} order by Oracle_val_row_count desc '.format(colName=col,tname=Oracle_table,measureCols=measure_cols_Oracle)
                    
                    # print("Running query:",query)
                    res_df=self.run_Oracle_query(query)
                    res_df.insert(0,'Oracle_ColName',col)
                    res_df.columns=['Oracle_ColName','Oracle_col_value','Oracle_val_row_count' ] + measure_cols_Oracle_names_list
                    df_Oracle=df_Oracle.append(res_df)
                    
                    
                try:
                    finalFD=df_Oracle.apply(pd.to_numeric, errors='coerce').fillna(df_Oracle)
                except:pass
                
                    
                            
                #    df_Oracle=df_Oracle.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                    
                #making FreqDistri
                
                print('Creating final freq distribution')
                    
                tableName=Oracle_table.split('.')[1:][0]
                finalFD.to_csv('Tab4_'+Oracle_table+'.csv',index=False)
                print('Frequency Distribution Tab (4th) created')
                self.writeIntoExcel_stats(tableName,tab1_df,row_counts,final_Distinct,finalFD)
                # try:
                #     os.remove('Tab1_'+Oracle_table+'.csv')
                #     os.remove('Tab2_'+Oracle_table+'.csv')
                #     os.remove('Tab3_'+Oracle_table+'.csv')
                #     os.remove('Tab4_'+Oracle_table+'.csv')
                # except:
                #     print("Could not delete intermediate tab csv files as file either doesn't exist or it's opened")

                print('Excel created!')
            except Exception as err:
                print('Not able to complete tab4 due to: ',str(err))
        except Exception as err:
            raise Exception('Download error: ',str(err))



    def DownloadAndGetCount_TD(self,td_table,td_condition,cols_to_ignore=[]):
        path=self.tableStatsResultDownloadLocationEntry_TD.get()
        os.chdir(path)
        try:
            ########################## BASIC INFO #################################
        
            #td_table=row[1]
            #td_condition=row[2]
            original_td_table=td_table
            if (str(td_condition)!='nan'):
                        td_table +=' where '+td_condition
            
            #sf_table=row[3]
            #sf_condition=row[4]
            
            
            td_table_cols=self.getCols_TD(original_td_table)
            
            
            print('***Basic info checked for:',original_td_table)
            
            ########################## TAB-1 SumOfValues #################################
            
            try:
                td_checksum=list()
                measure_cols_td=''
                measure_cols_td_names_list=list()
                for col in td_table_cols:
                    try:
                        t_q='select cast(sum({colName}) as float) as c from {tn}'.format(colName=col,tn=td_table)
                       # print("Running query: ",t_q)
                        res=self.run_td_query(t_q)
                        td_checksum.append(res.iloc[:,0][0])
                        measure_cols_td +=', cast(sum({colName}) as float) as TD_Sum_{colName} '.format(colName=col)
                        measure_cols_td_names_list.append('TD_Sum_'+col)
                    except:
                        td_checksum.append(np.nan)
                

                
                        
                tab1_df=pd.concat([pd.DataFrame({'TD_ColName':td_table_cols}),pd.DataFrame({'TD_CheckSum':td_checksum})],axis=1)
                

                tab1_df.to_csv('Tab1_'+original_td_table+'.csv',index=False)
                print('SumOfValues Tab (1st) created')

            except Exception as err:
                print('Not able to complete tab1 due to: ',str(err))

            ########################## TAB-2 Total Row counts #################################
            try:
                tq='select cast(count(*) as float) as TD_ROW_COUNT from  '+ td_table
                # print("Running query: ",tq)
                row_counts=self.run_td_query(tq)
                
                    
                
                #result
                row_counts.to_csv('Tab2_'+original_td_table+'.csv',index=False)    
                print('Row Count Tab (2nd) created')
            except Exception as err:
                print('Not able to complete tab2 due to: ',str(err))
            ########################## TAB-3 DISTINCT COUNT #################################
        
            #TD
            try:
                td_distinct_query_list=list()
                for col in td_table_cols:
                    td_distinct_query_list.append("cast(count(distinct({colName}) (CASESPECIFIC) ) as float) as {colName}".format(colName=col))
                
                td_distinct_query=",".join(td_distinct_query_list)
                
                finalDistinct_query='select ' + td_distinct_query + " from "+ td_table
                # print("Running query: ",finalDistinct_query)
                final_Distinct=self.run_td_query(finalDistinct_query)
                final_Distinct=final_Distinct.transpose().reset_index()
                final_Distinct.columns=['TD_ColName','TD_DistinctCount']
                
                    
                    
                
                
                    
                final_Distinct.to_csv('Tab3_'+original_td_table+'.csv',index=False) 
                print('Distinct Counts Tab (3rd) Created')
            except Exception as err:
                print('Not able to complete tab3 due to: ',str(err))    
            ########################## TAB-4 FREQUENCY DISTRIBUTION #################################
            
            #TD
            
            try:
                df_TD=pd.DataFrame()
                
                print("****TD queries")
                for col in td_table_cols: 
                    # if col.lower() not in cols_to_ignore:
                        query='select top 1000 {colName},cast(count({colName}) as float) as td_val_row_count {measureCols}  from {tname} group by {colName} order by td_val_row_count desc '.format(colName=col,tname=td_table,measureCols=measure_cols_td)
                        
                        # print("Running query:",query)
                        res_df=self.run_td_query(query)
                        res_df.insert(0,'TD_ColName',col)
                        res_df.columns=['TD_ColName','TD_col_value','TD_val_row_count' ] + measure_cols_td_names_list
                        df_TD=df_TD.append(res_df)
                    
                    
                try:
                    finalFD=df_TD.apply(pd.to_numeric, errors='coerce').fillna(df_TD)
                except:pass
                
                    
                            
                #    df_TD=df_TD.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                    
                #making FreqDistri
                
                print('Creating final freq distribution')
                    
                tableName=td_table.split('.')[1:][0]
                finalFD.to_csv('Tab4_'+original_td_table+'.csv',index=False)
                print('Frequency Distribution Tab (4th) created')
                self.writeIntoExcel_stats(tableName,tab1_df,row_counts,final_Distinct,finalFD)
                # try:
                #     os.remove('Tab1_'+td_table+'.csv')
                #     os.remove('Tab2_'+td_table+'.csv')
                #     os.remove('Tab3_'+td_table+'.csv')
                #     os.remove('Tab4_'+td_table+'.csv')
                # except:
                #     print("Could not delete intermediate tab csv files as file either doesn't exist or it's opened")

                print('Excel created!')
            except Exception as err:
                print('Not able to complete tab4 due to: ',str(err))
        except Exception as err:
            raise Exception('Download error: ',str(err))


    def getTimestampColumns(self):
        df_timestamps=pd.DataFrame(columns=['TableName','Timestamp_Cols'])

        # sf_tables=['R01_EDM_V3.EDM_V3_NOGBD.DDIM_PRL','R01_EDM_V3.EDM_V3_STG_NOGBD.PAS_EP']
        path=self.download_path_SF_timestamps.get()
        os.chdir(path)
        sf_tables=[]

        with open(self.query_SF_timestamps.get(),'r',encoding='utf-8') as file:
            sf_tables=(str(file.read())).split(',')
        

        absent_tables_ts=list()

        for table in sf_tables:
            try:
                table=table.strip()
                query='desc table '+ table 

                df=self.run_sf_query(query)
                df=df[df['type'].str.lower().str.contains('time') | df['type'].str.lower().str.contains('date')]
                ts=",".join(list(df['name']))
                df_timestamps.loc[len(df_timestamps.index)]=[table,ts]
            except:
                print(table, 'Not Found')
                absent_tables_ts.append(table)

        absent_tables_df=pd.DataFrame({"TableMissing":absent_tables_ts})
        fileName='timestamps_'+str(time.time())+'.xlsx'
        with pd.ExcelWriter(fileName) as writer:
                    df_timestamps.to_excel(writer,sheet_name='Timestamp',index = False,header=True) 
                    absent_tables_df.to_excel(writer,sheet_name='Absent Tables',index = False,header=True) 
                    
         
        print("Timestamps file saved as "+fileName+ 'in location',path)
        messagebox.showinfo('Done',"Timestamps file saved as "+fileName+ 'in location'+path)

    
    def DownloadAndGetCount(self,td_table,td_condition,sf_table,sf_condition,cols_to_sum,selected_cols_td,selected_cols_sf,ignore_case_during_freqDistri=False,ignore_cols_during_freq_distri=[]):
        ########################## BASIC INFO #################################
       
        path=self.tableStatsResultDownloadLocationEntry.get()
        os.chdir(path)
           #td_table=row[1]
           #td_condition=row[2]
        original_td_table=td_table
        original_sf_table=sf_table
        try:
            if (str(td_condition)!='nan'):
                        td_table +=' where '+td_condition
            
            #sf_table=row[3]
            #sf_condition=row[4]
            
            if (str(sf_condition)!='nan'):
                        sf_table +=' where '+sf_condition
            
                
            
            td_table_cols=self.getCols_TD(td_table,selected_cols_td)
            sf_cols=self.getCols_SF(original_sf_table,selected_cols_sf)

            if(selected_cols_td!='*'):
                td_table_cols=[x.strip().lower() for x in str(selected_cols_td).split(',')]
            if(selected_cols_sf!='*'):
                sf_cols=[x.strip().lower() for x in str(selected_cols_sf).split(',')]



            
            print('***Basic info checked for:',sf_table)
            measure_cols_td=''
            measure_cols_td_names_list=list()

            measure_cols_sf=''
            measure_cols_sf_names_list=list()

            
         

            
            
            
            ########################## TAB-1 SumOfValues #################################

            try:
                measure_cols_td=''
                measure_cols_td_names_list=list()

                measure_cols_sf=''
                measure_cols_sf_names_list=list()

                td_checksum=list()
                for col in td_table_cols:
                    try:
                        t_q='select cast(sum({colName}) as float )as c from {tn}'.format(colName=col,tn=td_table)
                        #    print("Running query: ",t_q)
                        res=self.run_td_query(t_q)
                        td_checksum.append(res.iloc[:,0][0])
                        # measure_cols_td +=', cast(sum({colName}) as double) as TD_Sum_{colName} '.format(colName=col)
                        # measure_cols_td_names_list.append(str(col).lower()+"_TD")
                    except:
                        td_checksum.append(np.nan)
                
                for col in cols_to_sum:
                    # if col in td_table:
                        measure_cols_td +=', cast(sum({colName}) as float) as TD_Sum_{colName} '.format(colName=col)
                        measure_cols_td_names_list.append(str(col).lower()+"_TD")

                sf_checksum=list()
                
                for col in sf_cols:
                    try:
                        s_q='select cast( sum({colName}) as double) as c from {tn}'.format(colName=col,tn=sf_table)
                       #    print("Running query: ",s_q)
                        res=self.run_sf_query(s_q)
                        sf_checksum.append(res.iloc[:,0][0])
                        # measure_cols_sf +=', cast(sum({colName}) as double) as SF_Sum_{colName} '.format(colName=col)
                        # measure_cols_sf_names_list.append(str(col).lower()+"_SF")
                    except:
                        sf_checksum.append(np.nan)

                db_name,schema_name,table_name=original_sf_table.split('.')         
                for col in cols_to_sum:
                    
                        measure_cols_sf +=', cast(sum({colName}) as double) as SF_Sum_{colName} '.format(colName=col)
                        measure_cols_sf_names_list.append(str(col).lower()+"_SF")
                
                clustering_key_sf1="SELECT  CLUSTERING_KEY from {db_name}.INFORMATION_SCHEMA.tables where table_name='{tn}' and table_schema='{ts}' ".format(db_name=db_name.strip().upper(),tn=table_name.strip().upper(),ts=schema_name.strip().upper())
                clustering_key_sf1_df=self.run_sf_query(clustering_key_sf1)
                try:
                    indexed_cols=[s.strip() for s in clustering_key_sf1_df['CLUSTERING_KEY'][0].replace(')','').replace('LINEAR(','').split(',')]
                except:indexed_cols=list()

                indexed_cols_yes_no=list()
                for col in sf_cols:
                    if col in indexed_cols:
                        indexed_cols_yes_no.append('Yes')
                    else: indexed_cols_yes_no.append('No')
                
                

                col_info_query="SELECT  COLUMN_NAME as SF_ColName ,CHARACTER_MAXIMUM_LENGTH as SF_ColumnLength  from {db_name}.INFORMATION_SCHEMA.columns where table_name='{tn}' and table_schema='{ts}' ".format(db_name=db_name.strip().upper(),tn=table_name.strip().upper(),ts=schema_name.strip().upper())
                
                sf_col_info_df=self.run_sf_query(col_info_query)
                sf_col_info_df.columns=['SF_ColName','SF_ColumnLength']
                   

                datatype_query='desc table '+ original_sf_table
                sf_datatype_df=self.run_sf_query(datatype_query)
                sf_datatype_df=sf_datatype_df[['name','type']]
                sf_datatype_df.columns=['SF_ColName','SF_Datatype']



               
                
                sf_indexed_cols_df=pd.DataFrame({'SF_ColName':sf_cols,'SF_Indexed':indexed_cols_yes_no})

                df2_sf=pd.concat([ pd.DataFrame({'SF_ColName':sf_cols}),pd.DataFrame({'SF_CheckSum':sf_checksum})],axis=1)
                df2_sf=df2_sf.merge(sf_col_info_df,left_on='SF_ColName',right_on='SF_ColName',suffixes=['_Base','_Release'])
                df2_sf=df2_sf.merge(sf_indexed_cols_df,left_on='SF_ColName',right_on='SF_ColName',suffixes=['_Base','_Release'])
                df2_sf=df2_sf.merge(sf_datatype_df,left_on='SF_ColName',right_on='SF_ColName')
                
                
                colLengthIndexed_query_td='help column '+ original_td_table+'.*'
                colLengthIndexed_query_td_df=self.run_td_query(colLengthIndexed_query_td)
                colLengthIndexed_query_td_df=colLengthIndexed_query_td_df[['Column Dictionary Name','Max Length','Indexed?']]
                colLengthIndexed_query_td_df.columns=['TD_ColName','TD_ColumnLength','TD_Indexed']


                query_creation_for_td_datatype=''
                for col in td_table_cols:
                    query_creation_for_td_datatype +=', TYPE({colName}) as  {colName}'.format(colName=col)

                if(len(query_creation_for_td_datatype)>0):
                    query_creation_for_td_datatype=query_creation_for_td_datatype[1:]


                td_datatype_query='Select Distinct {query} from {tn}'.format(query=query_creation_for_td_datatype,tn=original_td_table)
                td_datatype_query_df=self.run_td_query(td_datatype_query)
                td_datatype_query_df=td_datatype_query_df.transpose().reset_index()
                td_datatype_query_df.columns=['TD_ColName','TD_Datatype']

                df1_td=pd.concat([pd.DataFrame({'TD_ColName':td_table_cols}),pd.DataFrame({'TD_CheckSum':td_checksum})],axis=1)
                df1_td=df1_td.merge(colLengthIndexed_query_td_df,left_on='TD_ColName',right_on='TD_ColName')
                df1_td=df1_td.merge(td_datatype_query_df,left_on='TD_ColName',right_on='TD_ColName')
                
                
                left_indexes=['TD_ColName']
                right_indexes=['SF_ColName']
                 
                df1_td['TD_ColumnLength'] = pd.to_numeric(df1_td['TD_ColumnLength'],errors='coerce').astype('float64').fillna(df1_td['TD_ColumnLength']).tolist()
                df1_td['TD_CheckSum'] = pd.to_numeric(df1_td['TD_CheckSum'],errors='coerce').astype('float64').fillna(df1_td['TD_CheckSum']).tolist()
                df2_sf['SF_ColumnLength'] = pd.to_numeric(df2_sf['SF_ColumnLength'],errors='coerce').astype('float64').fillna(df2_sf['SF_ColumnLength']).tolist()
                df2_sf['SF_CheckSum'] = pd.to_numeric(df2_sf['SF_CheckSum'],errors='coerce').astype('float64').fillna(df2_sf['SF_CheckSum']).tolist()


                tab1_df=df1_td.merge(df2_sf,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_TD','_SF'])

                finalCols_tab1=['SF_ColName','SF_Datatype','SF_ColumnLength','SF_CheckSum','SF_Indexed',
                                'TD_ColName','TD_Datatype','TD_ColumnLength','TD_CheckSum','TD_Indexed']
                tab1_df=tab1_df[finalCols_tab1]


                tab1_df['ColNameCheck_Result']=tab1_df['SF_ColName']==tab1_df['TD_ColName']
                tab1_df['Datatype_Result']=tab1_df['SF_Datatype']==tab1_df['TD_Datatype']
                tab1_df['ColumnLength_Result']=tab1_df['SF_ColumnLength'].eq(tab1_df['TD_ColumnLength'])
                tab1_df['Checksum_Result']=tab1_df['SF_CheckSum'].eq(tab1_df['TD_CheckSum'])
                tab1_df['Indexed_Result']=tab1_df['SF_Indexed'].eq(tab1_df['TD_Indexed'])
                
                 
            #     tab1_df['diff_%']=100*tab1_df['diff']/tab1_df['TD_CheckSum']
                
                tab1_df.to_csv('Tab1_'+original_td_table+'.csv',index=False)
                print('SumOfValues Tab (1st) created')

            except Exception as err:
                print('Not able to complete tab1 due to: ',str(err))
            ########################## TAB-2 Total Row counts #################################

            try:
                tq='select cast(count(*) as float) as TD_ROW_COUNT from  '+ td_table
                #    print("Running query: ",tq)
                td_count=self.run_td_query(tq)
                
                sq='select cast(count(*) as double) as SF_ROW_COUNT from  '+ sf_table
                #    print("Running query: ",sq)
                sf_count=self.run_sf_query(sq)
                
                
                #result
                
                row_counts=pd.concat([td_count,sf_count],axis=1)
                row_counts['Difference(TD-SF)']=row_counts['TD_ROW_COUNT']-row_counts['SF_ROW_COUNT']
                row_counts['Difference_%']=100*row_counts['Difference(TD-SF)']/row_counts['TD_ROW_COUNT']
                
                row_counts.to_csv('Tab2_'+original_td_table+'.csv',index=False)
                print('Row Count Tab (2nd) created')
            except Exception as err:
                print('Not able to complete tab2 due to: ',str(err))
            
            ########################## TAB-3 DISTINCT COUNT #################################
            
            #TD
            try:
                #td_distinct_query_list=list()
                td_distinct_query_res=list()
                for col in td_table_cols:
                    try:
                        q="select cast(count(distinct({colName}) (CASESPECIFIC) ) as float) as {colName} from {tn}".format(colName=col,tn=td_table)
                        # td_distinct_query_list.append()
                        val_df=self.run_td_query(q)
                        val=val_df.iloc[0][0]
                        td_distinct_query_res.append(val)
                    except:
                        q="select cast(count(distinct({colName}) ) as float) as {colName} from {tn}".format(colName=col,tn=td_table)
                        # td_distinct_query_list.append()
                        val_df=self.run_td_query(q)
                        val=val_df.iloc[0][0]
                        td_distinct_query_res.append(val)

                
                # td_distinct_query=",".join(td_distinct_query_list)
                
                # finalDistinct_query='select ' + td_distinct_query + " from "+ td_table
                # #    print("Running query: ",finalDistinct_query)
                # td_res_df=self.run_td_query(finalDistinct_query)
                # td_res_df=td_res_df.transpose().reset_index()
                td_res_df=pd.DataFrame({'TD_ColName':td_table_cols,'TD_DistinctCount':td_distinct_query_res})
                td_res_df.columns=['TD_ColName','TD_DistinctCount']
                
                #SF
                sf_distinct_query_list=list()
                for col in sf_cols:
                    sf_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                
                sf_distinct_query=",".join(sf_distinct_query_list)
                
                finalDistinct_query='select ' + sf_distinct_query + " from "+ sf_table

                #    print("Running query: ",finalDistinct_query)

                sf_res_df=self.run_sf_query(finalDistinct_query)
                sf_res_df=sf_res_df.transpose().reset_index()
                sf_res_df.columns=['SF_ColName','SF_DistinctCount']
                
                
                
                
                #making finalDistinctRes
                left_indexes=['TD_ColName']
                right_indexes=['SF_ColName']
                td_res_df=td_res_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                sf_res_df=sf_res_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                
                final_Distinct=td_res_df.merge(sf_res_df,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_TD','_SF'])
                
                #final_Distinct=pd.concat([td_res_df,sf_res_df],axis=1)
                final_Distinct['Difference(TD-SF)']=final_Distinct['TD_DistinctCount']-final_Distinct['SF_DistinctCount']
                final_Distinct['Difference_%']=100*final_Distinct['Difference(TD-SF)']/final_Distinct['TD_DistinctCount']
                
                final_Distinct.to_csv('Tab3_'+original_sf_table+'.csv',index=False)
                print('Distinct Counts Tab (3rd) Created')

            except Exception as err:
                print('Not able to complete tab3 due to: ',str(err))    


            ########################## TAB-4 FREQUENCY DISTRIBUTION (td,sf)#################################
            
            #TD
            #  
            
            try:
                td_table_cols=[col.strip().upper() for col in td_table_cols]
                sf_cols=[col.strip().upper() for col in sf_cols]

                

                for col in str(ignore_cols_during_freq_distri).strip().split(','):
                    if col.strip().upper() in td_table_cols: td_table_cols.remove(col.strip().upper())
                    if col.strip().upper() in sf_cols: sf_cols.remove(col.strip().upper())

                common_cols=list()
                for col in td_table_cols:
                    if col in sf_cols:
                        common_cols.append(col)

                finalFD=pd.DataFrame()
                ts_cndtn=''
                for col in common_cols:
                    #td
                    try:
                        td_query='select top 1000 {colName} (CASESPECIFIC),cast(count({colName} (CASESPECIFIC))  as float) as td_val_row_count {measureCols} from {tname} group by {colName}(CASESPECIFIC) order by td_val_row_count, {colName}(CASESPECIFIC) desc '.format(colName=col,tname=td_table,measureCols=measure_cols_td)
                        td_res_df=self.run_td_query(td_query)
                        
                    except:
                        td_query='select top 1000 {colName},cast(count({colName} ) as float) as td_val_row_count {measureCols} from {tname} group by {colName} order by td_val_row_count, {colName} desc '.format(colName=col,tname=td_table,measureCols=measure_cols_td)
                        td_res_df=self.run_td_query(td_query)

                    try:
                        try:
                            int(td_res_df.iloc[0][0])
                        except:
                            
                            val_ts_n=td_res_df[col].isna().sum()
                            if val_ts_n>0: ts_cndtn=' or c1 is null '

                            x=pd.to_datetime()
                            td_res_df.dropna(inplace=True)
                             


                            td_res_df[col]=pd.to_datetime(td_res_df[col])
                    except: pass

                    # values=list(td_res_df[col].astype(str))
                    # # values=",".join(values)
                    # values=str(values)[1:-1]
                    values=str(list(td_res_df[col].astype(str)))[2:-2].replace('"',"'").replace("',",'#,').replace(", '","$^").replace("'", "\\'") 
                    values="'"+values.replace('#$^',"','")+"'"
                    


                    td_res_df.insert(0,'TD_ColName',col)
                    td_res_df.columns=['TD_ColName','TD_col_value','TD_val_row_count']  + measure_cols_td_names_list
                    #sf
                    try:
                        query='select cast({colName} as double) as c1,cast(count(c1) as double) as sf_val_row_count {measureCols}   from {tname}  group by c1 having c1 in ({values}) {cndtn} '.format(cndtn=ts_cndtn,colName=col,tname=sf_table,measureCols=measure_cols_sf,values=values)
                        #query='select cast({colName} as double) as c1,cast(count(c1) as double) as sf_val_row_count {measureCols}   from {tname} group by c1 order by sf_val_row_count, c1 desc limit 1000'.format(colName=col,tname=sf_table,measureCols=measure_cols_sf)
                        sf_res_df=self.run_sf_query(query)
                    except:
                        query='select {colName} as c1,cast(count(c1) as double) as sf_val_row_count {measureCols}   from {tname}  group by c1 having c1 in ({values}) {cndtn} '.format(cndtn=ts_cndtn,colName=col,tname=sf_table,measureCols=measure_cols_sf,values=values)
                        #query='select {colName},cast(count({colName}) as double) as sf_val_row_count {measureCols}   from {tname} group by {colName} order by sf_val_row_count, {colName} desc limit 1000'.format(colName=col,tname=sf_table,measureCols=measure_cols_sf)
                        sf_res_df=self.run_sf_query(query)

                    
                    
                    sf_res_df.insert(0,'SF_ColName',col)
                    sf_res_df.columns=['SF_ColName','SF_col_value','SF_val_row_count'] + measure_cols_sf_names_list

                    #combining
                    left_indexes=['TD_ColName','TD_col_value']
                    right_indexes=['SF_ColName','SF_col_value']
                    
                    td_res_df['TD_ColName']=td_res_df['TD_ColName'].astype(str)
                    sf_res_df['SF_ColName']=sf_res_df['SF_ColName'].astype(str)
                    
                    try:
                        # dt=str(datetime.datetime.fromtimestamp( float(tsVal))) # if format is epoch
                        # s, ms = divmod(float(tsVal), 1000)  # (1236472051, 807)
                        # '%s.%03d' % (time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(s)), ms)
                        values_dts=[str(x).strip() for x in list(td_res_df['TD_col_value'][:5])]
                        for val in values_dts:
                            if val in ['nan','NaN','None','na','']:continue
                            if '-' not in val:
                                if '/' not in val:
                                    raise Exception('ts')
                        td_res_df['TD_col_value'].apply(pd.to_datetime)

                    except:
                        td_res_df['TD_col_value']=td_res_df['TD_col_value'].astype(str)
                        # td_res_df['TD_col_value'] = pd.to_numeric(td_res_df['TD_col_value'],errors='coerce').astype('float64').fillna(td_res_df['TD_col_value']).tolist()
                        try:
                            td_res_df['TD_col_value']=td_res_df['TD_col_value'].map(lambda x: Decimal(x))
                        except: pass
                    

                    try:
                        # dt=str(datetime.datetime.fromtimestamp( float(tsVal))) # if format is epoch
                        # s, ms = divmod(float(tsVal), 1000)  # (1236472051, 807)
                        # '%s.%03d' % (time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(s)), ms)
                        values_dts=[str(x).strip() for x in list(sf_res_df['SF_col_value'][:5])]
                        for val in values_dts:
                            if val in ['nan','NaN','None','na','']:continue
                            if '-' not in val:
                                if '/' not in val:
                                    raise Exception('ts')
                                
                        sf_res_df['SF_col_value'].apply(pd.to_datetime)

                    except:
                        # sf_res_df['SF_col_value'] = pd.to_numeric(sf_res_df['SF_col_value'],errors='coerce').astype('float64').fillna(sf_res_df['SF_col_value']).tolist()
                        sf_res_df['SF_col_value']=sf_res_df['SF_col_value'].astype(str)
                        
                        try:
                            sf_res_df['SF_col_value']=sf_res_df['SF_col_value'].map(lambda x: Decimal(x))
                        except: pass

                    
                    

                    td_res_df['TD_col_value']=td_res_df['TD_col_value'].astype(str)
                    sf_res_df['SF_col_value']=sf_res_df['SF_col_value'].astype(str)

                    td_res_df['TD_val_row_count']=td_res_df['TD_val_row_count'].astype(str)
                    td_res_df['TD_val_row_count']=td_res_df['TD_val_row_count'].map(lambda x: Decimal(x))  

                    # td_res_df['TD_val_row_count'] = pd.to_numeric(td_res_df['TD_val_row_count'], errors='coerce').fillna(td_res_df['TD_val_row_count']).tolist()
                    
                    # sf_res_df['SF_val_row_count'] = pd.to_numeric(sf_res_df['SF_val_row_count'], errors='coerce').fillna(sf_res_df['SF_val_row_count']).tolist()
                    sf_res_df['SF_val_row_count']=sf_res_df['SF_val_row_count'].astype(str)
                    sf_res_df['SF_val_row_count']=sf_res_df['SF_val_row_count'].map(lambda x: Decimal(x)) 

                    td_res_df=td_res_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                    sf_res_df=sf_res_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                    #making FreqDistri

                    print(' not Creating sub-final freq distribution')
                    # sub_finalFD=td_res_df.merge(sf_res_df,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_TD','_SF'])
                    # sub_finalFD.to_csv('subfinal.csv',index=False) #debug

                    # reverse start

                    #sf
                    try:
                            query='select cast({colName} as double) as c1,cast(count(c1) as double) as sf_val_row_count {measureCols}   from {tname} group by c1 order by sf_val_row_count, c1 desc limit 1000'.format(colName=col,tname=sf_table,measureCols=measure_cols_sf)
                            sf_res_df2=self.run_sf_query(query)
                            sf_res_df2.columns=[col,'sf_val_row_count']
                    except:
                            query='select {colName},cast(count({colName}) as double) as sf_val_row_count {measureCols}   from {tname} group by {colName} order by sf_val_row_count, {colName} desc limit 1000'.format(colName=col,tname=sf_table,measureCols=measure_cols_sf)
                            sf_res_df2=self.run_sf_query(query)


                    # values=list(sf_res_df2[col].astype(str))   
                    # # values=",".join(values)
                    # values=str(values)[1:-1]
                    sf_cndtn=''
                    try:
                        try:
                            int(sf_res_df2.iloc[0][0])
                        except:
                            
                            val_ts_n=sf_res_df2[col].isna().sum()
                            if val_ts_n>0: sf_cndtn=' or c1 is null '

                            x=pd.to_datetime(sf_res_df2[col])
                            sf_res_df2.dropna(inplace=True)
                             


                            sf_res_df2[col]=pd.to_datetime(sf_res_df2[col])
                    except: pass

                    values=str(list(sf_res_df2[col].astype(str)))[2:-2].replace('"',"'").replace("',",'#,').replace(", '","$^").replace("'", "''") 
                    values="'"+values.replace('#$^',"','")+"'"

                    #td
                    # select cast({colName} AS TIMESTAMP FORMAT 'dd/mm/yyyyBhh:mi:SS.s(6)') from T16_ETL_VIEWS_PI_XM.MCT_RQST_RAR_MBR_DROP_DTL where cast(dcg_incrd_prd_bgn_dt AS TIMESTAMP FORMAT 'dd/mm/yyyyBhh:mi:SS.s(6)')>'2021-07-01 00:00:00'
                    
                    try:
                        td_query='select cast({colName}  AS TIMESTAMP FORMAT \'dd/mm/yyyyBhh:mi:SS.s(6)\')   as c1,cast(count(c1)  as float) as td_val_row_count {measureCols}  from {tname} group by c1 having c1 in ({values}) {cndtn}'.format(cndtn=sf_cndtn,colName=col,tname=td_table,measureCols=measure_cols_td,values=(values).strip())
                        td_res_df2=self.run_td_query(td_query)
                        td_res_df2.columns=[col,'td_val_row_count']
                    except:
                        try:
                            query='select cast({colName} as float) as c1,cast(count(c1) as float) as td_val_row_count {measureCols}   from {tname}  group by c1 having c1 in ({values}) {cndtn}'.format(cndtn=sf_cndtn,colName=col,tname=td_table,measureCols=measure_cols_td,values=(values).strip())
                            td_res_df2=self.run_td_query(query)
                        except:
                            try:
                                td_query='select {colName} (CASESPECIFIC) as c1,cast(count(c1)  as float) as td_val_row_count {measureCols} from {tname} group by c1 having c1 in ({values})  {cndtn} '.format(cndtn=sf_cndtn,colName=col,tname=td_table,measureCols=measure_cols_td,values=(values).strip())
                                td_res_df2=self.run_td_query(td_query)
                                
                            except Exception as err:
                                print('Intermediate Error: ',str(err))
                                td_query='select  {colName} ,cast(count({colName} )  as float) as td_val_row_count {measureCols} from {tname} group by {colName} having {colName} in ({values}) '.format(colName=col,tname=td_table,measureCols=measure_cols_td,values=(values).strip())
                                td_res_df2=self.run_td_query(td_query)

                   
                    td_res_df2.insert(0,'TD_ColName',col)
                    td_res_df2.columns=['TD_ColName','TD_col_value','TD_val_row_count']  + measure_cols_td_names_list
                    

                    
                    
                    sf_res_df2.insert(0,'SF_ColName',col)
                    sf_res_df2.columns=['SF_ColName','SF_col_value','SF_val_row_count'] + measure_cols_sf_names_list

                    #combining
                    left_indexes=['TD_ColName','TD_col_value']
                    right_indexes=['SF_ColName','SF_col_value']
                    
                    td_res_df2['TD_ColName']=td_res_df2['TD_ColName'].astype(str)
                    sf_res_df2['SF_ColName']=sf_res_df2['SF_ColName'].astype(str)
                    

                    try:
                        # dt=str(datetime.datetime.fromtimestamp( float(tsVal))) # if format is epoch
                        # s, ms = divmod(float(tsVal), 1000)  # (1236472051, 807)
                        # '%s.%03d' % (time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(s)), ms)
                        values_dts=[str(x).strip() for x in list(td_res_df2['TD_col_value'][:5])]
                        for val in values_dts:
                            if val in ['nan','NaN','None','na','']:continue
                            if '-' not in val:
                                if '/' not in val:
                                    raise Exception('ts')
                                
                        td_res_df2['TD_col_value'].apply(pd.to_datetime)
                        

                    except:
                        # td_res_df2['TD_col_value'] = pd.to_numeric(td_res_df2['TD_col_value'],errors='coerce').astype('float64').fillna(td_res_df2['TD_col_value']).tolist()
                        td_res_df2['TD_col_value']=td_res_df2['TD_col_value'].astype(str)
                        # td_res_df['TD_col_value'] = pd.to_numeric(td_res_df['TD_col_value'],errors='coerce').astype('float64').fillna(td_res_df['TD_col_value']).tolist()
                        try:
                            td_res_df2['TD_col_value']=td_res_df2['TD_col_value'].map(lambda x: Decimal(x))
                        except: pass

                    try:
                        # dt=str(datetime.datetime.fromtimestamp( float(tsVal))) # if format is epoch
                        # s, ms = divmod(float(tsVal), 1000)  # (1236472051, 807)
                        # '%s.%03d' % (time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(s)), ms)
                        values_dts=[str(x).strip() for x in list(sf_res_df2['SF_col_value'][:5])]
                        for val in values_dts:
                            if val in ['nan','NaN','None','na','']:continue
                            if '-' not in val:
                                if '/' not in val:
                                    raise Exception('ts')
                        sf_res_df2['SF_col_value'].apply(pd.to_datetime)
                    except:
                        # sf_res_df2['SF_col_value']= pd.to_numeric(sf_res_df2['SF_col_value'],errors='coerce').astype('float64').fillna(sf_res_df2['SF_col_value']).tolist()
                        sf_res_df2['SF_col_value']=sf_res_df2['SF_col_value'].astype(str)
                        
                        try:
                            sf_res_df2['SF_col_value']=sf_res_df2['SF_col_value'].map(lambda x: Decimal(x))
                        except: pass
                    


                    
                     
                    td_res_df2['TD_col_value']=td_res_df2['TD_col_value'].astype(str)
                    sf_res_df2['SF_col_value']=sf_res_df2['SF_col_value'].astype(str)

                    td_res_df2['TD_val_row_count']=td_res_df2['TD_val_row_count'].astype(str)
                    td_res_df2['TD_val_row_count']=td_res_df2['TD_val_row_count'].map(lambda x: Decimal(x))  

                    # td_res_df2['TD_val_row_count'] = pd.to_numeric(td_res_df2['TD_val_row_count'], errors='coerce').fillna(td_res_df2['TD_val_row_count']).tolist()
                    # sf_res_df2['SF_val_row_count'] = pd.to_numeric(sf_res_df2['SF_val_row_count'], errors='coerce').fillna(sf_res_df2['SF_val_row_count']).tolist()
                    sf_res_df2['SF_val_row_count']=sf_res_df2['SF_val_row_count'].astype(str)
                    sf_res_df2['SF_val_row_count']=sf_res_df2['SF_val_row_count'].map(lambda x: Decimal(x)) 


                    td_res_df2['TD_val_row_count']=td_res_df2['TD_val_row_count'].astype(str)
                    sf_res_df2['SF_val_row_count']=sf_res_df2['SF_val_row_count'].astype(str)



                    td_res_df2=td_res_df2.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                    sf_res_df2=sf_res_df2.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                    #making FreqDistri

                    print('Creating sub-final freq distribution')
                    # sf_res_df=sf_res_df+sf_res_df2
                    sf_res_df=pd.concat([sf_res_df,sf_res_df2],ignore_index=True)
                    td_res_df=pd.concat([td_res_df,td_res_df2],ignore_index=True)
                    # td_res_df=td_res_df+td_res_df2
                    
                    f1='td_res_df__'+str(col)+'.csv'
                    f2='sf_res_df__'+str(col)+'.csv'
                    
                    
                    td_res_df.to_csv(f1,index=False)
                    sf_res_df.to_csv(f2,index=False)
                    
                    


                    td_res_df=pd.read_csv(f1,dtype=str)
                    sf_res_df=pd.read_csv(f2,dtype=str)
                    
                    try:
                        os.remove(f1)
                        os.remove(f2)
                    except: pass

                    sub_finalFD=td_res_df.merge(sf_res_df,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_TD','_SF'])
                    # try:
                    #     sub_finalFD=sub_finalFD.append(sub_finalFD_rev)
                    # except:
                    #     sub_finalFD=pd.concat([sub_finalFD,sub_finalFD_rev])

                    #reverse ends
                    
                    sub_finalFD['TD_val_row_count']=sub_finalFD['TD_val_row_count'].map(lambda x: Decimal(x)) 
                    sub_finalFD['SF_val_row_count']=sub_finalFD['SF_val_row_count'].map(lambda x: Decimal(x)) 
                    # sub_finalFD['TD_val_row_count'] = pd.to_numeric(sub_finalFD['TD_val_row_count'], errors='coerce').fillna(sub_finalFD['TD_val_row_count']).tolist()
                    # sub_finalFD['SF_val_row_count'] = pd.to_numeric(sub_finalFD['SF_val_row_count'], errors='coerce').fillna(sub_finalFD['SF_val_row_count']).tolist()
                    
                    sub_finalFD['Difference(TD-SF)']=sub_finalFD['TD_val_row_count']-sub_finalFD['SF_val_row_count']
                    sub_finalFD['Difference_%']=100*sub_finalFD['Difference(TD-SF)']/sub_finalFD['TD_val_row_count']
                    
                    sub_finalFD=sub_finalFD.drop_duplicates()
                    # sub_finalFD.to_csv('subFinalFD___'+str(col)+'.csv',index=False)

                    finalFD=finalFD.append(sub_finalFD)
                  
                print('Creating final freq distribution')
                


                

                

                
                common_cols_td_sf=list()
                diff_cols=list()
                diff_percent_cols=list()
                diff_result_cols=list()
                td_cols=list()
                sf_cols=list()

                for col in measure_cols_sf_names_list:
                    for col_td in measure_cols_td_names_list:
                        if finalFD[col_td].dtype!=np.float64:
                             finalFD['col_td']=finalFD['col_td'].map(lambda x: Decimal(x))
                            #  finalFD[col_td] = pd.to_numeric( finalFD[col_td],errors='coerce').fillna( finalFD[col_td]).tolist()
                        if finalFD[col].dtype!=np.float64:
                            finalFD['col']=finalFD['col'].map(lambda x: Decimal(x))
                            # finalFD[col] = pd.to_numeric( finalFD[col],errors='coerce').fillna( finalFD[col]).tolist()

                        if col[0:-2]==col_td[0:-2]:
                            finalFD['Diff(TD-SF): '+col]=finalFD[col_td]-finalFD[col]
                            finalFD['Diff %: '+col[0:-2]]=100*finalFD['Diff(TD-SF): '+col]/finalFD[col_td]
                            finalFD['Result Diff: '+col[0:-2]]=(finalFD['Diff(TD-SF): '+col]==0)
                            common_cols_td_sf.append(col[0:-2])
                            diff_cols.append('Diff(TD-SF): '+col)
                            diff_percent_cols.append('Diff %: '+col[0:-2])
                            diff_result_cols.append('Result Diff: '+col[0:-2])
                            td_cols.append(col_td)
                            sf_cols.append(col)

            

                
                #finalFD=pd.concat([df_TD,df_SF],axis=1)

                main_cols=['TD_ColName','TD_col_value','TD_val_row_count','SF_ColName','SF_col_value','SF_val_row_count','Difference(TD-SF)','Difference_%']
                
                finalcols=main_cols

                for i in range(len(common_cols_td_sf)):
                    finalcols.append(td_cols[i])
                    finalcols.append(sf_cols[i])
                    finalcols.append(diff_cols[i])
                    finalcols.append(diff_percent_cols[i])
                    finalcols.append(diff_result_cols[i])




                remaining_uncommon_cols=list()
                for col in finalFD.columns:
                    if col not in finalcols:
                        remaining_uncommon_cols.append(col)

                finalcols=finalcols+remaining_uncommon_cols

                finalFD=finalFD[finalcols]
                
                
                tableName=td_table.split()[0]
                # finalFD.to_csv('Tab4__original_sf_table'+str(time.time())+'.csv',index=False)
                print('Frequency Distribution Tab (4th) created')

            except Exception as err:
                print('Not able to complete tab4 due to: ',str(err))
            
            
            
            self.writeIntoExcel_stats_td_sf(tableName,tab1_df,row_counts,final_Distinct,finalFD)

            try:
                os.remove('Tab1_'+original_td_table+'.csv')
                os.remove('Tab2_'+original_td_table+'.csv')
                os.remove('Tab3_'+original_sf_table+'.csv')
                # os.remove('Tab4_'+original_sf_table+'.csv')
            except:
                print("Could not delete intermediate tab csv files as file either doesn't exist or it's opened")

            print('Excel created!')
        except Exception as err:
            raise Exception(str(err))

     
    def DownloadAndGetCount_hive_sf(self,hive_table,hive_condition,sf_table,sf_condition,cols_to_sum,selected_cols_hive,selected_cols_sf):
        ########################## BASIC INFO #################################
       
        path=self.tableStatsResultDownloadLocationEntry_hive_to_sf.get()
        os.chdir(path)
           #hive_table=row[1]
           #hive_condition=row[2]
        original_hive_table=hive_table
        original_sf_table=sf_table
        try:
            if (str(hive_condition)!='nan'):
                        hive_table +=' where '+hive_condition
            
            #sf_table=row[3]
            #sf_condition=row[4]
            
            if (str(sf_condition)!='nan'):
                        sf_table +=' where '+sf_condition
            
                
            
            hive_table_cols=self.getCols_hive_forStats(hive_table,selected_cols_hive)
            sf_cols=self.getCols_SF(original_sf_table,selected_cols_sf)

            hive_table_cols=[x.split('.')[-1].strip().lower() for x in hive_table_cols]
            if(selected_cols_hive!='*'):
                hive_table_cols=[x.strip().lower() for x in selected_cols_hive]
            if(selected_cols_sf!='*'):
                sf_cols=[x.strip().lower() for x in selected_cols_sf]



            
            print('***Basic info checked for:',sf_table)
            
            ########################## TAB-1 SumOfValues ##############i###################

            try:
                measure_cols_hive=''
                measure_cols_hive_names_list=list()

                measure_cols_sf=''
                measure_cols_sf_names_list=list()

                hive_checksum=list()
                for col in hive_table_cols:
                    try:
                        t_q='select cast(sum({colName}) as double )as c from {tn}'.format(colName=col,tn=hive_table)
                        #    print("Running query: ",t_q)
                        res=self.run_hive_query(t_q)
                        hive_checksum.append(res.iloc[:,0][0])
                        # measure_cols_hive +=', cast(sum({colName}) as double) as hive_Sum_{colName} '.format(colName=col)
                        # measure_cols_hive_names_list.append(str(col).lower()+"_hive")
                    except:
                        hive_checksum.append(np.nan)
                
                for col in cols_to_sum:
                    # if col in hive_table:
                        measure_cols_hive +=', cast(sum({colName}) as double) as hive_Sum_{colName} '.format(colName=col)
                        measure_cols_hive_names_list.append(str(col).lower()+"_hive")

                sf_checksum=list()
                
                for col in sf_cols:
                    try:
                        s_q='select cast( sum({colName}) as double) as c from {tn}'.format(colName=col,tn=sf_table.strip().upper())
                        #    print("Running query: ",s_q)
                        res=self.run_sf_query(s_q)
                        sf_checksum.append(res.iloc[:,0][0])
                        # measure_cols_sf +=', cast(sum({colName}) as double) as SF_Sum_{colName} '.format(colName=col)
                        # measure_cols_sf_names_list.append(str(col).lower()+"_SF")
                    except:
                        sf_checksum.append(np.nan)

                db_name,schema_name,table_name=original_sf_table.split('.')         
                for col in cols_to_sum:
                    
                        measure_cols_sf +=', cast(sum({colName}) as double) as SF_Sum_{colName} '.format(colName=col)
                        measure_cols_sf_names_list.append(str(col).lower()+"_SF")
                
             
                

                col_info_query="SELECT  COLUMN_NAME as SF_ColName ,CHARACTER_MAXIMUM_LENGTH as SF_ColumnLength  from {db_name}.INFORMATION_SCHEMA.columns where table_name='{tn}' and table_schema='{ts}' ".format(db_name=db_name.strip().upper(),tn=table_name.strip().upper(),ts=schema_name.strip().upper())
                
                sf_col_info_df=self.run_sf_query(col_info_query)
                sf_col_info_df.columns=['SF_ColName','SF_ColumnLength']
                   

                datatype_query='desc table '+ original_sf_table
                sf_datatype_df=self.run_sf_query(datatype_query)
                sf_datatype_df=sf_datatype_df[['name','type']]
                sf_datatype_df.columns=['SF_ColName','SF_Datatype']



               
                
                # sf_indexed_cols_df=pd.DataFrame({'SF_ColName':sf_cols,'SF_Indexed':indexed_cols_yes_no})

                df2_sf=pd.concat([ pd.DataFrame({'SF_ColName':sf_cols}),pd.DataFrame({'SF_CheckSum':sf_checksum})],axis=1)
                df2_sf=df2_sf.merge(sf_col_info_df,left_on='SF_ColName',right_on='SF_ColName',suffixes=['_Base','_Release'])
                # df2_sf=df2_sf.merge(sf_indexed_cols_df,left_on='SF_ColName',right_on='SF_ColName',suffixes=['_Base','_Release'])
                df2_sf=df2_sf.merge(sf_datatype_df,left_on='SF_ColName',right_on='SF_ColName')
                df2_sf=df2_sf.applymap(lambda x: x.lower().strip() if isinstance(x, str) else x)
                
                
                # colLengthIndexed_query_hive='help column '+ original_hive_table+'.*'
                # colLengthIndexed_query_hive_df=self.run_hive_query(colLengthIndexed_query_hive)
                # colLengthIndexed_query_hive_df=colLengthIndexed_query_hive_df[['Column Dictionary Name','Max Length','Indexed?']]
                # colLengthIndexed_query_hive_df.columns=['hive_ColName','hive_ColumnLength','hive_Indexed']


                # query_creation_for_hive_datatype=''
                # for col in hive_table_cols:
                #     query_creation_for_hive_datatype +=', TYPE({colName}) as  {colName}'.format(colName=col)

                # if(len(query_creation_for_hive_datatype)>0):
                #     query_creation_for_hive_datatype=query_creation_for_hive_datatype[1:]


                hive_datatype_query='describe {tn}'.format(tn=hive_table)
                hive_datatype_query_df=self.run_hive_query(hive_datatype_query)
                # hive_datatype_query_df=hive_datatype_query_df.transpose().reset_index()
                hive_datatype_query_df.columns=['hive_ColName','hive_Datatype','hive_comment']
                hive_datatype_query_df=hive_datatype_query_df[['hive_ColName','hive_Datatype']]

                df1_hive=pd.concat([pd.DataFrame({'hive_ColName':hive_table_cols}),pd.DataFrame({'hive_CheckSum':hive_checksum})],axis=1)
                # df1_hive=df1_hive.merge(colLengthIndexed_query_hive_df,left_on='hive_ColName',right_on='hive_ColName')
                df1_hive=df1_hive.merge(hive_datatype_query_df,left_on='hive_ColName',right_on='hive_ColName')
                
                
                left_indexes=['hive_ColName']
                right_indexes=['SF_ColName']
                tab1_df=df1_hive.merge(df2_sf,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_hive','_SF'])

                finalCols_tab1=['SF_ColName','SF_Datatype','SF_CheckSum',
                                'hive_ColName','hive_Datatype','hive_CheckSum']
                tab1_df=tab1_df[finalCols_tab1]


                tab1_df['ColNameCheck_Result']=tab1_df['SF_ColName'].eq(tab1_df['hive_ColName'])
                tab1_df['Datatype_Result']=tab1_df['SF_Datatype'].eq(tab1_df['hive_Datatype'])
                # tab1_df['ColumnLength_Result']=tab1_df['SF_ColumnLength']==tab1_df['hive_ColumnLength']
                tab1_df['Checksum_Result']=tab1_df['SF_CheckSum'].eq(tab1_df['hive_CheckSum'])
                # tab1_df['Indexed_Result']=tab1_df['SF_Indexed']==tab1_df['hive_Indexed']
                
                 
            #     tab1_df['diff_%']=100*tab1_df['diff']/tab1_df['hive_CheckSum']
                
                tab1_df.to_csv('Tab1_'+original_hive_table+'.csv',index=False)
                print('SumOfValues Tab (1st) created')

            except Exception as err:
                print('Not able to complete tab1 due to: ',str(err))
            ########################## TAB-2 Total Row counts #################################

            try:
                tq='select cast(count(*) as double) as hive_ROW_COUNT from  '+ hive_table
                #    print("Running query: ",tq)
                hive_count=self.run_hive_query(tq)
                
                sq='select cast(count(*) as double) as SF_ROW_COUNT from  '+ sf_table
                #    print("Running query: ",sq)
                sf_count=self.run_sf_query(sq)
                
                
                #result
                
                row_counts=pd.concat([hive_count,sf_count],axis=1)
                row_counts['Difference(Hive-SF)']=row_counts['hive_row_count']-row_counts['SF_ROW_COUNT']
                row_counts['Difference_%']=100*row_counts['Difference(Hive-SF)']/row_counts['hive_row_count']
                
                row_counts.to_csv('Tab2_'+original_hive_table+'.csv',index=False)
                print('Row Count Tab (2nd) created')
            except Exception as err:
                print('Not able to complete tab2 due to: ',str(err))
            
            ########################## TAB-3 DISTINCT COUNT #################################
            
            #Hive
            try:
                # hive_distinct_query_list=list()
                hive_distinct_query_res=list()
                for col in hive_table_cols:
                    # hive_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                    q="select cast(count(distinct({colName})) as double) as {colName} from {ht}".format(colName=col,ht=hive_table)
                    val_df=self.run_hive_query(q)
                    val=val_df.iloc[0][0]
                    hive_distinct_query_res.append(val)

                
                
                hive_res_df=pd.DataFrame({'hive_ColName':hive_table_cols,'hive_DistinctCount':hive_distinct_query_res})
                hive_res_df.columns=['hive_ColName','hive_DistinctCount']
                
                #SF
                sf_distinct_query_list=list()
                for col in sf_cols:
                    sf_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                
                sf_distinct_query=",".join(sf_distinct_query_list)
                
                finalDistinct_query='select ' + sf_distinct_query + " from "+ sf_table

                #    print("Running query: ",finalDistinct_query)

                sf_res_df=self.run_sf_query(finalDistinct_query)
                sf_res_df=sf_res_df.transpose().reset_index()
                sf_res_df.columns=['SF_ColName','SF_DistinctCount']
                
                
                
                
                #making finalDistinctRes
                left_indexes=['hive_ColName']
                right_indexes=['SF_ColName']

                hive_res_df=hive_res_df.applymap(lambda x: x.lower().strip() if isinstance(x, str) else x)
                sf_res_df=sf_res_df.applymap(lambda x: x.lower().strip() if isinstance(x, str) else x)
                
                
                final_Distinct=hive_res_df.merge(sf_res_df,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_hive','_SF'])
                
                #final_Distinct=pd.concat([hive_res_df,sf_res_df],axis=1)
                final_Distinct['Difference(Hive-SF)']=final_Distinct['hive_DistinctCount']-final_Distinct['SF_DistinctCount']
                final_Distinct['Difference_%']=100*final_Distinct['Difference(Hive-SF)']/final_Distinct['hive_DistinctCount']
                
                final_Distinct.to_csv('Tab3_'+original_sf_table+'.csv',index=False)
                print('Distinct Counts Tab (3rd) Created')

            except Exception as err:
                print('Not able to complete tab3 due to: ',str(err))    
            ########################## TAB-4 FREQUENCY DISTRIBUTION (td,sf)#################################
            
            #TD
            
            try:
                df_hive=pd.DataFrame()
                
                print("****Hive queries")
                for col in hive_table_cols :
                    # if col.lower() not in cols_to_ignore:
                        query='select  {colName},cast(count({colName}) as double) as hive_val_row_count {measureCols} from {tname} group by {colName} order by hive_val_row_count, {colName} desc limit 1000 '.format(colName=col,tname=hive_table,measureCols=measure_cols_hive)
                        
                        #    print("Running query:",query)
                        res_df=self.run_hive_query(query)
                        res_df.insert(0,'hive_ColName',col)
                        res_df.columns=['hive_ColName','hive_col_value','hive_val_row_count']  + measure_cols_hive_names_list
                        df_hive=df_hive.append(res_df)
                    
                #SF
                df_SF=pd.DataFrame()
                
                print("****SF queries")
                for col in sf_cols:
                    #  if (col.lower() not in cols_to_ignore):
                        query='select {colName},cast(count({colName}) as double) as sf_val_row_count {measureCols}   from {tname} group by {colName} order by sf_val_row_count, {colName} desc limit 1000'.format(colName=col,tname=sf_table.strip().upper(),measureCols=measure_cols_sf)
                        
                #         print(query)
                        #    print("Running query:",query)
                        res_df=self.run_sf_query(query)
                        res_df.insert(0,'SF_ColName',col)
                        res_df.columns=['SF_ColName','SF_col_value','SF_val_row_count'] + measure_cols_sf_names_list
                        df_SF=df_SF.append(res_df)
                
                # try:
                #     # df_hive=df_hive.apply(pd.to_numeric, errors='coerce').fillna(df_hive)
                #     df_hive=df_hive.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                
                # except:
                #     print('not able to convert a column to int. too big string')
                    
                # try:
                #     # df_SF=df_SF.apply(pd.to_numeric, errors='coerce').fillna(df_SF)
                #     df_SF=df_SF.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                # except:
                #     print('not able to convert a column to int. too big string')
                
                left_indexes=['hive_ColName','hive_col_value']
                right_indexes=['SF_ColName','SF_col_value']
                

                

                for col in hive_table_cols:
                    if  col in left_indexes:
                            df_hive[col] = df_hive[col].astype(str)
                    else:
                        try:
                            df_hive[col] = pd.to_numeric(df_hive[col],errors='coerce').fillna(df_hive[col]).tolist()
                        except:pass
                    
                

                for col in sf_cols:
                    if  col in right_indexes:
                            df_SF[col] = df_SF[col].astype(str)
                    else:
                        try:
                            df_SF[col] = pd.to_numeric(df_SF[col], errors='coerce').fillna(df_SF[col]).tolist()
                        except:pass

                
                for  col in right_indexes:
                            df_SF[col] = df_SF[col].astype(str)   
                for  col in left_indexes:
                            df_hive[col] = df_hive[col].astype(str)       

                            
                df_hive=df_hive.applymap(lambda x: x.lower().strip() if isinstance(x, str) else x)
                df_SF=df_SF.applymap(lambda x: x.lower().strip() if isinstance(x, str) else x)
                #making FreqDistri

                print(df_hive.dtypes)
                print(df_SF.dtypes)
                print('Creating final freq distribution')
                finalFD=df_hive.merge(df_SF,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_hive','_SF'])


                

                

                finalFD['Difference(Hive-SF)']=finalFD['hive_val_row_count']-finalFD['SF_val_row_count']
                
                
                finalFD['Difference_%']=100*finalFD['Difference(Hive-SF)']/finalFD['hive_val_row_count']

                common_cols_hive_sf=list()
                diff_cols=list()
                diff_percent_cols=list()
                diff_result_cols=list()
                hive_cols=list()
                sf_cols=list()

                for col in measure_cols_sf_names_list:
                    for col_hive in measure_cols_hive_names_list:
                        if col[0:-2]==col_hive[0:-2]:
                            finalFD['Diff(Hive-SF): '+col]=finalFD[col_hive]-finalFD[col]
                            finalFD['Diff %: '+col[0:-2]]=100*finalFD['Diff(Hive-SF): '+col]/finalFD[col_hive]
                            finalFD['Result Diff: '+col[0:-2]]=(finalFD['Diff(Hive-SF): '+col]==0)
                            common_cols_hive_sf.append(col[0:-2])
                            diff_cols.append('Diff(Hive-SF): '+col)
                            diff_percent_cols.append('Diff %: '+col[0:-2])
                            diff_result_cols.append('Result Diff: '+col[0:-2])
                            hive_cols.append(col_hive)
                            sf_cols.append(col)

            

                
                #finalFD=pd.concat([df_hive,df_SF],axis=1)

                main_cols=['hive_ColName','hive_col_value','hive_val_row_count','SF_ColName','SF_col_value','SF_val_row_count','Difference(Hive-SF)','Difference_%']
                
                finalcols=main_cols

                for i in range(len(common_cols_hive_sf)):
                    finalcols.append(hive_cols[i])
                    finalcols.append(sf_cols[i])
                    finalcols.append(diff_cols[i])
                    finalcols.append(diff_percent_cols[i])
                    finalcols.append(diff_result_cols[i])




                remaining_uncommon_cols=list()
                for col in finalFD.columns:
                    if col not in finalcols:
                        remaining_uncommon_cols.append(col)

                finalcols=finalcols+remaining_uncommon_cols

                finalFD=finalFD[finalcols]
                
                
                tableName=hive_table.split('.')[1:][0]
                finalFD.to_csv('Tab4__original_sf_table'+str(time.time())+'.csv',index=False)
                print('Frequency Distribution Tab (4th) created')

            except Exception as err:
                print('Not able to complete tab4 due to: ',str(err))
            self.writeIntoExcel_stats_hive_sf(hive_table.split('.')[1:][0],tab1_df,row_counts,final_Distinct,finalFD)

            # try:
            #     os.remove('Tab1_'+hive_table+'.csv')
            #     os.remove('Tab2_'+hive_table+'.csv')
            #     os.remove('Tab3_'+hive_table+'.csv')
            #     os.remove('Tab4_'+hive_table+'.csv')
            # except:
            #     print("Could not delete intermediate tab csv files as file either doesn't exist or it's opened")

            print('Excel created!')
        except Exception as err:
            raise Exception(str(err))



    #   hive-sf stats fn end 
    # 
    #  
    def createRowDistinctFreqDistri_mssql_to_sf(self):
        
        df_path=(self.TableStatsFile_mssql_to_sf.get()).strip()
        print(df_path)
       
        df=pd.read_excel(df_path)
        
        for row in (df.itertuples()):
         ########################## BASIC INFO #################################
            # print(row[0])
            # print(row[1])
            # print(row[2])
            # print(row[3])
            # print(row[4])
            try:
               cols_to_sum=[]
               try:
                if (str(row[5])!='nan'):
                        cols_to_sum=[(col.lower()).strip() for col in str(row[5]).split(',')]
               except: pass
               self.DownloadAndGetCount_mssql_to_sf(row[1],row[2],row[3],row[4],cols_to_sum)
            except Exception as err:
                print("table ",row[1]," where ",row[2]," getting error: ",str(err))

    

    def createRowDistinctFreqDistri_sf_to_sf(self):
        
        df_path=(self.TableStatsFile_sf_to_sf.get()).strip()
        print(df_path)
       
        df=pd.read_excel(df_path)
        
        for row in (df.itertuples()):
         ########################## BASIC INFO #################################
            # print(row[0])
            # print(row[1])
            # print(row[2])
            # print(row[3])
            # print(row[4])
            try:
               cols_to_sum=[]
               try:
                if (str(row[5])!='nan'):
                        cols_to_sum=[(col.lower()).strip() for col in str(row[5]).split(',')]
               except: pass
               self.DownloadAndGetCount_sf_to_sf(row[1],row[2],row[3],row[4],cols_to_sum)
            except Exception as err:
                print("table ",row[1]," where ",row[2]," getting error: ",str(err))
    
    def performSanityCheck_td_sf(self):
        
        df_path=(self.sanity_checkFile_td_sf.get()).strip()
        print(df_path)
       
        df=pd.read_excel(df_path)
        self.datatype_df=pd.DataFrame()
        self.tablesNotFound_td_sf=pd.DataFrame(columns=['TableName','Found','DB']) # put it in calling function
        self.td_sf_sanity_check_checksum_df=pd.DataFrame()
        self.td_sf_sanity_check_rowcount_df=pd.DataFrame()
        self.td_sf_sanity_check_distinct_count_df=pd.DataFrame()
        curr_time=str(time.time())

        for row in (df.itertuples()):
         ########################## BASIC INFO #################################
            # print(row[0])
            # print(row[1])
            # print(row[2])
            # print(row[3])
            # print(row[4])
            try:
               cols_to_sum=[]
               
               ddl_check='y'
               row_count_check='y'
               distinct_check='y'
               SumOfValues='y'

               try:
                if(str(row[5]).lower().strip()=='n'):
                    ddl_check='n'
               except: pass 

               try:
                if(str(row[6]).lower().strip()=='n'):
                    row_count_check='n'
               except: pass 

               try:
                if(str(row[7]).lower().strip()=='n'):
                    distinct_check='n'
               except: pass 

               try:
                if(str(row[8]).lower().strip()=='n'):
                    SumOfValues='n'
               except: pass 


               self.sanity_check_td_sf(row[1],row[2],row[3],row[4],cols_to_sum,curr_time,ddl_check,row_count_check,distinct_check,SumOfValues)

            except Exception as err:
                print("table ",row[1]," where ",row[2]," getting error: ",str(err))

                
        path=self.sanity_checkResultDownloadLocationEntry_td_to_sf.get()
        os.chdir(path)
        #df_path=(self.TableStatsFile.get()).strip()
        #path=df_path.split('/')[0:-1]
        
        #fileName= "\\".join(path)+'\\'+tableName+'.xlsx'
        fileName='td_SF_Sanity_check_'+curr_time+'.xlsx'
        with pd.ExcelWriter(fileName) as writer:
                    self.datatype_df.to_excel(writer,sheet_name='DDL',index = False,header=True) 
                    self.td_sf_sanity_check_checksum_df.to_excel(writer,sheet_name='SumOfValues',index = False,header=True) 
                    self.td_sf_sanity_check_rowcount_df.to_excel(writer,sheet_name='Row Count',index = False,header=True) 
                    self.td_sf_sanity_check_distinct_count_df.to_excel(writer,sheet_name='Distinct',index = False,header=True)
                    self.tablesNotFound_td_sf.to_excel(writer,sheet_name='Missing Tables',index = False,header=True) 

        print('Excel writing done')

    def performSanityCheck_mssql_sf(self):
        
        df_path=(self.sanity_checkFile_mssql_sf.get()).strip()
        print(df_path)
       
        df=pd.read_excel(df_path)
        self.datatype_df=pd.DataFrame()
        self.missing_tables_mssql_sf_sanity_df=pd.DataFrame(columns=['TableName','DB'])
        self.mssql_sf_sanity_check_checksum_df=pd.DataFrame()
        self.mssql_sf_sanity_check_rowcount_df=pd.DataFrame()
        self.mssql_sf_sanity_check_distinct_count_df=pd.DataFrame()
        curr_time=str(time.time())

        for row in (df.itertuples()):
         ########################## BASIC INFO #################################
            # print(row[0])
            # print(row[1])
            # print(row[2])
            # print(row[3])
            # print(row[4])
            try:
               cols_to_sum=[]
               
               ddl_check='y'
               row_count_check='y'
               distinct_check='y'
               SumOfValues='y'

               try:
                if(str(row[5]).lower().strip()=='n'):
                    ddl_check='n'
               except: pass 

               try:
                if(str(row[6]).lower().strip()=='n'):
                    row_count_check='n'
               except: pass 

               try:
                if(str(row[7]).lower().strip()=='n'):
                    distinct_check='n'
               except: pass 

               try:
                if(str(row[7]).lower().strip()=='n'):
                    SumOfValues='n'
               except: pass 


               self.sanity_check_mssql_sf(row[1],row[2],row[3],row[4],cols_to_sum,curr_time,ddl_check,row_count_check,distinct_check,SumOfValues)

            except Exception as err:
                print("table ",row[1]," where ",row[2]," getting error: ",str(err))

                
        path=self.sanity_checkResultDownloadLocationEntry_mssql_to_sf.get()
        os.chdir(path)
        #df_path=(self.TableStatsFile.get()).strip()
        #path=df_path.split('/')[0:-1]
        
        #fileName= "\\".join(path)+'\\'+tableName+'.xlsx'
        fileName='mssql_SF_Sanity_check_'+curr_time+'.xlsx'
        with pd.ExcelWriter(fileName) as writer:
                    self.datatype_df.to_excel(writer,sheet_name='DDL',index = False,header=True) 
                    self.mssql_sf_sanity_check_checksum_df.to_excel(writer,sheet_name='SumOfValues',index = False,header=True) 
                    self.mssql_sf_sanity_check_rowcount_df.to_excel(writer,sheet_name='Row Count',index = False,header=True) 
                    self.mssql_sf_sanity_check_distinct_count_df.to_excel(writer,sheet_name='Distinct',index = False,header=True)
                    self.missing_tables_mssql_sf_sanity_df.to_excel(writer,sheet_name='Missing Tables',index = False,header=True) 

        print('Excel writing done')
        messagebox.showinfo('Done','Sanity check b/w MSSQL-SF done. ')


    def performSanityCheck_hive_sf(self):
        
        df_path=(self.sanity_checkFile_hive_sf.get()).strip()
        print(df_path)
       
        df=pd.read_excel(df_path)
        self.datatype_df=pd.DataFrame(columns=['Hive_ColName','Hive_Datatype','SF_ColName','SF_Datatype','Datatypes Same?','TableName'])
        self.missing_tables_hive_sf_sanity_df=pd.DataFrame(columns=['TableName','DB'])
        self.hive_sf_sanity_check_checksum_df=pd.DataFrame()
        self.hive_sf_sanity_check_rowcount_df=pd.DataFrame()
        self.hive_sf_sanity_check_distinct_count_df=pd.DataFrame()
        self.hive_sf_sanity_check_null_count_df=pd.DataFrame()
        curr_time=str(time.time())

        for row in (df.itertuples()):
         ########################## BASIC INFO #################################
            # print(row[0])
            # print(row[1])
            # print(row[2])
            # print(row[3])
            # print(row[4])
            try:
               cols_to_sum=[]
               
               ddl_check='y'
               row_count_check='y'
               distinct_check='y'
               SumOfValues='y'
               null_count_check='y'

               try:
                if(str(row[5]).lower().strip()=='n'):
                    ddl_check='n'
               except: pass 

               try:
                if(str(row[6]).lower().strip()=='n'):
                    row_count_check='n'
               except: pass 

               try:
                if(str(row[7]).lower().strip()=='n'):
                    distinct_check='n'
               except: pass 

               try:
                if(str(row[7]).lower().strip()=='n'):
                    SumOfValues='n'
               except: pass

               try:
                if(str(row[8]).lower().strip()=='n'):
                    null_count_check='n'
               except: pass 



               self.sanity_check_hive_sf(row[1],row[2],row[3],row[4],cols_to_sum,curr_time,ddl_check,row_count_check,distinct_check,SumOfValues,null_count_check)

            except Exception as err:
                print("table ",row[1]," where ",row[2]," getting error: ",str(err))

                
        path=self.sanity_checkResultDownloadLocationEntry_hive_to_sf.get()
        os.chdir(path)
        #df_path=(self.TableStatsFile.get()).strip()
        #path=df_path.split('/')[0:-1]
        
        #fileName= "\\".join(path)+'\\'+tableName+'.xlsx'
        fileName='Hive_SF_Sanity_check_'+curr_time+'.xlsx'
        with pd.ExcelWriter(fileName) as writer:
                    self.datatype_df.to_excel(writer,sheet_name='DDL',index = False,header=True) 
                    self.hive_sf_sanity_check_checksum_df.to_excel(writer,sheet_name='SumOfValues',index = False,header=True) 
                    self.hive_sf_sanity_check_rowcount_df.to_excel(writer,sheet_name='Row Count',index = False,header=True) 
                    self.hive_sf_sanity_check_distinct_count_df.to_excel(writer,sheet_name='Distinct',index = False,header=True)
                    self.hive_sf_sanity_check_null_count_df.to_excel(writer,sheet_name='Null Count',index = False,header=True)
                    self.missing_tables_hive_sf_sanity_df.to_excel(writer,sheet_name='Missing Tables',index = False,header=True) 

        print('Excel writing done')
        
    def performSanityCheck_sf1_sf2(self):
        
        df_path=(self.sanity_checkFile_sf1_sf2.get()).strip()
        print(df_path)
       
        df=pd.read_excel(df_path)
        self.datatype_df=pd.DataFrame(columns=['sf1_ColName','sf1_Datatype','sf2_ColName','sf2_Datatype','Datatypes Same?','TableName'])
        self.missing_tables_sf1_sf2_sanity_df=pd.DataFrame(columns=['TableName','DB'])
        self.sf1_sf2_sanity_check_checksum_df=pd.DataFrame()
        self.sf1_sf2_sanity_check_rowcount_df=pd.DataFrame()
        self.sf1_sf2_sanity_check_distinct_count_df=pd.DataFrame()
        self.sf1_sf2_sanity_check_null_count_df=pd.DataFrame()
        curr_time=str(time.time())

        for row in (df.itertuples()):
         ########################## BASIC INFO #################################
            # print(row[0])
            # print(row[1])
            # print(row[2])
            # print(row[3])
            # print(row[4])
            try:
               cols_to_sum=[]
               
               ddl_check='y'
               row_count_check='y'
               distinct_check='y'
               SumOfValues='y'
               null_count_check='y'

               try:
                if(str(row[5]).lower().strip()=='n'):
                    ddl_check='n'
               except: pass 

               try:
                if(str(row[6]).lower().strip()=='n'):
                    row_count_check='n'
               except: pass 

               try:
                if(str(row[7]).lower().strip()=='n'):
                    distinct_check='n'
               except: pass 

               try:
                if(str(row[8]).lower().strip()=='n'):
                    SumOfValues='n'
               except: pass

               try:
                if(str(row[9]).lower().strip()=='n'):
                    null_count_check='n'
               except: pass 



               self.sanity_check_sf1_sf2(row[1],row[2],row[3],row[4],cols_to_sum,curr_time,ddl_check,row_count_check,distinct_check,SumOfValues,null_count_check)

            except Exception as err:
                print("table ",row[1]," where ",row[2]," getting error: ",str(err))

                
        path=self.sanity_checkResultDownloadLocationEntry_sf1_to_sf2.get()
        os.chdir(path)
        #df_path=(self.TableStatsf2ile.get()).strip()
        #path=df_path.split('/')[0:-1]
        
        #fileName= "\\".join(path)+'\\'+tableName+'.xlsx'
        fileName='sf1_sf2_Sanity_check_'+curr_time+'.xlsx'
        with pd.ExcelWriter(fileName) as writer:
                    self.datatype_df.to_excel(writer,sheet_name='DDL',index = False,header=True) 
                    self.sf1_sf2_sanity_check_checksum_df.to_excel(writer,sheet_name='SumOfValues',index = False,header=True) 
                    self.sf1_sf2_sanity_check_rowcount_df.to_excel(writer,sheet_name='Row Count',index = False,header=True) 
                    self.sf1_sf2_sanity_check_distinct_count_df.to_excel(writer,sheet_name='Distinct',index = False,header=True)
                    self.sf1_sf2_sanity_check_null_count_df.to_excel(writer,sheet_name='Null Count',index = False,header=True)
                    self.missing_tables_sf1_sf2_sanity_df.to_excel(writer,sheet_name='Missing Tables',index = False,header=True) 

        print('Excel writing done')
        

    
    
    def sanity_check_td_sf(self,td_table,td_condition,sf_table,sf_condition,cols_to_sum,curr_time,ddl_check,row_count_check,distinct_check,SumOfValues):
        ########################## BASIC INFO #################################
        
       
        
    
        path=self.sanity_checkResultDownloadLocationEntry_td_to_sf.get()
        os.chdir(path)
        #td_table=row[1]
        #td_condition=row[2]
        original_td_table=td_table.strip()
        original_sf_table=sf_table.strip()
        try:
            if (str(td_condition)!='nan' and str(td_condition).strip()!=''):
                        td_table +=' where '+td_condition
            
            
            if (str(sf_condition)!='nan'  and str(sf_condition).strip()!=''):
                        sf_table +=' where '+sf_condition
            
                
            
            td_table_cols=self.getCols_TD(td_table,'*')
            sf_cols=self.getCols_SF(sf_table,'*')
            
            print('***checking datatype :',sf_table)
            #checking datatype`
            if ddl_check!='n':
                try:                  

                    datatypes_td_sf_mapping=dict()
                    datatypes_td_sf_mapping['string']='varchar'
                    datatypes_td_sf_mapping['timestamp']='timestamp_ntz'
                    datatypes_td_sf_mapping['int']='number'
                    datatypes_td_sf_mapping['decimal']='number'
                    datatypes_td_sf_mapping['float']='number'
                    datatypes_td_sf_mapping['date']='date'


                    colnames=['td_COLUMN_NAME','td_DATA_TYPE','td_IS_NULLABLE','td_CHARACTER_MAXIMUM_LENGTH','SF_COLUMN_NAME','SF_IS_NULLABLE','SF_CHARACTER_MAXIMUM_LENGTH','SF_DATA_TYPE','Mismatched_Attr','DDL Same?','TableName']

                    

                    
                    table_td=original_td_table
                    table_sf=original_sf_table
                    print("td table:",table_td," SF table:",table_sf)
                    

                    #td
                    colLengthIndexed_query_td_df=''
                    try:
                        colLengthIndexed_query_td='help column '+ table_td+'.*'
                        colLengthIndexed_query_td_df=self.run_td_query(colLengthIndexed_query_td)
                        colLengthIndexed_query_td_df=colLengthIndexed_query_td_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

                        colLengthIndexed_query_td_df=colLengthIndexed_query_td_df[['Column Dictionary Name','Type','Max Length','Nullable']]
                        colLengthIndexed_query_td_df['Type']=colLengthIndexed_query_td_df['Type'].replace({'A1':'ARRAY (one dimensional)','AN':'ARRAY (multidimensional)','I8':'BIGINT','BO':'BINARY LARGE OBJECT','BF':'BYTE','BV':'BYTE VARYING','I1':'BYTEINT','CF':'CHARACTER (fixed)','CV':'CHARACTER (varying)','CO':'CHARACTER LARGE OBJECT','D':'DECIMAL','DA':'DATE','F':'FLOAT','I':'INTEGER','DY':'INTERVAL DAY','DH':'INTERVAL DAY TO HOUR','DM':'INTERVAL DAY TO MINUTE','DS':'INTERVAL DAY TO SECOND','HR':'INTERVAL HOUR','HM':'INTERVAL HOUR TO MINUTE','HS':'INTERVAL HOUR TO SECOND','MI':'INTERVAL MINUTE','MS':'INTERVAL MINUTE TO SECOND','MO':'INTERVAL MONTH','SC':'INTERVAL SECOND','YR':'INTERVAL YEAR','YM':'INTERVAL YEAR TO MONTH','N':'NUMBER','D':'NUMERIC','PD':'PERIOD(DATE)','PT':'PERIOD(TIME(n))','PZ':'PERIOD(TIME(n) WITH TIME ZONE)','PS':'PERIOD(TIMESTAMP(n))','PM':'PERIOD(TIMESTAMP(n) WITH TIME ZONE)','F':'FLOAT','I2':'SMALLINT','AT':'TIME','TS':'TIMESTAMP','TZ':'TIME WITH TIME ZONE','SZ':'TIMESTAMP WITH TIME ZONE','UT':'USER DEFINED TYPE (all types)','XM':'XML'})
                        colLengthIndexed_query_td_df.columns=['td_COLUMN_NAME','td_DATA_TYPE','td_CHARACTER_MAXIMUM_LENGTH','td_IS_NULLABLE']
                    except:
                        self.tablesNotFound_td_sf.loc[len(self.tablesNotFound_td_sf.index)]=[table_td,'No','TD']


                
        
                    
                    #sf
                    sf_db_name,sf_schema,sf_tablename=table_sf.split('.')

                    info_schema_sf_df=''
                    try:
                        datatype_query='desc table '+ table_sf
                        sf_datatype_df=self.run_sf_query(datatype_query)
                        sf_datatype_df=sf_datatype_df[['name','type']]
                        sf_datatype_df.columns=['SF_COLUMN_NAME','SF_DATA_TYPE']
                    
                        info_schema_sf_query="select COLUMN_NAME as SF_COLUMN_NAME ,IS_NULLABLE as SF_IS_NULLABLE,CHARACTER_MAXIMUM_LENGTH as SF_CHARACTER_MAXIMUM_LENGTH from  {db}.INFORMATION_SCHEMA.columns  where table_name='{tn}' and table_schema='{ts}'".format(db=sf_db_name.strip().upper(),tn=sf_tablename.strip().upper(),ts=sf_schema.strip().upper())
                        info_schema_sf_df=self.run_sf_query(info_schema_sf_query)

                        info_schema_sf_df=info_schema_sf_df.merge(sf_datatype_df,on='SF_COLUMN_NAME')
                        info_schema_sf_df=info_schema_sf_df.applymap(lambda x: x.upper() if isinstance(x, str) else x)
                    except Exception as err:
                        print(err) 
                        self.tablesNotFound_td_sf.loc[len(self.tablesNotFound_td_sf.index)]=[table_sf,'No','SF']
                    
                    tempDf=colLengthIndexed_query_td_df.merge(info_schema_sf_df,left_on='td_COLUMN_NAME',right_on='SF_COLUMN_NAME',how='outer')
                    # tempDf.to_csv('td_sf_merged.csv')
                    
                    tn=str(table_sf).split('.')[-1]
                    tablename_list=list()
                    
                    attributes=['COLUMN_NAME','DATA_TYPE','IS_NULLABLE','CHARACTER_MAXIMUM_LENGTH']
                    mismatching_attribute_list=list()
                    result_datatype=list()

                    for row in (tempDf.itertuples()):
                        tablename_list.append(tn)
                        mismatches=list()
                        for attr in attributes:
                            td_col=str(getattr(row,'td_'+attr))
                            sf_col=str(getattr(row,'SF_'+attr))
                            

                            if attr=='DATA_TYPE':
                                td_dtype=td_col
                                sf_dtype=sf_col
                                if str(td_col)=='nan' or str(sf_col)=='nan':
                                    mismatching_attribute_list.append('Column Missing')
                                    result_datatype.append("No")
                                    break
                                
                                if not ( ( td_dtype==sf_dtype  ) or
                                    ( ('INT' in td_dtype) and ('NUMBER' in sf_dtype) ) or 
                                    ( ('SMALLINT' in td_dtype) and ('NUMBER' in sf_dtype) ) or 
                                    ( ('BYTEINT' in td_dtype) and ('NUMBER' in sf_dtype) ) or 
                                    ( ('DECIMAL' in td_dtype) and ('NUMBER' in sf_dtype) ) or 
                                    ( ('NUMERIC' in td_dtype) and ('NUMBER' in sf_dtype) ) or 
                                    ( ('FLOAT' in td_dtype) and ('NUMBER' in sf_dtype) ) or 
                                    ( ('CHAR' in td_dtype) and ('VARCHAR' in sf_dtype) ) or 
                                    ( ('DATE' in td_dtype) and ('DATE' in sf_dtype) ) or 
                                    ( (td_dtype in ['CHAR','VARCHAR']) and ('VARCHAR' in sf_dtype)  ) or
                                    ( ('TIMESTAMP' in td_dtype) and ('TIMESTAMP' in sf_dtype) )  ):
                                        mismatches.append('Datatype mismatch')
                                        
                                        
                            
                            elif attr=='IS_NULLABLE':
                                    if (td_col.upper()=='Y' and sf_col.upper()=='YES') or (td_col.upper()=='N' and sf_col.upper()=='NO'):
                                        continue
                                    mismatches.append(attr)

                            elif  attr=='CHARACTER_MAXIMUM_LENGTH':
                                try:
                                    a=int(td_col)
                                    b=int(sf_col)
                                    if a!=b:
                                        mismatches.append(attr)
                                except:
                                    if td_col.lower()=='nan' and sf_col.lower()!='nan':
                                        mismatches.append(attr)
                                    elif td_col.lower()!='nan' and sf_col.lower()=='nan':
                                        mismatches.append(attr)


                                        

                            elif td_col!=sf_col :
                                
                                    
                                    mismatches.append(attr)
                                    
                            if attr=='CHARACTER_MAXIMUM_LENGTH':
                                mismatching_attribute_list.append(','.join(mismatches))
                                if len(mismatches)>0:
                                    result_datatype.append("No")
                                else:
                                    result_datatype.append("Yes")

                        
                                
                    tempDf['Mismatched_Attr']=mismatching_attribute_list
                    tempDf['DDL Same?']=result_datatype
                    tempDf['TableName']=tablename_list
                    
                    
                    
                    tempDf.columns=['TD_ColName','TD_Datatype','TD_IS_NULLABLE','TD_MaxLen','SF_ColName','SF_IS_NULLABLE','SF_MaxLen','SF_Datatype','Mismatched_Attr','DDL Same?','TableName']
                    # finalDf.to_csv(res_name,index=False)
                    self.datatype_df=pd.concat([self.datatype_df,tempDf],ignore_index=True)

                except Exception as err:
                    print("In table ",row[0],'facing issue: ',str(err))

                
        
             
                 
            ########################## TAB-2 Total Row counts #################################

            if row_count_check!='n':
                try:
                    tq='select cast(count(*) as float) as td_ROW_COUNT from  '+ td_table
                    #    print("Running query: ",tq)
                    td_count=self.run_td_query(tq)
                    td_count=td_count.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    td_count.columns=['td_ROW_COUNT']
                    
                    sq='select cast(count(*) as double) as SF_ROW_COUNT from  '+ sf_table
                    #    print("Running query: ",sq)
                    sf_count=self.run_sf_query(sq)
                    sf_count=sf_count.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    sf_count.columns=['SF_ROW_COUNT']
                    
                    # result
                    
                    row_counts=pd.concat([td_count,sf_count],axis=1)
                    row_counts['Difference(td-SF)']=row_counts['td_ROW_COUNT']-row_counts['SF_ROW_COUNT']
                    row_counts['Difference_%']=100*row_counts['Difference(td-SF)']/row_counts['td_ROW_COUNT']
                    row_counts['TableName']=original_td_table
                    self.td_sf_sanity_check_rowcount_df=pd.concat([self.td_sf_sanity_check_rowcount_df,row_counts])

                    # row_counts.to_csv('Tab2_'+original_td_table+'.csv',index=False)
                    print('Row Count Tab (2nd) created')
                except Exception as err:
                    print('Not able to complete tab2 due to: ',str(err))

            ########################## TAB-1 SumOfValues #################################
            if SumOfValues!='n':
                try:
                    

                    td_checksum=list()
                    for col in td_table_cols:
                        try:
                            t_q='select cast(sum({colName}) as float )as c from {tn}'.format(colName=col,tn=td_table)
                            #    print("Running query: ",t_q)
                            res=self.run_td_query(t_q)
                            td_checksum.append(res.iloc[:,0][0])
                            # measure_cols_td +=', cast(sum({colName}) as double) as td_Sum_{colName} '.format(colName=col)
                            # measure_cols_td_names_list.append(str(col).lower()+"_td")
                        except:
                            td_checksum.append(np.nan)
                    

                    sf_checksum=list()
                    
                    for col in sf_cols:
                        try:
                            s_q='select cast( sum({colName}) as double) as c from {tn}'.format(colName=col,tn=sf_table.strip().upper())
                            #    print("Running query: ",s_q)
                            res=self.run_sf_query(s_q)
                            sf_checksum.append(res.iloc[:,0][0])
                            # measure_cols_sf +=', cast(sum({colName}) as double) as SF_Sum_{colName} '.format(colName=col)
                            # measure_cols_sf_names_list.append(str(col).lower()+"_SF")
                        except:
                            sf_checksum.append(np.nan)

                
                    

                    df2_sf=pd.concat([ pd.DataFrame({'SF_ColName':sf_cols}),pd.DataFrame({'SF_CheckSum':sf_checksum})],axis=1)
                    df2_sf=df2_sf.applymap(lambda x: x.lower() if isinstance(x, str) else x)

                    

                    df1_td=pd.concat([pd.DataFrame({'td_ColName':td_table_cols}),pd.DataFrame({'td_CheckSum':td_checksum})],axis=1)
                    df1_td=df1_td.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    left_indexes=['td_ColName']
                    right_indexes=['SF_ColName']
                    df1_td['TD_ColumnLength'] = pd.to_numeric(df1_td['TD_ColumnLength'],errors='coerce').astype('float64').fillna(df1_td['TD_ColumnLength']).tolist()
                    df1_td['TD_CheckSum'] = pd.to_numeric(df1_td['TD_CheckSum'],errors='coerce').astype('float64').fillna(df1_td['TD_CheckSum']).tolist()
                    df2_sf['SF_ColumnLength'] = pd.to_numeric(df2_sf['SF_ColumnLength'],errors='coerce').astype('float64').fillna(df2_sf['SF_ColumnLength']).tolist()
                    df2_sf['SF_CheckSum'] = pd.to_numeric(df2_sf['SF_CheckSum'],errors='coerce').astype('float64').fillna(df2_sf['SF_CheckSum']).tolist()


                    tab1_df=df1_td.merge(df2_sf,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_td','_SF'])

                    finalCols_tab1=['SF_ColName','SF_CheckSum','td_ColName','td_CheckSum']
                    tab1_df=tab1_df[finalCols_tab1]


                    
                    tab1_df['Checksum_Result']=tab1_df['SF_CheckSum'].eq(tab1_df['td_CheckSum'])
                    tab1_df['TableName']=original_td_table

                    self.td_sf_sanity_check_checksum_df=pd.concat([self.td_sf_sanity_check_checksum_df,tab1_df])
                    
                    
                #     tab1_df['diff_%']=100*tab1_df['diff']/tab1_df['td_CheckSum']
                    
                    # tab1_df.to_csv('Tab1_checksum_'+original_td_table+'.csv',index=False)
                    print('SumOfValues Tab (1st) created')

                except Exception as err:
                    print('Not able to complete tab1 due to: ',str(err))
            
            ########################## TAB-3 DISTINCT COUNT #################################
        
            #TD
            if distinct_check!='n':
                try:
                    #td_distinct_query_list=list()
                    td_distinct_query_res=list()
                    for col in td_table_cols:
                        try:
                            q="select cast(count(distinct({colName}) (CASESPECIFIC) ) as float) as {colName} from {tn}".format(colName=col,tn=td_table)
                            # td_distinct_query_list.append()
                            val_df=self.run_td_query(q)
                            val=val_df.iloc[0][0]
                            td_distinct_query_res.append(val)
                        except:
                            q="select cast(count(distinct({colName}) ) as float) as {colName} from {tn}".format(colName=col,tn=td_table)
                            # td_distinct_query_list.append()
                            val_df=self.run_td_query(q)
                            val=val_df.iloc[0][0]
                            td_distinct_query_res.append(val)

                    td_res_df=pd.DataFrame({'TD_ColName':td_table_cols,'TD_DistinctCount':td_distinct_query_res})   
                     
                    td_res_df.columns=['td_ColName','td_DistinctCount']
                    td_res_df=td_res_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    #SF
                    sf_distinct_query_list=list()
                    for col in sf_cols:
                        sf_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                    
                    sf_distinct_query=",".join(sf_distinct_query_list)
                    
                    finalDistinct_query='select ' + sf_distinct_query + " from "+ sf_table

                    #    print("Running query: ",finalDistinct_query)

                    sf_res_df=self.run_sf_query(finalDistinct_query)
                    sf_res_df=sf_res_df.transpose().reset_index()
                    sf_res_df.columns=['SF_ColName','SF_DistinctCount']
                    sf_res_df=sf_res_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    
                    
                    
                    
                    #making finalDistinctRes
                    left_indexes=['td_ColName']
                    right_indexes=['SF_ColName']
                    
                    final_Distinct=td_res_df.merge(sf_res_df,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_td','_SF'])
                    
                    #final_Distinct=pd.concat([td_res_df,sf_res_df],axis=1)
                    final_Distinct['Difference(TD-SF)']=final_Distinct['td_DistinctCount']-final_Distinct['SF_DistinctCount']
                    final_Distinct['Difference_%']=100*final_Distinct['Difference(TD-SF)']/final_Distinct['td_DistinctCount']
                    final_Distinct['TableName']=original_td_table
                    self.td_sf_sanity_check_distinct_count_df=pd.concat([self.td_sf_sanity_check_distinct_count_df,final_Distinct])

                    # final_Distinct.to_csv('Tab3_'+original_td_table+'.csv',index=False)
                    print('Distinct Counts Tab (3rd) Created')

                except Exception as err:
                    print('Not able to complete tab3 due to: ',str(err))    
            
            

            fileName='td_SF_Sanity_check_'+curr_time+'.xlsx'
            with pd.ExcelWriter(fileName) as writer:
                        self.datatype_df.to_excel(writer,sheet_name='DDL',index = False,header=True) 
                        self.td_sf_sanity_check_checksum_df.to_excel(writer,sheet_name='SumOfValues',index = False,header=True) 
                        self.td_sf_sanity_check_rowcount_df.to_excel(writer,sheet_name='Row Count',index = False,header=True) 
                        self.td_sf_sanity_check_distinct_count_df.to_excel(writer,sheet_name='Distinct',index = False,header=True)
                        self.tablesNotFound_td_sf.to_excel(writer,sheet_name='Missing Tables',index = False,header=True) 

            print(original_td_table,' completed...')
        except Exception as err:
            raise Exception(str(err))
      

    def sanity_check_hive_sf(self,hive_table,hive_condition,sf_table,sf_condition,cols_to_sum,curr_time,ddl_check,row_count_check,distinct_check,SumOfValues,null_count_check):
        ########################## BASIC INFO #################################
        
        datatypes_hive_sf_mapping=dict()
        datatypes_hive_sf_mapping['string']='varchar'
        datatypes_hive_sf_mapping['timestamp']='timestamp_ntz'
        datatypes_hive_sf_mapping['int']='number'
        datatypes_hive_sf_mapping['decimal']='number'
        datatypes_hive_sf_mapping['float']='number'
        datatypes_hive_sf_mapping['bigint']='number'
        datatypes_hive_sf_mapping['date']='date'
        datatypes_hive_sf_mapping['varchar']='varchar'
        datatypes_hive_sf_mapping['char']='varchar'
        


        
    
        path=self.sanity_checkResultDownloadLocationEntry_hive_to_sf.get()
        os.chdir(path)
        #hive_table=row[1]
        #hive_condition=row[2]
        original_hive_table=hive_table
        original_sf_table=sf_table
        try:
            if (str(hive_condition).strip().lower()!='nan'):
                        hive_table +=' where '+hive_condition
            
            
            if (str(sf_condition).strip().lower()!='nan'):
                        sf_table +=' where '+sf_condition
            
                
            
            hive_table_cols=self.getCols_Hive(hive_table)
            sf_cols=self.getCols_SF(sf_table)
            
            print('***checking datatype :',sf_table)
            #checking datatype`
            if ddl_check!='n':
                try:
                    table_hive=original_hive_table
                    table_sf=original_sf_table
                    print("Datatype check for:=> Hive table:",table_hive," SF table:",table_sf)

                    try:
                    #hive
                        desc_query='describe '+table_hive
                        desc_df=self.run_hive_query(desc_query)
                        desc_df.columns=[col.lower() for col in desc_df.columns]
                    
                        if(self.runFromImpala.get()==1):
                                desc_df=desc_df[['name','type']] 
                        else:
                                desc_df=desc_df[['col_name','data_type']] 
                         
                        desc_df.columns=['Hive_ColName','Hive_Datatype']
                        desc_df=desc_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    except Exception as err:
                        print(err)
                        self.missing_tables_hive_sf_sanity_df.loc[len(self.missing_tables_hive_sf_sanity_df.index)]=[table_hive,'Hive']
                    #sf

                    try:
                        datatype_query='desc table '+ original_sf_table
                        sf_datatype_df=self.run_sf_query(datatype_query)
                        sf_datatype_df=sf_datatype_df[['name','type']]
                        sf_datatype_df.columns=['SF_ColName','SF_Datatype']
                        sf_datatype_df=sf_datatype_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    except Exception as err:
                        print(err)
                        self.missing_tables_hive_sf_sanity_df.loc[len(self.missing_tables_hive_sf_sanity_df.index)]=[original_sf_table,'SF']

                    tempDf=desc_df.merge(sf_datatype_df,left_on='Hive_ColName',right_on='SF_ColName',how='outer')

                    result_datatype=list()
                    tn=str(table_sf).split('.')[-1]
                    tablename_list=list()

                    for row in (tempDf.itertuples()):
                        tablename_list.append(tn)
                        hive_dtype=(str(getattr(row,'Hive_Datatype'))).lower().strip()
                        
                        sf_dtype=(str(getattr(row,'SF_Datatype'))).lower().strip()
                        
                        if(hive_dtype=='nan' or sf_dtype=='nan'):
                            result_datatype.append("No")

                        elif    ( ( hive_dtype==sf_dtype  ) or
                                   ( ('int' in hive_dtype) and ('number' in sf_dtype) ) or 
                                    ( ('smallint' in hive_dtype) and ('number' in sf_dtype) ) or 
                                    ( ('byteint' in hive_dtype) and ('number' in sf_dtype) ) or 
                                    ( ('decimal' in hive_dtype) and ('number' in sf_dtype) ) or 
                                    ( ('numeric' in hive_dtype) and ('number' in sf_dtype) ) or 
                                    ( ('float' in hive_dtype) and ('number' in sf_dtype) ) or
                                    ( ('float' in hive_dtype) and ('double' in sf_dtype) ) or 
                                    ( ('double' in hive_dtype) and ('float' in sf_dtype) ) or
                                    ( ('double' in hive_dtype) and ('number' in sf_dtype) ) or
                                    ( ('char' in hive_dtype) and ('varchar' in sf_dtype) ) or
                                    ( ('bit' in hive_dtype) and ('boolean' in sf_dtype) ) or 
                                    ( ('bigint' in hive_dtype) and ('number' in sf_dtype) ) or #(NULL)
                                    
                                    ( (hive_dtype in ['char','varchar']) and ('varchar' in sf_dtype)  ) or
                                    ( ('varchar' in hive_dtype) and (('varchar' in sf_dtype)) ) or
                                    ( ('timestamp' in hive_dtype or 'datetime' in hive_dtype) and ('timestamp' in sf_dtype) )  ):result_datatype.append("Yes")
                                        

                        elif hive_dtype in  datatypes_hive_sf_mapping :
                            if datatypes_hive_sf_mapping[hive_dtype] in sf_dtype:
                                result_datatype.append("Yes")
                            else:
                                result_datatype.append("No")
                        else: result_datatype.append("Mapping not found")        


                    tempDf['Datatypes Same?']=result_datatype
                    tempDf['TableName']=tablename_list
                    
                    
                    # tempDf.to_csv('Datatype_result_sanity.csv')
                    self.datatype_df=pd.concat([self.datatype_df,tempDf],ignore_index=True)

                    ###############################SumOfValues############################################

                except Exception as err:
                    print("In table ",row[0],'facing issue: ',str(err))
             
             
                 
            ########################## TAB-2 Total Row counts #################################

            if row_count_check!='n':
                try:
                    tq='select cast(count(*) as double) as hive_ROW_COUNT from  '+ hive_table
                    #    print("Running query: ",tq)
                    hive_count=self.run_hive_query(tq)
                    hive_count=hive_count.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    hive_count.columns=['hive_ROW_COUNT']
                    
                    sq='select cast(count(*) as double) as SF_ROW_COUNT from  '+ sf_table
                    #    print("Running query: ",sq)
                    sf_count=self.run_sf_query(sq)
                    sf_count=sf_count.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    sf_count.columns=['SF_ROW_COUNT']
                    
                    # result
                    
                    row_counts=pd.concat([hive_count,sf_count],axis=1)
                    row_counts['Difference(hive-SF)']=row_counts['hive_ROW_COUNT']-row_counts['SF_ROW_COUNT']
                    row_counts['Difference_%']=100*row_counts['Difference(hive-SF)']/row_counts['hive_ROW_COUNT']
                    row_counts['TableName']=original_hive_table
                    self.hive_sf_sanity_check_rowcount_df=pd.concat([self.hive_sf_sanity_check_rowcount_df,row_counts])

                    # row_counts.to_csv('Tab2_'+original_hive_table+'.csv',index=False)
                    print('Row Count Tab (2nd) created')
                except Exception as err:
                    print('Not able to complete tab2 due to: ',str(err))

            ########################## TAB-1 SumOfValues #################################
            if SumOfValues!='n':
                try:
                    

                    hive_checksum=list()
                    for col in hive_table_cols:
                        try:
                            t_q='select cast(sum({colName}) as double )as c from {tn}'.format(colName=col,tn=hive_table)
                            #    print("Running query: ",t_q)
                            res=self.run_hive_query(t_q)
                            hive_checksum.append(res.iloc[:,0][0])
                            # measure_cols_hive +=', cast(sum({colName}) as double) as hive_Sum_{colName} '.format(colName=col)
                            # measure_cols_hive_names_list.append(str(col).lower()+"_hive")
                        except:
                            hive_checksum.append(np.nan)
                    

                    sf_checksum=list()
                    
                    for col in sf_cols:
                        try:
                            s_q='select cast( sum({colName}) as double) as c from {tn}'.format(colName=col,tn=sf_table.strip().upper())
                            #    print("Running query: ",s_q)
                            res=self.run_sf_query(s_q)
                            sf_checksum.append(res.iloc[:,0][0])
                            # measure_cols_sf +=', cast(sum({colName}) as double) as SF_Sum_{colName} '.format(colName=col)
                            # measure_cols_sf_names_list.append(str(col).lower()+"_SF")
                        except:
                            sf_checksum.append(np.nan)

                
                    

                    df2_sf=pd.concat([ pd.DataFrame({'SF_ColName':sf_cols}),pd.DataFrame({'SF_CheckSum':sf_checksum})],axis=1)
                    df2_sf=df2_sf.applymap(lambda x: x.lower() if isinstance(x, str) else x)

                    

                    df1_hive=pd.concat([pd.DataFrame({'hive_ColName':hive_table_cols}),pd.DataFrame({'hive_CheckSum':hive_checksum})],axis=1)
                    df1_hive=df1_hive.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    left_indexes=['hive_ColName']
                    right_indexes=['SF_ColName']
                    tab1_df=df1_hive.merge(df2_sf,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_hive','_SF'])

                    finalCols_tab1=['SF_ColName','SF_CheckSum','hive_ColName','hive_CheckSum']
                    tab1_df=tab1_df[finalCols_tab1]


                    
                    tab1_df['Checksum_Result']=tab1_df['SF_CheckSum'].eq(tab1_df['hive_CheckSum'])
                    tab1_df['TableName']=original_hive_table

                    self.hive_sf_sanity_check_checksum_df=pd.concat([self.hive_sf_sanity_check_checksum_df,tab1_df])
                    
                    
                #     tab1_df['diff_%']=100*tab1_df['diff']/tab1_df['hive_CheckSum']
                    
                    # tab1_df.to_csv('Tab1_checksum_'+original_hive_table+'.csv',index=False)
                    print('SumOfValues Tab (1st) created')

                except Exception as err:
                    print('Not able to complete tab1 due to: ',str(err))
            
            ########################## TAB-3 DISTINCT COUNT #################################
        
            #Hive
            if distinct_check!='n':
                try:
                    # hive_distinct_query_list=list()
                    hive_distinct_query_res=list()
                    for col in hive_table_cols:
                        # hive_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                        q="select cast(count(distinct({colName})) as double) as {colName} from {ht}".format(colName=col,ht=hive_table)
                        val_df=self.run_hive_query(q)
                        val=val_df.iloc[0][0]
                        hive_distinct_query_res.append(val)

                    
                    
                    hive_res_df=pd.DataFrame({'hive_ColName':hive_table_cols,'hive_DistinctCount':hive_distinct_query_res})
                    hive_res_df.columns=['hive_ColName','hive_DistinctCount']
                     
                    hive_res_df=hive_res_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    #SF
                    sf_distinct_query_list=list()
                    for col in sf_cols:
                        sf_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                    
                    sf_distinct_query=",".join(sf_distinct_query_list)
                    
                    finalDistinct_query='select ' + sf_distinct_query + " from "+ sf_table

                    #    print("Running query: ",finalDistinct_query)

                    sf_res_df=self.run_sf_query(finalDistinct_query)
                    sf_res_df=sf_res_df.transpose().reset_index()
                    sf_res_df.columns=['SF_ColName','SF_DistinctCount']
                    sf_res_df=sf_res_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    
                    
                    
                    
                    #making finalDistinctRes
                    left_indexes=['hive_ColName']
                    right_indexes=['SF_ColName']
                    
                    final_Distinct=hive_res_df.merge(sf_res_df,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_hive','_SF'])
                    
                    #final_Distinct=pd.concat([hive_res_df,sf_res_df],axis=1)
                    final_Distinct['Difference(TD-SF)']=final_Distinct['hive_DistinctCount']-final_Distinct['SF_DistinctCount']
                    final_Distinct['Difference_%']=100*final_Distinct['Difference(TD-SF)']/final_Distinct['hive_DistinctCount']
                    final_Distinct['TableName']=original_hive_table
                    self.hive_sf_sanity_check_distinct_count_df=pd.concat([self.hive_sf_sanity_check_distinct_count_df,final_Distinct])

                    # final_Distinct.to_csv('Tab3_'+original_hive_table+'.csv',index=False)
                    print('Distinct Counts Tab (3rd) Created')

                except Exception as err:
                    print('Not able to complete tab3 due to: ',str(err))    
            
            #######
            ########################## TAB-4 null COUNT #################################
        
            #Hive null count
             
            if null_count_check!='n':
                try:
                    # hive_distinct_query_list=list()
                    hive_null_query_res=list()
                    for col in hive_table_cols:
                        q="select cast(count(*) as double) from {ht} where {colName} is null or {colName}='' ".format(ht=hive_table,colName=col)
                        # hive_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                        # q="select cast(count(distinct({colName})) as double) as {colName} from {ht}".format(colName=col,ht=hive_table)
                        val_df=self.run_hive_query(q)
                        val=val_df.iloc[0][0]
                        hive_null_query_res.append(val)

                    
                    
                    hive_res_df=pd.DataFrame({'hive_ColName':hive_table_cols,'hive_NullCount':hive_null_query_res})
                    hive_res_df.columns=['hive_ColName','hive_NullCount']
                     
                    hive_res_df=hive_res_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    #SF
                    sf_null_query_res=list()
                    for col in sf_cols:
                        try:
                            q="select cast(count(*) as double) from {st} where {colName} is null or {colName}='' ".format(st=sf_table,colName=col)
                            # sf_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                            # q="select cast(count(distinct({colName})) as double) as {colName} from {ht}".format(colName=col,ht=sf_table)
                            val_df=self.run_sf_query(q)
                        except:
                            q="select cast(count(*) as double) from {st} where {colName} is null  ".format(st=sf_table,colName=col)
                            # sf_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                            # q="select cast(count(distinct({colName})) as double) as {colName} from {ht}".format(colName=col,ht=sf_table)
                            val_df=self.run_sf_query(q)

                        val=val_df.iloc[0][0]
                        sf_null_query_res.append(val)

                    
                    
                    sf_res_df=pd.DataFrame({'SF_ColName':sf_cols,'SF_NullCount':sf_null_query_res})
                    sf_res_df.columns=['SF_ColName','SF_NullCount']
                     
                    sf_res_df=sf_res_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    
                    
                    
                    
                    #making finalDistinctRes
                    left_indexes=['hive_ColName']
                    right_indexes=['SF_ColName']
                    
                    final_NullCount=hive_res_df.merge(sf_res_df,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_hive','_SF'])
                    
                    #final_NullCount=pd.concat([hive_res_df,sf_res_df],axis=1)
                    final_NullCount['Difference(Hive-SF)']=final_NullCount['hive_NullCount']-final_NullCount['SF_NullCount']
                    final_NullCount['Difference_%']=100*final_NullCount['Difference(Hive-SF)']/final_NullCount['hive_NullCount']
                    final_NullCount['TableName']=original_hive_table
                    self.hive_sf_sanity_check_null_count_df=pd.concat([self.hive_sf_sanity_check_null_count_df,final_NullCount])

                    # final_NullCount.to_csv('Tab3_'+original_hive_table+'.csv',index=False)
                    print('null Counts Tab (4th) Created')

                except Exception as err:
                    print('Not able to complete tab4(null count) due to: ',str(err))    
            

            #######
            

            fileName='Hive_SF_Sanity_check_'+curr_time+'.xlsx'
            with pd.ExcelWriter(fileName) as writer:
                        self.datatype_df.to_excel(writer,sheet_name='DDL',index = False,header=True) 
                        self.hive_sf_sanity_check_checksum_df.to_excel(writer,sheet_name='SumOfValues',index = False,header=True) 
                        self.hive_sf_sanity_check_rowcount_df.to_excel(writer,sheet_name='Row Count',index = False,header=True) 
                        self.hive_sf_sanity_check_distinct_count_df.to_excel(writer,sheet_name='Distinct',index = False,header=True)
                        self.missing_tables_hive_sf_sanity_df.to_excel(writer,sheet_name='Missing Tables',index = False,header=True) 

            print(original_hive_table,' completed...')
        except Exception as err:
            raise Exception(str(err))
       
    def sanity_check_sf1_sf2(self,sf1_table,sf1_condition,sf2_table,sf2_condition,cols_to_sum,curr_time,ddl_check,row_count_check,distinct_check,SumOfValues,null_count_check):
        ########################## BASIC INFO #################################
        
         
        


        
    
        path=self.sanity_checkResultDownloadLocationEntry_sf1_to_sf2.get()
        os.chdir(path)
        #sf1_table=row[1]
        #sf1_condition=row[2]
        original_sf1_table=sf1_table
        original_sf2_table=sf2_table
        try:
            if (str(sf1_condition).strip().lower()!='nan'):
                        sf1_table +=' where '+sf1_condition
            
            
            if (str(sf2_condition).strip().lower()!='nan'):
                        sf2_table +=' where '+sf2_condition
            
                
            
            sf1_table_cols=self.getCols_SF(sf1_table)
            sf2_cols=self.getCols_SF(sf2_table)
            
            print('***checking datatype :',sf2_table)
            #checking datatype`
            if ddl_check!='n':
                try:
                    table_sf1=original_sf1_table
                    table_sf2=original_sf2_table
                    print("Datatype check for:=> sf1 table:",table_sf1," sf2 table:",table_sf2)

                    try:
                    #sf1
                        datatype_query='desc table '+ original_sf1_table
                        desc_df=self.run_sf_query(datatype_query)
                        desc_df=desc_df[['name','type']]
                        desc_df.columns=['sf1_ColName','sf1_Datatype']
                        desc_df=desc_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    except Exception as err:
                        print(err)
                        self.missing_tables_sf1_sf2_sanity_df.loc[len(self.missing_tables_sf1_sf2_sanity_df.index)]=[table_sf1,'sf1']
                    #sf2

                    try:
                        datatype_query='desc table '+ original_sf2_table
                        sf2_datatype_df=self.run_sf_query(datatype_query)
                        sf2_datatype_df=sf2_datatype_df[['name','type']]
                        sf2_datatype_df.columns=['sf2_ColName','sf2_Datatype']
                        sf2_datatype_df=sf2_datatype_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    except Exception as err:
                        print(err)
                        self.missing_tables_sf1_sf2_sanity_df.loc[len(self.missing_tables_sf1_sf2_sanity_df.index)]=[original_sf2_table,'sf2']

                    tempDf=desc_df.merge(sf2_datatype_df,left_on='sf1_ColName',right_on='sf2_ColName',how='outer')

                    result_datatype=list()
                    tn=str(table_sf2).split('.')[-1]
                    

                    for row in (tempDf.itertuples()):
                         
                        sf1_dtype=(str(getattr(row,'sf1_Datatype'))).lower().strip()
                        
                        sf2_dtype=(str(getattr(row,'sf2_Datatype'))).lower().strip()
                        
                        if sf1_dtype!=sf2_dtype:result_datatype.append("No")
                        else: result_datatype.append("Yes")


                    tempDf['Datatypes Same?']=result_datatype
                    tempDf['TableName']=str(sf1_table)+'----'+str(sf2_table)
                    
                    
                    # tempDf.to_csv('Datatype_result_sanity.csv')
                    self.datatype_df=pd.concat([self.datatype_df,tempDf],ignore_index=True)

                    ###############################SumOfValues############################################

                except Exception as err:
                    print("In table ",row[0],'facing issue: ',str(err))
             
             
                 
            ########################## TAB-2 Total Row counts #################################

            if row_count_check!='n':
                try:
                    tq='select cast(count(*) as double) as sf1_ROW_COUNT from  '+ sf1_table
                    #    print("Running query: ",tq)
                    sf1_count=self.run_sf_query(tq)
                    sf1_count=sf1_count.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    sf1_count.columns=['sf1_ROW_COUNT']
                    
                    sq='select cast(count(*) as double) as sf2_ROW_COUNT from  '+ sf2_table
                    #    print("Running query: ",sq)
                    sf2_count=self.run_sf_query(sq)
                    sf2_count=sf2_count.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    sf2_count.columns=['sf2_ROW_COUNT']
                    
                    # result
                    
                    row_counts=pd.concat([sf1_count,sf2_count],axis=1)
                    row_counts['Difference(sf1-sf2)']=row_counts['sf1_ROW_COUNT']-row_counts['sf2_ROW_COUNT']
                    row_counts['Difference_%']=100*row_counts['Difference(sf1-sf2)']/row_counts['sf1_ROW_COUNT']
                    row_counts['TableName']=original_sf1_table
                    self.sf1_sf2_sanity_check_rowcount_df=pd.concat([self.sf1_sf2_sanity_check_rowcount_df,row_counts])

                    # row_counts.to_csv('Tab2_'+original_sf1_table+'.csv',index=False)
                    print('Row Count Tab (2nd) created')
                except Exception as err:
                    print('Not able to complete tab2 due to: ',str(err))

            ########################## TAB-1 SumOfValues #################################
            if SumOfValues!='n':
                try:
                    

                    sf1_checksum=list()
                    for col in sf1_table_cols:
                        try:
                            t_q='select cast(sum({colName}) as double )as c from {tn}'.format(colName=col,tn=sf1_table)
                            #    print("Running query: ",t_q)
                            res=self.run_sf_query(t_q)
                            sf1_checksum.append(res.iloc[:,0][0])
                            # measure_cols_sf1 +=', cast(sum({colName}) as double) as sf1_Sum_{colName} '.format(colName=col)
                            # measure_cols_sf1_names_list.append(str(col).lower()+"_sf1")
                        except:
                            sf1_checksum.append(np.nan)
                    

                    sf2_checksum=list()
                    
                    for col in sf2_cols:
                        try:
                            s_q='select cast( sum({colName}) as double) as c from {tn}'.format(colName=col,tn=sf2_table.strip().upper())
                            #    print("Running query: ",s_q)
                            res=self.run_sf_query(s_q)
                            sf2_checksum.append(res.iloc[:,0][0])
                            # measure_cols_sf2 +=', cast(sum({colName}) as double) as sf2_Sum_{colName} '.format(colName=col)
                            # measure_cols_sf2_names_list.append(str(col).lower()+"_sf2")
                        except:
                            sf2_checksum.append(np.nan)

                
                    

                    df2_sf2=pd.concat([ pd.DataFrame({'sf2_ColName':sf2_cols}),pd.DataFrame({'sf2_CheckSum':sf2_checksum})],axis=1)
                    df2_sf2=df2_sf2.applymap(lambda x: x.lower() if isinstance(x, str) else x)

                    

                    df1_sf1=pd.concat([pd.DataFrame({'sf1_ColName':sf1_table_cols}),pd.DataFrame({'sf1_CheckSum':sf1_checksum})],axis=1)
                    df1_sf1=df1_sf1.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    left_indexes=['sf1_ColName']
                    right_indexes=['sf2_ColName']
                    tab1_df=df1_sf1.merge(df2_sf2,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_sf1','_sf2'])

                    finalCols_tab1=['sf2_ColName','sf2_CheckSum','sf1_ColName','sf1_CheckSum']
                    tab1_df=tab1_df[finalCols_tab1]


                    
                    tab1_df['Checksum_Result']=tab1_df['sf2_CheckSum'].eq(tab1_df['sf1_CheckSum'])
                    tab1_df['TableName']=original_sf1_table

                    self.sf1_sf2_sanity_check_checksum_df=pd.concat([self.sf1_sf2_sanity_check_checksum_df,tab1_df])
                    
                    
                #     tab1_df['diff_%']=100*tab1_df['diff']/tab1_df['sf1_CheckSum']
                    
                    # tab1_df.to_csv('Tab1_checksum_'+original_sf1_table+'.csv',index=False)
                    print('SumOfValues Tab (1st) created')

                except Exception as err:
                    print('Not able to complete tab1 due to: ',str(err))
            
            ########################## TAB-3 DISTINCT COUNT #################################
        
            #sf1
            if distinct_check!='n':
                try:
                    # sf1_distinct_query_list=list()
                    sf1_distinct_query_res=list()
                    for col in sf1_table_cols:
                        # sf1_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                        q="select cast(count(distinct({colName})) as double) as {colName} from {ht}".format(colName=col,ht=sf1_table)
                        val_df=self.run_sf_query(q)
                        val=val_df.iloc[0][0]
                        sf1_distinct_query_res.append(val)

                    
                    
                    sf1_res_df=pd.DataFrame({'sf1_ColName':sf1_table_cols,'sf1_DistinctCount':sf1_distinct_query_res})
                    sf1_res_df.columns=['sf1_ColName','sf1_DistinctCount']
                     
                    sf1_res_df=sf1_res_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    #sf2
                    sf2_distinct_query_list=list()
                    for col in sf2_cols:
                        sf2_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                    
                    sf2_distinct_query=",".join(sf2_distinct_query_list)
                    
                    finalDistinct_query='select ' + sf2_distinct_query + " from "+ sf2_table

                    #    print("Running query: ",finalDistinct_query)

                    sf2_res_df=self.run_sf_query(finalDistinct_query)
                    sf2_res_df=sf2_res_df.transpose().reset_index()
                    sf2_res_df.columns=['sf2_ColName','sf2_DistinctCount']
                    sf2_res_df=sf2_res_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    
                    
                    
                    
                    #making finalDistinctRes
                    left_indexes=['sf1_ColName']
                    right_indexes=['sf2_ColName']
                    
                    final_Distinct=sf1_res_df.merge(sf2_res_df,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_sf1','_sf2'])
                    
                    #final_Distinct=pd.concat([sf1_res_df,sf2_res_df],axis=1)
                    final_Distinct['Difference(TD-sf2)']=final_Distinct['sf1_DistinctCount']-final_Distinct['sf2_DistinctCount']
                    final_Distinct['Difference_%']=100*final_Distinct['Difference(TD-sf2)']/final_Distinct['sf1_DistinctCount']
                    final_Distinct['TableName']=original_sf1_table
                    self.sf1_sf2_sanity_check_distinct_count_df=pd.concat([self.sf1_sf2_sanity_check_distinct_count_df,final_Distinct])

                    # final_Distinct.to_csv('Tab3_'+original_sf1_table+'.csv',index=False)
                    print('Distinct Counts Tab (3rd) Created')

                except Exception as err:
                    print('Not able to complete tab3 due to: ',str(err))    
            
            #######
            ########################## TAB-4 null COUNT #################################
        
            #sf1 null count
             
            if null_count_check!='n':
                try:
                    # sf1_distinct_query_list=list()
                    sf1_null_query_res=list()
                    for col in sf1_table_cols:
                        try:
                            q="select cast(count(*) as double) from {ht} where {colName} is null or {colName}='' ".format(ht=sf1_table,colName=col)
                            # sf1_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                            # q="select cast(count(distinct({colName})) as double) as {colName} from {ht}".format(colName=col,ht=sf1_table)
                            val_df=self.run_sf_query(q)
                        except:
                            q="select cast(count(*) as double) from {ht} where {colName} is null  ".format(ht=sf1_table,colName=col)
                            # sf1_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                            # q="select cast(count(distinct({colName})) as double) as {colName} from {ht}".format(colName=col,ht=sf1_table)
                            val_df=self.run_sf_query(q)

                        val=val_df.iloc[0][0]
                        sf1_null_query_res.append(val)

                    
                    
                    sf1_res_df=pd.DataFrame({'sf1_ColName':sf1_table_cols,'sf1_NullCount':sf1_null_query_res})
                    sf1_res_df.columns=['sf1_ColName','sf1_NullCount']
                     
                    sf1_res_df=sf1_res_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    #sf2
                    sf2_null_query_res=list()
                    for col in sf2_cols:
                        try:
                            q="select cast(count(*) as double) from {st} where {colName} is null or {colName}='' ".format(st=sf2_table,colName=col)
                            # sf2_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                            # q="select cast(count(distinct({colName})) as double) as {colName} from {ht}".format(colName=col,ht=sf2_table)
                            val_df=self.run_sf_query(q)
                        except:
                            q="select cast(count(*) as double) from {st} where {colName} is null ".format(st=sf2_table,colName=col)
                            # sf2_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                            # q="select cast(count(distinct({colName})) as double) as {colName} from {ht}".format(colName=col,ht=sf2_table)
                            val_df=self.run_sf_query(q)

                        val=val_df.iloc[0][0]
                        sf2_null_query_res.append(val)

                    
                    
                    sf2_res_df=pd.DataFrame({'sf2_ColName':sf2_cols,'sf2_NullCount':sf2_null_query_res})
                    sf2_res_df.columns=['sf2_ColName','sf2_NullCount']
                     
                    sf2_res_df=sf2_res_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    
                    
                    
                    
                    #making finalDistinctRes
                    left_indexes=['sf1_ColName']
                    right_indexes=['sf2_ColName']
                    
                    final_NullCount=sf1_res_df.merge(sf2_res_df,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_sf1','_sf2'])
                    
                    #final_NullCount=pd.concat([sf1_res_df,sf2_res_df],axis=1)
                    final_NullCount['Difference(sf1-sf2)']=final_NullCount['sf1_NullCount']-final_NullCount['sf2_NullCount']
                    final_NullCount['Difference_%']=100*final_NullCount['Difference(sf1-sf2)']/final_NullCount['sf1_NullCount']
                    final_NullCount['TableName']=original_sf1_table
                    self.sf1_sf2_sanity_check_null_count_df=pd.concat([self.sf1_sf2_sanity_check_null_count_df,final_NullCount])

                    # final_NullCount.to_csv('Tab3_'+original_sf1_table+'.csv',index=False)
                    print('null Counts Tab (4th) Created')

                except Exception as err:
                    print('Not able to complete tab4(null count) due to: ',str(err))    
            

            #######
            

            fileName='sf1_sf2_Sanity_check_'+curr_time+'.xlsx'
            with pd.ExcelWriter(fileName) as writer:
                        self.datatype_df.to_excel(writer,sheet_name='DDL',index = False,header=True) 
                        self.sf1_sf2_sanity_check_checksum_df.to_excel(writer,sheet_name='SumOfValues',index = False,header=True) 
                        self.sf1_sf2_sanity_check_rowcount_df.to_excel(writer,sheet_name='Row Count',index = False,header=True) 
                        self.sf1_sf2_sanity_check_distinct_count_df.to_excel(writer,sheet_name='Distinct',index = False,header=True)
                        self.missing_tables_sf1_sf2_sanity_df.to_excel(writer,sheet_name='Missing Tables',index = False,header=True) 

            print(original_sf1_table,' completed...')
        except Exception as err:
            raise Exception(str(err))
      

    def sanity_check_mssql_sf(self,mssql_table,mssql_condition,sf_table,sf_condition,cols_to_sum,curr_time,ddl_check,row_count_check,distinct_check,SumOfValues):
        ########################## BASIC INFO #################################
        
        datatypes_mssql_sf_mapping=dict()
        datatypes_mssql_sf_mapping['string']='varchar'
        datatypes_mssql_sf_mapping['timestamp']='timestamp_ntz'
        datatypes_mssql_sf_mapping['int']='number'
        datatypes_mssql_sf_mapping['decimal']='number'
        datatypes_mssql_sf_mapping['float']='number'
        datatypes_mssql_sf_mapping['bigint']='number'
        datatypes_mssql_sf_mapping['date']='date'
        datatypes_mssql_sf_mapping['varchar']='varchar'
        datatypes_mssql_sf_mapping['char']='varchar'
        


        
    
        path=self.sanity_checkResultDownloadLocationEntry_mssql_to_sf.get()
        os.chdir(path)
        #mssql_table=row[1]
        #mssql_condition=row[2]
        original_mssql_table=mssql_table
        original_sf_table=sf_table
        try:
            if (str(mssql_condition).strip().lower()!='nan'):
                        mssql_table +=' where '+mssql_condition
            
            
            if (str(sf_condition).strip().lower()!='nan'):
                        sf_table +=' where '+sf_condition
            
                
            
            mssql_table_cols=self.getCols_mssql(mssql_table)
            sf_cols=self.getCols_SF(sf_table)
            
            print('***checking datatype :',sf_table)
            #checking datatype`
            if ddl_check!='n':
                try:
                    table_mssql=original_mssql_table
                    table_sf=original_sf_table
                    print("Datatype check for:=> mssql table:",table_mssql," SF table:",table_sf)

                    try:
                    #mssql
                        mssql_db_name,mssql_schema,mssql_tablename=table_mssql.split('.')
                        #mssql
                        query_mssql= "select * from {db}.information_schema.columns  where table_name = '{tn}' and table_schema='{ts}' order by ordinal_position".format(db=mssql_db_name,tn=mssql_tablename,ts=mssql_schema)
                        
                        if self.useSqoopCommand.get()==1:
                            query_mssql_df=self.run_mssql_sqoop_query(query_mssql)
                        else:
                            query_mssql_df=self.run_mssql_query(query_mssql)
                        
                        
                        query_mssql_df.columns=[col.strip() for col in query_mssql_df.columns]
                        query_mssql_df=query_mssql_df[['COLUMN_NAME','DATA_TYPE','COLUMN_DEFAULT','IS_NULLABLE','CHARACTER_MAXIMUM_LENGTH']]
                        
                        query_mssql_df=query_mssql_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                        

                        query_mssql_df.columns=['mssql_COLUMN_NAME','mssql_DATA_TYPE','mssql_COLUMN_DEFAULT','mssql_IS_NULLABLE','mssql_CHARACTER_MAXIMUM_LENGTH']

                        query_mssql_df=query_mssql_df.applymap(lambda x: x.upper() if isinstance(x, str) else x)
                        # query_mssql_df=query_mssql_df.replace('(NULL)',np.NaN)

                    except Exception as err:
                        print(err)
                        self.missing_tables_mssql_sf_sanity_df.loc[len(self.missing_tables_mssql_sf_sanity_df.index)]=[table_mssql,'mssql']
                    #sf

                    try:
                        sf_db_name,sf_schema,sf_tablename=table_sf.split('.')

                        datatype_query='desc table '+ table_sf
                        sf_datatype_df=self.run_sf_query(datatype_query)
                        sf_datatype_df=sf_datatype_df[['name','type']]
                        sf_datatype_df.columns=['SF_COLUMN_NAME','SF_DATA_TYPE']
                    
                        info_schema_sf_query="select COLUMN_NAME as SF_COLUMN_NAME ,IS_NULLABLE as SF_IS_NULLABLE,COLUMN_DEFAULT as SF_COLUMN_DEFAULT,CHARACTER_MAXIMUM_LENGTH as SF_CHARACTER_MAXIMUM_LENGTH from  {db}.INFORMATION_SCHEMA.columns  where table_name='{tn}' and table_schema='{ts}'".format(db=sf_db_name.upper(),tn=sf_tablename.upper(),ts=sf_schema.upper())
                        info_schema_sf_df=self.run_sf_query(info_schema_sf_query)

                        info_schema_sf_df=info_schema_sf_df.merge(sf_datatype_df,on='SF_COLUMN_NAME')
                        info_schema_sf_df=info_schema_sf_df.applymap(lambda x: x.upper() if isinstance(x, str) else x)
                        
                    except Exception as err:
                        print(err)
                        self.missing_tables_mssql_sf_sanity_df.loc[len(self.missing_tables_mssql_sf_sanity_df.index)]=[original_sf_table,'SF']

                    tempDf=query_mssql_df.merge(info_schema_sf_df,left_on='mssql_COLUMN_NAME',right_on='SF_COLUMN_NAME',how='outer')

                    result_datatype=list()
                    tn=str(table_sf).split('.')[-1]
                    

                    attributes=['COLUMN_NAME','DATA_TYPE','COLUMN_DEFAULT','IS_NULLABLE','CHARACTER_MAXIMUM_LENGTH']
                    mismatching_attribute_list=list()
                    result_datatype=list()

                    for row in (tempDf.itertuples()):
                         

                        for attr in attributes:
                            mssql_col=str(getattr(row,'mssql_'+attr)).strip()
                            sf_col=str(getattr(row,'SF_'+attr)).strip()


                            try:
                                v1=str(float(mssql_col))
                                v2=str(float(sf_col))
                                mssql_col=v1
                                sf_col=v2
                                    
                            except: pass
                            # if ( ('(NULL)' in mssql_col) and ('None'==sf_col or 'nan'==sf_col) ):
                            #     continue
                            

                            if attr=='DATA_TYPE':
                                mssql_dtype=str(getattr(row,'mssql_DATA_TYPE'))
                        
                                sf_dtype=str(getattr(row,'SF_DATA_TYPE'))

                                if str(mssql_col)=='nan' or str(sf_col)=='nan':
                                    mismatching_attribute_list.append('Column Missing')
                                    result_datatype.append("No")
                                    continue
                                
                                if not ( ( mssql_dtype==sf_dtype  ) or
                                    ( ('INT' in mssql_dtype) and ('NUMBER' in sf_dtype) ) or 
                                    ( ('SMALLINT' in mssql_dtype) and ('NUMBER' in sf_dtype) ) or 
                                    ( ('BYTEINT' in mssql_dtype) and ('NUMBER' in sf_dtype) ) or 
                                    ( ('DECIMAL' in mssql_dtype) and ('NUMBER' in sf_dtype) ) or 
                                    ( ('NUMERIC' in mssql_dtype) and ('NUMBER' in sf_dtype) ) or 
                                    ( ('FLOAT' in mssql_dtype) and ('NUMBER' in sf_dtype) ) or 
                                    ( ('DECIMAL' in mssql_dtype) and ('FLOAT' in sf_dtype) ) or 
                                    ( ('CHAR' in mssql_dtype) and ('VARCHAR' in sf_dtype) ) or
                                    ( ('BIT' in mssql_dtype) and ('BOOLEAN' in sf_dtype) ) or 
                                    ( ('DATE' in mssql_dtype) and ('DATE' in sf_dtype) ) or #(NULL)
                                    
                                    ( (mssql_dtype in ['CHAR','VARCHAR']) and ('VARCHAR' in sf_dtype)  ) or
                                    ( ('TIMESTAMP' in mssql_dtype or 'DATETIME' in mssql_dtype) and ('TIMESTAMP' in sf_dtype) )  ):
                                        mismatching_attribute_list.append('Datatype mismatch')
                                        result_datatype.append("No")
                                        break
                            
                            elif mssql_col!=sf_col and not ( mssql_col=='(NULL)' and sf_col in ['nan','None']):
                                    mismatching_attribute_list.append(attr)
                                    result_datatype.append("No")
                                    break
                            if attr=='CHARACTER_MAXIMUM_LENGTH':
                                mismatching_attribute_list.append('')
                                result_datatype.append("Yes")

                        
                                
                    tempDf['Mismatched_Attr']=mismatching_attribute_list
                    tempDf['DDL Same?']=result_datatype
                    tempDf['TableName']=str(table_mssql) + " ---- " + str(table_sf)
                    
                    
                    
                    self.datatype_df=pd.concat([self.datatype_df,tempDf],ignore_index=True)


                    ###############################SumOfValues############################################

                except Exception as err:
                    print("In table ",row[0],'facing issue: ',str(err))
             
             
                 
            ########################## TAB-2 Total Row counts #################################

            if row_count_check!='n':
                try:
                    tq='select cast(count(*) as float) as mssql_ROW_COUNT from  '+ mssql_table
                    #    print("Running query: ",tq)
                    mssql_count=self.run_mssql_query(tq)
                    mssql_count=mssql_count.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    mssql_count.columns=['mssql_ROW_COUNT']
                    
                    sq='select cast(count(*) as float) as SF_ROW_COUNT from  '+ sf_table
                    #    print("Running query: ",sq)
                    sf_count=self.run_sf_query(sq)
                    sf_count=sf_count.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    sf_count.columns=['SF_ROW_COUNT']
                    
                    # result
                    
                    row_counts=pd.concat([mssql_count,sf_count],axis=1)
                    row_counts['Difference(mssql-SF)']=row_counts['mssql_ROW_COUNT']-row_counts['SF_ROW_COUNT']
                    row_counts['Difference_%']=100*row_counts['Difference(mssql-SF)']/row_counts['mssql_ROW_COUNT']
                    row_counts['TableName']=original_mssql_table
                    self.mssql_sf_sanity_check_rowcount_df=pd.concat([self.mssql_sf_sanity_check_rowcount_df,row_counts])

                    # row_counts.to_csv('Tab2_'+original_mssql_table+'.csv',index=False)
                    print('Row Count Tab (2nd) created')
                except Exception as err:
                    print('Not able to complete tab2 due to: ',str(err))

            ########################## TAB-1 SumOfValues #################################
            if SumOfValues!='n':
                try:
                    

                    mssql_checksum=list()
                    for col in mssql_table_cols:
                        try:
                            t_q='select cast(sum({colName}) as double )as c from {tn}'.format(colName=col,tn=mssql_table)
                            #    print("Running query: ",t_q)
                            res=self.run_mssql_query(t_q)
                            mssql_checksum.append(res.iloc[:,0][0])
                            # measure_cols_mssql +=', cast(sum({colName}) as double) as mssql_Sum_{colName} '.format(colName=col)
                            # measure_cols_mssql_names_list.append(str(col).lower()+"_mssql")
                        except:
                            mssql_checksum.append(np.nan)
                    

                    sf_checksum=list()
                    
                    for col in sf_cols:
                        try:
                            s_q='select cast( sum({colName}) as double) as c from {tn}'.format(colName=col,tn=sf_table.strip().upper())
                            #    print("Running query: ",s_q)
                            res=self.run_sf_query(s_q)
                            sf_checksum.append(res.iloc[:,0][0])
                            # measure_cols_sf +=', cast(sum({colName}) as double) as SF_Sum_{colName} '.format(colName=col)
                            # measure_cols_sf_names_list.append(str(col).lower()+"_SF")
                        except:
                            sf_checksum.append(np.nan)

                
                    

                    df2_sf=pd.concat([ pd.DataFrame({'SF_ColName':sf_cols}),pd.DataFrame({'SF_CheckSum':sf_checksum})],axis=1)
                    df2_sf=df2_sf.applymap(lambda x: x.lower() if isinstance(x, str) else x)

                    

                    df1_mssql=pd.concat([pd.DataFrame({'mssql_ColName':mssql_table_cols}),pd.DataFrame({'mssql_CheckSum':mssql_checksum})],axis=1)
                    df1_mssql=df1_mssql.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    left_indexes=['mssql_ColName']
                    right_indexes=['SF_ColName']
                    tab1_df=df1_mssql.merge(df2_sf,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_mssql','_SF'])

                    finalCols_tab1=['SF_ColName','SF_CheckSum','mssql_ColName','mssql_CheckSum']
                    tab1_df=tab1_df[finalCols_tab1]


                    
                    tab1_df['Checksum_Result']=tab1_df['SF_CheckSum'].eq(tab1_df['mssql_CheckSum'])
                    tab1_df['TableName']=original_mssql_table

                    self.mssql_sf_sanity_check_checksum_df=pd.concat([self.mssql_sf_sanity_check_checksum_df,tab1_df])
                    
                    
                #     tab1_df['diff_%']=100*tab1_df['diff']/tab1_df['mssql_CheckSum']
                    
                    # tab1_df.to_csv('Tab1_checksum_'+original_mssql_table+'.csv',index=False)
                    print('SumOfValues Tab (1st) created')

                except Exception as err:
                    print('Not able to complete tab1 due to: ',str(err))
            
            ########################## TAB-3 DISTINCT COUNT #################################
        
            #mssql
            if distinct_check!='n':
                try:
                    mssql_distinct_query_list=list()
                    for col in mssql_table_cols:
                        mssql_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                    
                    mssql_distinct_query=",".join(mssql_distinct_query_list)
                    
                    finalDistinct_query='select ' + mssql_distinct_query + " from "+ mssql_table
                   #    print("Running query: ",finalDistinct_query)
                    mssql_res_df=self.run_mssql_query(finalDistinct_query)
                    mssql_res_df=mssql_res_df.transpose().reset_index()
                    mssql_res_df.columns=['mssql_ColName','mssql_DistinctCount']
                    mssql_res_df=mssql_res_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    #SF
                    sf_distinct_query_list=list()
                    for col in sf_cols:
                        sf_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                    
                    sf_distinct_query=",".join(sf_distinct_query_list)
                    
                    finalDistinct_query='select ' + sf_distinct_query + " from "+ sf_table

                    #    print("Running query: ",finalDistinct_query)

                    sf_res_df=self.run_sf_query(finalDistinct_query)
                    sf_res_df=sf_res_df.transpose().reset_index()
                    sf_res_df.columns=['SF_ColName','SF_DistinctCount']
                    sf_res_df=sf_res_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
                    
                    
                    
                    
                    
                    #making finalDistinctRes
                    left_indexes=['mssql_ColName']
                    right_indexes=['SF_ColName']
                    
                    final_Distinct=mssql_res_df.merge(sf_res_df,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_mssql','_SF'])
                    
                    #final_Distinct=pd.concat([mssql_res_df,sf_res_df],axis=1)
                    final_Distinct['Difference(TD-SF)']=final_Distinct['mssql_DistinctCount']-final_Distinct['SF_DistinctCount']
                    final_Distinct['Difference_%']=100*final_Distinct['Difference(TD-SF)']/final_Distinct['mssql_DistinctCount']
                    final_Distinct['TableName']=original_mssql_table
                    self.mssql_sf_sanity_check_distinct_count_df=pd.concat([self.mssql_sf_sanity_check_distinct_count_df,final_Distinct])

                    # final_Distinct.to_csv('Tab3_'+original_mssql_table+'.csv',index=False)
                    print('Distinct Counts Tab (3rd) Created')

                except Exception as err:
                    print('Not able to complete tab3 due to: ',str(err))    
            
            

            fileName='mssql_SF_Sanity_check_'+curr_time+'.xlsx'
            with pd.ExcelWriter(fileName) as writer:
                        self.datatype_df.to_excel(writer,sheet_name='DDL',index = False,header=True) 
                        self.mssql_sf_sanity_check_checksum_df.to_excel(writer,sheet_name='SumOfValues',index = False,header=True) 
                        self.mssql_sf_sanity_check_rowcount_df.to_excel(writer,sheet_name='Row Count',index = False,header=True) 
                        self.mssql_sf_sanity_check_distinct_count_df.to_excel(writer,sheet_name='Distinct',index = False,header=True)
                        self.missing_tables_mssql_sf_sanity_df.to_excel(writer,sheet_name='Missing Tables',index = False,header=True) 

            print(original_mssql_table,' completed...')
        except Exception as err:
            raise Exception(str(err))
       


    
    def DownloadAndGetCount_mssql_to_sf(self,mssql_table,mssql_condition,sf_table,sf_condition,cols_to_sum):
        ########################## BASIC INFO #################################
       
           #mssql_table=row[1]
           #mssql_condition=row[2]
        path=self.tableStatsResultDownloadLocationEntry_mssql_to_sf.get()
        os.chdir(path)
        original_mssql_table=mssql_table
        original_sf_table=sf_table
        try:
            if (str(mssql_condition)!='nan'):
                        mssql_table +=' where '+mssql_condition
            
            #sf_table=row[3]
            #sf_condition=row[4]
            
            if (str(sf_condition)!='nan'):
                        sf_table +=' where '+sf_condition
            
            db_name,schema_name,table_name=original_mssql_table.split('.')
            getcol_mssql_query="select COLUMN_NAME from {db}.information_schema.columns  where table_name = '{tb}' and table_schema='{sch}' order by ordinal_position".format(db=db_name,tb=table_name,sch=schema_name)
            mssql_table_cols=list(self.run_mssql_query(getcol_mssql_query))
            sf_cols=self.getCols_SF(sf_table)
            
            print('***Basic info checked for:',sf_table)
            
            ########################## TAB-1 SumOfValues #################################
            try:
                measure_cols_mssql=''
                measure_cols_mssql_names_list=list()

                measure_cols_sf=''
                measure_cols_sf_names_list=list()

                mssql_checksum=list()
                for col in mssql_table_cols:
                    try:
                        t_q='select cast(sum({colName}) as double )as c from {tn}'.format(colName=col,tn=mssql_table)
                        #    print("Running query: ",t_q)
                        res=self.run_mssql_query(t_q)
                        mssql_checksum.append(res.iloc[:,0][0])
                        # measure_cols_mssql +=', cast(sum({colName}) as double) as mssql_Sum_{colName} '.format(colName=col)
                        # measure_cols_mssql_names_list.append(str(col).lower()+"_mssql")
                    except:
                        mssql_checksum.append(np.nan)

                for col in cols_to_sum:
                    measure_cols_mssql +=', cast(sum({colName}) as double) as mssql_Sum_{colName} '.format(colName=col)
                    measure_cols_mssql_names_list.append(str(col).lower()+"_mssql")

                
                sf_checksum=list()
                for col in sf_cols:
                    try:
                        s_q='select cast( sum({colName}) as double) as c from {tn}'.format(colName=col,tn=sf_table.strip().upper())
                        #    print("Running query: ",s_q)
                        res=self.run_sf_query(s_q)
                        sf_checksum.append(res.iloc[:,0][0])
                        # measure_cols_sf +=', cast(sum({colName}) as double) as sf_Sum_{colName} '.format(colName=col)
                        # measure_cols_sf_names_list.append(str(col).lower()+"_sf")
                    except:
                        sf_checksum.append(np.nan)

                for col in cols_to_sum:
                    measure_cols_sf +=', cast(sum({colName}) as double) as sf_Sum_{colName} '.format(colName=col)
                    measure_cols_sf_names_list.append(str(col).lower()+"_sf")

                # datatype_query='desc table '+original_mssql_table
                db_name,schema_name,table_name=original_mssql_table.split('.')
                
                clustering_key_mssql="SELECT  CLUSTERING_KEY from {db_name}.INFORMATION_SCHEMA.tables where table_name='{tn}' and table_schema='{ts}' ".format(db_name=db_name.strip().upper(),tn=table_name.strip().upper(),ts=schema_name.strip().upper())
                clustering_key_mssql_df=self.run_mssql_query(clustering_key_mssql)
                try:
                    indexed_cols=[s.strip() for s in clustering_key_mssql_df['CLUSTERING_KEY'][0].replace(')','').replace('LINEAR(','').split(',')]
                except:indexed_cols=list()

                indexed_cols_yes_no=list()
                for col in mssql_table_cols:
                    if col in indexed_cols:
                        indexed_cols_yes_no.append('Yes')
                    else: indexed_cols_yes_no.append('N/A')
                
                mssql_indexed_cols_df=pd.DataFrame({'Col_Name':mssql_table_cols,'SF_Base_Indexed':indexed_cols_yes_no})

                col_info_query="SELECT  COLUMN_NAME as SF_Base_ColName ,CHARACTER_MAXIMUM_LENGTH as SF_Base_ColumnLength,DATA_TYPE  from {db_name}.INFORMATION_SCHEMA.columns where table_name='{tn}' and table_schema='{ts}' ".format(db_name=db_name.strip().upper(),tn=table_name.strip().upper(),ts=schema_name.strip().upper())
                
                mssql_col_info_df=self.run_mssql_query(col_info_query)
                mssql_col_info_df.columns=['SF_Base_ColName','SF_Base_ColumnLength','SF_Base_Datatype']
                
                # datatype_query='desc table '+ original_mssql_table
                # sf_datatype_df=self.run_sf_query(datatype_query)
                # sf_datatype_df=sf_datatype_df[['name','type']]        
                # sf_datatype_df.columns=['SF_Base_ColName','SF_Base_Datatype']
                
                df1_mssql=pd.concat([pd.DataFrame({'SF_Base_ColName':mssql_table_cols}),pd.DataFrame({'SF_Base_CheckSum':mssql_checksum})],axis=1)
                df1_mssql=df1_mssql.merge(mssql_col_info_df,left_on='SF_Base_ColName',right_on='SF_Base_ColName',suffixes=['_Base','_Release'])
                df1_mssql=df1_mssql.merge(mssql_indexed_cols_df,left_on='SF_Base_ColName',right_on='Col_Name',suffixes=['_Base','_Release'])
                # df1_mssql=df1_mssql.merge(sf_datatype_df,left_on='SF_Base_ColName',right_on='SF_Base_ColName',suffixes=['_Base','_Release'])

                db_name,schema_name,table_name=original_sf_table.split('.')
                
                clustering_key_sf="SELECT  CLUSTERING_KEY as SF_Release_Indexed from {db_name}.INFORMATION_SCHEMA.tables where table_name='{tn}' and table_schema='{ts}' ".format(db_name=db_name.strip().upper(),tn=table_name.strip().upper(),ts=schema_name.strip().upper())
                clustering_key_sf_df=self.run_sf_query(clustering_key_sf)
                try:
                    indexed_cols=[s.strip() for s in clustering_key_sf_df['CLUSTERING_KEY'][0].replace(')','').replace('LINEAR(','').split(',')]
                except:indexed_cols=list()

                indexed_cols_yes_no=list()
                for col in sf_cols:
                    if col in indexed_cols:
                        indexed_cols_yes_no.append('Yes')
                    else: indexed_cols_yes_no.append('No')
                
                sf_indexed_cols_df=pd.DataFrame({'Col_Name':sf_cols,'SF_Release_Indexed':indexed_cols_yes_no})

                col_info_query="SELECT  COLUMN_NAME as SF_Release_ColName,CHARACTER_MAXIMUM_LENGTH as SF_Release_ColumnLength from {db_name}.INFORMATION_SCHEMA.columns where table_name='{tn}' and table_schema='{ts}' ".format(db_name=db_name.strip().upper(),tn=table_name.strip().upper(),ts=schema_name.strip().upper())
                sf_col_info_df=self.run_sf_query(col_info_query)
                sf_col_info_df.columns=['SF_Release_ColName','SF_Release_ColumnLength']

                datatype_query='desc table '+ original_sf_table
                sf_datatype_df=self.run_sf_query(datatype_query)
                sf_datatype_df=sf_datatype_df[['name','type']]        
                sf_datatype_df.columns=['SF_Release_ColName','SF_Release_Datatype']

                df2_sf=pd.concat([ pd.DataFrame({'SF_Release_ColName':sf_cols}),pd.DataFrame({'SF_Release_CheckSum':sf_checksum})],axis=1)
                df2_sf=df2_sf.merge(sf_col_info_df,left_on='SF_Release_ColName',right_on='SF_Release_ColName',suffixes=['_Base','_Release'])
                df2_sf=df2_sf.merge(sf_indexed_cols_df,left_on='SF_Release_ColName',right_on='Col_Name',suffixes=['_Base','_Release'])
                df2_sf=df2_sf.merge(sf_datatype_df,left_on='SF_Release_ColName',right_on='SF_Release_ColName',suffixes=['_Base','_Release'])

                left_indexes=['SF_Base_ColName']
                right_indexes=['SF_Release_ColName']
                tab1_df=df1_mssql.merge(df2_sf,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_Base','_Release'])
                
                
                finalCols_tab1=['SF_Base_ColName','SF_Base_Datatype','SF_Base_ColumnLength','SF_Base_CheckSum','SF_Base_Indexed',
                                'SF_Release_ColName','SF_Release_Datatype','SF_Release_ColumnLength','SF_Release_CheckSum','SF_Release_Indexed']
                tab1_df=tab1_df[finalCols_tab1]


                tab1_df['ColNameCheck_Result']=tab1_df['SF_Base_ColName']==tab1_df['SF_Release_ColName']
                tab1_df['Datatype_Result']=tab1_df['SF_Base_Datatype']==tab1_df['SF_Release_Datatype']
                tab1_df['ColumnLength_Result']=tab1_df['SF_Base_ColumnLength'].eq(tab1_df['SF_Release_ColumnLength'])
                tab1_df['Checksum_Result']=tab1_df['SF_Base_CheckSum'].eq(tab1_df['SF_Release_CheckSum'])
                tab1_df['Indexed_Result']=tab1_df['SF_Base_Indexed'].eq(tab1_df['SF_Release_Indexed'])
                
                
                

                tab1_df.to_csv('Tab1_'+original_mssql_table+'.csv',index=False)
                print('SumOfValues Tab (1st) created')

            except Exception as err:
                print('Not able to complete tab1 due to: ',str(err))
            ########################## TAB-2 Total Row counts #################################

            try:
                tq='select cast(count(*) as double) as mssql_ROW_COUNT from  '+ mssql_table
                #    print("Running query: ",tq)
                mssql_count=self.run_mssql_query(tq)
                
                sq='select cast(count(*) as double) as sf_ROW_COUNT from  '+ sf_table
                #    print("Running query: ",sq)
                sf_count=self.run_sf_query(sq)
                
                
                #result
                
                row_counts=pd.concat([mssql_count,sf_count],axis=1)
                row_counts['Difference(mssql-sf)']=row_counts['mssql_ROW_COUNT']-row_counts['sf_ROW_COUNT']
                row_counts['Difference_%']=100*row_counts['Difference(mssql-sf)']/row_counts['mssql_ROW_COUNT']
                
                row_counts.to_csv('Tab2_'+original_mssql_table+'.csv',index=False)
                print('Row Count Tab (2nd) created')
            except Exception as err:
                print('Not able to complete tab2 due to: ',str(err))
            ########################## TAB-3 DISTINCT COUNT #################################
        
            try:
                #mssql
                mssql_distinct_query_list=list()
                for col in mssql_table_cols:
                    mssql_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                
                mssql_distinct_query=",".join(mssql_distinct_query_list)
                
                finalDistinct_query='select ' + mssql_distinct_query + " from "+ mssql_table
                #    print("Running query: ",finalDistinct_query)
                mssql_res_df=self.run_mssql_query(finalDistinct_query)
                mssql_res_df=mssql_res_df.transpose().reset_index()
                mssql_res_df.columns=['mssql_ColName','mssql_DistinctCount']
                
                #sf
                sf_distinct_query_list=list()
                for col in sf_cols:
                    sf_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                
                sf_distinct_query=",".join(sf_distinct_query_list)
                
                finalDistinct_query='select ' + sf_distinct_query + " from "+ sf_table

                #    print("Running query: ",finalDistinct_query)

                sf_res_df=self.run_sf_query(finalDistinct_query)
                sf_res_df=sf_res_df.transpose().reset_index()
                sf_res_df.columns=['sf_ColName','sf_DistinctCount']
                
                
                
                
                #making finalDistinctRes
                left_indexes=['mssql_ColName']
                right_indexes=['sf_ColName']
                
                final_Distinct=mssql_res_df.merge(sf_res_df,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_mssql','_sf'])
                
                #final_Distinct=pd.concat([mssql_res_df,sf_res_df],axis=1)
                final_Distinct['Difference(mssql-sf)']=final_Distinct['mssql_DistinctCount']-final_Distinct['sf_DistinctCount']
                final_Distinct['Difference_%']=100*final_Distinct['Difference(mssql-sf)']/final_Distinct['mssql_DistinctCount']
                
                final_Distinct.to_csv('Tab3_'+original_mssql_table+'.csv',index=False)
                print('Distinct Counts Tab (3rd) Created')
            except Exception as err:
                print('Not able to complete tab3 due to: ',str(err))    
            ########################## TAB-4 FREQUENCY DISTRIBUTION #################################
            
            #mssql
            
            try:
                df_mssql=pd.DataFrame()
                
                print("****mssql queries")
                for col in mssql_table_cols: 
                    # if col.lower() not in cols_to_ignore:
                        query='select  {colName},cast(count({colName}) as double) as mssql_val_row_count {measureCols} from {tname} group by {colName} order by mssql_val_row_count desc limit 1000 '.format(colName=col,tname=mssql_table,measureCols=measure_cols_mssql)
                        
                        #    print("Running query:",query)
                        res_df=self.run_mssql_query(query)
                        res_df.insert(0,'mssql_ColName',col)
                        res_df.columns=['mssql_ColName','mssql_col_value','mssql_val_row_count'] + measure_cols_mssql_names_list
                        df_mssql=df_mssql.append(res_df)
                    
                #sf
                df_sf=pd.DataFrame()
                
                print("****sf queries")
                for col in sf_cols: 
                    # if col.lower() not in  cols_to_ignore:
                        query='select {colName},cast(count({colName}) as double) as sf_val_row_count  {measureCols}   from {tname} group by {colName} order by sf_val_row_count desc limit 1000'.format(colName=col,tname=sf_table,measureCols=measure_cols_sf)
                        
                #         print(query)
                        #    print("Running query:",query)
                        res_df=self.run_sf_query(query)
                        res_df.insert(0,'sf_ColName',col)
                        res_df.columns=['sf_ColName','sf_col_value','sf_val_row_count'] + measure_cols_sf_names_list
                        df_sf=df_sf.append(res_df)
                try:
                    df_mssql=df_mssql.apply(pd.to_numeric, errors='coerce').fillna(df_mssql)
                    df_sf=df_sf.apply(pd.to_numeric, errors='coerce').fillna(df_sf)
                except:pass
                
                
                
                
                left_indexes=['mssql_ColName','mssql_col_value']
                right_indexes=['sf_ColName','sf_col_value']
                
                for  col in left_indexes:
                            df_mssql[col] = df_mssql[col].astype(str)
                            

                for  col in right_indexes:
                            df_sf[col] = df_sf[col].astype(str)
                            
                df_mssql=df_mssql.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                df_sf=df_sf.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                #making FreqDistri
                
                print('Creating final freq distribution')
                finalFD=df_mssql.merge(df_sf,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_mssql','_sf'])
                
                #finalFD=pd.concat([df_mssql,df_sf],axis=1)
                print('Getting differences!')
                finalFD['mssql_val_row_count'].replace(np.NaN,0)
                finalFD['sf_val_row_count'].replace(np.NaN,0)
                finalFD['Difference(mssql-sf)']=finalFD['mssql_val_row_count']-finalFD['sf_val_row_count']
                finalFD['Difference_%']=100*finalFD['Difference(mssql-sf)']/finalFD['mssql_val_row_count']
                
                common_cols_mssql_sf=list()
                diff_cols=list()
                diff_percent_cols=list()
                diff_result_cols=list()
                mssql_cols=list()
                sf_cols=list()

                # finalFD.to_csv('abcd.csv')
                for col in measure_cols_sf_names_list:
                    for col_mssql in measure_cols_mssql_names_list:
                        if col[0:-2]==col_mssql[0:-2]:
                        
                            finalFD['Diff(mssql-sf): '+col]=finalFD[col_mssql]-finalFD[col]
                            finalFD['Diff %: '+col[0:-2]]=100*finalFD['Diff(mssql-sf): '+col]/finalFD[col_mssql]
                            finalFD['Result Diff: '+col[0:-2]]=(finalFD['Diff %: '+col[0:-2]]==0)
                            common_cols_mssql_sf.append(col[0:-2])
                            diff_cols.append('Diff(mssql-sf): '+col)
                            diff_percent_cols.append('Diff %: '+col[0:-2])
                            diff_result_cols.append('Result Diff: '+col[0:-2])
                            mssql_cols.append(col_mssql)
                            sf_cols.append(col)

            

                
                #finalFD=pd.concat([df_TD,df_SF],axis=1)

                main_cols=['mssql_ColName','mssql_col_value','mssql_val_row_count','sf_ColName','sf_col_value','sf_val_row_count','Difference(mssql-sf)','Difference_%']
                
                finalcols=main_cols

                for i in range(len(common_cols_mssql_sf)):
                    finalcols.append(mssql_cols[i])
                    finalcols.append(sf_cols[i])
                    finalcols.append(diff_cols[i])
                    finalcols.append(diff_percent_cols[i])
                    finalcols.append(diff_result_cols[i])




                remaining_uncommon_cols=list()
                for col in finalFD.columns:
                    if col not in finalcols:
                        remaining_uncommon_cols.append(col)

                finalcols=finalcols+remaining_uncommon_cols

                finalFD=finalFD[finalcols]
                
                



                tableName=mssql_table.split('.')[1:][-1]
                finalFD.to_csv('Tab4_'+original_mssql_table+'.csv',index=False)
                print('Frequency Distribution Tab (4th) created')
            except Exception as err:
                print('Not able to complete tab4 due to: ',str(err))
            self.writeIntoExcel_stats_sf_to_sf(tableName,tab1_df,row_counts,final_Distinct,finalFD)
            # try:
            #     os.remove('Tab1_'+mssql_table+'.csv')
            #     os.remove('Tab2_'+mssql_table+'.csv')
            #     os.remove('Tab3_'+mssql_table+'.csv')
            #     os.remove('Tab4_'+mssql_table+'.csv')
            # except:
            #     print("Could not delete intermediate tab csv files as file either doesn't exist or it's opened")

            print('Excel created!')
        except Exception as err:
            raise Exception(str(err))


    def DownloadAndGetCount_sf_to_sf(self,SF1_table,SF1_condition,SF2_table,SF2_condition,cols_to_sum):
        ########################## BASIC INFO #################################
       
           #SF1_table=row[1]
           #SF1_condition=row[2]
        path=self.tableStatsResultDownloadLocationEntry_sf_to_sf.get()
        os.chdir(path)
        original_SF1_table=SF1_table
        original_SF2_table=SF2_table
        try:
            if (str(SF1_condition)!='nan'):
                        SF1_table +=' where '+SF1_condition
            
            #SF2_table=row[3]
            #SF2_condition=row[4]
            
            if (str(SF2_condition)!='nan'):
                        SF2_table +=' where '+SF2_condition
            
                
             
            SF1_table_cols=self.getCols_SF(SF1_table)
            SF2_cols=self.getCols_SF(SF2_table)
            
            print('***Basic info checked for:',SF2_table)
            
            ########################## TAB-1 SumOfValues #################################
            try:
                measure_cols_SF1=''
                measure_cols_SF1_names_list=list()

                measure_cols_SF2=''
                measure_cols_SF2_names_list=list()

                SF1_checksum=list()
                for col in SF1_table_cols:
                    try:
                        t_q='select cast(sum({colName}) as double )as c from {tn}'.format(colName=col,tn=SF1_table)
                        #    print("Running query: ",t_q)
                        res=self.run_sf_query(t_q)
                        SF1_checksum.append(res.iloc[:,0][0])
                        # measure_cols_SF1 +=', cast(sum({colName}) as double) as SF1_Sum_{colName} '.format(colName=col)
                        # measure_cols_SF1_names_list.append(str(col).lower()+"_SF1")
                    except:
                        SF1_checksum.append(np.nan)

                for col in cols_to_sum:
                    measure_cols_SF1 +=', cast(sum({colName}) as double) as SF1_Sum_{colName} '.format(colName=col)
                    measure_cols_SF1_names_list.append(str(col).lower()+"_SF1")

                
                SF2_checksum=list()
                for col in SF2_cols:
                    try:
                        s_q='select cast( sum({colName}) as double) as c from {tn}'.format(colName=col,tn=SF2_table)
                        #    print("Running query: ",s_q)
                        res=self.run_sf_query(s_q)
                        SF2_checksum.append(res.iloc[:,0][0])
                        # measure_cols_SF2 +=', cast(sum({colName}) as double) as SF2_Sum_{colName} '.format(colName=col)
                        # measure_cols_SF2_names_list.append(str(col).lower()+"_SF2")
                    except:
                        SF2_checksum.append(np.nan)

                for col in cols_to_sum:
                    measure_cols_SF2 +=', cast(sum({colName}) as double) as SF2_Sum_{colName} '.format(colName=col)
                    measure_cols_SF2_names_list.append(str(col).lower()+"_SF2")

                # datatype_query='desc table '+original_SF1_table
                db_name,schema_name,table_name=original_SF1_table.split('.')
                
                clustering_key_sf1="SELECT  CLUSTERING_KEY from {db_name}.INFORMATION_SCHEMA.tables where table_name='{tn}' and table_schema='{ts}' ".format(db_name=db_name.strip().upper(),tn=table_name.strip().upper(),ts=schema_name.strip().upper())
                clustering_key_sf1_df=self.run_sf_query(clustering_key_sf1)
                try:
                    indexed_cols=[s.strip() for s in clustering_key_sf1_df['CLUSTERING_KEY'][0].replace(')','').replace('LINEAR(','').split(',')]
                except:indexed_cols=list()

                indexed_cols_yes_no=list()
                for col in SF1_table_cols:
                    if col in indexed_cols:
                        indexed_cols_yes_no.append('Yes')
                    else: indexed_cols_yes_no.append('No')
                
                SF1_indexed_cols_df=pd.DataFrame({'Col_Name':SF1_table_cols,'SF_Base_Indexed':indexed_cols_yes_no})

                col_info_query="SELECT  COLUMN_NAME as SF_Base_ColName ,CHARACTER_MAXIMUM_LENGTH as SF_Base_ColumnLength  from {db_name}.INFORMATION_SCHEMA.columns where table_name='{tn}' and table_schema='{ts}' ".format(db_name=db_name.strip().upper(),tn=table_name.strip().upper(),ts=schema_name.strip().upper())
                
                SF1_col_info_df=self.run_sf_query(col_info_query)
                SF1_col_info_df.columns=['SF_Base_ColName','SF_Base_ColumnLength']
                
                datatype_query='desc table '+ original_SF1_table
                sf_datatype_df=self.run_sf_query(datatype_query)
                sf_datatype_df=sf_datatype_df[['name','type']]        
                sf_datatype_df.columns=['SF_Base_ColName','SF_Base_Datatype']
                
                df1_SF1=pd.concat([pd.DataFrame({'SF_Base_ColName':SF1_table_cols}),pd.DataFrame({'SF_Base_CheckSum':SF1_checksum})],axis=1)
                df1_SF1=df1_SF1.merge(SF1_col_info_df,left_on='SF_Base_ColName',right_on='SF_Base_ColName',suffixes=['_Base','_Release'])
                df1_SF1=df1_SF1.merge(SF1_indexed_cols_df,left_on='SF_Base_ColName',right_on='Col_Name',suffixes=['_Base','_Release'])
                df1_SF1=df1_SF1.merge(sf_datatype_df,left_on='SF_Base_ColName',right_on='SF_Base_ColName',suffixes=['_Base','_Release'])

                db_name,schema_name,table_name=original_SF2_table.split('.')
                
                clustering_key_sf2="SELECT  CLUSTERING_KEY as SF_Release_Indexed from {db_name}.INFORMATION_SCHEMA.tables where table_name='{tn}' and table_schema='{ts}' ".format(db_name=db_name.strip().upper(),tn=table_name.strip().upper(),ts=schema_name.strip().upper())
                clustering_key_sf2_df=self.run_sf_query(clustering_key_sf2)
                try:
                    indexed_cols=[s.strip() for s in clustering_key_sf2_df['CLUSTERING_KEY'][0].replace(')','').replace('LINEAR(','').split(',')]
                except:indexed_cols=list()

                indexed_cols_yes_no=list()
                for col in SF2_cols:
                    if col in indexed_cols:
                        indexed_cols_yes_no.append('Yes')
                    else: indexed_cols_yes_no.append('No')
                
                SF2_indexed_cols_df=pd.DataFrame({'Col_Name':SF2_cols,'SF_Release_Indexed':indexed_cols_yes_no})

                col_info_query="SELECT  COLUMN_NAME as SF_Release_ColName,CHARACTER_MAXIMUM_LENGTH as SF_Release_ColumnLength from {db_name}.INFORMATION_SCHEMA.columns where table_name='{tn}' and table_schema='{ts}' ".format(db_name=db_name.strip().upper(),tn=table_name.strip().upper(),ts=schema_name.strip().upper())
                SF2_col_info_df=self.run_sf_query(col_info_query)
                SF2_col_info_df.columns=['SF_Release_ColName','SF_Release_ColumnLength']

                datatype_query='desc table '+ original_SF2_table
                sf_datatype_df=self.run_sf_query(datatype_query)
                sf_datatype_df=sf_datatype_df[['name','type']]        
                sf_datatype_df.columns=['SF_Release_ColName','SF_Release_Datatype']

                df2_SF2=pd.concat([ pd.DataFrame({'SF_Release_ColName':SF2_cols}),pd.DataFrame({'SF_Release_CheckSum':SF2_checksum})],axis=1)
                df2_SF2=df2_SF2.merge(SF2_col_info_df,left_on='SF_Release_ColName',right_on='SF_Release_ColName',suffixes=['_Base','_Release'])
                df2_SF2=df2_SF2.merge(SF2_indexed_cols_df,left_on='SF_Release_ColName',right_on='Col_Name',suffixes=['_Base','_Release'])
                df2_SF2=df2_SF2.merge(sf_datatype_df,left_on='SF_Release_ColName',right_on='SF_Release_ColName',suffixes=['_Base','_Release'])

                left_indexes=['SF_Base_ColName']
                right_indexes=['SF_Release_ColName']
                tab1_df=df1_SF1.merge(df2_SF2,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_Base','_Release'])
                
                
                finalCols_tab1=['SF_Base_ColName','SF_Base_Datatype','SF_Base_ColumnLength','SF_Base_CheckSum','SF_Base_Indexed',
                                'SF_Release_ColName','SF_Release_Datatype','SF_Release_ColumnLength','SF_Release_CheckSum','SF_Release_Indexed']
                tab1_df=tab1_df[finalCols_tab1]


                tab1_df['ColNameCheck_Result']=tab1_df['SF_Base_ColName']==tab1_df['SF_Release_ColName']
                tab1_df['Datatype_Result']=tab1_df['SF_Base_Datatype']==tab1_df['SF_Release_Datatype']
                tab1_df['ColumnLength_Result']=tab1_df['SF_Base_ColumnLength'].eq(tab1_df['SF_Release_ColumnLength'])
                tab1_df['Checksum_Result']=tab1_df['SF_Base_CheckSum'].eq(tab1_df['SF_Release_CheckSum'])
                tab1_df['Indexed_Result']=tab1_df['SF_Base_Indexed']==tab1_df['SF_Release_Indexed']
                
                
                

                tab1_df.to_csv('Tab1_'+original_SF1_table+'.csv',index=False)
                print('SumOfValues Tab (1st) created')

            except Exception as err:
                print('Not able to complete tab1 due to: ',str(err))
            ########################## TAB-2 Total Row counts #################################

            try:
                tq='select cast(count(*) as double) as SF1_ROW_COUNT from  '+ SF1_table
                #    print("Running query: ",tq)
                SF1_count=self.run_sf_query(tq)
                
                sq='select cast(count(*) as double) as SF2_ROW_COUNT from  '+ SF2_table
                #    print("Running query: ",sq)
                SF2_count=self.run_sf_query(sq)
                
                
                #result
                
                row_counts=pd.concat([SF1_count,SF2_count],axis=1)
                row_counts['Difference(SF1-SF2)']=row_counts['SF1_ROW_COUNT']-row_counts['SF2_ROW_COUNT']
                row_counts['Difference_%']=100*row_counts['Difference(SF1-SF2)']/row_counts['SF1_ROW_COUNT']
                
                row_counts.to_csv('Tab2_'+original_SF1_table+'.csv',index=False)
                print('Row Count Tab (2nd) created')
            except Exception as err:
                print('Not able to complete tab2 due to: ',str(err))
            ########################## TAB-3 DISTINCT COUNT #################################
        
            try:
                #SF1
                SF1_distinct_query_list=list()
                for col in SF1_table_cols:
                    SF1_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                
                SF1_distinct_query=",".join(SF1_distinct_query_list)
                
                finalDistinct_query='select ' + SF1_distinct_query + " from "+ SF1_table
                #    print("Running query: ",finalDistinct_query)
                SF1_res_df=self.run_sf_query(finalDistinct_query)
                SF1_res_df=SF1_res_df.transpose().reset_index()
                SF1_res_df.columns=['SF1_ColName','SF1_DistinctCount']
                
                #SF2
                SF2_distinct_query_list=list()
                for col in SF2_cols:
                    SF2_distinct_query_list.append("cast(count(distinct({colName})) as double) as {colName}".format(colName=col))
                
                SF2_distinct_query=",".join(SF2_distinct_query_list)
                
                finalDistinct_query='select ' + SF2_distinct_query + " from "+ SF2_table

                #    print("Running query: ",finalDistinct_query)

                SF2_res_df=self.run_sf_query(finalDistinct_query)
                SF2_res_df=SF2_res_df.transpose().reset_index()
                SF2_res_df.columns=['SF2_ColName','SF2_DistinctCount']
                
                
                
                
                #making finalDistinctRes
                left_indexes=['SF1_ColName']
                right_indexes=['SF2_ColName']
                
                final_Distinct=SF1_res_df.merge(SF2_res_df,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_SF1','_SF2'])
                
                #final_Distinct=pd.concat([SF1_res_df,SF2_res_df],axis=1)
                final_Distinct['Difference(SF1-SF2)']=final_Distinct['SF1_DistinctCount']-final_Distinct['SF2_DistinctCount']
                final_Distinct['Difference_%']=100*final_Distinct['Difference(SF1-SF2)']/final_Distinct['SF1_DistinctCount']
                
                final_Distinct.to_csv('Tab3_'+original_SF1_table+'.csv',index=False)
                print('Distinct Counts Tab (3rd) Created')
            except Exception as err:
                print('Not able to complete tab3 due to: ',str(err))    
            ########################## TAB-4 FREQUENCY DISTRIBUTION #################################
            
            #SF1
            
            try:
                df_SF1=pd.DataFrame()
                
                print("****SF1 queries")
                for col in SF1_table_cols: 
                    # if col.lower() not in cols_to_ignore:
                        query='select  {colName},cast(count({colName}) as double) as SF1_val_row_count {measureCols} from {tname} group by {colName} order by SF1_val_row_count desc limit 1000 '.format(colName=col,tname=SF1_table.strip().upper(),measureCols=measure_cols_SF1)
                        
                        #    print("Running query:",query)
                        res_df=self.run_sf_query(query)
                        res_df.insert(0,'SF1_ColName',col)
                        res_df.columns=['SF1_ColName','SF1_col_value','SF1_val_row_count'] + measure_cols_SF1_names_list
                        df_SF1=df_SF1.append(res_df)
                    
                #SF2
                df_SF2=pd.DataFrame()
                
                print("****SF2 queries")
                for col in SF2_cols: 
                    # if col.lower() not in  cols_to_ignore:
                        query='select {colName},cast(count({colName}) as double) as SF2_val_row_count  {measureCols}   from {tname} group by {colName} order by SF2_val_row_count desc limit 1000'.format(colName=col,tname=SF2_table.strip().upper(),measureCols=measure_cols_SF2)
                        
                #         print(query)
                        #    print("Running query:",query)
                        res_df=self.run_sf_query(query)
                        res_df.insert(0,'SF2_ColName',col)
                        res_df.columns=['SF2_ColName','SF2_col_value','SF2_val_row_count'] + measure_cols_SF2_names_list
                        df_SF2=df_SF2.append(res_df)
                try:
                    df_SF1=df_SF1.apply(pd.to_numeric, errors='coerce').fillna(df_SF1)
                    df_SF2=df_SF2.apply(pd.to_numeric, errors='coerce').fillna(df_SF2)
                except:pass
                
                
                
                
                left_indexes=['SF1_ColName','SF1_col_value']
                right_indexes=['SF2_ColName','SF2_col_value']
                
                for  col in left_indexes:
                            df_SF1[col] = df_SF1[col].astype(str)
                            

                for  col in right_indexes:
                            df_SF2[col] = df_SF2[col].astype(str)
                            
                df_SF1=df_SF1.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                df_SF2=df_SF2.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                #making FreqDistri
                
                print('Creating final freq distribution')
                finalFD=df_SF1.merge(df_SF2,left_on=left_indexes,right_on=right_indexes,how='outer',suffixes=['_SF1','_SF2'])
                
                #finalFD=pd.concat([df_SF1,df_SF2],axis=1)
                print('Getting differences!')
                finalFD['SF1_val_row_count'].replace(np.NaN,0)
                finalFD['SF2_val_row_count'].replace(np.NaN,0)
                finalFD['Difference(SF1-SF2)']=finalFD['SF1_val_row_count']-finalFD['SF2_val_row_count']
                finalFD['Difference_%']=100*finalFD['Difference(SF1-SF2)']/finalFD['SF1_val_row_count']
                
                common_cols_SF1_SF2=list()
                diff_cols=list()
                diff_percent_cols=list()
                diff_result_cols=list()
                sf1_cols=list()
                sf2_cols=list()

                # finalFD.to_csv('abcd.csv')
                for col in measure_cols_SF2_names_list:
                    for col_sf1 in measure_cols_SF1_names_list:
                        if col[0:-2]==col_sf1[0:-2]:
                        
                            finalFD['Diff(SF1-SF2): '+col]=finalFD[col_sf1]-finalFD[col]
                            finalFD['Diff %: '+col[0:-2]]=100*finalFD['Diff(SF1-SF2): '+col]/finalFD[col_sf1]
                            finalFD['Result Diff: '+col[0:-2]]=(finalFD['Diff %: '+col[0:-2]]==0)
                            common_cols_SF1_SF2.append(col[0:-2])
                            diff_cols.append('Diff(SF1-SF2): '+col)
                            diff_percent_cols.append('Diff %: '+col[0:-2])
                            diff_result_cols.append('Result Diff: '+col[0:-2])
                            sf1_cols.append(col_sf1)
                            sf2_cols.append(col)

            

                
                #finalFD=pd.concat([df_TD,df_SF],axis=1)

                main_cols=['SF1_ColName','SF1_col_value','SF1_val_row_count','SF2_ColName','SF2_col_value','SF2_val_row_count','Difference(SF1-SF2)','Difference_%']
                
                finalcols=main_cols

                for i in range(len(common_cols_SF1_SF2)):
                    finalcols.append(sf1_cols[i])
                    finalcols.append(sf2_cols[i])
                    finalcols.append(diff_cols[i])
                    finalcols.append(diff_percent_cols[i])
                    finalcols.append(diff_result_cols[i])




                remaining_uncommon_cols=list()
                for col in finalFD.columns:
                    if col not in finalcols:
                        remaining_uncommon_cols.append(col)

                finalcols=finalcols+remaining_uncommon_cols

                finalFD=finalFD[finalcols]
                
                



                tableName=SF1_table.split('.')[1:][-1]
                finalFD.to_csv('Tab4_'+original_SF1_table+'.csv',index=False)
                print('Frequency Distribution Tab (4th) created')
            except Exception as err:
                print('Not able to complete tab4 due to: ',str(err))
            self.writeIntoExcel_stats_sf_to_sf(tableName,tab1_df,row_counts,final_Distinct,finalFD)
            # try:
            #     os.remove('Tab1_'+SF1_table+'.csv')
            #     os.remove('Tab2_'+SF1_table+'.csv')
            #     os.remove('Tab3_'+SF1_table+'.csv')
            #     os.remove('Tab4_'+SF1_table+'.csv')
            # except:
            #     print("Could not delete intermediate tab csv files as file either doesn't exist or it's opened")

            print('Excel created!')
        except Exception as err:
            raise Exception(str(err))

    

    def writeIntoExcel_stats_hive_sf(self,tableName,tab1_df,row_counts,final_Distinct,finalFD):
        
        path=self.tableStatsResultDownloadLocationEntry_hive_to_sf.get()
        os.chdir(path)
        #df_path=(self.TableStatsFile.get()).strip()
        #path=df_path.split('/')[0:-1]
       
        #fileName= "\\".join(path)+'\\'+tableName+'.xlsx'
        tableName=tableName[:15]
        fileName=tableName+'.xlsx'
        with pd.ExcelWriter(fileName) as writer:
                    tab1_df.to_excel(writer,sheet_name='ColNames',index = False,header=True) 
                    row_counts.to_excel(writer,sheet_name='Total Row Count',index = False,header=True) 
                    final_Distinct.to_excel(writer,sheet_name='Distinct',index = False,header=True) 
                    finalFD.to_excel(writer,sheet_name='FreqDistribution',index = False,header=True) 
        
        # messagebox.showinfo('Done','Result excel created in ' + str(path) )
        wb = openpyxl.load_workbook(fileName)
        ws = wb['FreqDistribution'] #Name of the working sheet

        max_col = ws.max_column
        for i in range(9, max_col,5):

                gen_random_hexDec = ([random.choice('ABCDEF0123456789') for i in range(6)])
                hexadecimal = ''.join(gen_random_hexDec)
                fill_cell = PatternFill(patternType='solid', fgColor= hexadecimal)
                ws.cell(column=i, row=1).fill = fill_cell
                ws.cell(column=i+1, row=1).fill = fill_cell
                ws.cell(column=i+2, row=1).fill = fill_cell
                ws.cell(column=i+3, row=1).fill = fill_cell
                ws.cell(column=i+4, row=1).fill = fill_cell
            
        wb.save(fileName)


    def writeIntoExcel_stats_td_sf(self,tableName,tab1_df,row_counts,final_Distinct,finalFD):
        
        path=self.tableStatsResultDownloadLocationEntry.get()
        os.chdir(path)
        #df_path=(self.TableStatsFile.get()).strip()
        #path=df_path.split('/')[0:-1]
       
        #fileName= "\\".join(path)+'\\'+tableName+'.xlsx'
        tableName=tableName
        fileName=tableName+'.xlsx'
        with pd.ExcelWriter(fileName) as writer:
                    tab1_df.to_excel(writer,sheet_name='ColNames',index = False,header=True) 
                    row_counts.to_excel(writer,sheet_name='Total Row Count',index = False,header=True) 
                    final_Distinct.to_excel(writer,sheet_name='Distinct',index = False,header=True) 
                    finalFD.to_excel(writer,sheet_name='FreqDistribution',index = False,header=True) 
        
        # messagebox.showinfo('Done','Result excel created in ' + str(path) )
        wb = openpyxl.load_workbook(fileName)
        ws = wb['FreqDistribution'] #Name of the working sheet

        max_col = ws.max_column
        for i in range(9, max_col,5):

                gen_random_hexDec = ([random.choice('ABCDEF0123456789') for i in range(6)])
                hexadecimal = ''.join(gen_random_hexDec)
                fill_cell = PatternFill(patternType='solid', fgColor= hexadecimal)
                ws.cell(column=i, row=1).fill = fill_cell
                ws.cell(column=i+1, row=1).fill = fill_cell
                ws.cell(column=i+2, row=1).fill = fill_cell
                ws.cell(column=i+3, row=1).fill = fill_cell
                ws.cell(column=i+4, row=1).fill = fill_cell
            
        wb.save(fileName)


          
    def writeIntoExcel_stats_sf_to_sf(self,tableName,tab1_df,row_counts,final_Distinct,finalFD):
        
        path=self.tableStatsResultDownloadLocationEntry_sf_to_sf.get()
        os.chdir(path)
        #df_path=(self.TableStatsFile.get()).strip()
        #path=df_path.split('/')[0:-1]
       
        #fileName= "\\".join(path)+'\\'+tableName+'.xlsx'
        fileName=tableName+'.xlsx'
        with pd.ExcelWriter(fileName) as writer:
                    tab1_df.to_excel(writer,sheet_name='ColNames',index = False,header=True) 
                    row_counts.to_excel(writer,sheet_name='Total Row Count',index = False,header=True) 
                    final_Distinct.to_excel(writer,sheet_name='Distinct',index = False,header=True) 
                    finalFD.to_excel(writer,sheet_name='FreqDistribution',index = False,header=True) 
        
        # messagebox.showinfo('Done','Result excel created in ' + str(path) )
         
    
    def writeIntoExcel_stats(self,tableName,tab1_df,row_counts,final_Distinct,finalFD):
        
        path=self.tableStatsResultDownloadLocationEntry_TD.get()
        os.chdir(path)
        #df_path=(self.TableStatsFile.get()).strip()
        #path=df_path.split('/')[0:-1]
       
        #fileName= "\\".join(path)+'\\'+tableName+'.xlsx'
        fileName=tableName+'.xlsx'
        with pd.ExcelWriter(fileName) as writer:
                    tab1_df.to_excel(writer,sheet_name='ColNames',index = False,header=True) 
                    row_counts.to_excel(writer,sheet_name='Total Row Count',index = False,header=True) 
                    final_Distinct.to_excel(writer,sheet_name='Distinct',index = False,header=True) 
                    finalFD.to_excel(writer,sheet_name='FreqDistribution',index = False,header=True) 
        
        # messagebox.showinfo('Done','Result excel created in ' + str(path) )
                    
                    


    def getSourceFilesFromPrefix(self):
        try:
            source_prefix=self.batch_releasePrefix_entry.get()
            source_base_location=self.batch_Source_entry.get()
            all_ff=os.listdir(source_base_location)
            source_files_list=list()
            for file in all_ff:
                if file.split('.')[-1] in ['xlsx','csv'] and file.startswith(source_prefix):
                    source_files_list.append(file)

            return source_files_list
        except Exception as err:
            print('Error while reading source files from prefix ',str(err))
    


    def getTargetFilesFromPrefix(self):
        try:
            target_prefix=self.batch_releasePrefix_entry.get()
            target_base_location=self.batch_release_entry.get()
            all_ff=os.listdir(target_base_location)
            target_files_list=list()
            for file in all_ff:
                if file.split('.')[-1] in ['xlsx','csv'] and file.startswith(target_prefix):
                    target_files_list.append(file)

            return target_files_list

        except Exception as err:
            print('Error while reading target files from prefix ',str(err))




    def ddl_grid_forget(self):
        self.DDL_validation_hive_sf.grid_forget()
        self.DDL_validation_ddl_schema_sf_sf.grid_forget()
        self.DDL_validation_ddl_td_sf.grid_forget()
        self.DDL_validation_ddl_mssql_sf.grid_forget()
        self.DDL_validation_ddl_sf_sf.grid_forget()
        self.DDL_validation_ddl_mssql_td.grid_forget()
        self.tokenization_check_frame.grid_forget()
        self.View_Validation_SF.grid_forget()

    def table_stats_grid_forget(self):
        self.frame_stats_td_sf.grid_forget()
        self.frame_stats_sf_sf.grid_forget()
        self.frame_stats_hive_sf_sanity.grid_forget()
        self.frame_stats_sf1_sf2_sanity.grid_forget()
        self.frame_stats_td_sf_sanity.grid_forget()
        self.frame_stats_mssql_sf.grid_forget()
        self.frame_stats_hive_sf.grid_forget()
        self.frame_stats_mssql_sf_sanity.grid_forget()
        
    
    def selectDDLPair(self,eventObj=np.NAN):
        #  ['Hive-SF DDL check','SF-SF Schema & DDL check','TD-SF DDL check','MSSQL-SF DDL check','MSSQL-TD DDL check']
        self.ddl_grid_forget()
        if (self.selectDDL_validationDropDown.get()=='Hive-SF DDL check'):
            self.DDL_validation_hive_sf.grid(row=3,column=0)
        elif (self.selectDDL_validationDropDown.get()=='SF-SF Schema & DDL check'):
            self.DDL_validation_ddl_schema_sf_sf.grid(row=3,column=0)
        elif (self.selectDDL_validationDropDown.get()=='SF-SF DDL check'):
            self.DDL_validation_ddl_sf_sf.grid(row=3,column=0)
        elif (self.selectDDL_validationDropDown.get()=='TD-SF DDL check'):
            self.DDL_validation_ddl_td_sf.grid(row=3,column=0)
        elif (self.selectDDL_validationDropDown.get()=='MSSQL-SF DDL check'):
            self.DDL_validation_ddl_mssql_sf.grid(row=3,column=0)
        elif (self.selectDDL_validationDropDown.get()=='MSSQL-TD DDL check'):
            self.DDL_validation_ddl_mssql_td.grid(row=3,column=0)
        elif (self.selectDDL_validationDropDown.get()=='Tokenization Check'):
            self.tokenization_check_frame.grid(row=3,column=0)
        elif (self.selectDDL_validationDropDown.get()=='TD-SF casted col validation'):
            self.td_sf_casted_cols_frame.grid(row=3,column=0)
        elif (self.selectDDL_validationDropDown.get()=='SF View Validation'):
            self.View_Validation_SF.grid(row=3,column=0)

    
    def selectTableStatsPair(self,eventObj=np.NAN):
        #  ['Hive-SF DDL check','SF-SF Schema & DDL check','TD-SF DDL check','MSSQL-SF DDL check','MSSQL-TD DDL check']
        self.table_stats_grid_forget()
        if (self.selectframe_statsDropDown.get()=='TD-SF stats'):
            self.frame_stats_td_sf.grid(row=3,column=0)
        elif (self.selectframe_statsDropDown.get()=='SF-SF stats'):
            self.frame_stats_sf_sf.grid(row=3,column=0)
        elif (self.selectframe_statsDropDown.get()=='Hive-SF PostProd Check (DDL, Rowcount,SumOfValues,Rowcount)'):
            self.frame_stats_hive_sf_sanity.grid(row=3,column=0)
        elif (self.selectframe_statsDropDown.get()=='MSSQL-SF PostProd Check (DDL, Rowcount,SumOfValues,Rowcount)'):
            self.frame_stats_mssql_sf_sanity.grid(row=3,column=0)
        elif (self.selectframe_statsDropDown.get()=='SF-SF PostProd Check (DDL, Rowcount,SumOfValues,Rowcount)'):
            self.frame_stats_sf1_sf2_sanity.grid(row=3,column=0)
        elif (self.selectframe_statsDropDown.get()=='TD-SF PostProd Check (DDL, Rowcount,SumOfValues,Rowcount)'):
            self.frame_stats_td_sf_sanity.grid(row=3,column=0)
        elif (self.selectframe_statsDropDown.get()=='MSSQL-SF stats'):
            self.frame_stats_mssql_sf.grid(row=3,column=0)
        elif (self.selectframe_statsDropDown.get()=='Hive-SF stats'):
            self.frame_stats_hive_sf.grid(row=3,column=0)





    def selectTestMethod(self,eventObj=np.NAN):
        # print("inside select test method",self.selected.get())
        self.forget_frames()
        self.ddl_grid_forget()
        self.table_stats_grid_forget()
        
        if(self.selected.get()=='Batch Comparison'):
            
            self.frameBatch.grid(row=3,column=1)
           
        elif (self.selected.get()=='Individual Comparison'):
            self.frameIndividual.grid(row=1)
                 
            
        
        elif (self.selected.get()=='TableStats'):
            self.frame_stats.grid(row=1)
            
        elif (self.selected.get()=='Utilities'):
            self.frameUtilities.grid(row=1)
        elif (self.selected.get()=='PdfCompare'):
            self.framePdfCompare.grid(row=1)
        elif (self.selected.get()=='DDL Validation'):
            self.DDL_validation.grid(row=1)
        elif (self.selected.get()=='Sqoop Batch Execution'):
            self.Sqoop_batch_exection.grid(row=1)
            
        if (self.selected.get()=='DB Connect'):
            self.frame_database_connect.grid(row=3)
            
            if (self.selectDataBaseDropDown.get()=='Teradata'):
                self.td_frame.grid(row=3,column=0)
            elif (self.selectDataBaseDropDown.get()=='Snowflake'):
                self.sf_frame.grid(row=3,column=0)
            
            elif (self.selectDataBaseDropDown.get()=='Snowflake_Prod'):
                self.sf_prod_frame.grid(row=3,column=0)

                
            elif (self.selectDataBaseDropDown.get()=='Oracle'):
                
                self.Oracle_frame.grid(row=3,column=0)
            elif (self.selectDataBaseDropDown.get()=='Hive'):
                
                self.Hive_frame.grid(row=3,column=0)
            elif (self.selectDataBaseDropDown.get()=='SSH'):
                
                self.SSH_frame.grid(row=3,column=0)
            elif (self.selectDataBaseDropDown.get()=='MySQL WorkBench'):
                
                self.MySQLWorkBench_frame.grid(row=3,column=0)

            elif (self.selectDataBaseDropDown.get()=='MS-SQL Server'):
                self.mssql_frame.grid(row=3,column=0)
                # PostgresSQL
            elif (self.selectDataBaseDropDown.get()=='PostgresSQL'):
                self.PostgresSQL_frame.grid(row=3,column=0)


    

     # tk.Label(frame0, text="Batch Comparison", font=("Arial", self.fontSize+5)).grid(row=1,column=1,padx=(0,0),sticky='w')
    # r1.config(activebackground="#0C7A79")
    # r1.config()
    def forget_frames(self):
                self.Sqoop_batch_exection.grid_forget()
                self.sf_prod_frame.grid_forget()
                self.td_frame.grid_forget()
                self.frameBatch.grid_forget()
                self.frame_database_connect.grid_forget()
                self.td_sf_casted_cols_frame.grid_forget()
                try:
                    self.resetButton.grid_forget()
                except:pass

                try:
                    self.frame_stats.grid_forget()
                except:pass
                self.frameIndividual.grid_forget()
                self.frameUtilities.grid_forget()
                self.framePdfCompare.grid_forget()
                self.sf_frame.grid_forget()
                self.Oracle_frame.grid_forget()
                self.Hive_frame.grid_forget() 
                self.SSH_frame.grid_forget() 
                self.MySQLWorkBench_frame.grid_forget()     
                self.DDL_validation.grid_forget()
                self.mssql_frame.grid_forget()
                self.PostgresSQL_frame.grid_forget()


    def convert_batch_Excel_to_csv(self):
        # self.schemaAndDDLverification_sf()
        # self.establish_connection_SF_prod()
        # query="select * from LandingZone.information_schema.columns where table_name = 'LZ_HIP_ContactHistory' and table_schema='dbo' order by ordinal_position;"
        # df=self.run_mssql_sqoop_query(query)
       # print(df.head())
        # return
        try:
            path=self.batch_ExcelFolder_entry.get()
            files=os.listdir(path)
           
            try:
                os.mkdir(path+'/CSV_files')
            except: pass

            os.chdir(path)
            files_xls=[f for f in files if f[-4:]=='xlsx']
            for f in files_xls:
                df=pd.read_excel(f)
                df.reset_index(drop=True,inplace=True)
                df.to_csv('CSV_files/'+f.split('.xlsx')[0]+'.csv',index=False)
            messagebox.showinfo('Task Complete','Convertion to csv done')
        except Exception as err:
            raise Exception('Error while getting xlsx files: ',str(err))


    def select_batch_Excel_folder_location(self):
        global batch_Excel_location_adrs
        self.batch_Excel_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select Excel directory')
        self.batch_ExcelFolder_entry.delete(0,END)
        self.batch_ExcelFolder_entry.insert(0,self.batch_Excel_location_adrs)
        print(str(datetime.datetime.now()),": Batch File Excel Location selected: ",self.batch_Excel_location_adrs)
                
    def select_pdf_comparison_result_folder_location(self):
        
        pdfComparisonResultFolder_entry=fd.askdirectory(parent=self,initialdir="/",title='Please select Excel directory')
        self.pdfComparisonResultFolder_entry.delete(0,END)
        self.pdfComparisonResultFolder_entry.insert(0,pdfComparisonResultFolder_entry)
        print(str(datetime.datetime.now()),": Batch File Excel Location selected: ",self.batch_Excel_location_adrs)

    def select_pdf_comparison_Source(self):
        
        pdfComparisonFolder_entry=fd.askopenfilename(title='Select Source PDF file') 
        self.pdfComparisonSource_entry.delete(0,END)
        self.pdfComparisonSource_entry.insert(0,pdfComparisonFolder_entry)
        print(str(datetime.datetime.now()),": PDF comparison result Location selected: ",self.batch_Excel_location_adrs)

    def select_pdf_comparison_Target(self):
        
        pdfComparisonFolder_entry=fd.askopenfilename(title='Select Target PDF file') 
        self.pdfComparisonTarget_entry.delete(0,END)
        self.pdfComparisonTarget_entry.insert(0,pdfComparisonFolder_entry)
        print(str(datetime.datetime.now()),": PDF comparison result Location selected: ",self.batch_Excel_location_adrs)
                

        
    def select_batch_Source_folder_location(self):
        global batch_Source_location_adrs
        self.batch_Source_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select Source File')
        self.batch_Source_entry.delete(0,END)
        self.batch_Source_entry.insert(0,self.batch_Source_location_adrs)
        print(str(datetime.datetime.now()),": Batch File Source Location selected: ",self.batch_Source_location_adrs)

    
        
    def select_batch_release_folder_location(self):
        global batch_release_location_adrs
        batch_release_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select Target directory')
        self.batch_release_entry.delete(0,END)
        self.batch_release_entry.insert(0,batch_release_location_adrs)
        print(str(datetime.datetime.now()),": Batch File Target Location selected: ",batch_release_location_adrs)

    



    #select result_storage_path
    def select_batch_Result_folder_location(self):
        global batch_Result_location_adrs
        self.batch_Result_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select Result directory')
        self.batch_Result_entry.delete(0,END)
        self.batch_Result_entry.insert(0,self.batch_Result_location_adrs)
        print(str(datetime.datetime.now()),": Batch Result Folder selected: ",self.batch_release_location_adrs)

    
    # frame5=Frame(self.frameIndividual, highlightcolor="yellow",bg="#0C7A79", borderwidth=10, relief=RIDGE,padx=1,pady=1)
    # frame5.grid(row=2,column=0,sticky='e')

    def resetEverything(self):
        self.hl.delete(0,END)
        self.sl.delete(0,END)
        self.bl.delete(0,END)
        self.resultFileName.delete(0,END)
        self.pk.delete(0,END)
        self.skipCols.delete(0,END)
        self.skipColsDuringFullCompare.delete(0,END)
        self.batchExcelPath.delete(0,END)
        print(str(datetime.datetime.now()),": All fields resetted")
        




    
    # //add timestamp cols
    def add_primary_key_cols(self):
        newcol=self.commonCols_combo.get()
        oldCols=(self.pk.get()).strip()
        if len(oldCols)==0:
            combinedCol=newcol
        else:
            combinedCol=oldCols +","+newcol
        self.pk.delete(0,END)
        self.pk.insert(0,combinedCol)


    


    def add_all_common_cols(self):

        # for pairCol in str(self.colsMapping.get()).split(','):
        #         try:
        #             Source_col,release_col=pairCol.split('->')
        #             colsDict[(str(Source_col).lower().strip()).strip()]=(str(release_col).lower().strip()).strip()
        #         except:
        #             pass

        # all_cols=colsDict.keys()
        all_cols=cols
        # print('666 ',all_cols)
        combinedCol=",".join(all_cols)
        self.pk.delete(0,END)
        self.pk.insert(0,combinedCol)
        
    

    

    def startPoint(self):
        excelFile= fd.askopenfilename(title='Input Excel File')
        self.batchExcelPath.delete(0,END)
        self.batchExcelPath.insert(0,excelFile)
        
    

    def lastStepConsolidation(self):
    
        print(str(datetime.datetime.now()),": Consolidation Started")
        try:
        #     print ('Source File Path:')
            # path = input()
            path=self.batch_Result_entry.get()
            os.chdir(path)
            files = os.listdir(path)
            files_xls = [f for f in files if f[-4:] == 'xlsx']
        #     print(files_xls)

            df_fullcompare = pd.DataFrame()
            df_fullcompare_tableNames=list()
            
            df_columnwise=pd.DataFrame()
            df_columnwise_tableNames=list()

            df_tablefacts=pd.DataFrame()
            df_tablefacts_tableNames=list()

            

            for f in files_xls:
                    print(str(datetime.datetime.now()),": Writing=>",f)
                    
                    nlst=list()
                    nlst.append((f.split(".")[0] + f.split(".")[1]))
                    
                    header_names=['ColumnName', 'Source_Null_Values','Release_Null_Values',
                                'Result_Null','Source_LeadingTrailing', 'Release_LeadingTrailing Space-Target',
                                'Result_LeadingTrailingSpace','Source_UniqueValues','Release_UniqueValues','Result_UniqueValues','SumOfValues-Source','SumOfValues-Target','Result_Checksum',
                                'Source_NotAvailable','Release_NotAvailable','Result_NotAvailable']
                    
                    # df_columnwise_sheet = pd.read_excel(f, 'column-wise', header=None,skiprows=1,names=header_names)
                    df_columnwise_sheet = pd.read_excel(f, 'column-wise')
                    
                    df_columnwise_sheet=df_columnwise_sheet[ (df_columnwise_sheet['Result_Null']=='FAIL' )| (df_columnwise_sheet['Result_LeadingTrailingSpace']=='FAIL') | (df_columnwise_sheet['Release_UniqueValues']=='FAIL') | (df_columnwise_sheet['Result_Checksum']=='FAIL') | (df_columnwise_sheet['Result_NotAvailable']=='FAIL')  ]
                    df_columnwise = df_columnwise.append(df_columnwise_sheet)
                    df_columnwise_tableNames.extend(nlst*(df_columnwise_sheet.shape[0]))


                    df_tablefacts_sheet = pd.read_excel(f, 'Target-table-facts', header=None,skiprows=1)
                    df_tablefacts_sheet=df_tablefacts_sheet[~((df_tablefacts_sheet[1]==0) | (df_tablefacts_sheet[1]=='Yes') | (df_tablefacts_sheet[1]=='[]') | (df_tablefacts_sheet[0]=='Common Column Count') | ((df_tablefacts_sheet[0]=='Row Count Result') & ( df_tablefacts_sheet[1]=='Matching')  )  ) | (( df_tablefacts_sheet[0]=='Row Count Result') & (df_tablefacts_sheet[1]==0 )) ]
                    df_tablefacts = df_tablefacts.append(df_tablefacts_sheet) 
                    df_tablefacts_tableNames.extend(nlst*(df_tablefacts_sheet.shape[0]))

                    
                    dataFullComparison = pd.read_excel(f, 'Data-Comparison-Detail', header=None,names=['Primary Key', 'Column','Source-Value','Target-Value'],skiprows=1)
                    df_fullcompare = df_fullcompare.append(dataFullComparison)
                    

                    df_fullcompare_tableNames.extend(nlst*(dataFullComparison.shape[0]))
                    
                    # ColumnName         Source_Null_Values              Release_Null_Values      Result_Null        Source_LeadingTrailing Space    Release_LeadingTrailing Space-Target              Result_LeadingTrailingSpace          'Source_UniqueValues','Release_UniqueValues','Result_UniqueValues'              SumOfValues-Source           SumOfValues-Target            Result_Checksum              Source_NotAvailable      Release_NotAvailable              Result_NotAvailable

                    

                    
                
                
            
            df_tablefacts.columns=['Condition','Result','Remark']
            fileName=path +'\FinalReport'

            try:
                os.mkdir(fileName)
            except Exception as err :
                print(str(datetime.datetime.now()),': Creating report Into ',fileName)

            fileName = fileName + '\ConsolidatedReport.xlsx'
    
            df_fullcompare['Result FileName']=df_fullcompare_tableNames
            
            df_columnwise['Result FileName']=df_columnwise_tableNames
            df_tablefacts['Result FileName']=df_tablefacts_tableNames
            

            with pd.ExcelWriter(fileName) as writer:
                
                df_columnwise.to_excel(writer,sheet_name='Column-Wise',index = False,header=True) 
                df_tablefacts.to_excel(writer,sheet_name='Table-Facts',index = False,header=True) 
                df_fullcompare.to_excel(writer,sheet_name='FullDataComparison',index = False,header=True) 

            print (str(datetime.datetime.now()),': Completed Report is stored in' + path + '\FinalReport\ConsolidatedReport.xlsx')
            self.reset_fields()
            print(str(datetime.datetime.now()),": Results consolidated.")
        except Exception as err:
            print(str(err))
    


    def initializeAllFromExcel(self):
        
        # source_files_list=self.getSourceFilesFromPrefix()
        # target_files_list=self.getTargetFilesFromPrefix()

        # if len(source_files_list)!=len(target_files_list):
        #     raise Exception('Target file list size not equal to Source files list')



        excelPath=self.batchExcelPath.get()

        try:

            df=pd.read_excel(excelPath)
        except Exception as err:
            messagebox.showerror("Error while reading batch excel file : ", str(err))
    # df=df.reset_index()
        # df=df.drop(columns=['index'])
        # print(df) itertuples add a index..need to ignore first col
        #  ['Source Path+FileName0', 'Target Path+ File Name1', 'Result File Path2 ',
        #    'Result File Name3', 'Primary Key4', 'Tokenized5', 'All6', 'Null Check7',
        #    'Blank Check8', 'TimeStamp Check9', 'RowCount Check10',
        #    'Duplicate Rows Check11', 'Column Name check12',
        #    'Latitude and Longitude Check13', 'Complete Record Check14',
        #    'Unique Count Check15', 'Leading Trailing Space check16', 'CheckSum17']
        print(str(datetime.datetime.now()),": Batch Processing started")

        
        for row in (df.itertuples()):
            # global countOfSuccesreleaseullyTestedTables
            # if(row[0]<countOfSuccesreleaseullyTestedTables): continue
            print(row[0],"**************************")
             

            startIndex=1
            
    # originally reading path from first column
           # self.Source_file_name=row[startIndex]
            # if(str(self.Source_file_name).strip() not in ['NaN','','nan']):
            #     hl.delete(0,END)
            #     hl.insert(0,self.Source_file_name)
            # else: hl.delete(0,END)

            # startIndex =startIndex+1
        
        #    source file name
            self.Source_file_name=str(self.batch_Source_entry.get()).strip()+"/"+str(row[startIndex])
            # print("bathc Source",self.Source_file_name)
            if(str(self.Source_file_name).strip() not in ['NaN','','nan']):
                self.hl.delete(0,END)
                self.hl.insert(0,self.Source_file_name)
            else: self.hl.delete(0,END)
        
        #    originally reading Target filename
            # startIndex = startIndex+1
            # global self.release_file_name
           # self.release_file_name=row[startIndex]
            # if(str(self.release_file_name).strip() not in ['NaN','','nan']):
           #     sl.delete(0,END)
            #     sl.insert(0,self.release_file_name)
            # else: sl.delete(0,END)

    # target file name
            startIndex = startIndex+1
             
        
            self.release_file_name=str(self.batch_release_entry.get()).strip()+"/"+row[startIndex]
            if(str(self.release_file_name).strip() not in ['NaN','','nan']):
                self.sl.delete(0,END)
                self.sl.insert(0,self.release_file_name)
            else: self.sl.delete(0,END)

            print(str(datetime.datetime.now()),":  Testing file ",self.Source_file_name,"______",self.release_file_name)
    # result path
            # startIndex = startIndex+1
            global Source_location_adrs
            # Source_location_adrs=row[startIndex]
            Source_location_adrs=self.batch_Result_entry.get()

            if(str(Source_location_adrs).strip() not in ['NaN','','nan']):
                self.bl.delete(0,END)
                self.bl.insert(0,Source_location_adrs)
            else: self.bl.delete(0,END)

            # startIndex = startIndex+1            
            # resultFileNameFromExcel=row[startIndex]
            # resultFileNameFromExcel=
            # if(str(resultFileNameFromExcel).strip() not in ['NaN','','nan']):
            #     self.resultFileName.delete(0,END)
            #     self.resultFileName.insert(0,resultFileNameFromExcel)
            # else: self.resultFileName.delete(0,END)

    # primary keys
            startIndex = startIndex+1            
            primaryKeys=row[startIndex]
            if(str(primaryKeys).strip() not in ['NaN','','nan']):
                self.pk.delete(0,END)
                self.pk.insert(0,primaryKeys)
            else: 
                self.pk.delete(0,END)
                # add_all_common_cols()
            # print("PRIMARY KEY: ",primaryKeys)


    # skip cols
            startIndex = startIndex+1 
            skipColsId=row[startIndex]

            if(str(skipColsId).strip() not in ['NaN','','nan']):
                self.skipCols.delete(0,END)
                self.skipCols.insert(0,skipColsId)
            else: self.skipCols.delete(0,END)




   #  skip while full compare (tokenized col)
            startIndex = startIndex+1                    
            tokenizedCols_skipDuringFullCompare=row[startIndex]

            if(str(tokenizedCols_skipDuringFullCompare).strip() not in ['NaN','','nan']):
                self.skipColsDuringFullCompare.delete(0,END)
                self.skipColsDuringFullCompare.insert(0,tokenizedCols_skipDuringFullCompare)

    # timstamp columns names
            startIndex = startIndex+1   
            timeStampCols=row[startIndex]
            if(str(timeStampCols).strip() not in ['NaN','','nan']):
                self.timest.delete(0,END)
                self.timest.insert(0,timeStampCols)

    #longitude column names
            startIndex = startIndex+1   
            longitudeColumns=row[startIndex]
            if(str(longitudeColumns).strip() not in ['NaN','','nan']):
                self.longi.delete(0,END)
                self.longi.insert(0,longitudeColumns)


    #latitude column names
            startIndex = startIndex+1   
            latitudeColumns=row[startIndex]
            if(str(latitudeColumns).strip() not in ['NaN','','nan']):
                self.lati.delete(0,END)
                self.lati.insert(0,latitudeColumns)


    #column mapping 
            startIndex = startIndex+1   
            columnMapping=row[startIndex]
            if(str(columnMapping).strip() not in ['NaN','','nan']):
                self.colsMapping.delete(0,END)
                self.colsMapping.insert(0,columnMapping)

    # compare complete record check(Full comparison)
            startIndex = startIndex+1            
            completeRecordCheckId=row[startIndex]
            if (str(completeRecordCheckId).strip()).lower()=='y':
                self.doCompareDataCheck.set(True)
            elif (str(completeRecordCheckId).strip()).lower()=='n': 
                self.doCompareDataCheck.set(False)

# do case insensitive check
            startIndex = startIndex+1            
            caseInsensitiveCheck=row[startIndex]
            if (str(caseInsensitiveCheck).strip()).lower()=='y':
                self.doCaseInsensitiveCheck.set(True)
            else: self.doCaseInsensitiveCheck.set(False)

    # do all check checkbox
            startIndex = startIndex+1            
            doAllCheckId=row[startIndex]
            if (str(doAllCheckId).strip()).lower()=='y':
                self.doAllChecks.set(True)
                self.checkAll()


                




    # null check 
            startIndex = startIndex+1            
            nullCheckId=row[startIndex]
            if (str(nullCheckId).strip()).lower()=='y':
                self.doNullCheck.set(True)
            elif (str(nullCheckId).strip()).lower()=='n': 
                self.doNullCheck.set(False)

    # blank check
            startIndex = startIndex+1            
            blankCheckId=row[startIndex]
            if (str(blankCheckId).strip()).lower()=='y':
                self.doBlankCheck.set(True)
            elif (str(blankCheckId).strip()).lower()=='n': 
                self.doBlankCheck.set(False)

    # timstamp check
            startIndex = startIndex+1            
            timestampCheckId=row[startIndex]
            if (str(timestampCheckId).strip()).lower()=='y':
                self.doTimeStCheck.set(True)
            elif (str(timestampCheckId).strip()).lower()=='n': 
                self.doTimeStCheck.set(False)

    # rowcount check
            startIndex = startIndex+1            
            rowCountCheckId=row[startIndex]
            if (str(rowCountCheckId).strip()).lower()=='y':
                self.doRowCountCheck.set(True)
            elif (str(rowCountCheckId).strip()).lower()=='n': 
                self.doRowCountCheck.set(False)

    # duplicate row check
            startIndex = startIndex+1            
            duplicateRowCheckId=row[startIndex]
            if (str(duplicateRowCheckId).strip()).lower()=='y':
                self.doDuplicateRowCheck.set(True)
            elif (str(duplicateRowCheckId).strip()).lower()=='n': 
                self.doDuplicateRowCheck.set(False)

    # column name check
            startIndex = startIndex+1            
            columnNameCheckId=row[startIndex]
            if (str(columnNameCheckId).strip()).lower()=='y':
                self.doColNameCheck.set(True)
            elif (str(columnNameCheckId).strip()).lower()=='n': 
                self.doColNameCheck.set(False)

    # latitude longitude check
            startIndex = startIndex+1            
            latLongCheckId=row[startIndex]
            if (str(latLongCheckId).strip()).lower()=='y':
                self.doLatLongCheck.set(True)
            elif (str(latLongCheckId).strip()).lower()=='n': 
                self.doLatLongCheck.set(False)


            
    # unique count check
            startIndex = startIndex+1            
            uniqueCountCheckId=row[startIndex]
            if (str(uniqueCountCheckId).strip()).lower()=='y':
                self.doUniqueValueCountCheck.set(True)
            elif (str(uniqueCountCheckId).strip()).lower()=='n': 
                self.doUniqueValueCountCheck.set(False)

    # leading trailing cheheck 
            startIndex = startIndex+1            
            leadingTrailingCheckId=row[startIndex]
            if (str(leadingTrailingCheckId).strip()).lower()=='y':
               self.doLeadingTrailingSpaceCheck.set(True)
            elif (str(leadingTrailingCheckId).strip()).lower()=='n': 
                self.doLeadingTrailingSpaceCheck.set(False)

    # SumOfValues
            startIndex = startIndex+1            
            checksumCheckId=row[startIndex]
            if (str(checksumCheckId).strip()).lower()=='y':
                self.doCheckSumCheck.set(True)
            elif (str(checksumCheckId).strip()).lower()=='n': 
               self.doCheckSumCheck.set(False)
            
            
            self.add_common_cols()
            self.initialize()
        
           # testThread = threading.Thread(target=self.initialize)
        
            # # starting thread 1
            # testThread.start()

            
        # global countOfSuccesreleaseullyTestedTables
        # countOfSuccesreleaseullyTestedTables=0
        self.lastStepConsolidation()
        messagebox.showinfo("Done","All specified testing & consolidation of reports is done. Check Result folder")
        

            
            


        # traverse row by row
        # use values, update all fields then click test
        # get the results 

        

    
    def checkAll(self):
        if(self.doAllChecks.get()==1):
            self.doNullCheck.set(True)
            self.doBlankCheck.set(True)
            self.doTimeStCheck.set(True)
            self.doRowCountCheck.set(True)
            self.doDuplicateRowCheck.set(True)
            self.doColNameCheck.set(True)
            self.doLatLongCheck.set(True)
            self.doCompareDataCheck.set(True)
            self.doUniqueValueCountCheck.set(True)
            self.doLeadingTrailingSpaceCheck.set(True)
            self.doCheckSumCheck.set(True)
        else:
            
            self.doNullCheck.set(False)
            self.doBlankCheck.set(False)
            self.doTimeStCheck.set(False)
            self.doRowCountCheck.set(False)
            self.doDuplicateRowCheck.set(False)
            self.doColNameCheck.set(False)
            self.doLatLongCheck.set(False)
            self.doCompareDataCheck.set(False)
            self.doUniqueValueCountCheck.set(False)
            self.doLeadingTrailingSpaceCheck.set(False)
            self.doCheckSumCheck.set(False)

        

    
    def search(event):
        value=event.widget.get().strip()
        
        if value=='':
            self.commonCols_combo['values']=cols
        else:
            data=[]
            
            for col in cols:
                if value.lower() in col.lower():
                    data.append(col)
            self.commonCols_combo['values']=data 
    #         commonCols_combo.event_generate('<Down>')
            self.commonCols_combo.focus()
            
    

    # def primary_key_added(event):
    #     """ handle the month changed event """
    #     keysAlready=(pk.get()).strip()
    #     newkey=commonCols_combo.get()
        
    #     if len(keysAlready)==0:
    #         combined_key=newkey
    #     else:
    #         combined_key=keysAlready +","+ newkey
            
    #     pk.delete(0,END)
    #     pk.insert(0,combined_key)
        
    # commonCols_combo.bind('<<ComboboxSelected>>', primary_key_added)



    

    global fdf
    global singleFactsDf
    global cols
   
    global Source_location_adrs
    global mergedTableDf
    global skipColsList
    global unCommonSourceCols
    global unCommonreleaseCols
    global colsDict
    # global countOfSuccesreleaseullyTestedTables
    # countOfSuccesreleaseullyTestedTables=0
    # colsDict=dict()
            
    def sanitizeColNames(self,df):
        try:
            cls=list()
            for col in df.columns:
                if('.' in col): cls.append(((col.split('.')[1]).lower()).strip())
                else: cls.append((col.lower()).strip())
                # cls.append((col.lower()).strip())
            print(str(datetime.datetime.now()),': Column names sanitized without removing database names..')
            return cls
            
        except  Exception as err:
            raise Exception("Error occurred while sanitizing column names. Is the column names any other format than tablename.columnname? MSG:",str(err))
            return

    def checkNull(self):
        try:
    #         //issue
            lst_Source=list()
            lst_release=list()
            for Source_col in colsDict:
                release_col=colsDict[Source_col]
                lst_Source.append(self.df_Source[Source_col].isnull().sum())
                lst_release.append(self.df_release[release_col].isnull().sum())

            fdf['Source_Null_Values']=pd.DataFrame(lst_Source) 
            fdf['Release_Null_Values']=pd.DataFrame(lst_release)
            fdf['Result_Null']=(fdf['Source_Null_Values']==fdf['Release_Null_Values'])
            fdf['Result_Null'].replace(True,'PASS',inplace=True)
            fdf['Result_Null'].replace(False,'FAIL',inplace=True)
            print(str(datetime.datetime.now()),": Null check done")
        except Exception as err:
            raise Exception("Error occurred while checking for null values: ", str(err))
            return

    def getUniqueValueCount(self):
        try:
            lst_Source=list()
            lst_release=list()
            for Source_col in colsDict:
                release_col=colsDict[Source_col]
                lst_Source.append(len(self.df_Source[Source_col].unique()))
                lst_release.append(len(self.df_release[release_col].unique()))
            fdf['Source_UniqueValues']=pd.DataFrame(lst_Source)    
            fdf['Release_UniqueValues']=pd.DataFrame(lst_release)  
            fdf['Result_UniqueValues']=fdf['Source_UniqueValues']==fdf['Release_UniqueValues']
            fdf['Result_UniqueValues'].replace(True,'PASS',inplace=True)
            fdf['Result_UniqueValues'].replace(False,'FAIL',inplace=True)
            print(str(datetime.datetime.now()),": Unique value count check done")
        except Exception as err:
            raise Exception("Error getting unique count. Msg: ",str(err))
            
            
                

    def checkRowCount(self):
        try:
    #         fdf['rowCount_Source']=self.df_Source.shape[0]
    #         fdf['rowCount_release']=self.df_release.shape[0]
    #         fdf['result_rowCount']=(fdf['rowCount_Source']==fdf['rowCount_release'])
            row_count_source=self.df_Source.shape[0]
            row_count_tgt=self.df_release.shape[0]
            # singleFactsDf.loc[len(singleFactsDf.index)]=['Source: Row Count',self.df_Source.shape[0],'']  
            # singleFactsDf.loc[len(singleFactsDf.index)]=['Target: Row Count',self.df_release.shape[0],'']
            res=''
            if(row_count_source==row_count_tgt):
                res='Matching'
            else:res='Not Matching'


            singleFactsDf.loc[len(singleFactsDf.index)]=['Row Count Result',res,'Source:'+str(row_count_source)+', '+'Target:'+str(row_count_tgt)]  
            print(str(datetime.datetime.now()),": Row count check done")
        except Exception as err:
           raise Exception("Error occurred while checking row count ", str(err))
            


    def checkBlank(self):
        try:
    #        issue of uncommon columns
            lst_Source=list()
            lst_release=list()
            for Source_col in colsDict:
                release_col=colsDict[Source_col]
                lst_Source.append(self.df_Source[Source_col].isna().sum())
                lst_release.append(self.df_release[release_col].isna().sum())
            
            fdf['Source_NotAvailable']=pd.DataFrame(lst_Source)
            fdf['Release_NotAvailable']=pd.DataFrame(lst_release)
            fdf['Result_NotAvailable']=(fdf['Source_NotAvailable']==fdf['Release_NotAvailable']) 
            fdf['Result_NotAvailable'].replace(True,'PASS',inplace=True)
            fdf['Result_NotAvailable'].replace(False,'FAIL',inplace=True)
            print(str(datetime.datetime.now()),": Blank check done")
        except Exception as err:
            raise Exception("Error occurred while checking blank "+ str(err))
            return
        
    def is_leap_year(self,year):
            year=int(year)
            return year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)
    
        
    def validateTS(self,tsVal):
        try:
            try:
                dt=str(datetime.datetime.fromtimestamp( float(tsVal))) # if format is epoch
                s, ms = divmod(float(tsVal), 1000)  # (1236472051, 807)
                '%s.%03d' % (time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(s)), ms)
                return True 
            except:
                try:
                    parser.parse(tsVal)
                    return True
                except: pass
                
            
            len=tsVal.split(' ').__len__()
            if(len<2):
                return False
            elif(len==3):
                date,time,offset=tsVal.split(' ')

            else: 
                date,time=tsVal.split(' ')
            if '/' in  date:
                year,month,day=date.split('/')
            else:
                year,month,day=date.split('-')
                
            year=int(year)
            month=int(month)
            day=int(day)
            
            if(year<0 or month>12 or month<1 or day>31 or day<1):
                return False
            if(month==2 and (day>29)):
                return False
            if((not self.is_leap_year(year)) and day>28 and month==2):
                return False
            hours,mins,sec=time.split(':')
            sec=float(sec)
            hours=int(hours)
            mins=int(mins)

            if(hours<0 or hours>=24 or mins<0 or mins>=60 or sec<0 or sec>=60):
                return False
            return True
        except: return False


    def isTimeStampValid(self,df,tsColName,tableIdentifier):
        try:
            if tsColName not in colsDict.keys():
                raise Exception(tsColName+': Timestamp column not present in both tables or skipped')
            
            tss=list(df[tsColName])
    #         print('tss: ')
    #         print(tss)
            invalidTsList=list()
            for ts in tss:
                try:
                    if ts!='nan' and ts!='None' and ts!='none' and ts!='NaN' and self.validateTS((ts.strip()).lower())==False:
                        invalidTsList.append(ts)
                except:continue
                    
            
    #         lst=list(df.apply(lambda row:if validateTS(row,tsColName)==False:,axis=1))
    #         lst=[i for i in lst]
            singleFactsDf.loc[len(singleFactsDf.index)]=[tableIdentifier+'InvalidTimestamps: '+ tsColName,len(invalidTsList),str(invalidTsList)]
            print(str(datetime.datetime.now()),": Timestamp check done: ",tsColName)
    #         fdf['ValidTimestamp: '+ tsColName]=df.apply(lambda row:validateTS(row,tsColName),axis=1)
        except Exception as err:
            raise Exception( "Error occurred while timestamp "+str(err))
            return

    def countDuplicateRows(self,df):
        try:
            df_unique=df.drop_duplicates()
            print(str(datetime.datetime.now()),": Duplicate count checking done")
            return abs(df_unique.shape[0]-df.shape[0])  

        except Exception as err:
            raise Exception("Error occurred while duplicate rows "+str(err))
            return
    
        
    def areColNamesSame(self):
        source_cols=list(self.df_Source.columns)
        target_cols=list(self.df_release.columns)
        # source_cols.sort()
        # target_cols.sort()
        lst_uncommon_in_source=list()
        lst_uncommon_in_target=list()
        skipColsList=[str(i).strip().lower() for i in str(self.skipCols.get()).strip().split(',')]
        # print(str(skipColsList))
        for c in source_cols:
           # if(c not in cols and c not in skipColsList): lst_uncommon_in_source.append(c)
           if(c not in colsDict.keys()) and (c not in skipColsList): lst_uncommon_in_source.append(c)
        
        for c in target_cols:
            # if(c not in cols and c not in skipColsList): lst_uncommon_in_target.append(c)
            if(c not in colsDict.values()) and (c not in skipColsList): lst_uncommon_in_target.append(c)
                
        if len(lst_uncommon_in_source)==0 and len(lst_uncommon_in_target)==0 :
            singleFactsDf.loc[len(singleFactsDf.index)]=['Are Column Names Same','Yes','All cols name are same']  
        else:
            singleFactsDf.loc[len(singleFactsDf.index)]=['Count of different cols in Source (Cols present in source but not target)',len(lst_uncommon_in_source),str(lst_uncommon_in_source)]  
            singleFactsDf.loc[len(singleFactsDf.index)]=['Count of different cols in Target (Cols present in target but not in source)',len(lst_uncommon_in_target),str(lst_uncommon_in_target)]  
        print(str(datetime.datetime.now()),": Col names comparison done")
        
    def countInvalidLatLong(self,df,longCol,latCol):
        try:
            lg=0
            lgList=list()
            lt=0
            ltList=list()

            if longCol.lower()  in df.columns:
                for val in df[longCol.lower()].values:
                    try:
        #                 val=float(val)
                        if float(val)>180 or float(val)<-180:
                            lg=lg+1
                            lgList.append(str(val))
                    except:
                        lg=lg+1
                        lgList.append(str(val))


                sl=str(','.join(lgList))
                singleFactsDf.loc[len(singleFactsDf.index)]=['Invalid_longitudes: '+longCol,lg,sl]
        
            
            if latCol.lower()  in df.columns:
                for val in df[latCol.lower()].values:
                    try:
        #                 val=float(val)
                        if float(val)>90 or float(val)<-90:
                            lt=lt+1
                            ltList.append(str(val))
                    except:
                        lt=lt+1
                        ltList.append(str(val))

                sl=','.join(ltList)
                singleFactsDf.loc[len(singleFactsDf.index)]=['Invalid_latitudes '+latCol,lt,sl]
                print(str(datetime.datetime.now()),": Lati/Longi check done")
        except Exception as err:
            raise Exception("Error occurred while checking longitude/latitude "+ str(err))
            
            
    def checkLeadingAndTrailingSpace(self):
        try:
            lst_Source=list()
            lst_release=list()

            for Source_col in colsDict:
                release_col=colsDict[Source_col]
                hl=0;
                sl=0;
                for val in self.df_Source[Source_col]:
                    if str(val).strip() != str(val):
                        hl=hl+1
                for val in self.df_release[release_col]:
                    if str(val).strip() != str(val):
                        sl=sl+1
                lst_Source.append(hl)
                lst_release.append(sl)
            fdf['Source_LeadingTrailing Space']=pd.DataFrame(lst_Source)
            fdf['Release_LeadingTrailing Space-Target']=pd.DataFrame(lst_release)
            fdf['Result_LeadingTrailingSpace']=fdf['Source_LeadingTrailing Space']==fdf['Release_LeadingTrailing Space-Target']
            fdf['Result_LeadingTrailingSpace'].replace(True,'PASS',inplace=True)
            fdf['Result_LeadingTrailingSpace'].replace(False,'FAIL',inplace=True)
            print(str(datetime.datetime.now()),": Leading trailing check done...")
        except Exception as err:
            raise Exception('Error during Leading/Trailing space check Msg: '+str(err))
            
    def SumOfValues(self):
        try:
            lst_Source=list()
            lst_release=list()
            for Source_col in colsDict:
                release_col=colsDict[Source_col]
                try:
                    lst_Source.append(self.df_Source[Source_col].astype(float).sum())
                except:
                    lst_Source.append('NaN')
                try:
                    lst_release.append(self.df_release[release_col].astype(float).sum())
                except:
                    lst_release.append('NaN')
                    
            fdf['SumOfValues-Source']=pd.DataFrame(lst_Source)
            fdf['SumOfValues-Target']=pd.DataFrame(lst_release)
            fdf['Result_Checksum']=pd.DataFrame( fdf['SumOfValues-Source'].eq(fdf['SumOfValues-Target']))
            fdf['Result_Checksum'].replace(True,'PASS',inplace=True)
            fdf['Result_Checksum'].replace(False,'FAIL',inplace=True)
            print(str(datetime.datetime.now()),": SumOfValues check done..")
        except Exception as err:
            raise Exception("Error during SumOfValues msg: "+str(err))
            
            
            
    def compareDataIndexed_tuples(self ,dataComparisonDF,tsColList):

        print(str(datetime.datetime.now()),": ****Full data comparison started.....")
        try:
            
            dfLenId=len(dataComparisonDF.index)
     
        

                    
            try:
       

                timestampCol=(self.timest.get()).strip()
                tsColList=list()
                if len(timestampCol)>0 and self.doTimeStCheck.get()==1:
                    for ts in timestampCol.split(','):
                        trimmedTS=(ts.strip()).lower()
                        if trimmedTS not in colsDict.keys():
                            # raise Exception('Timestamp columns not present in both tables or delimiter is incorrect  ')
                            print(trimmedTS," column not present in both tables")
                        else:
                            tsColList.append(trimmedTS)


                for col in tsColList:
                    # if col is epoch then it will convert
                    # else if it is normal timestamp 
                    if col in self.df_Source.columns:

                        # for epochs
                        source_ts_cols=list()
                        for val in self.df_Source[col]:
                                    try:
                                       
                                        source_ts_cols.append(str(datetime.datetime.fromtimestamp( float(val))))
                                        
                                    except:
                                        try:
                                            dt=(parser.parse(val))
                                            dt=str(dt.replace(tzinfo=None))
                                            source_ts_cols.append(dt)
                                        except: 
                                            try:
                                                s, ms = divmod(float(val), 1000)  # (1236472051, 807)
                                                dt='%s.%03d' % (time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(s)), ms)
                                                source_ts_cols.append(dt)
                                            except:
                                                source_ts_cols.append(val)

                        self.df_Source[col]=source_ts_cols


                        release_ts_cols=list()
                        for val in self.df_release[col]:
                                    try:
                                        release_ts_cols.append(str(datetime.datetime.fromtimestamp( float(val))))
                                    except:
                                        try:
                                            dt=(parser.parse(val))
                                            dt=str(dt.replace(tzinfo=None))
                                            release_ts_cols.append(dt)
                                        except:
                                            try:
                                                    s, ms = divmod(float(val), 1000)  # (1236472051, 807)
                                                    dt='%s.%03d' % (time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(s)), ms)
                                                    release_ts_cols.append(dt)
                                            except:
                                                    release_ts_cols.append(val)
                                        

                        self.df_release[col]=release_ts_cols
                       
                        # self.df_release[col] = pd.to_datetime(self.df_release[col],format='%Y-%m-%d %H:%M:%S.%f')
                        # self.df_Source[col] = pd.to_datetime(self.df_Source[col],format='%Y-%m-%d %H:%M:%S.%f')

                        # for normal timstamp

                        ts_col_normal_list=list()
                        for val in self.df_Source[col]:
                            try:
                                offset=val.split(':')[-1]
                                
                                ts_col_normal_list.append(":".join(val.split(':')[0:-1])+":"+str(float(offset)))
                                
                            except:
                                ts_col_normal_list.append(val)

                        self.df_Source[col]=ts_col_normal_list

                        ts_col_normal_list.clear()
                        for val in self.df_release[col]:
                            try:
                                offset=val.split(':')[-1]
                                
                                ts_col_normal_list.append(":".join(val.split(':')[0:-1])+":"+str(float(offset)))
                                
                            except:
                                ts_col_normal_list.append(val)

                        self.df_release[col]=ts_col_normal_list
                                
                        
                          
                            

                   
                    
                    

                for  col in self.df_Source.columns:
                #         # self.df_Source[col]=self.df_Source[col].astype(object)
                #         self.df_Source[col].fillna("Empty cell", inplace = True)
                        
                        try:
                            self.df_Source[col] = pd.Series(self.df_Source[col],dtype='float')
                        except: pass
                #         self.df_Source[col] =pd.Series(self.df_Source[col],dtype='string')

                for  col in self.df_release.columns:
                        try:
                            self.df_release[col] = pd.Series(self.df_release[col],dtype='float')
                        except: pass

                skipDuringFullCompareList=[ i.lower().strip() for i in str(self.skipColsDuringFullCompare.get()).strip().split(',')]  

                # print(cols)
                Source_cols_list=list()
                release_cols_list=list()
                for Source_col in colsDict :

                    if Source_col not in skipDuringFullCompareList:

                        Source_cols_list.append(Source_col)

                        release_cols_list.append(colsDict[Source_col])


                for  col in self.df_release.columns:
                        self.df_release[col].fillna("n/a", inplace = True)


                for  col in self.df_Source.columns:
                        self.df_Source[col].fillna("n/a", inplace = True)

                self.df_release = self.df_release.replace( ['nan.0','nan',':nan','NA','null'],'n/a')
                self.df_Source = self.df_Source.replace( ['nan.0','nan',':nan','NA','null'],'n/a')
                
                if self.doCaseInsensitiveCheck.get()==1:
                    print(str(datetime.datetime.now()),": Making everything lowercase.....")
                    self.df_Source= self.df_Source.applymap(lambda s:s.lower() if isinstance(s, str) else s)
                    self.df_release= self.df_release.applymap(lambda s:s.lower() if isinstance(s, str) else s)
                    print(str(datetime.datetime.now()),": Lowercase conversion done.....")
                    
                df_Source1=self.df_Source.groupby(Source_cols_list,dropna=False).size().reset_index().rename(columns={0:'NoOfRecords'}) 
                df_release1=self.df_release.groupby(release_cols_list, dropna=False).size().reset_index().rename(columns={0:'NoOfRecords'}) 
        
                try:
                     
                    Source1_new_cols=list()

                    
                    leftkeys=self.indexVal
                    rightkeys=[colsDict[col] for col in self.indexVal]
                     

                    for col in df_Source1.columns:
                        if col in self.indexVal: Source1_new_cols.append(col)
                        else: Source1_new_cols.append(col+"_Source")
                    
                    
                    release1_new_cols=list()

                    for col in df_release1.columns:
                        if col in self.indexVal: release1_new_cols.append(col)
                        else: release1_new_cols.append(col+"_release")

                    df_Source1.columns=Source1_new_cols
                    df_release1.columns=release1_new_cols
                    for col in str(self.skipColsDuringFullCompare.get()).strip().split(','):
                            if col.lower() in self.indexVal:

                                self.indexVal.remove(col.lower())
                    # print("BERFORE MERGE")

                    for  col in df_release1.columns:
                       df_release1[col] = df_release1[col].astype(str)

                    for  col in df_Source1.columns:
                       df_Source1[col] = df_Source1[col].astype(str)


                    

                    print(str(datetime.datetime.now()),": Table merging started.....")


                    df=df_Source1.merge(df_release1,how='outer',on=self.indexVal,indicator=True,validate='1:1').set_index(self.indexVal) 
                    # df=df_Source1.merge(df_release1,how='outer',left_on=leftkeys,right_on=rightkeys,indicator=True)
                    
                    # df.to_excel('merged'+str(time.time())+".xlsx",index=False,header=True)
                    print(str(datetime.datetime.now()),": Tables merged.....")
                    
                except Exception as err:
                    raise Exception('Non-unique key given Msg: ',str(err))
    #             global mergedTableDf
    #             mergedTableDf=df 
            except Exception as err:
                raise Exception('Exception during merging: ',str(err))

            mergeIndex=len(df.columns)
            
            
            # print("MEERGED TABLES*******************************")
    #         print(df)
            
            
            # print("tscollist:1146",tsColList)
            # Source_cols=[i for i in colsDict.keys()]
            # newCols=Source_cols+['NoOfRecords']
            # colsDict['NoOfRecords']='NoOfRecords'
            # for skipFullCompareCol in str(self.skipColsDuringFullCompare.get()).strip().split(','):
            #           if skipFullCompareCol.strip().lower() in newCols:
            #                         newCols.remove(skipFullCompareCol.strip().lower())
            total_no_of_rows=df.shape[0]
            row_executing=0
            odd_names=[x for x in filter(lambda x:'->' in x and x not in self.indexVal,df.columns)]
            df.columns=[p for p in map(lambda x: x.replace('->','map_999999_') if x in odd_names else x,df.columns)]
            
            print('Full data comparison status in %:')

            for row in (df.itertuples(index=self.indexVal)):

                    row_executing =row_executing+1
                    # print(str(row_executing)+"/"+str(total_no_of_rows)+"#", end='\r')
                    percentage=100*row_executing/total_no_of_rows
                    print(str(percentage)+"%",end='\r')


                    if(row[mergeIndex]!='both'): 
                        if(row[mergeIndex]=='left_only'):
                            dataComparisonDF.loc[len(dataComparisonDF.index)]=[row.Index,'Target Row missing','Present','Absent']
                            dfLenId=dfLenId+1
                        else:
                            dataComparisonDF.loc[len(dataComparisonDF.index)]=[row.Index,'Source Row missing','Absent','Present']
                            dfLenId=dfLenId+1
                    else:
                        
                        # print("INDEX",self.indexVal)
                        Source_cols=[i for i in colsDict.keys()]
                        newCols=Source_cols+['NoOfRecords']
                        colsDict['NoOfRecords']='NoOfRecords'
                        # print(newCols)
                        # print(str(self.skipColsDuringFullCompare.get()).strip().split(','))
                        for skipFullCompareCol in str(self.skipColsDuringFullCompare.get()).strip().split(','):
                                if skipFullCompareCol.strip().lower() in newCols:
                                                newCols.remove(skipFullCompareCol.strip().lower())
                        
                        for col in newCols:
                            

                            if col in self.indexVal: continue

                            

                            hc=col+'_Source'
                            sc=colsDict[col]+'_release'

                            if '->' in hc:
                                hc=hc.replace('->','map_999999_')
                            if '->' in sc:
                                sc=sc.replace('->','map_999999_')


                            if str(getattr(row,hc)).strip()!=str((getattr(row,sc))).strip() and getattr(row,hc)!=getattr(row ,sc):
                                
                                val1=str(getattr(row,hc))
                                val2=str(getattr(row,sc))
                                
                                

                                try:
                                    if float(val1)==float(val2): pass
                                    else: dataComparisonDF.loc[len(dataComparisonDF.index)]=[row.Index,col,val1,val2]
                                    # print("try:",val1,val2)
                                    dfLenId=dfLenId+1
                                except:
                                    
                                        
                                    if val1.lower() in ['nan.0','nan',"NaN","<NA>","none",'None',"<na>","n/a","nat",":nan",'NA','null']:
                                        dataComparisonDF.loc[len(dataComparisonDF.index)]=[row.Index,col,"-Empty cell-",val2.strip()]
                                    
                                    elif val2.lower()  in ['nan.0','nan',"NaN","<NA>","none",'None',"<na>","n/a","nat",":nan",'NA','null']:
                                        dataComparisonDF.loc[len(dataComparisonDF.index)]=[row.Index,col,val1.strip(),"-Empty cell-"]
                                    else:
                                        # print('woegnwo;n+++++++',col)
                                        if col in tsColList:
                                            try:
                                                try:
                                                    # print(val1,val2)
                                                    dt1=str(datetime.datetime.fromtimestamp( float(val1))) # if epoch format (val1 in epoch format and val2 in normat datetime)
                                                    # print(dt1,"from here")
                                                    if(dt1==val2):
                                                        continue
                                                except:
                                                    # print(val1,val2)
                                                    dt2=str(datetime.datetime.fromtimestamp( float(val2))) # if epoch format (val2 in epoch format and val1 in normat datetime)
                                                    # print(dt2,type(dt2),"from there",val1,type(val1))
                                                    if(dt2==val1): 
                                                        continue

                                            except: pass
                                            # print("VAL1:",val1)
                                            # print("VAL2:",val2)
                                                
                                            try:
                                                offset1=val1.split('.')[1]
                                                offset2=val2.split('.')[1]
                                            except:
                                                dataComparisonDF.loc[len(dataComparisonDF.index)]=[row.Index,col,val1.strip(),val2.strip()]
                                                continue


                                            try:
                                                dt1=(parser.parse(val1))
                                                dt1=str(dt1.replace(tzinfo=None))
                                                parsed_val1=parser.parse(dt1)
                                                dt2=(parser.parse(val2))
                                                dt2=str(dt2.replace(tzinfo=None))
                                                parsed_val2=parser.parse(dt2)
                                                if(parsed_val1==parsed_val2): continue
                                                else:
                                                    dataComparisonDF.loc[len(dataComparisonDF.index)]=[row.Index,col,val1.strip(),val2.strip()]
                                                    continue
                                            except:pass

                                            try:
                                                # print("hereeeeeeeeeeeeeeeeeee",offset1,offset2)
                                                if float(offset1)==float(offset2) and val1.split('.')[0]==val2.split('.')[0]: continue 

                                                    
                                            except: pass
                                            if offset1!=offset2:
                                                dataComparisonDF.loc[len(dataComparisonDF.index)]=[row.Index,col,val1.strip(),val2.strip()]
                                            else:
                                                if(val1.split('.')[0])!=(val2.split('.')[0]):
                                                    dataComparisonDF.loc[len(dataComparisonDF.index)]=[row.Index,col,val1.strip(),val2.strip()]
                                        else:
                                            dataComparisonDF.loc[len(dataComparisonDF.index)]=[row.Index,col,val1.strip(),val2.strip()]
                                                
                                        

                                    dfLenId=dfLenId+1
                                    # print("catch",val1,val2)
                        colsDict.pop('NoOfRecords')
            dataComparisonDF.sort_values(by=dataComparisonDF.columns[0],inplace=True)
            # dataComparisonDF.to_csv('Full_Data_Comparison'+str(time.time())+'.csv')
            print(str(datetime.datetime.now()),": **** Full Data Comparison Done...")                              
        except Exception as err:
                raise Exception("Error while comparing files "+ str(err))

        
    


    def checkTokenization(self):
        try:
            print(str(datetime.datetime.now()),": Tokenization check started...")
            tokenizedCols=[ (col.strip()).lower() for col in self.skipColsDuringFullCompare.get().split(',')]
            nonTokenizedColsList=list()

            for col in tokenizedCols:
                if col not in colsDict:
                    print(col,"not present in both tables")
                    continue
                else:
                    val_lst_1=set(self.df_Source[col].to_list())
                    val_lst_2=set(self.df_release[col].to_list())
                    common_values=val_lst_1.intersection(val_lst_2)
                    for val in common_values:
                        if val not in [np.NaN,'n/a','NaN','nan','null']:
                            nonTokenizedColsList.append(col)
                            break

                   

            
            if(len(nonTokenizedColsList)>0):
                msg='Some values from this column is matching in both table'
                singleFactsDf.loc[len(singleFactsDf.index)]=['Invalid Tokenized columns',msg,str(nonTokenizedColsList)]
            properly_tokenized_col=list()
            msg=''
            for col in tokenizedCols:
                if col not in nonTokenizedColsList:
                    properly_tokenized_col.append(col)

            singleFactsDf.loc[len(singleFactsDf.index)]=['Valid Tokenized columns',msg,str(properly_tokenized_col)]
            
                
            

            
            

            print(str(datetime.datetime.now()),": Tokenization check finished")
        except Exception as err:
            raise Exception('Error during tokenization check ',str(err))
        
        
    def getPrimaryKeys(self):
        df=self.df_Source;
        t=time.time()
        try:
            full_list = chain.from_iterable(combinations(df, i) for i in range(1, len(df.columns)+1))

            n = len(df.index)

            res = []
            for cols in full_list:
                cols = list(cols)
                if len(df[cols].drop_duplicates().index) == n:
                    res.append(cols)
        
            messagebox.showinfo("Done",time.time()-t)
            
        except Exception as err:
            raise Exception('Error in getting primary keys',str(err))


            
    
    def writeIntoExcel(self,dataComparisonDF):
        try:
            
            print(str(datetime.datetime.now()),": Writing into excel started...")
            fileNameFromUser=str(self.resultFileName.get()).strip() 
            
            if fileNameFromUser=='':
                try:
                    sourceFileName=(((self.Source_file_name.split('/'))[-1]).split('.'))[0]
                    targetFileName=(((self.release_file_name.split('/'))[-1]).split('.'))[0]
                    s=str(datetime.datetime.now())
                    s=s.replace(' ','__')
                    s=s.replace('-','__')
                    s=s.replace(':','_')
                
                    fileNameFromUser=sourceFileName+"__"+targetFileName+"_Result" +"_"  + s  
                except Exception as err:
                    print('Warning: during excel writing',str(err))  
            
                     

            if(fileNameFromUser!=''):
                fileName=fileNameFromUser + '.xlsx'
            else:
                fileName='TkintertableName_TestResult' + str(time.time()) +'.xlsx'

            with pd.ExcelWriter(fileName) as writer:
                fdf.to_excel(writer,sheet_name='column-wise',index=False,header=True)
                singleFactsDf.to_excel(writer,sheet_name='Target-table-facts',index=False,header=True)
                dataComparisonDF.to_excel(writer,sheet_name='Data-Comparison-Detail',index=False,header=True)
                # mergedTableDf.to_excel(writer,sheet_name='MergedTable',index=False,header=True)
                

            # self.separate_result_excel_3rdTab(fileName)
            wb = openpyxl.load_workbook(fileName)
            ws = wb['column-wise'] #Name of the working sheet

            max_col = ws.max_column
            max_row=ws.max_row
            my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
            fill_cell = PatternFill(patternType='solid', fgColor=my_red)

            for i in range(1,max_col+1):
                for j in range(1,max_row+1):
                    # print(ws.cell(column=i,row=j).value)
                    if ws.cell(column=i,row=j).value=='FAIL':
                        ws.cell(column=i, row=j).fill = fill_cell


                    # gen_random_hexDec = ([random.choice('ABCDEF0123456789') for i in range(6)])
                    # hexadecimal = ''.join(gen_random_hexDec)
                    
                    # ws.cell(column=i, row=1).fill = fill_cell
                    # ws.cell(column=i+1, row=1).fill = fill_cell
                    # ws.cell(column=i+2, row=1).fill = fill_cell
                    # ws.cell(column=i+3, row=1).fill = fill_cell
                    # ws.cell(column=i+4, row=1).fill = fill_cell
               
            wb.save(fileName)
            print(str(datetime.datetime.now()),": Done writing in excel")
            
        except Exception as err:
            raise Exception("Error writing into excel: "+ str(err))
    

    def separate_result_excel_3rdTab(self,fileName):
            import xlrd

            # diff code for full data comparison tab
            
            header_style = xlwt.easyxf('font: name Arial,bold on')
            style0 = xlwt.easyxf('font: name Arial, color-index red')
            style1 = xlwt.easyxf(num_format_str='D-MMM-YY')

            wb = xlwt.Workbook()
            sheet1 = wb.add_sheet('Diff', cell_overwrite_ok=True)
            print(datetime.datetime.now())

            dataframe1 = pd.read_excel(fileName,sheet_name='Data-Comparison-Detail')
            header_data1 = dataframe1.columns.ravel()
            header_data2 = header_data1.tolist()
            header_data3 = ', '.join(map(str, header_data2))

            raw_data = []
            raw_header = []

            raw_data1 = dataframe1.to_numpy()
            raw_data = raw_data1.tolist()
            counter = 0

            counters = counter-1

            row_cnt = 0
            for m in range(0, len(dataframe1. index), 2):
                row_cnt = row_cnt + 1
                rawitem1 = ', '.join(map(str, raw_data[m]))
                rawitem2 = ', '.join(map(str, raw_data[m+1]))

                item = tuple(map(str, rawitem1.split(', ')))
                item1 = tuple(map(str, rawitem2.split(', ')))

                for j in range(0, len(item)):
                    #print(j)
                    if item[j] == item1[j]:
                        sheet1.write(row_cnt, j, item[j])

                    if item[j] != item1[j]:
                        sheet1.write(row_cnt, j, "S:" + item[j] + " T:" + item1[j], style0)


            item = tuple(map(str, header_data3.split(', ')))
            for k in range(0, len(item)):
                sheet1.write(0, k, item[k], header_style)
            print(datetime.datetime.now())

            

            
            with pd.ExcelWriter(fileName) as writer:
                fdf.to_excel(writer,sheet_name='column-wise',index=False,header=True)

            fileName_diff='Diff__'+fileName[0:-5]+'.xls'
            wb.save(fileName_diff)

            book = xlrd.open_workbook(fileName_diff)

            df = pd.read_excel(book)
            df.to_excel(fileName_diff+'x')
            os.remove(fileName_diff)
            self.combine_sheets(fileName,fileName_diff+'x')

    
    def combine_sheets(self,file1,file2):
        #=====New Code====#
        from openpyxl import Workbook, load_workbook
        
         

        dest_wb = Workbook()

        for  file  in [file1,file2]:
             
                filed= file.split('.')[0]
                # Absolute Path for Excel files
                # file_path = os.path.abspath(os.path.join(base, file))

                # Create new sheet in destination Workbook
                dest_wb.create_sheet(filed)
                dest_ws = dest_wb[filed]

                # =====New Code====#

                # Read source data
                source_wb = load_workbook(file)
                source_sheet = source_wb.active
                for row in source_sheet.rows:
                    for cell in row:
                        dest_ws[cell.coordinate] = cell.value
                # =================#

         
        dest_wb.save(file1)

    def all_checks(self,latiColsList,longiColsList ,tsColList,dataComparisonDF):
        



        if self.doNullCheck.get()==1:
            self.checkNull()
    #         //add checkbox 
        if self.doLeadingTrailingSpaceCheck.get()==1:
            self.checkLeadingAndTrailingSpace()

        self.df_release=self.df_release.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        self.df_Source=self.df_Source.applymap(lambda x: x.strip() if isinstance(x, str) else x)

        self.df_Source.replace(['NULL','','(null)','nan',':nan'],np.NaN,inplace=True)
        self.df_release.replace(['NULL','','(null)','nan',':nan'],np.NaN,inplace=True)

        if self.doUniqueValueCountCheck.get()==1:
            self.getUniqueValueCount()
            
       
            
            
        if self.doCheckSumCheck.get()==1:
            self.SumOfValues()
        
        if self.doRowCountCheck.get()==1:
            self.checkRowCount()
            
        if self.doBlankCheck.get()==1:
            self.checkBlank()
        
        if self.doTimeStCheck.get()==1:
            for tsCol in tsColList:
                if tsCol in self.df_release.columns:
                    self.isTimeStampValid(self.df_release,tsCol,"Target File") 
                if tsCol in self.df_Source.columns:                    
                    self.isTimeStampValid(self.df_Source,tsCol,"Source File") 
        
        if self.doDuplicateRowCheck.get()==1:
            numOfDupRows_Source=self.countDuplicateRows(self.df_Source)
        
            numOfDupRows_release=self.countDuplicateRows(self.df_release)
        
            singleFactsDf.loc[len(singleFactsDf.index)]=['DuplicateRowCount-Target',numOfDupRows_release,'']
            singleFactsDf.loc[len(singleFactsDf.index)]=['DuplicateRowCount-Source',numOfDupRows_Source,'']
        
        if self.doColNameCheck.get()==1:
            self.areColNamesSame()
        
    #     print('*********************************************************\nbefore Source:')
        #print(self.df_Source) 
    #     print(self.df_Source.info())
    #     print('before- Target')
    #     print(self.df_release.head())
    #     print(self.df_release.info())
        if self.doLatLongCheck.get()==1: 
            for longiCol in longiColsList:
                self.countInvalidLatLong(self.df_release,longiCol,"none")
                
            for latiCol in latiColsList:
                self.countInvalidLatLong(self.df_release,"none",latiCol)
            
    #     for  col in self.df_release.columns:
    #                self.df_release[col] = pd.Series(self.df_release[col],dtype='string')
    # #                print(self.df_release[col].dtype)

    #     for  col in self.df_Source.columns:
    #                self.df_Source[col] =pd.Series(self.df_Source[col],dtype='string')
                
                
    #     print(self.df_release.info())
    #     print(self.df_release)
        
    #     stringcols = self.df_Source.select_dtypes(include='object').columns
    #     self.df_Source[stringcols] = self.df_Source[stringcols].fillna('').astype(str)
    
    #     print("SAMOS",str(tsColList))
        # for col in tsColList:
    #         release_list=[]
    #         # print("here")
    #         for val in self.df_release[col]:
    #             left=str(val).split('.')[0]
                
    #             try:
    #                 right=str(val).split('.')[1]
    #                 release_list.append(left+"."+str(int(right)))
    # #                 print(right)
    #             except:
    #                 release_list.append(left)

                    
    #         Source_list=[]
    #         for val in self.df_Source[col]:
    #             left=str(val).split('.')[0]
                
                # try:
                #     right=str(val).split('.')[1]
                #     Source_list.append(left+"."+str(int(right)))
                # except:
                #     Source_list.append(left)

                
                
            # self.df_release[col]=pd.Series(release_list)
            # self.df_Source[col]=pd.Series(Source_list)
    #         print(self.df_release[col])
    #         print("()))))))))))))()")
    #         print(self.df_Source[col])
                    
        if len(str(self.skipColsDuringFullCompare.get()).strip())>0:
            self.checkTokenization()

        if  self.doCompareDataCheck.get()==1:  
            self.compareDataIndexed_tuples( dataComparisonDF,tsColList)
        
        if( self.doNullCheck.get()==1 or self.doBlankCheck.get()==1 or self.doTimeStCheck.get()==1 or self.doRowCountCheck.get()==1
            or self.doDuplicateRowCheck.get() or self.doColNameCheck.get()==1 or self.doLatLongCheck.get()==1 or self.doCompareDataCheck.get()==1):   
            self.writeIntoExcel(dataComparisonDF)
        else: raise Exception('No checks are selected!')

        

    def initialize(self):
        
        try:
            
            
    #         self.df_Source,self.df_release=add_common_cols(Source_file,release_file)
            print(str(datetime.datetime.now()),": Initializing values...")
            if len((self.bl.get()).strip())==0 or len((self.hl.get()).strip())==0 or len((self.sl.get()).strip())==0 or len((self.delimit.get()).strip())==0:
                raise Exception('Input Source location, Target filename,Source filename and delimiter before testing.')
            
            global fdf
            fdf=pd.DataFrame()
            
            # global colsDict
            naming_dictionary=dict()

            for pairCol in str(self.colsMapping.get()).split(','):
                try:
                    Source_col,release_col=pairCol.split('->')
                    source_col_name=(str(Source_col).lower().strip()).strip() 
                    target_col_name=(str(release_col).lower().strip()).strip()
                    cols.append(source_col_name+'->'+ target_col_name)
                    naming_dictionary[source_col_name]=source_col_name+'->'+ target_col_name
                    naming_dictionary[target_col_name]=source_col_name+'->'+ target_col_name
                    colsDict[source_col_name+'->'+ target_col_name]=source_col_name+'->'+ target_col_name
                except:
                    pass

            
            
            self.df_Source.rename(naming_dictionary,axis=1,inplace=True)
            self.df_release.rename(naming_dictionary,axis=1,inplace=True)

    #         for x, y in colsDict.items():
    #                             print(x, y)

            global singleFactsDf
            singleFactsDf=pd.DataFrame(columns=['Condition','Result','Remark']) 
            
            
            singleFactsDf.loc[len(singleFactsDf.index)]=['Common Column Count',len(set(cols)),list(set(cols))] 
            
            for skipcol in str(self.skipCols.get()).strip().split(','):

                if skipcol.lower() in cols: cols.remove(skipcol.lower())
            
            for skipcol in str(self.skipCols.get()).strip().split(','):
                if skipcol.lower() in cols: cols.remove(skipcol.lower())
            
            # print("skipcols1409: ",str(skipCols.get()).strip().split(','))
        
            
            primary_keys=(self.pk.get())
            if primary_keys=='':
                # raise Exception("Please input primary keys")
                # return

                # pk_string=",".join(cols)
                pk_string=",".join(list(colsDict.keys()))
                self.pk.delete(0,END)
                self.pk.insert(0,pk_string)
                primary_keys=(self.pk.get())


            self.indexVal=[(i.lower().strip()) for i in primary_keys.split(',')]
        
            # print("INDEX: ",self.indexVal)
            # skipColsList=str(skipCols.get()).strip().split(',')
            skipColsList=[str(i).strip().lower() for i in str(self.skipCols.get()).strip().split(',')]

            
            for skipcol in skipColsList:
                skipcol=str(skipcol).strip().lower()
                if skipcol in colsDict.keys():
                    colsDict.pop(skipcol)


            # print("skipcolslist: ",skipColsList)
            for idCol in self.indexVal:
                if idCol in skipColsList:
                    self.indexVal.remove(idCol)
                
    #         //put into add-common col list

            latiColsList=list()
            latiCols=(self.lati.get()).strip()
            
            if(len(latiCols)>0 and self.doLatLongCheck.get()==1 ):
                for latitude in latiCols.split(','):
                    trimmedLati=(latitude.strip()).lower()
                    if trimmedLati not in colsDict.keys():
                        # raise Exception('latitude columns not present in both tables or delimiter is incorrect ')
                        print(trimmedLati,"not present in both tables...")
                        pass
                    else:
                        latiColsList.append(trimmedLati)
            
            if(self.doLatLongCheck.get()==1 and len(latiColsList)==0):
                # raise Exception( "Please input Latitude column name or uncheck the latitude checkbox")
                self.doLatLongCheck.set(False)
                
            
            longiColsList=list()
            longiCols=(self.longi.get()).strip()
            
            if len(longiCols)>0 and self.doLatLongCheck.get()==1:
                for longitude in longiCols.split(','):
                   trimmedLongi=(longitude.strip()).lower()
                   if trimmedLongi not in colsDict.keys():
                        raise Exception('Longitude columns not present in both tables or delimiter is incorrect ')
                   else:
                        longiColsList.append(trimmedLongi)
                
            
            if(self.doLatLongCheck.get()==1 and len(longiColsList)==0):
                # raise Exception( "Please input Longitude column name or uncheck longitude checkbox ")
                self.doLatLongCheck.set(False)
                

            timestampColList=list()
            
            timestampCol=(self.timest.get()).strip()
            
            if len(timestampCol)>0 and self.doTimeStCheck.get()==1:
                for ts in timestampCol.split(','):
                    trimmedTS=(ts.strip()).lower()
                    if trimmedTS not in colsDict.keys():
                       # raise Exception('Timestamp columns not present in both tables or delimiter is incorrect  ')
                        print(trimmedTS," not present in both tables")
                    else:
                        timestampColList.append(trimmedTS)
            
            if(self.doTimeStCheck.get()==1 and len(timestampColList)==0):
                # raise Exception( "Please input  column name containing timestamp or uncheck Timestamp check")
                self.doTimeStCheck.set(False)

#             datetime issue 
            for col in timestampColList:

                lst=list()

                for val in self.df_Source[col]:
                    try:
                        lst.append(datetime.datetime.strptime(str(val), '%Y-%m-%d').strftime('%Y-%m-%d %H:%M:%S') )
                    except:
                        try:
                            lst.append(datetime.datetime.strptime(str(val), '%Y-%m-%d.0').strftime('%Y-%m-%d %H:%M:%S') )
                        except: lst.append(val)


                self.df_Source[col]=lst

                lst.clear()

                for val in self.df_release[col]:
                    try:
                        lst.append(datetime.datetime.strptime(str(val), '%Y-%m-%d').strftime('%Y-%m-%d %H:%M:%S') )
                    except:
                        try:
                            lst.append(datetime.datetime.strptime(str(val), '%Y-%m-%d.0').strftime('%Y-%m-%d %H:%M:%S') )
                        except: lst.append(val)


                self.df_release[col]=lst


            
            # print("self.indexVal",self.indexVal)
            # print("map",colsDict)
            self.indexVal=[i.strip().lower() for i in self.indexVal]

            for idx in self.indexVal:
                if idx not in colsDict.keys():
                    raise Exception( " Cannot find given key in the tables or delimiter is incorrect.")
                    return
                
            fdf['ColumnName']=[i for i in colsDict.keys()]
            self.fullComparisonIndexes=[i for i  in self.indexVal]

            for col in str(self.skipColsDuringFullCompare.get()).strip().split(','):

                            if col.lower() in self.fullComparisonIndexes:

                                self.fullComparisonIndexes.remove(col.lower())
            # print("TSCOLIST",str(timestampColList))
            dataComparisonDF=pd.DataFrame(columns=[str(self.fullComparisonIndexes),'column','Source (Source)','Target (Target)']) 
            print(str(datetime.datetime.now()),": Initializations done..")

            

            self.all_checks(latiColsList,longiColsList ,timestampColList,dataComparisonDF)
            if self.selected.get()=='Individual Comparison':
                messagebox.showinfo("Done","All specified testing is done. Check Source folder for result")
            # try:
            #     countOfSuccesreleaseullyTestedTables =countOfSuccesreleaseullyTestedTables+1
            # except:
            #     pass  # case where single run happens

        except Exception as err:
            # print("ERROR MSG: countOfSuccesreleaseullyTestedTables ",countOfSuccesreleaseullyTestedTables )
            messagebox.showerror("Error", err)
            return




    def open_Source_file(self):
         
        self.Source_file_name= fd.askopenfilename(title='Select Source file') 
        self.hl.delete(0,END)
        self.hl.insert(0,self.Source_file_name)
        
    def open_release_file(self):
         
        self.release_file_name=fd.askopenfilename(title='Select Target file')
        self.sl.delete(0,END)
        self.sl.insert(0,self.release_file_name)
        
    def select_Source_location(self):
        global Source_location_adrs
        Source_location_adrs=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        self.bl.delete(0,END)
        self.bl.insert(0,Source_location_adrs)
        
    def select_file(self,title,entry):
        file_name=fd.askopenfilename(title=title)
        entry.delete(0,END)
        entry.insert(0,file_name)

    def select_folder(self,entry):
        location=fd.askdirectory(parent=self,initialdir="/",title='Please select a directory')
        entry.delete(0,END)
        entry.insert(0,location)




    def convertNumbersToFloatAndTrimSpaces(self):
        try:
            self.df_Source=self.df_Source.apply(pd.to_numeric, errors='coerce').fillna(self.df_Source)
            
            self.df_release=self.df_release.apply(pd.to_numeric, errors='coerce').fillna(self.df_release)
        except:pass
    #     for col in self.df_Source.columns:
    #         for val in self.df_Source[col]:
    #             try:


    def add_common_cols(self):
        try:
        
            
            global cols
            cols=list()
            global colsDict
            colsDict=dict()
            
            
            Source_location= (self.bl.get())
            if Source_location=='':
                raise Exception( "Please input Source location")
                return
            
            
            try:
                os.chdir(Source_location)
            except Exception as err:
                raise Exception( "Cannot find given Source location." + str(err))
                
            
            Source_file=(self.hl.get())
            if Source_file=='':
                raise Exception( "Please input Source location")
                return

            release_file=(self.sl.get())
            if release_file=='':
                raise Exception( "Please input Target file full name")
                return
            
            delimitBy=str(self.delimit.get()).strip()
            if delimitBy=='':
                raise Exception("Please input delimiter")
                return
            
#             messagebox.showinfo('d',Source_file.split('.')[-1])
            try:
                # print('while reading Source',Source_file)
                 
                if(Source_file.split('.')[-1]=='xlsx'):
#                     messagebox.showinfo('d',Source_file)
                    self.df_Source=pd.read_excel(Source_file,na_values=[''], keep_default_na=False,index_col=False)
                elif (Source_file.split('.')[-1]=='csv'):
                    self.df_Source=pd.read_csv(Source_file,delimiter=',',encoding='utf8',na_values=[''], keep_default_na=False,dtype=str,index_col=False)
                elif (Source_file.split('.')[-1]=='tsv'):
                    self.df_Source=pd.read_csv(Source_file,delimiter='\t',encoding='utf8',na_values=[''], keep_default_na=False,dtype=str,index_col=False)
                elif (str(delimitBy).strip())=='|':
                    self.df_Source=pd.read_csv(Source_file,delimiter='|',encoding='utf8',skiprows=[0,2],na_values=[''], keep_default_na=False,dtype=str,index_col=False)
                    self.df_Source=self.df_Source.iloc[0:-1,1:-1]
                else: 
                    
                    self.df_Source=pd.read_csv(Source_file,delimiter=delimitBy,encoding='utf8',na_values=[''], keep_default_na=False,dtype=str,index_col=False)
                
            except Exception as err:
                raise Exception(  "Please check if delimiter is correct? Cannot find given Source file in Source location. Make sure to include file extension as well in the name", str(err))
                return

            try:
                 
                if(release_file.split('.')[-1]=='xlsx'):
                    self.df_release=pd.read_excel(release_file,na_values=[''], keep_default_na=False,index_col=False)
                elif (release_file.split('.')[-1]=='csv'):
                    self.df_release=pd.read_csv(release_file,delimiter=',',encoding='utf8',na_values=[''], keep_default_na=False,dtype=str,index_col=False)
                elif (release_file.split('.')[-1]=='tsv'):
                    self.df_release=pd.read_csv(release_file,delimiter='\t',encoding='utf8',na_values=[''], keep_default_na=False,dtype=str,index_col=False)

                elif (str(delimitBy).strip())=='|':
                    self.df_release=pd.read_csv(release_file,delimiter='|',encoding='utf8',skiprows=[0,2],na_values=[''], keep_default_na=False,dtype=str,index_col=False)
                    self.df_release=self.df_release.iloc[0:-1,1:-1]
                else: self.df_release=pd.read_csv(release_file,delimiter=delimitBy,encoding='utf8',na_values=[''], keep_default_na=False,dtype=str,index_col=False)

            except:
                raise Exception( "Cannot find given Target file in Source location. Make sure to include file extension as well in the name")
                return
            
            
            
            self.df_Source.columns = self.sanitizeColNames(self.df_Source)
            self.df_release.columns = self.sanitizeColNames(self.df_release)



            for  col in self.df_release.columns:
                       self.df_release[col] = self.df_release[col].astype(str)

            for  col in self.df_Source.columns:
                       self.df_Source[col] = self.df_Source[col].astype(str)


            print(str(datetime.datetime.now()),": Converting empty values to NaN")
            self.df_Source.replace(r'^\s*$',np.NaN,regex=True,inplace=True)
            self.df_release.replace(r'^\s*$',np.NaN,regex=True,inplace=True)
            print(str(datetime.datetime.now()),": Converted empty values to NaN")

             



            print(str(datetime.datetime.now()),": Converting 'NULL','','(null)','nan' values to NaN")
            self.df_Source.replace(['NULL','','(null)','nan',':nan'],np.NaN,inplace=True)
            self.df_release.replace(['NULL','','(null)','nan',':nan'],np.NaN,inplace=True)
            print(str(datetime.datetime.now()),": Converted 'NULL','','(null)','nan' values to NaN")
            
            

            # self.convertNumbersToFloatAndTrimSpaces()
            print(str(datetime.datetime.now()),": Converting possible numbers to float")
            try:
                self.df_Source=self.df_Source.apply(pd.to_numeric, errors='coerce').fillna(self.df_Source)
            
                self.df_release=self.df_release.apply(pd.to_numeric, errors='coerce').fillna(self.df_release)
                print(str(datetime.datetime.now()),": Converted possible numbers to float")
            except:pass
            
            
            l1=list(self.df_Source.columns)
            l2=list(self.df_release.columns)
            global skipColsList
            # skipColsList=[str(i).strip().lower() for i in str(skipCols.get()).strip().split(',')]
            skipColsList=str(self.skipCols.get()).strip().split(',')

            # print("skipcolslist1608: ",skipColsList)
            # global colsDict
            for col in l1:
                if( (col  in l2) and (col not  in skipColsList)) :
                            col=(str(col)).strip()
                            cols.append(col)
                            colsDict[col]=col

            # print("From commoncols")
            # for x, y in colsDict.items():
            #                     print(x, y)
            print("Column Dictionary",str(colsDict))
            
            self.commonCols_combo.set('')
            self.commonCols_combo['values'] = [i for i in cols]

            SourceUncommonColsList=list()
            for col in l1:
                if  col not in cols: SourceUncommonColsList.append(col)
            
            releaseUncommonColsList=list()
            for col in l2:
                if  col not in cols: releaseUncommonColsList.append(col)

            self.unCommonCol_Source_combo.set('')
            self.unCommonCol_Source_combo['values']=[i for i in SourceUncommonColsList]

            
            self.unCommonCol_Release_combo.set('')
            self.unCommonCol_Release_combo['values']=[i for i in releaseUncommonColsList]

            print(str(datetime.datetime.now()),": Common cols added ")
            if self.selected.get()=='Individual Comparison':
                messagebox.showinfo("Added dropdown", "Common cols added")
        except Exception as err:
            messagebox.showerror("Error", err)
            return
        
    #         return self.df_Source,self.df_release
            



        
        
        
    




    # //add latitude cols
    def add_latitude_cols(self):
        newcol=self.commonCols_combo.get()
        oldCols=(self.lati.get()).strip()
        if len(oldCols)==0:
            combinedCol=newcol
        else:
            combinedCol=oldCols +","+newcol
        self.lati.delete(0,END)
        self.lati.insert(0,combinedCol)

    


    # //add longitude cols
    def add_longitude_cols(self):
        newcol=self.commonCols_combo.get()
        oldCols=(self.longi.get()).strip()
        if len(oldCols)==0:
            combinedCol=newcol
        else:
            combinedCol=oldCols +","+newcol
        self.longi.delete(0,END)
        self.longi.insert(0,combinedCol)

       
        
    # //add timestamp cols
    def add_timestamp_cols(self):
        newcol=self.commonCols_combo.get()
        oldCols=(self.timest.get()).strip()
        if len(oldCols)==0:
            combinedCol=newcol
        else:
            combinedCol=oldCols +","+newcol
        self.timest.delete(0,END)
        self.timest.insert(0,combinedCol)

    

    
    # //add timestamp cols
    def add_skip_cols(self):
        newcol=self.commonCols_combo.get()
        oldCols=(self.skipCols.get()).strip()
        if len(oldCols)==0:
            combinedCol=newcol
        else:
            combinedCol=oldCols +","+newcol
        self.skipCols.delete(0,END)
        self.skipCols.insert(0,combinedCol)

    


    # //add timestamp cols
    def add_skipFullCompare_cols(self):
        newcol=self.commonCols_combo.get()
        oldCols=(self.skipColsDuringFullCompare.get()).strip()
        if len(oldCols)==0:
            combinedCol=newcol
        else:
            combinedCol=oldCols +","+newcol
        self.skipColsDuringFullCompare.delete(0,END)
        self.skipColsDuringFullCompare.insert(0,combinedCol)

    


    def getAllSFtables(self):
        pass



    # add column map

    def add_col_map(self):
        Source_col=str(self.unCommonCol_Source_combo.get()).strip() 
        release_col=str(self.unCommonCol_Release_combo.get()).strip()
        col_map_old=str(self.colsMapping.get()).strip()
        combinedCol=''
        if(col_map_old==''):
            combinedCol=(Source_col+"->"+release_col)
        else:
            combinedCol=col_map_old + ","+Source_col+"->"+release_col
    
        self.colsMapping.delete(0,END)
        self.colsMapping.insert(0,combinedCol)
        

    


    # s   # self.mainloop()

if __name__ == "__main__":
  app = App()
#   app.setDaemon(True)
  
  app.start()
  app.mainloop()  
